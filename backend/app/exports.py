from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pymupdf as fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from .config import Settings
from .models import Answer, Artifact, Job, QuestionnaireGroup, ReviewEvent
from .storage import file_sha256


SHEET_NAMES = [
    "Questionnaires",
    "Long_Answers",
    "Page_Extracts",
    "Conflicts",
    "Failed_Jobs",
    "QA_Summary",
    "Data_Analysis",
    "Run_Log",
    "Reasonableness",
    "Review_Audit",
]


def answer_dict(answer: Answer) -> dict[str, Any]:
    return {
        "id": answer.id,
        "group_id": answer.group_id,
        "page_number": answer.page_number,
        "question_id": answer.question_id,
        "question_text": answer.question_text,
        "section": answer.section,
        "answer_type": answer.answer_type,
        "allowed_options": answer.allowed_options,
        "selected_options": answer.selected_options,
        "qwen_value": answer.qwen_value,
        "yolo_value": answer.yolo_value,
        "scanner_value": answer.scanner_value,
        "scanner_confidence": answer.scanner_confidence,
        "fusion_reason": answer.fusion_reason,
        "evidence": answer.evidence,
        "reasonableness_status": answer.reasonableness_status,
        "judge_suggestion": answer.judge_suggestion,
        "judge_reason": answer.judge_reason,
        "judge_confidence": answer.judge_confidence,
        "rule_refs": answer.rule_refs,
        "final_value": answer.final_value,
        "final_source": answer.final_source,
        "review_status": answer.review_status,
        "reviewer_id": answer.reviewer_id,
        "reviewed_at": answer.reviewed_at.isoformat() if answer.reviewed_at else None,
        "review_comment": answer.review_comment,
    }


def result_payload(db: Session, job: Job) -> dict[str, Any]:
    groups = db.scalars(
        select(QuestionnaireGroup).where(QuestionnaireGroup.job_id == job.id).order_by(QuestionnaireGroup.group_index)
    ).all()
    answers = db.scalars(
        select(Answer).where(Answer.job_id == job.id).order_by(Answer.group_id, Answer.page_number, Answer.question_id)
    ).all()
    return {
        "schema_version": "2.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "job": {
            "id": job.id,
            "filename": job.filename,
            "status": job.status,
            "page_count": job.page_count,
            "profile_snapshot": job.profile_snapshot,
            "created_at": job.created_at.isoformat(),
        },
        "groups": [
            {
                "id": group.id,
                "group_index": group.group_index,
                "start_page": group.start_page,
                "end_page": group.end_page,
                "participant_id": group.participant_id,
            }
            for group in groups
        ],
        "answers": [answer_dict(answer) for answer in answers],
        "unresolved_count": sum(answer.review_status == "pending" for answer in answers),
    }


def _value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _write_rows(sheet, rows: list[dict[str, Any]]) -> None:  # type: ignore[no-untyped-def]
    if not rows:
        sheet.append(["No data"])
        return
    headers = list(rows[0])
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="18312E")
        cell.alignment = Alignment(wrap_text=True)
    for row in rows:
        sheet.append([_value(row.get(header)) for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(55, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_excel(db: Session, job: Job, path: Path, draft: bool) -> None:
    groups = db.scalars(select(QuestionnaireGroup).where(QuestionnaireGroup.job_id == job.id)).all()
    answers = db.scalars(select(Answer).where(Answer.job_id == job.id)).all()
    group_map = {group.id: group for group in groups}
    review_events = db.scalars(
        select(ReviewEvent).join(Answer, ReviewEvent.answer_id == Answer.id).where(Answer.job_id == job.id)
    ).all()
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {name: workbook.create_sheet(name) for name in SHEET_NAMES}
    _write_rows(
        sheets["Questionnaires"],
        [
            {
                "Export_Status": "DRAFT — REVIEW REQUIRED" if draft else "FINAL",
                "Source_PDF": job.filename,
                "Questionnaire_Index": group.group_index,
                "Source_Pages": f"{group.start_page}-{group.end_page}",
                "Participant_ID": group.participant_id or "N/A",
                "Needs_Review": any(a.review_status == "pending" for a in answers if a.group_id == group.id),
            }
            for group in groups
        ],
    )
    long_rows = []
    for answer in answers:
        group = group_map[answer.group_id]
        long_rows.append(
            {
                "Participant_ID": group.participant_id or "N/A",
                "Page": answer.page_number,
                "Question_ID": answer.question_id,
                "Section": answer.section,
                "Question": answer.question_text,
                "Answer_Type": answer.answer_type,
                "Allowed_Options": answer.allowed_options,
                "Selected_Options": answer.selected_options,
                "Qwen_Value": answer.qwen_value,
                "YOLO_Value": answer.yolo_value,
                "Scanner_Value_Immutable": answer.scanner_value,
                "Scanner_Confidence": answer.scanner_confidence,
                "Fusion_Reason": answer.fusion_reason,
                "Qwen_Judge_Status": answer.reasonableness_status,
                "Qwen_Judge_Suggestion": answer.judge_suggestion,
                "Qwen_Judge_Reason": answer.judge_reason,
                "Final_Value": answer.final_value,
                "Final_Source": answer.final_source,
                "Review_Status": answer.review_status,
                "Reviewer": answer.reviewer_id or "",
            }
        )
    _write_rows(sheets["Long_Answers"], long_rows)
    _write_rows(sheets["Page_Extracts"], long_rows)
    _write_rows(
        sheets["Conflicts"],
        [row for row in long_rows if "disagree" in str(row["Fusion_Reason"]).casefold() or row["Review_Status"] == "pending"],
    )
    _write_rows(sheets["Failed_Jobs"], [] if not job.error else [{"Job": job.id, "Error": job.error}])
    _write_rows(
        sheets["QA_Summary"],
        [
            {"Metric": "Answers", "Value": len(answers)},
            {"Metric": "Pending review", "Value": sum(a.review_status == "pending" for a in answers)},
            {"Metric": "Qwen corrections", "Value": sum(a.final_source == "qwen_judge" for a in answers)},
            {"Metric": "Qwen/YOLO agreements", "Value": sum("Qwen and YOLO agree" in a.fusion_reason for a in answers)},
        ],
    )
    _write_rows(sheets["Data_Analysis"], [{"Status": "Analysis-ready rows are available in Long_Answers"}])
    _write_rows(sheets["Run_Log"], [{"Time": job.updated_at.isoformat(), "Stage": job.stage_message, "Status": job.status}])
    _write_rows(
        sheets["Reasonableness"],
        [
            {
                "Question_ID": answer.question_id,
                "Original_Scanner_Value": answer.scanner_value,
                "Status": answer.reasonableness_status,
                "Suggestion": answer.judge_suggestion,
                "Reason": answer.judge_reason,
                "Confidence": answer.judge_confidence,
                "Rule_References": answer.rule_refs,
                "Current_Final": answer.final_value,
            }
            for answer in answers
        ],
    )
    answer_map = {answer.id: answer for answer in answers}
    _write_rows(
        sheets["Review_Audit"],
        [
            {
                "Question_ID": answer_map[event.answer_id].question_id,
                "Action": event.action,
                "Previous_Value": event.previous_value,
                "New_Value": event.new_value,
                "Reviewer": event.reviewer_id,
                "Comment": event.comment,
                "Time": event.created_at.isoformat(),
            }
            for event in review_events
            if event.answer_id in answer_map
        ],
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_annotated_pdf(db: Session, job: Job, path: Path, draft: bool) -> None:
    source = Path(job.stored_path)
    if source.suffix.casefold() == ".pdf":
        document = fitz.open(source)
    else:
        document = fitz.open()
        image_document = fitz.open(source)
        pdf_bytes = image_document.convert_to_pdf()
        image_document.close()
        document.insert_pdf(fitz.open("pdf", pdf_bytes))
    answers = db.scalars(select(Answer).where(Answer.job_id == job.id)).all()
    for page_number in range(1, document.page_count + 1):
        page = document.load_page(page_number - 1)
        if draft:
            page.insert_text((24, 24), "DRAFT — REVIEW MAY BE REQUIRED", fontsize=9, color=(0.65, 0.34, 0.05))
        for answer in [item for item in answers if item.page_number == page_number]:
            for evidence in answer.evidence:
                bbox = evidence.get("bbox")
                if not isinstance(bbox, list) or len(bbox) != 4:
                    continue
                rect = fitz.Rect(
                    bbox[0] * page.rect.width,
                    bbox[1] * page.rect.height,
                    bbox[2] * page.rect.width,
                    bbox[3] * page.rect.height,
                )
                color = (0.11, 0.55, 0.40) if evidence.get("source") == "yolo" else (0.15, 0.35, 0.72)
                page.draw_rect(rect, color=color, width=1.2)
                page.insert_text((rect.x0, max(8, rect.y0 - 2)), answer.question_id[:28], fontsize=6.5, color=color)
    path.parent.mkdir(parents=True, exist_ok=True)
    document.save(path, garbage=3, deflate=True)
    document.close()


def generate_artifacts(db: Session, job: Job, settings: Settings, *, draft: bool) -> list[Artifact]:
    target = settings.artifacts_dir / job.id / ("draft" if draft else "final")
    target.mkdir(parents=True, exist_ok=True)
    stem = Path(job.filename).stem
    paths = {
        "json": target / f"{stem}_{'draft' if draft else 'final'}.json",
        "excel": target / f"{stem}_{'draft' if draft else 'final'}.xlsx",
        "annotated_pdf": target / f"{stem}_{'draft' if draft else 'final'}_annotated.pdf",
    }
    paths["json"].write_text(json.dumps(result_payload(db, job), ensure_ascii=False, indent=2), encoding="utf-8")
    write_excel(db, job, paths["excel"], draft)
    write_annotated_pdf(db, job, paths["annotated_pdf"], draft)
    db.execute(delete(Artifact).where(Artifact.job_id == job.id, Artifact.draft.is_(draft)))
    artifacts: list[Artifact] = []
    for kind, path in paths.items():
        artifact = Artifact(
            job_id=job.id,
            kind=kind,
            draft=draft,
            filename=path.name,
            stored_path=str(path.resolve()),
            sha256=file_sha256(path),
        )
        db.add(artifact)
        artifacts.append(artifact)
    db.commit()
    return artifacts
