# -*- coding: utf-8 -*-
"""
Chinese Questionnaire Full-Page Local VLM Extractor v7
-----------------------------------------------------
For LM Studio local Vision LLM / VLM.

Main idea:
    scanned Chinese questionnaire PDF -> full page image -> local VLM -> JSON -> Excel.

v2 improvements over the earlier generic version:
    - Supports participant IDs like CSA122 as well as A001/B002/C003.
    - Uses page-specific prompts for the 6-page Chinese questionnaire in the sample.
    - Strongly reduces false positives by asking only for SELECTED/CIRCLED answers.
    - Second-pass verification no longer blindly overwrites the first pass.
      It keeps first-pass values when both passes conflict, records conflicts, and can run a
      targeted conflict resolver using the image.
    - Debug page images are optional and OFF by default to save disk space.
    - Debug JSON/raw model outputs are optional and ON by default.
    - Adds a Conflicts sheet to help you audit model disagreements.

v3 improvements:
    - Resume/skip completed questionnaires across multiple PDFs.
    - Checkpoint NDJSON written after every questionnaire.
    - Failed jobs sheet + failed_jobs.jsonl.
    - Save every N questionnaires to reduce Excel I/O.
    - Review-only rerun mode for rows marked needs_review/REVIEW_PID/ERROR.
    - Post-validation: ambiguous scale answers like 0/1/2 are rechecked; if unresolved, a deterministic random allowed value is chosen and flagged.

v4 improvements:
    - AI-only zoom table pass for pages 3-6: crops table areas, sends them to the same local VLM, and merges answers safely.
    - Conflict resolver no longer auto-overwrites scale/table answers. Scale conflicts are rechecked; if still unresolved, a deterministic random allowed value is chosen and flagged.
    - Reduces false conflicts with simple Traditional/Simplified/common text normalization.
    - Adds QA_Summary sheet and review highlighting in Excel.

v5 improvements:
    - No UNCLEAR mode for scale/table answers: if a scale answer is unclear/conflicting, the app asks the VLM one more targeted time; if still unresolved, it chooses a deterministic random allowed value and records it in uncertain_fields.

v6 improvements:
    - Voting/audit strategy for scale answers: full-page, zoom crop, verify, and tiebreak values are treated as candidates.
    - Zoom crop no longer blindly overwrites a valid full-page answer. Disagreements trigger a tiebreak/vote and are recorded.
    - Removes forced-answer output: final answers contain digits or N/A only.
    - N/A is allowed for genuinely blank/unanswered fields; UNCLEAR is not allowed.

Install:
    pip install pymupdf pillow requests openpyxl

Package:
    auto-py-to-exe -> script = this file -> One File -> Window Based

Recommended LM Studio settings:
    Base URL: http://127.0.0.1:1234/v1
    Model: qwen2.5-vl-7b-instruct or the exact API Model Identifier shown in LM Studio
    For accuracy, use image max side 2800-3200 and temperature 0.
"""

import base64
import io
import json
import os
import queue
import re
import threading
import time
import traceback
import unicodedata
import hashlib
import random
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import fitz  # PyMuPDF
import requests
from PIL import Image, ImageDraw, ImageOps
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, filedialog, messagebox


# ============================================================
# Constants
# ============================================================
APP_TITLE = "Chinese Questionnaire Full-Page LM Studio Extractor v7 - No UNCLEAR, N/A OK"
DEFAULT_BASE_URL = "http://127.0.0.1:1234/v1"
DEFAULT_MODEL_ID = "qwen2.5-vl-7b-instruct"
DEFAULT_DPI = 240
DEFAULT_PAGES_PER_QUESTIONNAIRE = 6
DEFAULT_MAX_TOKENS = 4096
DEFAULT_TIMEOUT = 360
DEFAULT_IMAGE_MAX_SIDE = 3000
DEFAULT_SAVE_EVERY_N = 5
DEFAULT_TABLE_ZOOM_PASS = True
DEFAULT_NO_UNCLEAR_FOR_SCALE = True

SHEET_MAIN = "Questionnaires"
SHEET_LONG = "Long_Answers"
SHEET_CONFLICTS = "Conflicts"
SHEET_LOG = "Run_Log"
SHEET_FAILED = "Failed_Jobs"
SHEET_QA = "QA_Summary"

# Accept both the old A/B/C format and this Chinese questionnaire's CSA### format.
PID_RE = re.compile(r"\b(CSA|[ABCabc])\s*[-_ ]?\s*0*(\d{1,4})\b", re.I)
CONTROL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


# ============================================================
# Basic helpers
# ============================================================
def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def safe_mkdir(path: str) -> None:
    if path:
        os.makedirs(path, exist_ok=True)


def clean_text(value: Any) -> str:
    if value is None:
        return "N/A"
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s)
    if not s or s.lower() in {"n/a", "na", "none", "null", "unknown", "unreadable", "blank"}:
        return "N/A"
    return s


def clean_for_excel(value: Any) -> str:
    if value is None:
        return "N/A"
    if isinstance(value, (dict, list)):
        s = json.dumps(value, ensure_ascii=False)
    else:
        s = str(value)
    s = CONTROL_CHARS_RE.sub(" ", s)
    if len(s) > 32000:
        s = s[:32000] + " ...[TRUNCATED]"
    return s


def normalize_pid(value: Any) -> str:
    """Normalize participant ID.

    Examples:
        CSA122 -> CSA122
        A1 -> A001
        B005 -> B005
    """
    if value is None:
        return "N/A"
    s = str(value).strip().upper()
    if not s:
        return "N/A"
    s = s.replace("Ｃ", "C").replace("Ｓ", "S").replace("Ａ", "A")
    s = s.replace("－", "-").replace("_", " ").replace("-", " ")
    m = PID_RE.search(s)
    if not m:
        return "N/A"
    prefix = m.group(1).upper()
    n = int(m.group(2))
    if n == 0:
        return "N/A"
    if prefix == "CSA":
        return f"CSA{n}"
    return f"{prefix}{n:03d}" if n <= 999 else f"{prefix}{n}"


def sanitize_key(value: Any, max_len: int = 90) -> str:
    s = clean_text(value)
    if s == "N/A":
        return "unknown"
    s = re.sub(r"[\s\n\r\t]+", "_", s)
    s = re.sub(r"[^0-9A-Za-z_\u4e00-\u9fff\u3400-\u4dbf]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return (s or "unknown")[:max_len]


def normalize_value_for_compare(value: Any) -> str:
    """Normalize answers for comparing extraction vs verification.

    It intentionally normalizes common Traditional/Simplified variants and
    unordered checkbox/list outputs so harmless differences do not trigger review.
    """
    if value is None:
        return "N/A"
    if isinstance(value, list):
        normed = [normalize_value_for_compare(v) for v in value if normalize_value_for_compare(v) != "N/A"]
        return "|".join(sorted(normed)) if normed else "N/A"
    if isinstance(value, dict):
        # Normalize dict values too, then sort keys for stable comparison.
        norm = {str(k): normalize_value_for_compare(v) for k, v in value.items() if normalize_value_for_compare(v) != "N/A"}
        return json.dumps(norm, ensure_ascii=False, sort_keys=True)
    s = normalize_text_for_compare_string(value)
    if not s or s.lower() in {"n/a", "na", "none", "null", "unknown", "blank", "unreadable"}:
        return "N/A"
    return s

def is_missing(value: Any) -> bool:
    return normalize_value_for_compare(value) in {"N/A", "", "[]", "{}"}


def bool_from_cell(value: Any) -> bool:
    s = str(value).strip().lower()
    return s in {"true", "1", "yes", "y", "是", "需要", "review"}


# Small text-normalization table to reduce false conflicts such as 鍾/锺 or 視/视.
# This is not a full translator; it is only for comparing model outputs.
ZH_COMPARE_MAP = str.maketrans({
    "锺": "鍾", "钟": "鍾",
    "视": "視", "远": "遠", "药": "藥", "医": "醫", "师": "師",
    "问": "問", "题": "題", "选": "選", "项": "項", "请": "請", "岁": "歲",
    "测": "測", "验": "驗", "体": "體", "数": "數", "写": "寫", "质": "質",
    "过": "過", "现": "現", "觉": "覺", "实": "實", "际": "際", "亲": "親",
})

def normalize_text_for_compare_string(s: Any) -> str:
    """Normalize human text for conflict comparison only."""
    if s is None:
        return "N/A"
    out = unicodedata.normalize("NFKC", str(s)).strip()
    out = out.translate(ZH_COMPARE_MAP)
    out = out.replace(" ", "").replace("　", "")
    out = out.replace("，", ",").replace("、", ",").replace("：", ":")
    out = out.replace("（", "(").replace("）", ")")
    # Very common equivalent short forms in this questionnaire.
    equiv = {
        "母親": "母", "媽媽": "母", "媽": "母",
        "父親": "父", "爸爸": "父", "爸": "父",
        "沒有": "否", "無": "否",
        "女生": "女", "女性": "女", "F": "女", "female": "女",
        "男生": "男", "男性": "男", "M": "男", "male": "男",
    }
    return equiv.get(out, out)


def is_scale_answer_key(page_no: int, key: Any) -> bool:
    return str(key) in expected_scale_keys_for_page(page_no)


def append_ndjson(path: str, obj: Dict[str, Any]) -> None:
    safe_mkdir(os.path.dirname(path) or ".")
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


def expected_scale_keys_for_page(page_no: int) -> Dict[str, set]:
    if page_no == 3:
        return {f"P3_Q{i}": set("01234") for i in range(1, 20)}
    if page_no == 4:
        return {f"P4_Q{i}": set("012") for i in range(1, 21)}
    if page_no == 5:
        return {f"P5_Q{i}": set("012") for i in range(1, 21)}
    if page_no == 6:
        out = {f"P6_Q5{ch}": set("0123") for ch in "abcdefghij"}
        out.update({
            "P6_Q6_sleep_medicine": set("0123"),
            "P6_Q7_daytime_sleepiness": set("0123"),
            "P6_Q8_enthusiasm_problem": set("0123"),
            "P6_Q9_sleep_quality": set("0123"),
        })
        return out
    return {}


def normalize_scale_answer(value: Any, allowed: set) -> Tuple[str, bool, str]:
    """Return (cleaned_value, needs_review, reason).

    Final scale answers may be one legal digit or N/A. Literal UNCLEAR is not a
    permitted final value. N/A is treated as a real blank / unanswered value.
    """
    if value is None:
        return "N/A", False, "blank_or_missing_scale_answer"
    s = str(value).strip()
    if not s or s.upper() in {"N/A", "NA", "NULL", "NONE", "BLANK", "空白", "未填"}:
        return "N/A", False, "blank_scale_answer"
    s2 = s.replace("０", "0").replace("１", "1").replace("２", "2").replace("３", "3").replace("４", "4").replace("５", "5")
    s2 = s2.replace(" ", "")
    if s2.upper() == "UNCLEAR":
        return "N/A", True, "model_unclear_converted_to_na"
    # If model returns schema examples or multiple candidates, recheck before choosing.
    if any(sep in s2 for sep in ["/", ",", "，", "、", "|"]):
        return "N/A", True, f"ambiguous_scale_answer:{s}"
    if re.fullmatch(r"\d", s2) and s2 in allowed:
        return s2, False, ""
    return "N/A", True, f"invalid_scale_answer:{s}"



def extract_scale_candidates(value: Any, allowed: set) -> List[str]:
    """Extract possible single-digit candidates from a model value.

    For example, "0/1/2" returns ["0", "1", "2"]. This is only used as a
    last-resort candidate pool when the user has requested no UNCLEAR values.
    """
    if value is None:
        return []
    s = str(value).strip()
    s = s.replace("０", "0").replace("１", "1").replace("２", "2").replace("３", "3").replace("４", "4").replace("５", "5")
    digits = re.findall(r"\d", s)
    out = []
    for d in digits:
        if d in allowed and d not in out:
            out.append(d)
    return out


def stable_random_choice(options: List[str], seed_text: str) -> str:
    """Deterministic random choice so reruns are reproducible for the same file/key."""
    opts = sorted([str(x) for x in options if str(x)])
    if not opts:
        return "0"
    h = hashlib.sha256(seed_text.encode("utf-8", errors="ignore")).hexdigest()
    idx = int(h[:12], 16) % len(opts)
    return opts[idx]


def choose_scale_answer_no_unclear_na_ok(
    key: str,
    page_no: int,
    allowed: set,
    candidate_values: List[Any],
    seed_text: str,
) -> Tuple[str, str]:
    """Return (chosen_value, method).

    The returned value is never UNCLEAR. It is either one legal digit or N/A.
    Priority:
      1. Majority among valid single digits from extraction/verify/zoom/tiebreak.
      2. A single valid digit if only one exists.
      3. Deterministic random among observed candidates only.
      4. N/A if no observed digit candidate exists, interpreted as blank/unanswered.
    """
    valid_votes: List[str] = []
    observed_candidates: List[str] = []
    saw_explicit_na = False
    for v in candidate_values:
        clean, review, _ = normalize_scale_answer(v, allowed)
        if clean == "N/A" and not review:
            saw_explicit_na = True
        if not review and clean in allowed:
            valid_votes.append(clean)
            if clean not in observed_candidates:
                observed_candidates.append(clean)
        for c in extract_scale_candidates(v, allowed):
            if c not in observed_candidates:
                observed_candidates.append(c)
    if valid_votes:
        counts = {d: valid_votes.count(d) for d in sorted(set(valid_votes))}
        best_count = max(counts.values())
        best = [d for d, c in counts.items() if c == best_count]
        if len(best) == 1 and best_count >= 2:
            return best[0], "majority_vote"
        if len(best) == 1 and len(valid_votes) == 1 and not observed_candidates:
            return best[0], "single_valid_vote"
    if observed_candidates:
        return stable_random_choice(observed_candidates, f"{seed_text}|{key}|{','.join(observed_candidates)}"), "deterministic_random_observed"
    return "N/A", "blank_no_candidate" if saw_explicit_na else "na_no_candidate"


def scale_tiebreak_prompt(page_no: int, keys: List[str]) -> str:
    allowed = "0-4" if page_no == 3 else ("0-2" if page_no in {4, 5} else "0-3")
    return f"""
你正在作最後一次覆核。只看圖中指定題目被圈起來/勾選的答案。
非常重要：UNCLEAR 不允許。每一題只能輸出一個單一數字 {allowed}，或者在完全沒有圈選/空白時輸出 N/A。
如果兩個格都像被圈，請先選更像真正答案的一個；如果仍然不確定，也要在看得到的候選數字中選一個最可能的數字。
不要輸出 0/1/2、0,2、UNCLEAR、空白或解釋。

只輸出 JSON：
{{
  "answers": {{
    {", ".join([f'"{k}": "0"' for k in keys])}
  }}
}}
""".strip()


def replace_unclear_text_with_na(obj: Any, skip_keys: Optional[set] = None) -> Any:
    """Replace literal UNCLEAR in non-scale/text fields with N/A.

    Scale answers are handled by force_no_unclear_scale_answers; this function is
    just to keep final Excel free of the literal word UNCLEAR in other text fields.
    """
    skip_keys = skip_keys or set()
    if isinstance(obj, dict):
        out = {}
        for k, v in obj.items():
            if str(k) in skip_keys:
                out[k] = v
            else:
                out[k] = replace_unclear_text_with_na(v, skip_keys)
        return out
    if isinstance(obj, list):
        return [replace_unclear_text_with_na(x, skip_keys) for x in obj]
    if isinstance(obj, str) and obj.strip().upper() == "UNCLEAR":
        return "N/A"
    return obj

def postprocess_page_json(page_json: Dict[str, Any], page_no: int) -> Dict[str, Any]:
    """Safety pass after local VLM extraction. Does not use CV; only validates schema values."""
    if not isinstance(page_json, dict):
        return page_json
    answers = page_json.get("answers")
    if not isinstance(answers, dict):
        page_json["answers"] = {}
        page_json["needs_review"] = True
        page_json["uncertain_fields"] = ["answers_not_dict"]
        return page_json

    uncertain = page_json.get("uncertain_fields") or []
    if isinstance(uncertain, str):
        uncertain = [uncertain]
    if not isinstance(uncertain, list):
        uncertain = []

    expected = expected_scale_keys_for_page(page_no)
    for key, allowed in expected.items():
        cleaned, review, reason = normalize_scale_answer(answers.get(key), allowed)
        if review:
            if key in answers and str(answers.get(key)).strip().upper() != "UNCLEAR":
                answers[f"{key}_raw_model_value"] = answers.get(key)
            uncertain.append(f"{key}:{reason}")
            page_json["needs_review"] = True
        answers[key] = cleaned

    # Normalize known identity fields.
    if page_no == 1:
        ident = page_json.get("identity") or {}
        if isinstance(ident, dict):
            g = clean_text(ident.get("gender"))
            if g in {"男", "M", "male", "Male"}:
                ident["gender"] = "男"
            elif g in {"女", "F", "female", "Female"}:
                ident["gender"] = "女"
            elif g not in {"N/A", "UNCLEAR"}:
                ident["gender"] = "N/A"
                uncertain.append("identity.gender:ambiguous_gender_forced_NA")
                page_json["needs_review"] = True
            page_json["identity"] = ident

    page_json["answers"] = answers
    page_json["uncertain_fields"] = list(dict.fromkeys(map(str, uncertain)))
    # Keep final output free of literal UNCLEAR for non-scale/text fields.
    # Scale/table fields may still be UNCLEAR here, but v5 force_no_unclear_scale_answers
    # will replace them before final writing when enabled.
    return replace_unclear_text_with_na(page_json, skip_keys=set(expected.keys()))


def image_to_b64(img: Image.Image, max_side: int, fmt: str = "PNG") -> Tuple[str, str]:
    img = resize_keep_aspect(img, max_side)
    if img.mode != "RGB":
        img = img.convert("RGB")
    buf = io.BytesIO()
    if fmt.upper() == "JPEG":
        img.save(buf, format="JPEG", quality=92, optimize=True)
        mime = "image/jpeg"
    else:
        img.save(buf, format="PNG", optimize=True)
        mime = "image/png"
    return base64.b64encode(buf.getvalue()).decode("utf-8"), mime


def resize_keep_aspect(img: Image.Image, max_side: int) -> Image.Image:
    if max_side <= 0:
        return img
    w, h = img.size
    if max(w, h) <= max_side:
        return img
    scale = max_side / float(max(w, h))
    new_size = (max(1, int(w * scale)), max(1, int(h * scale)))
    return img.resize(new_size, Image.Resampling.LANCZOS)


def enhance_page_image(img: Image.Image) -> Image.Image:
    """Light preprocessing only. This is NOT answer extraction, just improving visibility."""
    img = ImageOps.exif_transpose(img)
    gray = ImageOps.grayscale(img)
    gray = ImageOps.autocontrast(gray, cutoff=1)
    return gray.convert("RGB")


def render_pdf_page(doc: fitz.Document, page_index0: int, dpi: int, enhance: bool) -> Image.Image:
    page = doc.load_page(page_index0)
    mat = fitz.Matrix(dpi / 72.0, dpi / 72.0)
    pix = page.get_pixmap(matrix=mat, alpha=False)
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    img = ImageOps.exif_transpose(img)
    if enhance:
        img = enhance_page_image(img)
    return img


def rotate_image(img: Image.Image, degrees: int) -> Image.Image:
    degrees = degrees % 360
    if degrees == 0:
        return img
    return img.rotate(degrees, expand=True)


def crop_norm(img: Image.Image, box: Tuple[float, float, float, float], upscale_factor: float = 1.8) -> Image.Image:
    """Crop by normalized coordinates and enlarge for VLM reading.

    This does not detect answers; it only gives the local VLM a larger view of
    dense tables. It keeps the workflow AI-only for answer extraction.
    """
    w, h = img.size
    x1 = max(0, min(w - 1, int(box[0] * w)))
    y1 = max(0, min(h - 1, int(box[1] * h)))
    x2 = max(x1 + 1, min(w, int(box[2] * w)))
    y2 = max(y1 + 1, min(h, int(box[3] * h)))
    out = img.crop((x1, y1, x2, y2))
    if upscale_factor and upscale_factor > 1.0:
        ow, oh = out.size
        out = out.resize((max(1, int(ow * upscale_factor)), max(1, int(oh * upscale_factor))), Image.Resampling.LANCZOS)
    return out


# Normalized crop boxes tuned for the 6-page sample layout.
# The boxes are deliberately a bit wide to tolerate scan shifts.
TABLE_CROP_SPECS: Dict[int, List[Tuple[str, Tuple[float, float, float, float], List[str]]]] = {
    3: [("COVID_QOL_table", (0.09, 0.12, 0.86, 0.70), [f"P3_Q{i}" for i in range(1, 20)])],
    4: [
        ("PedEyeQ_page4_upper_Q1_Q10", (0.08, 0.23, 0.88, 0.58), [f"P4_Q{i}" for i in range(1, 11)]),
        ("PedEyeQ_page4_lower_Q11_Q20", (0.08, 0.56, 0.88, 0.90), [f"P4_Q{i}" for i in range(11, 21)]),
    ],
    5: [
        ("PedEyeQ_page5_upper_Q1_Q10", (0.08, 0.13, 0.88, 0.51), [f"P5_Q{i}" for i in range(1, 11)]),
        ("PedEyeQ_page5_lower_Q11_Q20", (0.08, 0.46, 0.88, 0.82), [f"P5_Q{i}" for i in range(11, 21)]),
    ],
    6: [("PSQI_scale_table_Q5_Q9", (0.04, 0.23, 0.94, 0.92), [f"P6_Q5{ch}" for ch in "abcdefghij"] + ["P6_Q6_sleep_medicine", "P6_Q7_daytime_sleepiness", "P6_Q8_enthusiasm_problem", "P6_Q9_sleep_quality"])],
}


# ============================================================
# JSON parsing and prompts
# ============================================================
def extract_json_from_text(text: str) -> Dict[str, Any]:
    raw = (text or "").strip()
    raw = re.sub(r"^\s*```(?:json)?\s*", "", raw, flags=re.I)
    raw = re.sub(r"\s*```\s*$", "", raw)
    start = raw.find("{")
    end = raw.rfind("}")
    if start >= 0 and end > start:
        raw = raw[start:end + 1]
    raw = raw.replace("\u201c", '"').replace("\u201d", '"').replace("\u2018", "'").replace("\u2019", "'")
    raw = re.sub(r",\s*([}\]])", r"\1", raw)
    try:
        obj = json.loads(raw)
        if isinstance(obj, dict):
            return obj
        return {"value": obj}
    except Exception as e:
        raise ValueError(f"Invalid JSON from model. First 900 chars:\n{text[:900] if text else ''}") from e


def strict_json_rule() -> str:
    return (
        "只輸出 valid JSON。不要 markdown，不要 ```json，不要解釋。"
        "所有 key 必須用英文穩定 key；看不到/空白用 N/A；一般文字看不清用 N/A；量表/圈選數字題不要用 UNCLEAR，必須選最可能的一個數字。"
    )


def page_prompt(page_no: int, pages_per_q: int, verify_mode: bool = False) -> str:
    mode = "覆核模式：請獨立重新讀圖，不要假設第一次答案是對的。" if verify_mode else "提取模式：請直接讀圖。"
    common = f"""
你是非常仔細的中文問卷資料輸入員。這是同一份問卷的第 {page_no}/{pages_per_q} 頁。
{mode}
重點：只提取填寫者真正填上的答案，包括剔號、圈選、手寫圈、打叉、手寫文字。
不要把印刷的選項清單全部當成答案。沒有被選中的選項不要輸出。
{strict_json_rule()}
""".strip()

    if page_no == 1:
        return common + r'''

這頁通常是「個人資料 + 一般健康/眼睛健康/家族眼疾」頁。
請特別小心：很多方格只是印刷選項，只有被剔/圈/打叉的才算答案。
性別只可填「男」「女」「N/A」，不要填「男/女」或「UNCLEAR」。
右上角若有 CSA122 這類手寫 ID，participant_id 要填 CSA122。
家族眼疾選項下面的手寫「母/父/兄弟姊妹」是 relation，不是 parent_guardian_name。

輸出 JSON schema：
{
  "page_type": "basic_info_health_family",
  "participant_id": "CSA122 或 A001/B001/C001 或 N/A",
  "identity": {
    "service_school": "N/A",
    "service_date": "N/A",
    "name": "N/A",
    "date_of_birth": "N/A",
    "gender": "男/女/N/A",
    "check_number": "N/A"
  },
  "answers": {
    "P1_residence_district": [],
    "P1_Q1_general_health_selected": [],
    "P1_Q1_general_health_other_text": "N/A",
    "P1_Q2_longterm_medicine": "有/否/N/A",
    "P1_Q2_medicine_text": "N/A",
    "P1_Q3_eye_health_selected": [],
    "P1_Q3_eye_health_other_text": "N/A",
    "P1_Q4_family_eye_health_selected": [
      {"option": "例如 深近視（六百度或以上）", "relation": "例如 母/父/N/A"}
    ]
  },
  "visible_handwriting": [],
  "needs_review": false,
  "uncertain_fields": []
}
'''.strip()

    if page_no == 2:
        return common + r'''

這頁通常是健康症狀、藥物敏感、眼部手術、眼睛檢查、眼鏡/OK鏡、眼藥、每日活動時間。
只輸出被選中的選項。若「其他」旁邊有手寫，例如「沒有」，請保留在 text。

輸出 JSON schema：
{
  "page_type": "health_symptoms_devices",
  "participant_id": "N/A",
  "identity": {},
  "answers": {
    "P2_Q5_recent_symptoms_selected": [{"option": "其他/眼乾澀/...", "text": "N/A"}],
    "P2_Q6_drug_allergy": "有/否/N/A",
    "P2_Q6_drug_allergy_text": "N/A",
    "P2_Q7_eye_injury_surgery": "有/否/N/A",
    "P2_Q7_eye_injury_surgery_text": "N/A",
    "P2_Q8_eye_exam": "有/否/N/A",
    "P2_Q8_eye_exam_date": "N/A",
    "P2_Q8_dilated_pupil": "有/否/N/A",
    "P2_Q9_glasses_or_contacts_habit": "N/A",
    "P2_Q10_eye_medicine_use": "有/沒有使用任何眼部藥物/N/A",
    "P2_Q10_eye_medicine_type": [],
    "P2_Q10_eye_medicine_name": "N/A",
    "P2_Q10_eye_medicine_source": [],
    "P2_Q11_tv_hours": "N/A",
    "P2_Q11_phone_tablet_hours": "N/A",
    "P2_Q11_reading_hours": "N/A"
  },
  "visible_handwriting": [],
  "needs_review": false,
  "uncertain_fields": []
}
'''.strip()

    if page_no == 3:
        return common + r'''

這頁是 COVID-QOL 19 行量表。每一行有 0/1/2/3/4 五個數字，請讀出被圈起來/圈住的數字。
重要：如果圈在最左邊 0，就輸出 "0"。不要因為很多行相似而套用上一行答案。
必須輸出 P3_Q1 到 P3_Q19，共 19 個 key。若該題完全沒有圈選/空白，輸出 N/A；若有多個圈或不清楚但有痕跡，必須選最可能的一個數字，不可寫 UNCLEAR。

輸出 JSON schema：
{
  "page_type": "COVID_QOL",
  "participant_id": "CSA122 或 N/A",
  "identity": {},
  "answers": {
    "P3_Q1": "0",
    "P3_Q2": "0",
    "P3_Q3": "0",
    "P3_Q4": "0",
    "P3_Q5": "0",
    "P3_Q6": "0",
    "P3_Q7": "0",
    "P3_Q8": "0",
    "P3_Q9": "0",
    "P3_Q10": "0",
    "P3_Q11": "0",
    "P3_Q12": "0",
    "P3_Q13": "0",
    "P3_Q14": "0",
    "P3_Q15": "0",
    "P3_Q16": "0",
    "P3_Q17": "0",
    "P3_Q18": "0",
    "P3_Q19": "0"
  },
  "visible_handwriting": [],
  "needs_review": false,
  "uncertain_fields": []
}
'''.strip()

    if page_no == 4:
        return common + r'''

這頁是 PedEyeQ 12-17 years old，第 1-20 題。每行有 0/1/2 三個選項，欄名通常是 從不/偶爾/經常。
上半部分 1-10 題 = P4_Q1 到 P4_Q10；下半部分 1-10 題 = P4_Q11 到 P4_Q20。
請只讀每一行被圈選的數字。每行正常只能有一個答案。
如果完全沒有圈選/空白，輸出 N/A；如果看不清楚或有多個可能但有痕跡，仍然必須選最可能的一個數字。不可輸出 UNCLEAR、0/1/2 或 0,2。
嚴禁輸出 "0/1/2" 這種 schema 範例字串作為答案。
必須輸出 P4_Q1 到 P4_Q20，共 20 個 key。

輸出 JSON schema：
{
  "page_type": "PedEyeQ_page4",
  "participant_id": "CSA122 或 N/A",
  "identity": {},
  "answers": {
    "P4_Q1": "0", "P4_Q2": "0", "P4_Q3": "0", "P4_Q4": "0", "P4_Q5": "0",
    "P4_Q6": "0", "P4_Q7": "0", "P4_Q8": "0", "P4_Q9": "0", "P4_Q10": "0",
    "P4_Q11": "0", "P4_Q12": "0", "P4_Q13": "0", "P4_Q14": "0", "P4_Q15": "0",
    "P4_Q16": "0", "P4_Q17": "0", "P4_Q18": "0", "P4_Q19": "0", "P4_Q20": "0"
  },
  "visible_handwriting": [],
  "needs_review": false,
  "uncertain_fields": []
}
'''.strip()

    if page_no == 5:
        return common + r'''

這頁是 PedEyeQ 表格，第 21-40 題，可用 P5_Q1 到 P5_Q20 表示本頁 20 行。
每行有 0/1/2 三個選項，欄名通常是 從不/偶爾/經常。
只讀每一行被圈選的數字。每行正常只能有一個答案。
如果完全沒有圈選/空白，輸出 N/A；如果看不清楚或有多個可能但有痕跡，仍然必須選最可能的一個數字。不可輸出 UNCLEAR、0/1/2 或 0,2。
嚴禁輸出 "0/1/2" 這種 schema 範例字串作為答案。
必須輸出 P5_Q1 到 P5_Q20，共 20 個 key。

輸出 JSON schema：
{
  "page_type": "PedEyeQ_page5",
  "participant_id": "CSA122 或 N/A",
  "identity": {},
  "answers": {
    "P5_Q1": "0", "P5_Q2": "0", "P5_Q3": "0", "P5_Q4": "0", "P5_Q5": "0",
    "P5_Q6": "0", "P5_Q7": "0", "P5_Q8": "0", "P5_Q9": "0", "P5_Q10": "0",
    "P5_Q11": "0", "P5_Q12": "0", "P5_Q13": "0", "P5_Q14": "0", "P5_Q15": "0",
    "P5_Q16": "0", "P5_Q17": "0", "P5_Q18": "0", "P5_Q19": "0", "P5_Q20": "0"
  },
  "visible_handwriting": [],
  "needs_review": false,
  "uncertain_fields": []
}
'''.strip()

    if page_no == 6:
        return common + r'''

這頁是 PSQI 睡眠品質表。
上方 Q1-Q4 是手寫時間/分鐘/小時。
中間 Q5a-Q5j 是 0/1/2/3 圈選表格；Q6-Q7 是 0/1/2/3；Q8 是 0/1/2/3；Q9 是 0/1/2/3。
請特別避免把相鄰欄的圓圈誤讀成多選；如果只有 0 被圈，就只輸出 "0"，不要輸出 "0,2"。
如果完全沒有圈選/空白，輸出 N/A；如果不能肯定唯一答案但有痕跡，仍然必須選最可能的一個數字。不可輸出 UNCLEAR、0/1/2 或 0,2。
嚴禁輸出 "0/1/2/3" 這種 schema 範例字串作為答案。

輸出 JSON schema：
{
  "page_type": "PSQI",
  "participant_id": "CSA122 或 N/A",
  "identity": {},
  "answers": {
    "P6_Q1_bedtime": "N/A",
    "P6_Q2_minutes_to_sleep": "N/A",
    "P6_Q3_wakeup_time": "N/A",
    "P6_Q4_actual_sleep_hours": "N/A",
    "P6_Q5a": "0",
    "P6_Q5b": "0",
    "P6_Q5c": "0",
    "P6_Q5d": "0",
    "P6_Q5e": "0",
    "P6_Q5f": "0",
    "P6_Q5g": "0",
    "P6_Q5h": "0",
    "P6_Q5i": "0",
    "P6_Q5j": "0",
    "P6_Q5j_text": "N/A",
    "P6_Q6_sleep_medicine": "0",
    "P6_Q7_daytime_sleepiness": "0",
    "P6_Q8_enthusiasm_problem": "0",
    "P6_Q9_sleep_quality": "0"
  },
  "visible_handwriting": [],
  "needs_review": false,
  "uncertain_fields": []
}
'''.strip()

    # Fallback for non-standard page counts.
    return common + r'''

這頁不是標準 1-6 頁之一。請完整提取所有可見填寫答案。
輸出 JSON schema：
{
  "page_type": "other",
  "participant_id": "CSA122 或 A001/B001/C001 或 N/A",
  "identity": {},
  "answers": {},
  "visible_handwriting": [],
  "needs_review": false,
  "uncertain_fields": []
}
'''.strip()


def orientation_prompt() -> str:
    return (
        "You will receive four versions of the same scanned questionnaire page: original, 90, 180, 270 degree rotations. "
        "Pick the image where Chinese/English text is upright and easiest to read. "
        "Return ONLY JSON: {\"best_image\":1, \"rotation_degrees\":0, \"confidence\":0.0, \"reason\":\"...\"}. "
        "rotation_degrees means the rotation applied to original image and must be 0,90,180,270."
    )


def vision_test_prompt() -> str:
    return "Read the image. Return ONLY JSON with keys text and ok. ok must be true if you can see VISION TEST 123."


def conflict_resolution_prompt(page_no: int, conflicts: List[Dict[str, Any]]) -> str:
    compact = json.dumps(conflicts, ensure_ascii=False, indent=2)
    if len(compact) > 7000:
        compact = compact[:7000] + "\n...[TRUNCATED]"
    return f"""
你正在解決中文問卷第 {page_no} 頁的答案衝突。請只根據圖片判斷，不要猜。
下面是有衝突的欄位及兩次 local VLM 結果。請逐項重新看圖片，選出真正被填寫/圈選/剔選的答案。
如果完全沒有圈選/空白，輸出 N/A；如果看不清楚但有痕跡，仍然必須選最可能的一個答案，不可輸出 UNCLEAR。

衝突欄位：
{compact}

請只輸出 JSON，schema：
{{
  "resolved_answers": {{
    "欄位key": "修正答案"
  }},
  "still_uncertain": ["欄位key"]
}}
""".strip()


def table_zoom_prompt(page_no: int, crop_name: str, keys: List[str]) -> str:
    """Prompt for AI-only enlarged table crop extraction."""
    allowed = "0,1,2,3,4" if page_no == 3 else ("0,1,2,3" if page_no == 6 else "0,1,2")
    keys_txt = ", ".join([f'"{k}"' for k in keys])
    return f"""
你正在讀取一個中文問卷表格的放大裁剪圖，crop 名稱：{crop_name}。
這張圖只用來讀取圈選/勾選的量表答案。請不要解釋，不要輸出題目全文。

規則：
1. 每一題只能輸出一個單一數字：{allowed}，或者在完全沒有圈選/空白時輸出 N/A，絕不可輸出 UNCLEAR。
2. 如果一行不清楚、圈線壓住兩格、或你不肯定，仍然必須選最可能的一個數字。
3. 嚴禁輸出 "0/1/2"、"0,2"、"0或1" 這種多候選格式。
4. 不要猜；只根據圖片中真正被圈住/勾選的數字。
5. 必須輸出以下 keys：{keys_txt}

只輸出 valid JSON：
{{
  "answers": {{
    {", ".join([f'"{k}": "0"' for k in keys])}
  }},
  "needs_review": false,
  "uncertain_fields": []
}}
""".strip()


# ============================================================
# LM Studio client
# ============================================================
class LMStudioClient:
    def __init__(self, base_url: str, model: str, api_key: str = "", timeout_s: int = DEFAULT_TIMEOUT):
        b = (base_url or DEFAULT_BASE_URL).strip().rstrip("/")
        self.base_url = b if b.endswith("/v1") else b + "/v1"
        self.model = (model or "").strip()
        self.api_key = (api_key or "").strip()
        self.timeout_s = max(20, int(timeout_s))

    def _headers(self) -> Dict[str, str]:
        h = {"Content-Type": "application/json"}
        if self.api_key:
            h["Authorization"] = f"Bearer {self.api_key}"
        return h

    def list_models(self) -> List[str]:
        r = requests.get(f"{self.base_url}/models", headers=self._headers(), timeout=20)
        if r.status_code != 200:
            raise RuntimeError(f"HTTP {r.status_code}: {r.text[:500]}")
        data = r.json()
        return [str(x.get("id")) for x in data.get("data", []) if isinstance(x, dict) and x.get("id")]

    def chat_json(
        self,
        prompt: str,
        images: List[Tuple[str, Image.Image]],
        max_tokens: int,
        retries: int = 1,
        image_max_side: int = DEFAULT_IMAGE_MAX_SIDE,
        temperature: float = 0.0,
    ) -> Tuple[Dict[str, Any], str]:
        if not self.model:
            raise RuntimeError("Model ID is empty. Use the exact LM Studio API Model Identifier.")
        content: List[Dict[str, Any]] = [{"type": "text", "text": prompt}]
        for label, img in images:
            if label:
                content.append({"type": "text", "text": label})
            b64, mime = image_to_b64(img, image_max_side, fmt="PNG")
            content.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}})

        payload = {
            "model": self.model,
            "messages": [{"role": "user", "content": content}],
            "temperature": float(temperature),
            "max_tokens": int(max_tokens),
        }
        last_err: Optional[Exception] = None
        for attempt in range(retries + 1):
            try:
                r = requests.post(
                    f"{self.base_url}/chat/completions",
                    headers=self._headers(),
                    json=payload,
                    timeout=self.timeout_s,
                )
                ctype = (r.headers.get("Content-Type") or "").lower()
                if "json" not in ctype:
                    raise RuntimeError(f"Non-JSON HTTP response {r.status_code}: {r.text[:700]}")
                data = r.json()
                if r.status_code >= 400:
                    raise RuntimeError(f"HTTP {r.status_code}: {json.dumps(data, ensure_ascii=False)[:700]}")
                choices = data.get("choices") or []
                if not choices:
                    raise RuntimeError(f"No choices in response: {json.dumps(data, ensure_ascii=False)[:700]}")
                msg = choices[0].get("message", {})
                out = msg.get("content", "")
                if isinstance(out, list):
                    raw = "\n".join(str(x.get("text", "")) for x in out if isinstance(x, dict))
                else:
                    raw = str(out)
                return extract_json_from_text(raw), raw
            except Exception as e:
                last_err = e
                if attempt < retries:
                    # Strengthen JSON-only instruction for retry.
                    payload["messages"][0]["content"][0]["text"] = (
                        prompt + "\n\nIMPORTANT: Previous attempt failed JSON parsing. Output one complete valid JSON object only."
                    )
                    time.sleep(0.5)
        raise RuntimeError(str(last_err))


# ============================================================
# Excel writer
# ============================================================
class DynamicExcelWriter:
    def __init__(self, path: str):
        self.path = path
        safe_mkdir(os.path.dirname(path) or ".")
        self.wb = self._load_or_create(path)
        self.ws_main = self._ensure_sheet(SHEET_MAIN)
        self.ws_long = self._ensure_sheet(SHEET_LONG)
        self.ws_conflicts = self._ensure_sheet(SHEET_CONFLICTS)
        self.ws_log = self._ensure_sheet(SHEET_LOG)
        self.ws_failed = self._ensure_sheet(SHEET_FAILED)
        self.ws_qa = self._ensure_sheet(SHEET_QA)
        self.main_headers = self._ensure_headers(self.ws_main, [
            "participant_id", "source_pdf", "questionnaire_index_in_pdf", "source_pages", "pages_per_questionnaire",
            "extraction_time", "needs_review", "uncertain_fields", "conflict_count", "raw_json_file",
        ])
        self.long_headers = self._ensure_headers(self.ws_long, [
            "participant_id", "source_pdf", "questionnaire_index_in_pdf", "page_no_in_questionnaire", "key",
            "selected_value", "source", "needs_review",
        ])
        self.conflict_headers = self._ensure_headers(self.ws_conflicts, [
            "participant_id", "source_pdf", "questionnaire_index_in_pdf", "page_no_in_questionnaire", "key",
            "first_value", "verify_value", "chosen_value", "resolver_value", "status",
        ])
        self.log_headers = self._ensure_headers(self.ws_log, ["time", "level", "message"])
        self.failed_headers = self._ensure_headers(self.ws_failed, [
            "time", "source_pdf", "questionnaire_index_in_pdf", "source_pages", "error", "traceback", "status"
        ])
        self.qa_headers = self._ensure_headers(self.ws_qa, ["metric", "value"])
        self._style_all()
        self.save()

    def _load_or_create(self, path: str):
        if os.path.exists(path):
            return load_workbook(path)
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_MAIN
        return wb

    def _ensure_sheet(self, name: str):
        if name in self.wb.sheetnames:
            return self.wb[name]
        return self.wb.create_sheet(name)

    def _ensure_headers(self, ws, required: List[str]) -> List[str]:
        if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
            for c, h in enumerate(required, 1):
                ws.cell(1, c, h)
            return list(required)
        headers = [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]
        for h in required:
            if h not in headers:
                headers.append(h)
                ws.cell(1, len(headers), h)
        return headers

    def _append_dynamic(self, ws, headers: List[str], row: Dict[str, Any]) -> List[str]:
        for k in row.keys():
            if k not in headers:
                headers.append(k)
                ws.cell(1, len(headers), k)
                self._style_header_cell(ws.cell(1, len(headers)))
        r = ws.max_row + 1
        for c, h in enumerate(headers, 1):
            if h in row:
                ws.cell(r, c, clean_for_excel(row[h]))
        return headers

    def append_main(self, row: Dict[str, Any]) -> None:
        self.main_headers = self._append_dynamic(self.ws_main, self.main_headers, row)

    def append_long(self, rows: List[Dict[str, Any]]) -> None:
        for row in rows:
            self.long_headers = self._append_dynamic(self.ws_long, self.long_headers, row)

    def append_conflicts(self, rows: List[Dict[str, Any]]) -> None:
        for row in rows:
            self.conflict_headers = self._append_dynamic(self.ws_conflicts, self.conflict_headers, row)

    def append_failed(self, row: Dict[str, Any]) -> None:
        self.failed_headers = self._append_dynamic(self.ws_failed, self.failed_headers, row)

    def _header_map(self, ws) -> Dict[str, int]:
        return {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

    def _key_from_row(self, ws, r: int) -> Optional[Tuple[str, int]]:
        h = self._header_map(ws)
        if "source_pdf" not in h or "questionnaire_index_in_pdf" not in h:
            return None
        src = ws.cell(r, h["source_pdf"]).value
        q = ws.cell(r, h["questionnaire_index_in_pdf"]).value
        if src is None or q is None:
            return None
        try:
            qn = int(float(str(q)))
        except Exception:
            return None
        return (str(src), qn)

    def get_completed_keys(self) -> set:
        """Successful rows: not review and not ERROR/REVIEW_PID."""
        h = self._header_map(self.ws_main)
        out = set()
        for r in range(2, self.ws_main.max_row + 1):
            key = self._key_from_row(self.ws_main, r)
            if not key:
                continue
            pid = str(self.ws_main.cell(r, h.get("participant_id", 1)).value or "")
            needs = bool_from_cell(self.ws_main.cell(r, h.get("needs_review", 1)).value) if "needs_review" in h else False
            if (not needs) and pid not in {"ERROR", "REVIEW_PID", ""}:
                out.add(key)
        return out

    def get_review_keys(self) -> set:
        """Rows needing rerun: needs_review TRUE or ERROR/REVIEW_PID."""
        h = self._header_map(self.ws_main)
        out = set()
        for r in range(2, self.ws_main.max_row + 1):
            key = self._key_from_row(self.ws_main, r)
            if not key:
                continue
            pid = str(self.ws_main.cell(r, h.get("participant_id", 1)).value or "")
            needs = bool_from_cell(self.ws_main.cell(r, h.get("needs_review", 1)).value) if "needs_review" in h else False
            if needs or pid in {"ERROR", "REVIEW_PID", ""}:
                out.add(key)
        return out

    def _delete_rows_for_key(self, ws, source_pdf: str, qidx: int) -> None:
        for r in range(ws.max_row, 1, -1):
            key = self._key_from_row(ws, r)
            if key == (source_pdf, int(qidx)):
                ws.delete_rows(r, 1)

    def replace_long_for_key(self, source_pdf: str, qidx: int, rows: List[Dict[str, Any]]) -> None:
        self._delete_rows_for_key(self.ws_long, source_pdf, qidx)
        self.append_long(rows)

    def replace_conflicts_for_key(self, source_pdf: str, qidx: int, rows: List[Dict[str, Any]]) -> None:
        self._delete_rows_for_key(self.ws_conflicts, source_pdf, qidx)
        self.append_conflicts(rows)

    def upsert_main(self, row: Dict[str, Any]) -> None:
        source_pdf = str(row.get("source_pdf", ""))
        try:
            qidx = int(row.get("questionnaire_index_in_pdf", 0))
        except Exception:
            qidx = 0
        target_r = None
        for r in range(2, self.ws_main.max_row + 1):
            if self._key_from_row(self.ws_main, r) == (source_pdf, qidx):
                target_r = r
                break
        for k in row.keys():
            if k not in self.main_headers:
                self.main_headers.append(k)
                self.ws_main.cell(1, len(self.main_headers), k)
                self._style_header_cell(self.ws_main.cell(1, len(self.main_headers)))
        if target_r is None:
            self.append_main(row)
        else:
            for c, h in enumerate(self.main_headers, 1):
                if h in row:
                    self.ws_main.cell(target_r, c, clean_for_excel(row[h]))

    def append_log(self, message: str, level: str = "INFO") -> None:
        self.log_headers = self._append_dynamic(self.ws_log, self.log_headers, {
            "time": now_str(), "level": level, "message": message
        })

    def _style_header_cell(self, cell) -> None:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _style_all(self) -> None:
        for ws in [self.ws_main, self.ws_long, self.ws_conflicts, self.ws_log, self.ws_failed, self.ws_qa]:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                self._style_header_cell(cell)
            # Reasonable default widths.
            for idx in range(1, min(ws.max_column, 50) + 1):
                ws.column_dimensions[get_column_letter(idx)].width = 18

    def _cell_bool_by_header(self, ws, row_idx: int, header: str) -> bool:
        h = self._header_map(ws)
        if header not in h:
            return False
        return bool_from_cell(ws.cell(row_idx, h[header]).value)

    def _cell_text_by_header(self, ws, row_idx: int, header: str) -> str:
        h = self._header_map(ws)
        if header not in h:
            return ""
        return str(ws.cell(row_idx, h[header]).value or "")

    def _style_review_rows(self) -> None:
        review_fill = PatternFill("solid", fgColor="FFF2CC")
        error_fill = PatternFill("solid", fgColor="F4CCCC")
        for r in range(2, self.ws_main.max_row + 1):
            pid = self._cell_text_by_header(self.ws_main, r, "participant_id")
            needs = self._cell_bool_by_header(self.ws_main, r, "needs_review")
            fill = error_fill if pid == "ERROR" else (review_fill if needs or pid == "REVIEW_PID" else None)
            if fill:
                for c in range(1, self.ws_main.max_column + 1):
                    self.ws_main.cell(r, c).fill = fill
        for r in range(2, self.ws_conflicts.max_row + 1):
            for c in range(1, self.ws_conflicts.max_column + 1):
                self.ws_conflicts.cell(r, c).fill = review_fill
        for r in range(2, self.ws_failed.max_row + 1):
            for c in range(1, self.ws_failed.max_column + 1):
                self.ws_failed.cell(r, c).fill = error_fill

    def update_qa_summary(self) -> None:
        # Rebuild QA summary sheet from current workbook state.
        if self.ws_qa.max_row > 1:
            self.ws_qa.delete_rows(2, self.ws_qa.max_row - 1)
        h = self._header_map(self.ws_main)
        total = max(0, self.ws_main.max_row - 1)
        needs_review = 0
        errors = 0
        review_pid = 0
        uncertain_counter: Dict[str, int] = {}
        for r in range(2, self.ws_main.max_row + 1):
            pid = str(self.ws_main.cell(r, h.get("participant_id", 1)).value or "")
            if pid == "ERROR":
                errors += 1
            if pid == "REVIEW_PID":
                review_pid += 1
            if "needs_review" in h and bool_from_cell(self.ws_main.cell(r, h["needs_review"]).value):
                needs_review += 1
            if "uncertain_fields" in h:
                txt = str(self.ws_main.cell(r, h["uncertain_fields"]).value or "")
                for item in [x.strip() for x in txt.split(";") if x.strip()]:
                    key = item.split(":", 2)[0:2]
                    k = ":".join(key) if key else item[:60]
                    uncertain_counter[k] = uncertain_counter.get(k, 0) + 1
        conflict_count = max(0, self.ws_conflicts.max_row - 1)
        failed_count = max(0, self.ws_failed.max_row - 1)
        completed_no_review = max(0, total - needs_review - errors)
        review_rate = (needs_review / total) if total else 0.0
        metrics = [
            ("last_updated", now_str()),
            ("total_questionnaires_in_excel", total),
            ("completed_without_review", completed_no_review),
            ("needs_review_count", needs_review),
            ("needs_review_rate", f"{review_rate:.2%}"),
            ("error_rows", errors),
            ("review_pid_rows", review_pid),
            ("conflict_rows", conflict_count),
            ("failed_jobs_rows", failed_count),
        ]
        for k, v in sorted(uncertain_counter.items(), key=lambda kv: (-kv[1], kv[0]))[:20]:
            metrics.append((f"top_uncertain:{k}", v))
        for row_idx, (metric, value) in enumerate(metrics, 2):
            self.ws_qa.cell(row_idx, 1, metric)
            self.ws_qa.cell(row_idx, 2, value)

    def save(self) -> None:
        self.update_qa_summary()
        self._style_all()
        self._style_review_rows()
        self.wb.save(self.path)

    def close(self) -> None:
        self.save()
        self.wb.close()


# ============================================================
# Extraction logic
# ============================================================
@dataclass
class RunConfig:
    pdf_paths: List[str]
    output_excel: str
    base_url: str
    model_id: str
    api_key: str = ""
    dpi: int = DEFAULT_DPI
    first_page_1based: int = 1
    pages_per_questionnaire: int = DEFAULT_PAGES_PER_QUESTIONNAIRE
    max_tokens: int = DEFAULT_MAX_TOKENS
    timeout_s: int = DEFAULT_TIMEOUT
    image_max_side: int = DEFAULT_IMAGE_MAX_SIDE
    enhance_image: bool = True
    auto_orient: bool = False
    verify_pass: bool = False
    resolve_conflicts: bool = True
    table_zoom_pass: bool = DEFAULT_TABLE_ZOOM_PASS
    save_debug_json: bool = True
    save_debug_images: bool = False
    save_every_n: int = DEFAULT_SAVE_EVERY_N
    resume_skip_completed: bool = True
    review_only_rerun: bool = False
    checkpoint_ndjson: bool = True
    no_unclear_for_scale: bool = DEFAULT_NO_UNCLEAR_FOR_SCALE


class FullPageExtractorV4:
    def __init__(self, cfg: RunConfig, log_func):
        self.cfg = cfg
        self.log = log_func
        self.client = LMStudioClient(cfg.base_url, cfg.model_id, cfg.api_key, cfg.timeout_s)
        self.output_dir = os.path.dirname(cfg.output_excel) or "."
        self.debug_dir = os.path.join(self.output_dir, "debug_fullpage_v5")
        if cfg.save_debug_json or cfg.save_debug_images:
            safe_mkdir(self.debug_dir)

    def test_connection(self) -> List[str]:
        return self.client.list_models()

    def test_vision(self) -> Dict[str, Any]:
        img = Image.new("RGB", (700, 260), "white")
        draw = ImageDraw.Draw(img)
        draw.text((50, 95), "VISION TEST 123", fill="black")
        data, raw = self.client.chat_json(vision_test_prompt(), [("test image", img)], max_tokens=120, retries=0, image_max_side=700)
        return {"data": data, "raw": raw}

    def orient_page(self, img: Image.Image, stem: str) -> Tuple[Image.Image, Dict[str, Any]]:
        if not self.cfg.auto_orient:
            return img, {"rotation_degrees": 0, "confidence": 1.0, "disabled": True}
        variants = [
            ("Image 1 original", img),
            ("Image 2 rotated 90", rotate_image(img, 90)),
            ("Image 3 rotated 180", rotate_image(img, 180)),
            ("Image 4 rotated 270", rotate_image(img, 270)),
        ]
        try:
            data, raw = self.client.chat_json(orientation_prompt(), variants, max_tokens=180, retries=1, image_max_side=900)
            self._save_text(f"{stem}_orientation_raw.txt", raw)
            rot = int(data.get("rotation_degrees", 0))
            if rot not in {0, 90, 180, 270}:
                rot = {1: 0, 2: 90, 3: 180, 4: 270}.get(int(data.get("best_image", 1)), 0)
            return rotate_image(img, rot), data
        except Exception as e:
            self.log(f"[WARN] Orientation failed for {stem}: {e}. Using original.")
            return img, {"rotation_degrees": 0, "confidence": 0.0, "error": str(e)}

    def extract_page_once(self, img: Image.Image, page_no: int, stem: str, verify_mode: bool) -> Tuple[Dict[str, Any], str]:
        prompt = page_prompt(page_no, self.cfg.pages_per_questionnaire, verify_mode=verify_mode)
        return self.client.chat_json(
            prompt,
            [(f"Full page image, questionnaire page {page_no}", img)],
            max_tokens=self.cfg.max_tokens,
            retries=1,
            image_max_side=self.cfg.image_max_side,
            temperature=0.0,
        )

    def apply_table_zoom_pass(
        self,
        page_json: Dict[str, Any],
        img: Image.Image,
        page_no: int,
        stem: str,
        pdf_name: str,
        qidx: int,
    ) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        """AI-only enlarged table pass for dense scale tables.

        It does not use CV to decide answers. It only crops and enlarges known table
        regions, then asks the same local VLM to read the crop. Zoom answers are
        preferred when valid, but disagreements are recorded and marked review.
        """
        conflicts: List[Dict[str, Any]] = []
        if not self.cfg.table_zoom_pass or page_no not in TABLE_CROP_SPECS:
            return page_json, conflicts
        answers = page_json.get("answers")
        if not isinstance(answers, dict):
            answers = {}
            page_json["answers"] = answers
        uncertain = page_json.get("uncertain_fields") or []
        if isinstance(uncertain, str):
            uncertain = [uncertain]
        if not isinstance(uncertain, list):
            uncertain = []
        expected = expected_scale_keys_for_page(page_no)

        for crop_name, box, keys in TABLE_CROP_SPECS[page_no]:
            crop_img = crop_norm(img, box, upscale_factor=2.2 if page_no in {4, 6} else 1.9)
            if self.cfg.save_debug_images:
                self._save_image(f"{stem}_page{page_no:02d}_{crop_name}.png", crop_img, max_side=2200)
            try:
                prompt = table_zoom_prompt(page_no, crop_name, keys)
                data, raw = self.client.chat_json(
                    prompt,
                    [(f"AI-only zoom crop: {crop_name}", crop_img)],
                    max_tokens=min(self.cfg.max_tokens, 1600),
                    retries=1,
                    image_max_side=self.cfg.image_max_side,
                    temperature=0.0,
                )
                self._save_text(f"{stem}_page{page_no:02d}_{crop_name}_zoom_raw.txt", raw)
                self._save_json(f"{stem}_page{page_no:02d}_{crop_name}_zoom.json", data)
            except Exception as e:
                self.log(f"[WARN] AI table zoom pass failed for {stem} {crop_name}: {e}")
                uncertain.append(f"table_zoom_failed:{crop_name}:{e}")
                page_json["needs_review"] = True
                continue

            z_answers = data.get("answers") if isinstance(data, dict) else {}
            if not isinstance(z_answers, dict):
                uncertain.append(f"table_zoom_answers_not_dict:{crop_name}")
                page_json["needs_review"] = True
                continue

            for key in keys:
                if key not in expected:
                    continue
                old_value = answers.get(key)
                old_clean, old_review, _ = normalize_scale_answer(old_value, expected[key])
                zoom_value = z_answers.get(key)
                zoom_clean, zoom_review, zoom_reason = normalize_scale_answer(zoom_value, expected[key])

                if not zoom_review:
                    if old_review or old_clean == "UNCLEAR":
                        # Full-page was unclear/missing: accept the specialized zoom crop,
                        # but audit it as a low-confidence filled value.
                        answers[key] = zoom_clean
                        answers[f"{key}_zoom_candidate"] = zoom_clean
                        uncertain = [u for u in uncertain if not str(u).startswith(f"{key}:")]
                        uncertain.append(f"{key}:filled_from_zoom_after_fullpage_unclear:{clean_for_excel(old_value)}->{zoom_clean}")
                        page_json["needs_review"] = True
                    elif old_clean != zoom_clean:
                        # v6 safety rule: a zoom crop is a candidate, not an automatic overwrite.
                        # Keep the full-page answer for now, request a tiebreak later, and audit.
                        c = self.conflict_row(pdf_name, qidx, page_no, key, old_clean, zoom_clean, old_clean, "zoom_disagreed_kept_fullpage_pending_vote")
                        conflicts.append(c)
                        answers[key] = old_clean
                        answers[f"{key}_zoom_candidate"] = zoom_clean
                        answers[f"{key}_needs_tiebreak"] = True
                        uncertain.append(f"zoom_conflict:{key}:{old_clean}|{zoom_clean}")
                        page_json["needs_review"] = True
                    else:
                        answers[key] = old_clean
                        answers[f"{key}_zoom_candidate"] = zoom_clean
                else:
                    if old_review or old_clean == "UNCLEAR":
                        answers[key] = "UNCLEAR"
                        uncertain.append(f"{key}:zoom_{zoom_reason}")
                        page_json["needs_review"] = True
                    else:
                        # Keep a valid full-page value when zoom crop failed/unclear.
                        answers[key] = old_clean

        page_json["answers"] = answers
        page_json["uncertain_fields"] = list(dict.fromkeys(map(str, uncertain)))
        return postprocess_page_json(page_json, page_no), conflicts

    def force_no_unclear_scale_answers(
        self,
        page_json: Dict[str, Any],
        img: Image.Image,
        page_no: int,
        stem: str,
        pdf_name: str,
        qidx: int,
        conflicts: Optional[List[Dict[str, Any]]] = None,
    ) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        """Ensure scale/table answers contain no UNCLEAR.

        It asks the VLM one more targeted time for only the unresolved fields.
        If the response is still invalid, it uses a deterministic random choice.
        The forced fields are kept in uncertain_fields for audit, but the Excel
        data cells contain a valid numeric value instead of UNCLEAR.
        """
        if not self.cfg.no_unclear_for_scale:
            return page_json, conflicts or []
        expected = expected_scale_keys_for_page(page_no)
        if not expected:
            return page_json, conflicts or []
        answers = page_json.get("answers")
        if not isinstance(answers, dict):
            answers = {}
            page_json["answers"] = answers
        uncertain = page_json.get("uncertain_fields") or []
        if isinstance(uncertain, str):
            uncertain = [uncertain]
        if not isinstance(uncertain, list):
            uncertain = []
        unresolved: List[str] = []
        for key, allowed in expected.items():
            cleaned, review, _ = normalize_scale_answer(answers.get(key), allowed)
            # v6: also tiebreak valid values when full-page and zoom disagreed.
            needs_tiebreak = str(answers.get(f"{key}_needs_tiebreak", "")).strip().lower() in {"true", "1", "yes"}
            if review or cleaned == "UNCLEAR" or needs_tiebreak:
                unresolved.append(key)
        if not unresolved:
            return page_json, conflicts or []

        # One more targeted look using the full page image. This is deliberately
        # narrow: only unresolved fields are requested, with no UNCLEAR allowed.
        tiebreak_answers: Dict[str, Any] = {}
        try:
            prompt = scale_tiebreak_prompt(page_no, unresolved)
            data, raw = self.client.chat_json(
                prompt,
                [(f"No-UNCLEAR tie-break page {page_no}", img)],
                max_tokens=min(self.cfg.max_tokens, 1000),
                retries=1,
                image_max_side=self.cfg.image_max_side,
                temperature=0.0,
            )
            self._save_text(f"{stem}_page{page_no:02d}_no_unclear_tiebreak_raw.txt", raw)
            self._save_json(f"{stem}_page{page_no:02d}_no_unclear_tiebreak.json", data)
            ta = data.get("answers") if isinstance(data, dict) else {}
            if isinstance(ta, dict):
                tiebreak_answers = ta
        except Exception as e:
            self.log(f"[WARN] No-UNCLEAR tiebreak failed for {stem} p{page_no}: {e}")
            uncertain.append(f"no_unclear_tiebreak_failed:{e}")

        # Build a lookup of conflict values so forced choices can use all observed candidates.
        by_key_conflicts: Dict[str, List[Any]] = {}
        for c in conflicts or []:
            k = str(c.get("key"))
            by_key_conflicts.setdefault(k, [])
            by_key_conflicts[k].extend([c.get("first_value"), c.get("verify_value"), c.get("chosen_value"), c.get("resolver_value")])

        for key in unresolved:
            allowed = expected[key]
            candidates = [
                answers.get(key),
                answers.get(f"{key}_raw_model_value"),
                answers.get(f"{key}_zoom_candidate"),
                tiebreak_answers.get(key),
            ]
            candidates.extend(by_key_conflicts.get(key, []))
            chosen, method = choose_scale_answer_no_unclear_na_ok(
                key,
                page_no,
                allowed,
                candidates,
                seed_text=f"{pdf_name}|q{qidx}|p{page_no}|{key}",
            )
            old = answers.get(key)
            answers[key] = chosen
            answers[f"{key}_needs_tiebreak"] = False
            # No forced/audit output: final cell is a digit or N/A. Keep conflicts sheet only.
            uncertain.append(f"{key}:{method}:{clean_for_excel(old)}->{chosen}")
            page_json["needs_review"] = True
            # Update conflict chosen_value if this field was in conflicts.
            for c in conflicts or []:
                if str(c.get("key")) == key:
                    c["chosen_value"] = chosen
                    c["resolver_value"] = tiebreak_answers.get(key, "N/A") if key in tiebreak_answers else "N/A"
                    c["status"] = f"{c.get('status','')}|no_unclear_{method}"
        page_json["answers"] = answers
        page_json["uncertain_fields"] = list(dict.fromkeys(map(str, uncertain)))
        return page_json, conflicts or []

    def extract_page_json(self, img: Image.Image, page_no: int, stem: str, pdf_name: str, qidx: int) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        first, raw1 = self.extract_page_once(img, page_no, stem, verify_mode=False)
        first = postprocess_page_json(first, page_no)
        self._save_text(f"{stem}_page{page_no:02d}_extract_raw.txt", raw1)
        self._save_json(f"{stem}_page{page_no:02d}_extract.json", first)
        conflicts: List[Dict[str, Any]] = []

        # v4: AI-only enlarged table pass for dense scale pages.
        first, zoom_conflicts = self.apply_table_zoom_pass(first, img, page_no, stem, pdf_name, qidx)
        conflicts.extend(zoom_conflicts)
        self._save_json(f"{stem}_page{page_no:02d}_after_zoom.json", first)

        if not self.cfg.verify_pass:
            first, conflicts = self.force_no_unclear_scale_answers(first, img, page_no, stem, pdf_name, qidx, conflicts)
            first = postprocess_page_json(first, page_no)
            self._save_json(f"{stem}_page{page_no:02d}_final_no_unclear.json", first)
            return first, conflicts

        try:
            second, raw2 = self.extract_page_once(img, page_no, stem, verify_mode=True)
            second = postprocess_page_json(second, page_no)
            self._save_text(f"{stem}_page{page_no:02d}_verify_raw.txt", raw2)
            self._save_json(f"{stem}_page{page_no:02d}_verify.json", second)
        except Exception as e:
            self.log(f"[WARN] Verification failed for {stem} p{page_no}: {e}. Keeping first pass.")
            return first, conflicts

        merged, conflicts = self.safe_merge_page(first, second, page_no, img, stem, pdf_name, qidx)
        merged = postprocess_page_json(merged, page_no)
        merged, conflicts = self.force_no_unclear_scale_answers(merged, img, page_no, stem, pdf_name, qidx, conflicts)
        merged = postprocess_page_json(merged, page_no)
        self._save_json(f"{stem}_page{page_no:02d}_merged.json", merged)
        return merged, conflicts

    def safe_merge_page(
        self,
        first: Dict[str, Any],
        second: Dict[str, Any],
        page_no: int,
        img: Image.Image,
        stem: str,
        pdf_name: str,
        qidx: int,
    ) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        merged = dict(first)
        conflicts: List[Dict[str, Any]] = []

        # Merge top-level participant_id: prefer valid PID; if conflict, keep first but log.
        fpid = normalize_pid(first.get("participant_id"))
        spid = normalize_pid(second.get("participant_id"))
        if fpid == "N/A" and spid != "N/A":
            merged["participant_id"] = spid
        elif fpid != "N/A":
            merged["participant_id"] = fpid
        elif spid != "N/A":
            merged["participant_id"] = spid

        # Merge identity fields.
        merged_identity = dict(first.get("identity") or {})
        for k, sv in (second.get("identity") or {}).items():
            fv = merged_identity.get(k)
            if is_missing(fv) and not is_missing(sv):
                merged_identity[k] = sv
            elif not is_missing(fv) and not is_missing(sv) and normalize_value_for_compare(fv) != normalize_value_for_compare(sv):
                conflicts.append(self.conflict_row(pdf_name, qidx, page_no, f"identity.{k}", fv, sv, fv, "kept_first_identity_conflict"))
        merged["identity"] = merged_identity

        # Merge answers safely.
        f_answers = first.get("answers") or {}
        s_answers = second.get("answers") or {}
        if not isinstance(f_answers, dict):
            f_answers = {}
        if not isinstance(s_answers, dict):
            s_answers = {}
        merged_answers = dict(f_answers)
        answer_conflicts_for_resolver: List[Dict[str, Any]] = []

        for key in sorted(set(f_answers.keys()) | set(s_answers.keys())):
            fv = f_answers.get(key)
            sv = s_answers.get(key)
            nf = normalize_value_for_compare(fv)
            ns = normalize_value_for_compare(sv)
            if nf == ns:
                merged_answers[key] = fv if not is_missing(fv) else sv
            elif is_missing(fv) and not is_missing(sv):
                merged_answers[key] = sv
            elif not is_missing(fv) and is_missing(sv):
                merged_answers[key] = fv
            else:
                # v4 safety rule: scale/table conflicts must NOT be auto-overwritten by
                # a generic conflict resolver. If the zoom pass has not already solved it,
                # mark the value UNCLEAR so it is visible in review.
                if is_scale_answer_key(page_no, key):
                    merged_answers[key] = "UNCLEAR"
                    c = self.conflict_row(pdf_name, qidx, page_no, key, fv, sv, "UNCLEAR", "scale_conflict_set_unclear_no_resolver")
                    conflicts.append(c)
                else:
                    merged_answers[key] = fv
                    c = self.conflict_row(pdf_name, qidx, page_no, key, fv, sv, fv, "kept_first_conflict")
                    conflicts.append(c)
                    answer_conflicts_for_resolver.append({"key": key, "first_value": fv, "verify_value": sv})

        if self.cfg.resolve_conflicts and answer_conflicts_for_resolver:
            try:
                rprompt = conflict_resolution_prompt(page_no, answer_conflicts_for_resolver)
                resolved, rraw = self.client.chat_json(
                    rprompt,
                    [(f"Full page image for conflict resolution page {page_no}", img)],
                    max_tokens=min(self.cfg.max_tokens, 2200),
                    retries=1,
                    image_max_side=self.cfg.image_max_side,
                    temperature=0.0,
                )
                self._save_text(f"{stem}_page{page_no:02d}_conflict_resolve_raw.txt", rraw)
                self._save_json(f"{stem}_page{page_no:02d}_conflict_resolve.json", resolved)
                resolved_answers = resolved.get("resolved_answers") or {}
                if isinstance(resolved_answers, dict):
                    for c in conflicts:
                        key = c.get("key")
                        if is_scale_answer_key(page_no, key):
                            # Never let the resolver overwrite scale/table answers.
                            continue
                        if key in resolved_answers and not is_missing(resolved_answers[key]):
                            merged_answers[key] = resolved_answers[key]
                            c["resolver_value"] = resolved_answers[key]
                            c["chosen_value"] = resolved_answers[key]
                            c["status"] = "resolved_by_vlm_non_scale"
            except Exception as e:
                self.log(f"[WARN] Conflict resolver failed for {stem} p{page_no}: {e}")

        merged["answers"] = merged_answers

        # Merge uncertainty.
        u = []
        for obj in [first, second]:
            val = obj.get("uncertain_fields") or []
            if isinstance(val, str):
                val = [val]
            if isinstance(val, list):
                u.extend([x for x in val if clean_text(x) != "N/A"])
        if conflicts:
            u.extend([f"conflict:{c['key']}" for c in conflicts[:50]])
        merged["uncertain_fields"] = list(dict.fromkeys(map(str, u)))
        merged["needs_review"] = bool(merged.get("needs_review") or conflicts or u)
        return merged, conflicts

    def conflict_row(self, pdf_name: str, qidx: int, page_no: int, key: str, first_value: Any, verify_value: Any, chosen: Any, status: str) -> Dict[str, Any]:
        return {
            "participant_id": "PENDING",
            "source_pdf": pdf_name,
            "questionnaire_index_in_pdf": qidx,
            "page_no_in_questionnaire": page_no,
            "key": key,
            "first_value": first_value,
            "verify_value": verify_value,
            "chosen_value": chosen,
            "resolver_value": "N/A",
            "status": status,
        }

    def extract_questionnaire(
        self,
        pdf_name: str,
        qidx: int,
        block_start0: int,
        page_images: List[Image.Image],
    ) -> Tuple[Dict[str, Any], List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
        stem = f"{Path(pdf_name).stem}_q{qidx:04d}_pages_{block_start0 + 1}-{block_start0 + len(page_images)}"
        page_results: List[Dict[str, Any]] = []
        orientation_results: List[Dict[str, Any]] = []
        all_conflicts: List[Dict[str, Any]] = []

        for i, img in enumerate(page_images, 1):
            page_stem = f"{stem}_p{i:02d}"
            oriented, orient_info = self.orient_page(img, page_stem)
            orientation_results.append({"page": i, **orient_info})
            if self.cfg.save_debug_images:
                self._save_image(f"{page_stem}_oriented.png", oriented, max_side=1800)
            self.log(f"[INFO] Extracting {pdf_name} q{qidx} page {i}/{len(page_images)}")
            try:
                page_json, conflicts = self.extract_page_json(oriented, i, page_stem, pdf_name, qidx)
                all_conflicts.extend(conflicts)
            except Exception as e:
                self.log(f"[ERROR] Extraction failed: {pdf_name} q{qidx} p{i}: {e}")
                page_json = {
                    "page_type": "ERROR",
                    "participant_id": "N/A",
                    "identity": {},
                    "answers": {},
                    "visible_handwriting": [],
                    "needs_review": True,
                    "uncertain_fields": [f"page_{i}_error:{e}"],
                    "error": str(e),
                }
            page_results.append(page_json)

        participant_json = {
            "source_pdf": pdf_name,
            "questionnaire_index_in_pdf": qidx,
            "source_pages": f"{block_start0 + 1}-{block_start0 + len(page_images)}",
            "pages_per_questionnaire": len(page_images),
            "orientation_results": orientation_results,
            "pages": page_results,
        }
        pid = self.choose_participant_id(page_results, pdf_name)
        for c in all_conflicts:
            c["participant_id"] = pid
        main_row, long_rows = self.flatten_participant(participant_json, pid, all_conflicts)
        raw_json_file = ""
        if self.cfg.save_debug_json:
            raw_json_file = os.path.join(self.debug_dir, f"{stem}_participant_raw.json")
            with open(raw_json_file, "w", encoding="utf-8") as f:
                json.dump(participant_json, f, ensure_ascii=False, indent=2)
        main_row["raw_json_file"] = raw_json_file
        return main_row, long_rows, all_conflicts, participant_json

    def choose_participant_id(self, pages: List[Dict[str, Any]], pdf_name: str) -> str:
        candidates: List[str] = []
        fname_pid = normalize_pid(Path(pdf_name).stem)
        if fname_pid != "N/A":
            candidates.append(fname_pid)
        for page in pages:
            candidates.append(normalize_pid(page.get("participant_id")))
            ident = page.get("identity") or {}
            if isinstance(ident, dict):
                for key in ["participant_id", "id", "id_code", "check_number", "問卷編號", "檢查編號"]:
                    if key in ident:
                        candidates.append(normalize_pid(ident.get(key)))
            text = json.dumps(page, ensure_ascii=False)
            for m in PID_RE.finditer(text):
                candidates.append(normalize_pid(m.group(0)))
        valid = [c for c in candidates if c != "N/A"]
        if not valid:
            return "REVIEW_PID"
        scores: Dict[str, int] = {}
        for c in valid:
            scores[c] = scores.get(c, 0) + 1
        return sorted(scores.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]

    def flatten_participant(self, participant_json: Dict[str, Any], pid: str, conflicts: List[Dict[str, Any]]) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        pdf_name = participant_json["source_pdf"]
        qidx = participant_json["questionnaire_index_in_pdf"]
        main: Dict[str, Any] = {
            "participant_id": pid,
            "source_pdf": pdf_name,
            "questionnaire_index_in_pdf": qidx,
            "source_pages": participant_json["source_pages"],
            "pages_per_questionnaire": participant_json["pages_per_questionnaire"],
            "extraction_time": now_str(),
            "needs_review": False,
            "uncertain_fields": "",
            "conflict_count": len(conflicts),
        }
        long_rows: List[Dict[str, Any]] = []
        uncertain: List[str] = []
        needs_review = bool(conflicts or pid == "REVIEW_PID")

        for orient in participant_json.get("orientation_results", []):
            p = orient.get("page")
            if p:
                main[f"p{int(p):02d}_rotation"] = orient.get("rotation_degrees", "N/A")
                main[f"p{int(p):02d}_orientation_confidence"] = orient.get("confidence", "N/A")

        for page_no, page in enumerate(participant_json.get("pages", []), 1):
            pfx = f"p{page_no:02d}"
            main[f"{pfx}_page_type"] = page.get("page_type", "N/A")
            seen_pid = normalize_pid(page.get("participant_id"))
            if seen_pid != "N/A":
                main[f"{pfx}_participant_id_seen"] = seen_pid

            ident = page.get("identity") or {}
            if isinstance(ident, dict):
                for k, v in ident.items():
                    main[f"{pfx}_identity_{sanitize_key(k)}"] = v

            answers = page.get("answers") or {}
            if isinstance(answers, dict):
                for k, v in answers.items():
                    ks = str(k)
                    # v7: hide internal resolution metadata from the final Excel answer columns.
                    if ks.endswith(("_needs_tiebreak", "_zoom_candidate", "_raw_model_value")) or "_forced_" in ks:
                        continue
                    col = f"{pfx}_{sanitize_key(k)}"
                    main[col] = v
                    long_rows.append({
                        "participant_id": pid,
                        "source_pdf": pdf_name,
                        "questionnaire_index_in_pdf": qidx,
                        "page_no_in_questionnaire": page_no,
                        "key": k,
                        "selected_value": v,
                        "source": "answers",
                        "needs_review": page.get("needs_review", False),
                    })

            handwriting = page.get("visible_handwriting") or []
            if handwriting:
                main[f"{pfx}_visible_handwriting"] = handwriting

            u = page.get("uncertain_fields") or []
            if isinstance(u, str):
                u = [u]
            if isinstance(u, list):
                for item in u:
                    if clean_text(item) != "N/A":
                        uncertain.append(f"{pfx}:{item}")
            if page.get("needs_review"):
                needs_review = True
            if page.get("error"):
                needs_review = True
                uncertain.append(f"{pfx}:ERROR:{page.get('error')}")

        for c in conflicts:
            uncertain.append(f"p{int(c.get('page_no_in_questionnaire', 0)):02d}:conflict:{c.get('key')}")

        main["needs_review"] = bool(needs_review or uncertain)
        main["uncertain_fields"] = "; ".join(list(dict.fromkeys(map(str, uncertain)))[:250])
        return main, long_rows

    def _save_text(self, filename: str, text: str) -> None:
        if not self.cfg.save_debug_json:
            return
        safe_mkdir(self.debug_dir)
        with open(os.path.join(self.debug_dir, filename), "w", encoding="utf-8") as f:
            f.write(text or "")

    def _save_json(self, filename: str, data: Dict[str, Any]) -> None:
        if not self.cfg.save_debug_json:
            return
        safe_mkdir(self.debug_dir)
        with open(os.path.join(self.debug_dir, filename), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _save_image(self, filename: str, img: Image.Image, max_side: Optional[int] = None) -> None:
        if not self.cfg.save_debug_images:
            return
        safe_mkdir(self.debug_dir)
        out = resize_keep_aspect(img, max_side) if max_side else img
        out.save(os.path.join(self.debug_dir, filename))


# ============================================================
# GUI
# ============================================================
class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1280x880")
        self.msg_q: queue.Queue = queue.Queue()
        self.pdf_paths: List[str] = []
        self.running = False

        self.base_url_var = tk.StringVar(value=DEFAULT_BASE_URL)
        self.model_var = tk.StringVar(value=DEFAULT_MODEL_ID)
        self.api_key_var = tk.StringVar(value="")
        self.output_var = tk.StringVar(value=str(Path.cwd() / "chinese_questionnaire_fullpage_v5_output.xlsx"))
        self.dpi_var = tk.StringVar(value=str(DEFAULT_DPI))
        self.first_page_var = tk.StringVar(value="1")
        self.pages_per_q_var = tk.StringVar(value=str(DEFAULT_PAGES_PER_QUESTIONNAIRE))
        self.max_tokens_var = tk.StringVar(value=str(DEFAULT_MAX_TOKENS))
        self.timeout_var = tk.StringVar(value=str(DEFAULT_TIMEOUT))
        self.image_max_side_var = tk.StringVar(value=str(DEFAULT_IMAGE_MAX_SIDE))
        self.enhance_image_var = tk.BooleanVar(value=True)
        self.auto_orient_var = tk.BooleanVar(value=False)
        self.verify_pass_var = tk.BooleanVar(value=False)
        self.resolve_conflicts_var = tk.BooleanVar(value=True)
        self.table_zoom_pass_var = tk.BooleanVar(value=True)
        self.save_debug_json_var = tk.BooleanVar(value=True)
        self.save_debug_images_var = tk.BooleanVar(value=False)
        self.save_every_n_var = tk.StringVar(value=str(DEFAULT_SAVE_EVERY_N))
        self.resume_skip_completed_var = tk.BooleanVar(value=True)
        self.review_only_rerun_var = tk.BooleanVar(value=False)
        self.checkpoint_ndjson_var = tk.BooleanVar(value=True)
        self.no_unclear_for_scale_var = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="Ready")
        self.progress_var = tk.DoubleVar(value=0.0)

        self._build_ui()
        self.root.after(120, self.poll_messages)

    def _build_ui(self) -> None:
        outer = tk.Frame(self.root)
        outer.pack(fill="both", expand=True, padx=10, pady=8)

        settings = tk.LabelFrame(outer, text="LM Studio + Extraction Settings")
        settings.pack(fill="x", pady=6)

        r1 = tk.Frame(settings)
        r1.pack(fill="x", padx=6, pady=4)
        tk.Label(r1, text="Base URL", width=14, anchor="w").pack(side="left")
        tk.Entry(r1, textvariable=self.base_url_var, width=42).pack(side="left", padx=4)
        tk.Label(r1, text="Model ID", width=10, anchor="w").pack(side="left", padx=(12, 0))
        tk.Entry(r1, textvariable=self.model_var, width=36).pack(side="left", padx=4)
        tk.Label(r1, text="API Key", width=8, anchor="w").pack(side="left", padx=(12, 0))
        tk.Entry(r1, textvariable=self.api_key_var, width=18, show="*").pack(side="left", padx=4)
        tk.Button(r1, text="Test Connection", command=self.test_connection).pack(side="left", padx=6)
        tk.Button(r1, text="Test Vision", command=self.test_vision).pack(side="left", padx=4)

        r2 = tk.Frame(settings)
        r2.pack(fill="x", padx=6, pady=4)
        tk.Label(r2, text="Output Excel", width=14, anchor="w").pack(side="left")
        tk.Entry(r2, textvariable=self.output_var, width=110).pack(side="left", padx=4, fill="x", expand=True)
        tk.Button(r2, text="Browse", command=self.pick_output).pack(side="left", padx=6)

        r3 = tk.Frame(settings)
        r3.pack(fill="x", padx=6, pady=4)
        for label, var, width in [
            ("DPI", self.dpi_var, 7),
            ("First page", self.first_page_var, 7),
            ("Pages/questionnaire", self.pages_per_q_var, 7),
            ("Max tokens", self.max_tokens_var, 8),
            ("Timeout sec", self.timeout_var, 8),
            ("Image max side", self.image_max_side_var, 8),
            ("Save every N", self.save_every_n_var, 7),
        ]:
            tk.Label(r3, text=label, anchor="w").pack(side="left", padx=(8, 2))
            tk.Entry(r3, textvariable=var, width=width).pack(side="left", padx=(0, 6))

        r4 = tk.Frame(settings)
        r4.pack(fill="x", padx=6, pady=4)
        ttk.Checkbutton(r4, text="Enhance image contrast", variable=self.enhance_image_var).pack(side="left", padx=8)
        ttk.Checkbutton(r4, text="VLM auto-orientation (slower)", variable=self.auto_orient_var).pack(side="left", padx=8)
        ttk.Checkbutton(r4, text="Second-pass verification", variable=self.verify_pass_var).pack(side="left", padx=8)
        ttk.Checkbutton(r4, text="Resolve non-scale conflicts with VLM", variable=self.resolve_conflicts_var).pack(side="left", padx=8)
        ttk.Checkbutton(r4, text="AI-only table zoom pass", variable=self.table_zoom_pass_var).pack(side="left", padx=8)
        ttk.Checkbutton(r4, text="Save debug JSON/raw", variable=self.save_debug_json_var).pack(side="left", padx=8)
        ttk.Checkbutton(r4, text="Save debug page images", variable=self.save_debug_images_var).pack(side="left", padx=8)

        r5 = tk.Frame(settings)
        r5.pack(fill="x", padx=6, pady=4)
        ttk.Checkbutton(r5, text="Resume: skip completed questionnaires", variable=self.resume_skip_completed_var).pack(side="left", padx=8)
        ttk.Checkbutton(r5, text="Review-only rerun existing REVIEW/ERROR rows", variable=self.review_only_rerun_var).pack(side="left", padx=8)
        ttk.Checkbutton(r5, text="Checkpoint NDJSON after each questionnaire", variable=self.checkpoint_ndjson_var).pack(side="left", padx=8)
        ttk.Checkbutton(r5, text="No UNCLEAR for scale answers; allow N/A when blank", variable=self.no_unclear_for_scale_var).pack(side="left", padx=8)

        files = tk.LabelFrame(outer, text="PDF files")
        files.pack(fill="both", expand=False, pady=6)
        btns = tk.Frame(files)
        btns.pack(fill="x", padx=6, pady=4)
        tk.Button(btns, text="Add PDFs", command=self.add_pdfs).pack(side="left", padx=4)
        tk.Button(btns, text="Remove Selected", command=self.remove_selected).pack(side="left", padx=4)
        tk.Button(btns, text="Clear", command=self.clear_pdfs).pack(side="left", padx=4)
        self.files_list = tk.Listbox(files, selectmode=tk.EXTENDED, height=8)
        self.files_list.pack(fill="both", expand=True, padx=6, pady=6)

        run_frame = tk.Frame(outer)
        run_frame.pack(fill="x", pady=6)
        self.start_btn = tk.Button(run_frame, text="Start Extraction", command=self.start_worker, width=18)
        self.start_btn.pack(side="left", padx=4)
        self.pb = ttk.Progressbar(run_frame, orient="horizontal", mode="determinate", variable=self.progress_var)
        self.pb.pack(side="left", fill="x", expand=True, padx=8)
        tk.Label(outer, textvariable=self.status_var, anchor="w").pack(fill="x")

        logf = tk.LabelFrame(outer, text="Log")
        logf.pack(fill="both", expand=True, pady=6)
        self.log_list = tk.Listbox(logf, width=180, height=24)
        ys = tk.Scrollbar(logf, orient="vertical", command=self.log_list.yview)
        self.log_list.configure(yscrollcommand=ys.set)
        self.log_list.pack(side="left", fill="both", expand=True)
        ys.pack(side="right", fill="y")

    def queue_msg(self, kind: str, payload: Any) -> None:
        self.msg_q.put((kind, payload))

    def poll_messages(self) -> None:
        while True:
            try:
                kind, payload = self.msg_q.get_nowait()
            except queue.Empty:
                break
            if kind == "log":
                self.log_list.insert(tk.END, payload)
                self.log_list.yview_moveto(1.0)
            elif kind == "status":
                self.status_var.set(payload)
            elif kind == "progress":
                self.progress_var.set(float(payload))
            elif kind == "done_ok":
                self.start_btn.config(state="normal")
                self.running = False
                messagebox.showinfo("Done", payload)
            elif kind == "done_err":
                self.start_btn.config(state="normal")
                self.running = False
                messagebox.showerror("Error", payload)
        self.root.after(120, self.poll_messages)

    def add_pdfs(self) -> None:
        paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
        for p in paths:
            if p not in self.pdf_paths:
                self.pdf_paths.append(p)
                self.files_list.insert(tk.END, p)

    def remove_selected(self) -> None:
        idxs = list(self.files_list.curselection())[::-1]
        for i in idxs:
            self.files_list.delete(i)
            del self.pdf_paths[i]

    def clear_pdfs(self) -> None:
        self.files_list.delete(0, tk.END)
        self.pdf_paths = []

    def pick_output(self) -> None:
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if p:
            self.output_var.set(p)

    def make_cfg(self) -> RunConfig:
        return RunConfig(
            pdf_paths=list(self.pdf_paths),
            output_excel=self.output_var.get().strip(),
            base_url=self.base_url_var.get().strip(),
            model_id=self.model_var.get().strip(),
            api_key=self.api_key_var.get().strip(),
            dpi=max(120, int(self.dpi_var.get().strip())),
            first_page_1based=max(1, int(self.first_page_var.get().strip())),
            pages_per_questionnaire=max(1, int(self.pages_per_q_var.get().strip())),
            max_tokens=max(512, int(self.max_tokens_var.get().strip())),
            timeout_s=max(30, int(self.timeout_var.get().strip())),
            image_max_side=max(800, int(self.image_max_side_var.get().strip())),
            enhance_image=bool(self.enhance_image_var.get()),
            auto_orient=bool(self.auto_orient_var.get()),
            verify_pass=bool(self.verify_pass_var.get()),
            resolve_conflicts=bool(self.resolve_conflicts_var.get()),
            table_zoom_pass=bool(self.table_zoom_pass_var.get()),
            save_debug_json=bool(self.save_debug_json_var.get()),
            save_debug_images=bool(self.save_debug_images_var.get()),
            save_every_n=max(1, int(self.save_every_n_var.get().strip())),
            resume_skip_completed=bool(self.resume_skip_completed_var.get()),
            review_only_rerun=bool(self.review_only_rerun_var.get()),
            checkpoint_ndjson=bool(self.checkpoint_ndjson_var.get()),
            no_unclear_for_scale=bool(self.no_unclear_for_scale_var.get()),
        )

    def test_connection(self) -> None:
        try:
            cfg = self.make_cfg()
            client = LMStudioClient(cfg.base_url, cfg.model_id, cfg.api_key, cfg.timeout_s)
            models = client.list_models()
            exact = cfg.model_id in models
            msg = "LM Studio reachable.\n\nModels found:\n" + ("\n".join(models) if models else "(none)")
            if cfg.model_id:
                msg += f"\n\nSelected model: {cfg.model_id}\nExact match in list: {'YES' if exact else 'NO'}"
                if not exact:
                    msg += "\n\nWarning: use the exact API Model Identifier shown by LM Studio."
            messagebox.showinfo("Connection Test", msg)
        except Exception as e:
            messagebox.showerror("Connection failed", str(e))

    def test_vision(self) -> None:
        try:
            cfg = self.make_cfg()
            ext = FullPageExtractorV4(cfg, lambda m: None)
            result = ext.test_vision()
            messagebox.showinfo("Vision Test", json.dumps(result["data"], ensure_ascii=False, indent=2))
        except Exception as e:
            messagebox.showerror("Vision test failed", str(e))

    def start_worker(self) -> None:
        if self.running:
            return
        if not self.pdf_paths:
            messagebox.showerror("Error", "Please add at least one PDF.")
            return
        try:
            cfg = self.make_cfg()
        except Exception:
            messagebox.showerror("Error", "Please check numeric settings: DPI, pages, tokens, timeout, image max side.")
            return
        self.running = True
        self.start_btn.config(state="disabled")
        self.progress_var.set(0.0)
        self.status_var.set("Starting...")
        threading.Thread(target=self.run_extraction, args=(cfg,), daemon=True).start()

    def build_jobs(self, cfg: RunConfig, writer: DynamicExcelWriter, log) -> List[Dict[str, Any]]:
        completed_keys = writer.get_completed_keys() if cfg.resume_skip_completed else set()
        review_keys = writer.get_review_keys() if cfg.review_only_rerun else set()
        jobs: List[Dict[str, Any]] = []
        for pdf_path in cfg.pdf_paths:
            pdf_name = os.path.basename(pdf_path)
            try:
                doc = fitz.open(pdf_path)
                page_count = doc.page_count
                doc.close()
            except Exception as e:
                log(f"Cannot open PDF {pdf_path}: {e}", "ERROR")
                continue
            start0 = cfg.first_page_1based - 1
            qidx = 0
            for block_start0 in range(start0, page_count, cfg.pages_per_questionnaire):
                if block_start0 + cfg.pages_per_questionnaire > page_count:
                    remaining = page_count - block_start0
                    if remaining > 0:
                        log(f"Skipping incomplete block in {pdf_name} at page {block_start0 + 1}, remaining pages={remaining}", "WARN")
                    break
                qidx += 1
                key = (pdf_name, qidx)
                if cfg.review_only_rerun:
                    if key not in review_keys:
                        continue
                elif cfg.resume_skip_completed and key in completed_keys:
                    log(f"Resume skip: {pdf_name} q{qidx} already completed", "INFO")
                    continue
                jobs.append({
                    "pdf_path": pdf_path,
                    "pdf_name": pdf_name,
                    "qidx": qidx,
                    "block_start0": block_start0,
                    "source_pages": f"{block_start0 + 1}-{block_start0 + cfg.pages_per_questionnaire}",
                })
        return jobs

    def run_extraction(self, cfg: RunConfig) -> None:
        writer: Optional[DynamicExcelWriter] = None
        log_txt_path = os.path.splitext(cfg.output_excel)[0] + "_runlog.txt"
        checkpoint_path = os.path.splitext(cfg.output_excel)[0] + "_checkpoint.ndjson"
        failed_path = os.path.splitext(cfg.output_excel)[0] + "_failed_jobs.jsonl"
        log_lines: List[str] = []

        def log(msg: str, level: str = "INFO") -> None:
            line = f"{now_str()} [{level}] {msg}"
            log_lines.append(line)
            self.queue_msg("log", line)
            try:
                with open(log_txt_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(log_lines))
            except Exception:
                pass
            if writer is not None:
                try:
                    writer.append_log(msg, level)
                    if cfg.save_every_n <= 1:
                        writer.save()
                except Exception:
                    pass

        try:
            writer = DynamicExcelWriter(cfg.output_excel)
            extractor = FullPageExtractorV4(cfg, lambda m: log(m, "INFO"))
            models = extractor.test_connection()
            log(f"LM Studio reachable. Models: {', '.join(models) if models else '(none)'}")
            if cfg.model_id not in models:
                log(f"Selected model_id '{cfg.model_id}' is not an exact match in /models list. It may still work if LM Studio accepts aliases.", "WARN")

            jobs = self.build_jobs(cfg, writer, log)
            if not jobs:
                log("No jobs to process. If this is unexpected, turn off Resume or Review-only mode.", "WARN")
                writer.save()
                writer.close()
                self.queue_msg("status", "Done")
                self.queue_msg("done_ok", f"No jobs to process.\nExcel:\n{cfg.output_excel}\nLog:\n{log_txt_path}")
                return

            log(f"Total jobs to process: {len(jobs)}")
            done_q = 0
            pending_save_count = 0
            open_docs: Dict[str, fitz.Document] = {}

            try:
                for job in jobs:
                    pdf_path = job["pdf_path"]
                    pdf_name = job["pdf_name"]
                    qidx = int(job["qidx"])
                    block_start0 = int(job["block_start0"])
                    self.queue_msg("status", f"{pdf_name} questionnaire {qidx}")
                    if pdf_path not in open_docs:
                        open_docs[pdf_path] = fitz.open(pdf_path)
                    doc = open_docs[pdf_path]
                    page_images: List[Image.Image] = []
                    log(f"Rendering {pdf_name} questionnaire {qidx}, pages {job['source_pages']}")
                    for pi in range(block_start0, block_start0 + cfg.pages_per_questionnaire):
                        img = render_pdf_page(doc, pi, cfg.dpi, cfg.enhance_image)
                        page_images.append(img)

                    try:
                        main_row, long_rows, conflict_rows, participant_json = extractor.extract_questionnaire(pdf_name, qidx, block_start0, page_images)
                        main_row["run_mode"] = "review_only_rerun" if cfg.review_only_rerun else ("resume" if cfg.resume_skip_completed else "normal")
                        main_row["checkpoint_time"] = now_str()
                        # Replace existing rows for this PDF/qidx to avoid duplicate results after rerun.
                        writer.upsert_main(main_row)
                        writer.replace_long_for_key(pdf_name, qidx, long_rows)
                        writer.replace_conflicts_for_key(pdf_name, qidx, conflict_rows)
                        pending_save_count += 1

                        if cfg.checkpoint_ndjson:
                            append_ndjson(checkpoint_path, {
                                "time": now_str(),
                                "status": "ok",
                                "source_pdf": pdf_name,
                                "questionnaire_index_in_pdf": qidx,
                                "source_pages": job["source_pages"],
                                "participant_id": main_row.get("participant_id"),
                                "needs_review": main_row.get("needs_review"),
                                "conflict_count": len(conflict_rows),
                                "main_row": main_row,
                                "participant_json": participant_json,
                            })
                        log(f"Saved q{qidx} from {pdf_name} as participant_id={main_row.get('participant_id')} conflicts={len(conflict_rows)} review={main_row.get('needs_review')}")
                    except Exception as e:
                        tb = traceback.format_exc()
                        log(f"Questionnaire extraction failed for {pdf_name} q{qidx}: {e}\n{tb}", "ERROR")
                        err_row = {
                            "participant_id": "ERROR",
                            "source_pdf": pdf_name,
                            "questionnaire_index_in_pdf": qidx,
                            "source_pages": job["source_pages"],
                            "pages_per_questionnaire": cfg.pages_per_questionnaire,
                            "extraction_time": now_str(),
                            "needs_review": True,
                            "uncertain_fields": str(e),
                            "conflict_count": 0,
                            "run_mode": "error",
                        }
                        writer.upsert_main(err_row)
                        failed = {
                            "time": now_str(),
                            "source_pdf": pdf_name,
                            "questionnaire_index_in_pdf": qidx,
                            "source_pages": job["source_pages"],
                            "error": str(e),
                            "traceback": tb,
                            "status": "failed",
                        }
                        writer.append_failed(failed)
                        append_ndjson(failed_path, failed)
                        if cfg.checkpoint_ndjson:
                            append_ndjson(checkpoint_path, {"time": now_str(), "status": "failed", **failed})
                        pending_save_count += 1

                    done_q += 1
                    self.queue_msg("progress", min(100.0, done_q * 100.0 / len(jobs)))
                    if pending_save_count >= max(1, cfg.save_every_n):
                        writer.save()
                        log(f"Checkpoint save completed after {done_q}/{len(jobs)} jobs")
                        pending_save_count = 0
            finally:
                for d in open_docs.values():
                    try:
                        d.close()
                    except Exception:
                        pass

            if writer:
                writer.close()
            self.queue_msg("status", "Done")
            self.queue_msg("done_ok", f"Extraction finished.\nExcel saved to:\n{cfg.output_excel}\nLog saved to:\n{log_txt_path}\nCheckpoint:\n{checkpoint_path}\nFailed jobs:\n{failed_path}")
        except Exception as e:
            if writer:
                try:
                    writer.close()
                except Exception:
                    pass
            self.queue_msg("done_err", str(e))
            try:
                with open(log_txt_path, "a", encoding="utf-8") as f:
                    f.write("\nFATAL ERROR:\n" + traceback.format_exc())
            except Exception:
                pass


def main() -> None:
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
