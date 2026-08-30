from __future__ import annotations

from collections import Counter
from datetime import datetime, timezone
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from backend.app.models import Answer, Job, LocalBatch, LocalBatchItem, QuestionnaireGroup, ReviewEvent
from backend.app.scanner.fusion import normalized


SHEETS = (
    "Questionnaires",
    "Responses",
    "Long_Answers",
    "Page_Extracts",
    "Conflicts",
    "Failed_Jobs",
    "Grouping",
    "Data_Dictionary",
    "QA_Summary",
    "Data_Analysis",
    "Run_Log",
    "Run_Settings",
    "Reasonableness",
    "Review_Audit",
)


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _language_for(answers: Iterable[Answer]) -> str:
    text = " ".join(f"{item.question_text} {item.final_value or ''}" for item in answers)
    has_cjk = any("\u3400" <= char <= "\u9fff" for char in text)
    has_latin = any(char.isascii() and char.isalpha() for char in text)
    if has_cjk and has_latin:
        return "Mixed / 中英混合"
    if has_cjk:
        return "Chinese / 中文"
    if has_latin:
        return "English / 英文"
    return "Unknown / 未知"


def _flag_for(answer: Answer) -> str:
    flags: list[str] = []
    if answer.review_status == "pending":
        flags.append("REVIEW_REQUIRED")
    if answer.final_source == "qwen_judge":
        flags.append("QWEN_CORRECTION_PENDING_REVIEW")
    if answer.reasonableness_status not in {None, "reasonable", "not_checked"}:
        flags.append(str(answer.reasonableness_status).upper())
    if (
        answer.verifier_value is not None
        and answer.qwen_value is not None
        and normalized(answer.verifier_value) != normalized(answer.qwen_value)
    ):
        flags.append("PRIMARY_VERIFIER_CONFLICT")
    return "; ".join(dict.fromkeys(flags)) or "OK"


def _is_conflict(answer: Answer) -> bool:
    """Keep the conflict sheet for evidence disagreements, not skipped verification."""

    if answer.question_id.startswith("PAGE-") and answer.question_id.endswith("-EXTRACTION-ERROR"):
        return True
    primary = answer.qwen_value
    verifier = answer.verifier_value
    if primary is not None and verifier is not None and normalized(primary) != normalized(verifier):
        return True
    geometry = answer.geometry_value
    if geometry is not None and answer.scanner_value is not None:
        if normalized(geometry) != normalized(answer.scanner_value):
            return True
    reason = (answer.fusion_reason or "").casefold()
    return any(word in reason for word in ("disagree", "conflict", "multiple mark", "overwrite"))


def _write_rows(sheet, headers: list[str], rows: Iterable[dict[str, Any]]) -> int:
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="17365D")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    count = 0
    for row in rows:
        sheet.append([_cell_value(row.get(header)) for header in headers])
        count += 1

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(headers, 1):
        values = [str(header)]
        for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2, max_row=min(sheet.max_row, 202)):
            values.extend("" if item.value is None else str(item.value) for item in cell)
        width = min(55, max(10, max(len(value) for value in values) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width
    return count


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        # Questionnaire/model text is untrusted. Prevent Excel formula injection while
        # keeping the displayed value intact.
        return "'" + value
    return value


def _answer_row(
    *,
    item: LocalBatchItem,
    job: Job,
    group: QuestionnaireGroup,
    answer: Answer,
    language: str,
    series_label: str,
    series_questionnaire_index: int,
) -> dict[str, Any]:
    evidence = answer.evidence or []
    primary_evidence = [
        entry
        for entry in evidence
        if entry.get("source") in {"qwen", "primary_vision", "primary_adjudicator"}
    ]
    verifier_evidence = [entry for entry in evidence if entry.get("source") == "verifier_vision"]
    return {
        "Series_Label": series_label,
        "Series_Questionnaire_Index": series_questionnaire_index,
        "Source_File": Path(item.original_path).name,
        "Source_File_Index": item.order_index + 1,
        "Questionnaire_Index": group.group_index + 1,
        "Page_Range": f"{group.start_page}-{group.end_page}",
        "Participant_ID": group.participant_id,
        "Detected_Language": language,
        "Answer_ID": answer.id,
        "Answer_Key": answer.answer_key,
        "Template_Question_ID": answer.template_question_id,
        "Question_ID": answer.question_id,
        "Question_Text": answer.question_text,
        "Answer_Type": answer.answer_type,
        "Allowed_Options": answer.allowed_options,
        "Selected_Options": answer.selected_options,
        "Primary_Model_ID": (job.profile_snapshot or {}).get("extractor_model_id"),
        "Verifier_Model_ID": answer.verifier_model_id
        or (job.profile_snapshot or {}).get("verifier_model_id"),
        "Primary_Model_Value": answer.qwen_value,
        "Verifier_Model_Value": answer.verifier_value,
        "Geometry_Value": answer.geometry_value,
        "Geometry_Confidence": answer.geometry_confidence,
        "Primary_Model_Evidence": primary_evidence,
        "Verifier_Model_Evidence": verifier_evidence,
        "Scanner_Value_Immutable": answer.scanner_value,
        "Scanner_Confidence": answer.scanner_confidence,
        "Fusion_Reason": answer.fusion_reason,
        "Reasonableness_Status": answer.reasonableness_status,
        "Judge_Suggestion": answer.judge_suggestion,
        "Judge_Reason": answer.judge_reason,
        "Rule_References": answer.rule_refs,
        "Final_Value": answer.final_value,
        "Final_Source": answer.final_source,
        "Review_Status": answer.review_status,
        "Flag_Status": _flag_for(answer),
        "Source_Page": answer.page_number,
        "Page_Ordinal": answer.page_ordinal,
        "Updated_At": answer.updated_at,
    }


def write_series_excel(
    db: Session,
    batch: LocalBatch,
    items: list[LocalBatchItem],
    series_label: str,
    destination: str | Path,
) -> dict[str, Any]:
    """Atomically write one workbook containing every source in a labelled series."""

    if not items:
        raise ValueError("A series workbook requires at least one source item")
    items = sorted(items, key=lambda value: value.order_index)

    destination = Path(destination).expanduser().resolve()
    if destination.suffix.lower() != ".xlsx":
        destination = destination.with_suffix(".xlsx")
    destination.parent.mkdir(parents=True, exist_ok=True)

    item_by_job = {item.job_id: item for item in items if item.job_id}
    jobs = list(db.scalars(select(Job).where(Job.id.in_(list(item_by_job)))).all()) if item_by_job else []
    jobs.sort(key=lambda job: item_by_job[job.id].order_index)

    questionnaire_rows: list[dict[str, Any]] = []
    answer_rows: list[dict[str, Any]] = []
    page_rows: list[dict[str, Any]] = []
    conflict_rows: list[dict[str, Any]] = []
    flagged_rows: list[dict[str, Any]] = []
    reason_rows: list[dict[str, Any]] = []
    audit_rows: list[dict[str, Any]] = []
    grouping_rows: list[dict[str, Any]] = []
    response_records: list[dict[str, Any]] = []
    dictionary_by_id: dict[str, dict[str, Any]] = {}
    source_stats: Counter[str] = Counter()
    series_questionnaire_index = 0

    for item in items:
        item_job = next((candidate for candidate in jobs if candidate.id == item.job_id), None)
        item_groups = (
            list(
                db.scalars(
                    select(QuestionnaireGroup)
                    .where(QuestionnaireGroup.job_id == item_job.id)
                    .order_by(QuestionnaireGroup.group_index.asc())
                ).all()
            )
            if item_job
            else []
        )
        if item_groups:
            for group in item_groups:
                grouping_rows.append(
                    {
                        "Series_Label": series_label,
                        "Source_File": Path(item.original_path).name,
                        "Source_File_Index": item.order_index + 1,
                        "Expected_Questionnaires": item.expected_questionnaires,
                        "Detected_Questionnaires": len(item_groups),
                        "Detected_Cycle_Pages": item.pages_per_questionnaire,
                        "Questionnaire_Index": group.group_index + 1,
                        "Page_Range": f"{group.start_page}-{group.end_page}",
                        "Participant_ID": group.participant_id,
                        "Grouping_Confidence": item.grouping_confidence or group.confidence,
                        "Template_Variant": item.template_variant,
                        "Status": item.status,
                        "Reason_or_Skip": item.error or item.grouping_reason or group.reason,
                    }
                )
        else:
            grouping_rows.append(
                {
                    "Series_Label": series_label,
                    "Source_File": Path(item.original_path).name,
                    "Source_File_Index": item.order_index + 1,
                    "Expected_Questionnaires": item.expected_questionnaires,
                    "Detected_Questionnaires": 0,
                    "Detected_Cycle_Pages": item.pages_per_questionnaire,
                    "Questionnaire_Index": None,
                    "Page_Range": None,
                    "Participant_ID": None,
                    "Grouping_Confidence": item.grouping_confidence,
                    "Template_Variant": item.template_variant,
                    "Status": item.status,
                    "Reason_or_Skip": item.error or item.grouping_reason,
                }
            )

    for job in jobs:
        item = item_by_job[job.id]
        groups = list(
            db.scalars(
                select(QuestionnaireGroup)
                .where(QuestionnaireGroup.job_id == job.id)
                .order_by(QuestionnaireGroup.group_index.asc())
            ).all()
        )
        for group in groups:
            series_questionnaire_index += 1
            answers = list(
                db.scalars(
                    select(Answer)
                    .where(Answer.group_id == group.id)
                    .order_by(Answer.page_number.asc(), Answer.question_id.asc())
                ).all()
            )
            language = _language_for(answers)
            flags = sum(_flag_for(answer) != "OK" for answer in answers)
            source_name = Path(item.original_path).name
            source_stats[source_name] += len(answers)
            questionnaire_rows.append(
                {
                    "Series_Label": series_label,
                    "Series_Questionnaire_Index": series_questionnaire_index,
                    "Source_File": source_name,
                    "Source_File_Index": item.order_index + 1,
                    "Questionnaire_Index": group.group_index + 1,
                    "Page_Range": f"{group.start_page}-{group.end_page}",
                    "Participant_ID": group.participant_id,
                    "Detected_Language": language,
                    "Answers": len(answers),
                    "Flagged_Answers": flags,
                    "Job_Status": job.status,
                    "Template_Variant": item.template_variant,
                    "Pages_Per_Questionnaire": group.end_page - group.start_page + 1,
                }
            )
            response_record: dict[str, Any] = {
                "Series_Label": series_label,
                "Series_Questionnaire_Index": series_questionnaire_index,
                "Source_File": source_name,
                "Source_File_Index": item.order_index + 1,
                "Questionnaire_Index": group.group_index + 1,
                "Page_Range": f"{group.start_page}-{group.end_page}",
                "Participant_ID": group.participant_id,
                "Detected_Language": language,
                "Flagged_Answers": flags,
            }
            for answer in answers:
                row = _answer_row(
                    item=item,
                    job=job,
                    group=group,
                    answer=answer,
                    language=language,
                    series_label=series_label,
                    series_questionnaire_index=series_questionnaire_index,
                )
                answer_rows.append(row)
                stable_id = answer.template_question_id or answer.question_id
                response_record[stable_id] = answer.scanner_value
                dictionary_by_id.setdefault(
                    stable_id,
                    {
                        "Template_Question_ID": stable_id,
                        "Page_Ordinal": answer.page_ordinal,
                        "Question_ID": answer.question_id,
                        "Question_Text": answer.question_text,
                        "Answer_Type": answer.answer_type,
                        "Allowed_Values": answer.allowed_options,
                        "Template_Variant": item.template_variant,
                    },
                )
                page_rows.append(
                    {
                        key: row[key]
                        for key in (
                            "Series_Label",
                            "Series_Questionnaire_Index",
                            "Source_File",
                            "Source_File_Index",
                            "Questionnaire_Index",
                            "Page_Range",
                            "Participant_ID",
                            "Detected_Language",
                            "Source_Page",
                            "Question_ID",
                            "Question_Text",
                            "Primary_Model_Evidence",
                            "Verifier_Model_Evidence",
                            "Scanner_Value_Immutable",
                            "Final_Value",
                            "Flag_Status",
                        )
                    }
                )
                if row["Flag_Status"] != "OK":
                    flagged_rows.append(row)
                if _is_conflict(answer):
                    conflict_rows.append(row)
                reason_rows.append(
                    {
                        key: row[key]
                        for key in (
                            "Series_Label",
                            "Series_Questionnaire_Index",
                            "Source_File",
                            "Source_File_Index",
                            "Questionnaire_Index",
                            "Participant_ID",
                            "Question_ID",
                            "Question_Text",
                            "Scanner_Value_Immutable",
                            "Reasonableness_Status",
                            "Judge_Suggestion",
                            "Judge_Reason",
                            "Rule_References",
                            "Final_Value",
                            "Final_Source",
                            "Flag_Status",
                        )
                    }
                )

            response_records.append(response_record)

            answer_by_id = {answer.id: answer for answer in answers}
            events = (
                list(
                    db.scalars(
                        select(ReviewEvent)
                        .where(ReviewEvent.answer_id.in_(list(answer_by_id)))
                        .order_by(ReviewEvent.created_at.asc())
                    ).all()
                )
                if answer_by_id
                else []
            )
            for event in events:
                answer = answer_by_id.get(event.answer_id)
                audit_rows.append(
                    {
                        "Series_Label": series_label,
                        "Series_Questionnaire_Index": series_questionnaire_index,
                        "Source_File": source_name,
                        "Source_File_Index": item.order_index + 1,
                        "Questionnaire_Index": group.group_index + 1,
                        "Participant_ID": group.participant_id,
                        "Question_ID": answer.question_id if answer else None,
                        "Action": event.action,
                        "Old_Value": event.previous_value,
                        "New_Value": event.new_value,
                        "Comment": event.comment,
                        "Created_At": event.created_at,
                    }
                )

    failed_rows = [
        {
            "Series_Label": series_label,
            "Source_File": Path(item.original_path).name,
            "Source_File_Index": item.order_index + 1,
            "Status": item.status,
            "Error": item.error,
            "Updated_At": item.updated_at,
        }
        for item in items
        if item.status in {"failed", "export_failed", "skipped_grouping"}
    ]

    unresolved = len(flagged_rows)
    skipped_inputs = sum(item.status == "skipped_grouping" for item in items)
    incomplete = sum(
        item.status not in {"completed", "failed", "export_failed", "skipped_grouping"}
        for item in items
    )
    status_label = (
        "IN PROGRESS — PARTIAL CHECKPOINT"
        if incomplete
        else "COMPLETED — FLAGS PRESENT"
        if unresolved or failed_rows
        else "COMPLETED"
    )
    expected_questionnaires = sum(
        item.expected_questionnaires or 0 for item in items if item.status != "skipped_grouping"
    )
    expected_questionnaires = expected_questionnaires or len(questionnaire_rows)
    extraction_error_answers = [
        row for row in answer_rows if str(row.get("Question_ID") or "").endswith("-EXTRACTION-ERROR")
    ]
    successful_pages = len(
        {
            (row["Source_File_Index"], row["Source_Page"])
            for row in answer_rows
            if not str(row.get("Question_ID") or "").endswith("-EXTRACTION-ERROR")
        }
    )
    failed_pages = len(
        {(row["Source_File_Index"], row["Source_Page"]) for row in extraction_error_answers}
    )
    expected_pages = sum(
        group.end_page - group.start_page + 1
        for job in jobs
        for group in db.scalars(
            select(QuestionnaireGroup).where(QuestionnaireGroup.job_id == job.id)
        ).all()
    )
    normal_answers = [
        row for row in answer_rows if not str(row.get("Question_ID") or "").endswith("-EXTRACTION-ERROR")
    ]
    verifier_values = sum(row.get("Verifier_Model_Value") is not None for row in normal_answers)
    expected_answers = 0
    for job in jobs:
        snapshot = dict(job.profile_snapshot or {})
        pages = (snapshot.get("series_template_v1") or {}).get("pages") or {}
        schema_answer_count = sum(len(values) for values in pages.values() if isinstance(values, list))
        if schema_answer_count:
            job_group_count = len(
                db.scalars(select(QuestionnaireGroup).where(QuestionnaireGroup.job_id == job.id)).all()
            )
            expected_answers += schema_answer_count * job_group_count
    expected_answers = expected_answers or len(normal_answers)
    runtime_seconds = 0.0
    for item in items:
        if item.started_at and item.finished_at:
            runtime_seconds += max(0.0, (item.finished_at - item.started_at).total_seconds())
    qa_rows = [
        {"Metric": "Workbook_Status", "Value": status_label},
        {"Metric": "Series_Label", "Value": series_label},
        {"Metric": "Source_Files", "Value": len(items)},
        {"Metric": "Expected_Questionnaires", "Value": expected_questionnaires},
        {"Metric": "Questionnaires", "Value": len(questionnaire_rows)},
        {"Metric": "Successful_Pages", "Value": successful_pages},
        {"Metric": "Failed_Pages", "Value": failed_pages},
        {"Metric": "Expected_Pages", "Value": expected_pages},
        {"Metric": "Expected_Answers", "Value": expected_answers},
        {"Metric": "Answers", "Value": len(normal_answers)},
        {"Metric": "Flagged_Answers", "Value": unresolved},
        {"Metric": "Genuine_Conflicts", "Value": len(conflict_rows)},
        {"Metric": "Verifier_Answers", "Value": verifier_values},
        {
            "Metric": "Verifier_Coverage_Percent",
            "Value": round(100 * verifier_values / max(1, len(normal_answers)), 2),
        },
        {"Metric": "Failed_Inputs", "Value": len(failed_rows)},
        {"Metric": "Skipped_Inputs", "Value": skipped_inputs},
        {"Metric": "Incomplete_Inputs", "Value": incomplete},
        {"Metric": "Reasonableness_Suggestions", "Value": sum(row.get("Judge_Suggestion") is not None for row in answer_rows)},
        {"Metric": "Runtime_Seconds", "Value": round(runtime_seconds, 1)},
        {
            "Metric": "Throughput_Pages_Per_Minute",
            "Value": round(successful_pages / (runtime_seconds / 60), 2) if runtime_seconds else None,
        },
        {"Metric": "Batch_ID", "Value": batch.id},
        {"Metric": "Generated_At_UTC", "Value": _utcnow().isoformat()},
        {"Metric": "Open_Responses", "Value": "Go to questionnaire rows"},
        {"Metric": "Open_Grouping", "Value": "Go to grouping audit"},
        {"Metric": "Open_Conflicts", "Value": "Go to genuine disagreements"},
    ]
    analysis_rows = [
        {"Source_File": name, "Answer_Count": count}
        for name, count in sorted(source_stats.items(), key=lambda pair: pair[0].casefold())
        if count > 0
    ]
    run_rows = [
        {
            "Batch_ID": batch.id,
            "Series_Label": series_label,
            "Source_File": Path(item.original_path).name,
            "Source_File_Index": item.order_index + 1,
            "Status": item.status,
            "Error": item.error,
            "Started_At": item.started_at,
            "Finished_At": item.finished_at,
            "Vision_Model": batch.extractor_model_id,
            "Verifier_Model": batch.verifier_model_id,
            "Judge_Model": batch.judge_model_id,
            "LM_Studio": batch.lmstudio_base_url,
        }
        for item in items
    ]
    settings_rows = [
        {"Setting": "FormSight_Version", "Value": "0.6.0"},
        {"Setting": "Processing_Mode", "Value": batch.processing_mode},
        {"Setting": "Automatic_Safe_Skip", "Value": not batch.review_groups},
        {"Setting": "LM_Studio", "Value": batch.lmstudio_base_url},
        {"Setting": "Primary_Vision_Model", "Value": batch.extractor_model_id},
        {"Setting": "Selective_Verifier_Model", "Value": batch.verifier_model_id},
        {"Setting": "Reasonableness_Model", "Value": batch.judge_model_id},
        {"Setting": "Reasonableness_Policy", "Value": "FLAG_ONLY"},
        {"Setting": "Max_Page_Attempts", "Value": 1},
        {"Setting": "Normal_Request_Timeout_Seconds", "Value": 90},
    ]

    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEETS:
        wb.create_sheet(name)
    wb.properties.title = f"FormSight Local — {status_label}"
    wb.properties.subject = f"Questionnaire series: {series_label}"
    wb.properties.creator = "FormSight Local"

    questionnaire_headers = [
        "Series_Label",
        "Series_Questionnaire_Index",
        "Source_File",
        "Source_File_Index",
        "Questionnaire_Index",
        "Page_Range",
        "Participant_ID",
        "Detected_Language",
        "Answers",
        "Flagged_Answers",
        "Job_Status",
        "Template_Variant",
        "Pages_Per_Questionnaire",
    ]
    answer_headers = [
        "Series_Label",
        "Series_Questionnaire_Index",
        "Source_File",
        "Source_File_Index",
        "Questionnaire_Index",
        "Page_Range",
        "Participant_ID",
        "Detected_Language",
        "Answer_ID",
        "Answer_Key",
        "Template_Question_ID",
        "Question_ID",
        "Question_Text",
        "Answer_Type",
        "Allowed_Options",
        "Selected_Options",
        "Primary_Model_ID",
        "Verifier_Model_ID",
        "Primary_Model_Value",
        "Verifier_Model_Value",
        "Geometry_Value",
        "Geometry_Confidence",
        "Primary_Model_Evidence",
        "Verifier_Model_Evidence",
        "Scanner_Value_Immutable",
        "Scanner_Confidence",
        "Fusion_Reason",
        "Reasonableness_Status",
        "Judge_Suggestion",
        "Judge_Reason",
        "Rule_References",
        "Final_Value",
        "Final_Source",
        "Review_Status",
        "Flag_Status",
        "Source_Page",
        "Page_Ordinal",
        "Updated_At",
    ]
    page_headers = [
        "Series_Label",
        "Series_Questionnaire_Index",
        "Source_File",
        "Source_File_Index",
        "Questionnaire_Index",
        "Page_Range",
        "Participant_ID",
        "Detected_Language",
        "Source_Page",
        "Question_ID",
        "Question_Text",
        "Primary_Model_Evidence",
        "Verifier_Model_Evidence",
        "Scanner_Value_Immutable",
        "Final_Value",
        "Flag_Status",
    ]
    reason_headers = [
        "Series_Label",
        "Series_Questionnaire_Index",
        "Source_File",
        "Source_File_Index",
        "Questionnaire_Index",
        "Participant_ID",
        "Question_ID",
        "Question_Text",
        "Scanner_Value_Immutable",
        "Reasonableness_Status",
        "Judge_Suggestion",
        "Judge_Reason",
        "Rule_References",
        "Final_Value",
        "Final_Source",
        "Flag_Status",
    ]
    response_base_headers = [
        "Series_Label",
        "Series_Questionnaire_Index",
        "Source_File",
        "Source_File_Index",
        "Questionnaire_Index",
        "Page_Range",
        "Participant_ID",
        "Detected_Language",
        "Flagged_Answers",
    ]
    stable_headers = [
        key
        for key, _ in sorted(
            dictionary_by_id.items(),
            key=lambda pair: (
                pair[1].get("Page_Ordinal") or 999999,
                str(pair[0]).casefold(),
            ),
        )
    ]
    grouping_headers = [
        "Series_Label",
        "Source_File",
        "Source_File_Index",
        "Expected_Questionnaires",
        "Detected_Questionnaires",
        "Detected_Cycle_Pages",
        "Questionnaire_Index",
        "Page_Range",
        "Participant_ID",
        "Grouping_Confidence",
        "Template_Variant",
        "Status",
        "Reason_or_Skip",
    ]
    dictionary_headers = [
        "Template_Question_ID",
        "Page_Ordinal",
        "Question_ID",
        "Question_Text",
        "Answer_Type",
        "Allowed_Values",
        "Template_Variant",
    ]
    _write_rows(wb["Questionnaires"], questionnaire_headers, questionnaire_rows)
    _write_rows(wb["Responses"], response_base_headers + stable_headers, response_records)
    _write_rows(wb["Long_Answers"], answer_headers, answer_rows)
    _write_rows(wb["Page_Extracts"], page_headers, page_rows)
    _write_rows(wb["Conflicts"], answer_headers, conflict_rows)
    _write_rows(
        wb["Failed_Jobs"],
        ["Series_Label", "Source_File", "Source_File_Index", "Status", "Error", "Updated_At"],
        failed_rows,
    )
    _write_rows(wb["Grouping"], grouping_headers, grouping_rows)
    _write_rows(
        wb["Data_Dictionary"],
        dictionary_headers,
        [dictionary_by_id[key] for key in stable_headers],
    )
    _write_rows(wb["QA_Summary"], ["Metric", "Value"], qa_rows)
    _write_rows(wb["Data_Analysis"], ["Source_File", "Answer_Count"], analysis_rows)
    _write_rows(
        wb["Run_Log"],
        [
            "Batch_ID",
            "Series_Label",
            "Source_File",
            "Source_File_Index",
            "Status",
            "Error",
            "Started_At",
            "Finished_At",
            "Vision_Model",
            "Verifier_Model",
            "Judge_Model",
            "LM_Studio",
        ],
        run_rows,
    )
    _write_rows(wb["Run_Settings"], ["Setting", "Value"], settings_rows)
    _write_rows(wb["Reasonableness"], reason_headers, reason_rows)
    _write_rows(
        wb["Review_Audit"],
        [
            "Series_Label",
            "Series_Questionnaire_Index",
            "Source_File",
            "Source_File_Index",
            "Questionnaire_Index",
            "Participant_ID",
            "Question_ID",
            "Action",
            "Old_Value",
            "New_Value",
            "Comment",
            "Created_At",
        ],
        audit_rows,
    )

    # Compact navigation and visible flag styling make the workbook usable without
    # exposing the raw per-job profile JSON in the main data sheets.
    navigation = {
        "Open_Responses": "Responses",
        "Open_Grouping": "Grouping",
        "Open_Conflicts": "Conflicts",
    }
    for row_index in range(2, wb["QA_Summary"].max_row + 1):
        metric = wb["QA_Summary"].cell(row_index, 1).value
        target = navigation.get(str(metric))
        if target:
            cell = wb["QA_Summary"].cell(row_index, 2)
            cell.hyperlink = f"#'{target}'!A1"
            cell.style = "Hyperlink"

    warning_fill = PatternFill("solid", fgColor="FFF0D5")
    conflict_fill = PatternFill("solid", fgColor="FDE2E1")
    for sheet_name in ("Long_Answers", "Reasonableness"):
        sheet = wb[sheet_name]
        header_map = {cell.value: cell.column for cell in sheet[1]}
        flag_column = header_map.get("Flag_Status")
        if flag_column:
            for row_index in range(2, sheet.max_row + 1):
                if sheet.cell(row_index, flag_column).value not in {None, "OK"}:
                    sheet.cell(row_index, flag_column).fill = warning_fill
                    sheet.cell(row_index, flag_column).font = Font(color="9C5700", bold=True)
    for row in wb["Conflicts"].iter_rows(min_row=2):
        for cell in row:
            cell.fill = conflict_fill
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.row_dimensions[1].height = 32
    wb["QA_Summary"].sheet_properties.tabColor = "117D65"
    wb["Conflicts"].sheet_properties.tabColor = "C43D3D"
    wb["Grouping"].sheet_properties.tabColor = "3973B8"

    if analysis_rows:
        chart = BarChart()
        chart.title = "Answers by source file"
        chart.y_axis.title = "Answers"
        chart.x_axis.title = "Source file"
        data = Reference(wb["Data_Analysis"], min_col=2, min_row=1, max_row=len(analysis_rows) + 1)
        labels = Reference(wb["Data_Analysis"], min_col=1, min_row=2, max_row=len(analysis_rows) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.height = 7
        chart.width = 14
        wb["Data_Analysis"].add_chart(chart, "D2")

    temp_path = destination.with_name(f".{destination.stem}.writing.xlsx")
    wb.save(temp_path)
    temp_path.replace(destination)
    return {
        "path": str(destination),
        "status": status_label,
        "source_files": len(items),
        "questionnaires": len(questionnaire_rows),
        "answers": len(answer_rows),
        "flags": unresolved,
        "failed": len(failed_rows),
    }


def write_source_excel(
    db: Session,
    batch: LocalBatch,
    item: LocalBatchItem,
    destination: str | Path,
) -> dict[str, Any]:
    """Backward-compatible wrapper for one-source workbooks."""

    return write_series_excel(
        db,
        batch,
        [item],
        item.series_label or Path(item.original_path).stem,
        destination,
    )
