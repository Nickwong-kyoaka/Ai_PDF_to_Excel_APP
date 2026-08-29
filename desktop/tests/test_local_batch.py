from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
from sqlalchemy import select

from backend.app.models import Answer, LocalBatch, LocalBatchItem, QuestionnaireGroup
from desktop.model_discovery import DetectedModel, DiscoveryResult
from desktop.runner import LocalBatchRunner
from desktop.runtime import create_runtime


def ready_discovery() -> DiscoveryResult:
    vision = DetectedModel(
        api_id="qwen-vision-loaded",
        key="qwen/qwen3-vl-8b",
        display_name="Qwen3 VL 8B",
        architecture="qwen3_vl",
        quantization="Q4",
        context_length=32768,
        vision=True,
        score=400,
    )
    verifier = DetectedModel(
        api_id="gemma-verifier-loaded",
        key="google/gemma-3-4b",
        display_name="Gemma 3 4B",
        architecture="gemma3",
        quantization="Q4",
        context_length=8192,
        vision=True,
        score=10,
    )
    return DiscoveryResult(
        status="ready",
        base_url="http://127.0.0.1:1234",
        port=1234,
        vision_models=[vision, verifier],
        selected_vision=vision,
        selected_verifier=verifier,
        selected_judge=vision,
    )


class FakeYolo:
    def release(self) -> None:
        pass


class FakeExtractor:
    def __init__(self, settings, profile, weights=None, **kwargs):
        assert profile["extractor_model_id"] == "qwen-vision-loaded"
        assert profile["verifier_model_id"] == "gemma-verifier-loaded"
        assert profile["judge_model_id"] == "qwen-vision-loaded"
        self.yolo = FakeYolo()

    def extract_job(self, db, job) -> None:
        group = db.scalar(
            select(QuestionnaireGroup).where(QuestionnaireGroup.job_id == job.id)
        )
        db.add(
            Answer(
                job_id=job.id,
                group_id=group.id,
                page_number=1,
                question_id="Q1",
                question_text="你今天好嗎？ / Are you well today?",
                answer_type="single_choice",
                allowed_options=["Yes", "No"],
                selected_options=["Yes"],
                qwen_value="Yes",
                yolo_value=None,
                verifier_value="Yes",
                verifier_model_id="gemma-verifier-loaded",
                scanner_value="Yes",
                scanner_confidence=0.88,
                fusion_reason="Primary and independent verifier vision models agree",
                final_value="Yes",
                final_source="scanner",
                review_status="pending",
            )
        )
        db.commit()

    def judge_job(self, db, job) -> None:
        answer = db.scalar(select(Answer).where(Answer.job_id == job.id))
        answer.reasonableness_status = "review_required"
        answer.judge_reason = "Subjective answer retained"
        db.commit()


def test_mixed_batch_writes_one_corresponding_workbook_per_input(tmp_path, monkeypatch) -> None:
    local_app_data = tmp_path / "local-data"
    monkeypatch.setenv("LOCALAPPDATA", str(local_app_data))
    runtime = create_runtime("http://127.0.0.1:1234")
    monkeypatch.setattr("desktop.runner.QuestionnaireExtractor", FakeExtractor)

    first = tmp_path / "中文問卷.png"
    second = tmp_path / "english.jpg"
    broken = tmp_path / "broken.pdf"
    Image.new("RGB", (120, 160), "white").save(first)
    Image.new("RGB", (120, 160), "white").save(second)
    broken.write_bytes(b"this is not a pdf")
    output = tmp_path / "輸出結果"

    runner = LocalBatchRunner(runtime)
    batch_id = runner.create_batch(
        [first, second, broken], output, ready_discovery(), review_groups=False
    )
    result = runner.execute_batch(batch_id)

    assert result is not None
    assert result["output_directory"] == str(output.resolve())
    assert result["source_files"] == 3
    assert result["workbooks"] == 3
    assert result["failed"] == 1
    workbooks = sorted(output.glob("*.xlsx"))
    assert [path.name for path in workbooks] == [
        "broken_FormSight.xlsx",
        "english_FormSight.xlsx",
        "中文問卷_FormSight.xlsx",
    ]

    workbook = load_workbook(output / "中文問卷_FormSight.xlsx", data_only=True)
    assert workbook.sheetnames == [
        "Questionnaires",
        "Long_Answers",
        "Page_Extracts",
        "Conflicts",
        "Failed_Jobs",
        "QA_Summary",
        "Data_Analysis",
        "Run_Log",
        "Reasonableness",
        "Review_Audit",
    ]
    qa = {row[0]: row[1] for row in workbook["QA_Summary"].iter_rows(min_row=2, values_only=True)}
    assert qa["Workbook_Status"] == "COMPLETED — FLAGS PRESENT"
    headers = [cell.value for cell in workbook["Long_Answers"][1]]
    scanner_col = headers.index("Scanner_Value_Immutable") + 1
    source_col = headers.index("Source_File") + 1
    verifier_col = headers.index("Verifier_Model_Value") + 1
    assert workbook["Long_Answers"].cell(2, scanner_col).value == "Yes"
    assert workbook["Long_Answers"].cell(2, verifier_col).value == "Yes"
    assert "YOLO_Value" not in headers
    sources = {
        row[source_col - 1]
        for row in workbook["Long_Answers"].iter_rows(min_row=2, values_only=True)
    }
    assert sources == {first.name}

    english_workbook = load_workbook(output / "english_FormSight.xlsx", data_only=True)
    english_sources = {
        row[source_col - 1]
        for row in english_workbook["Long_Answers"].iter_rows(min_row=2, values_only=True)
    }
    assert english_sources == {second.name}

    failed_workbook = load_workbook(output / "broken_FormSight.xlsx", data_only=True)
    failed_rows = list(failed_workbook["Failed_Jobs"].iter_rows(min_row=2, values_only=True))
    assert len(failed_rows) == 1
    assert failed_rows[0][0] == broken.stem
    assert failed_rows[0][1] == broken.name

    with runtime.sessions() as db:
        batch = db.get(LocalBatch, batch_id)
        assert batch.status == "completed"
        items = list(
            db.scalars(
                select(LocalBatchItem)
                .where(LocalBatchItem.batch_id == batch_id)
                .order_by(LocalBatchItem.order_index)
            ).all()
        )
        assert [item.status for item in items] == ["completed", "completed", "failed"]
        assert all(item.output_path for item in items)


def test_same_stem_inputs_receive_distinct_stable_output_names(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("LOCALAPPDATA", str(tmp_path / "local-data"))
    runtime = create_runtime("http://127.0.0.1:1234")
    first = tmp_path / "survey.png"
    second = tmp_path / "survey.jpg"
    Image.new("RGB", (80, 100), "white").save(first)
    Image.new("RGB", (80, 100), "white").save(second)

    batch_id = LocalBatchRunner(runtime).create_batch(
        [first, second], tmp_path / "outputs", ready_discovery(), review_groups=False
    )

    with runtime.sessions() as db:
        items = list(
            db.scalars(
                select(LocalBatchItem)
                .where(LocalBatchItem.batch_id == batch_id)
                .order_by(LocalBatchItem.order_index)
            ).all()
        )
        assert [Path(item.output_path).name for item in items] == [
            "survey_FormSight.xlsx",
            "survey_2_FormSight.xlsx",
        ]


def test_same_series_label_consolidates_multiple_sources_into_one_workbook(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("LOCALAPPDATA", str(tmp_path / "local-data"))
    runtime = create_runtime("http://127.0.0.1:1234")
    monkeypatch.setattr("desktop.runner.QuestionnaireExtractor", FakeExtractor)
    first = tmp_path / "part-1.png"
    second = tmp_path / "part-2.jpg"
    broken = tmp_path / "part-3.pdf"
    Image.new("RGB", (80, 100), "white").save(first)
    Image.new("RGB", (80, 100), "white").save(second)
    broken.write_bytes(b"not a PDF")
    output = tmp_path / "outputs"

    runner = LocalBatchRunner(runtime)
    batch_id = runner.create_batch(
        [first, second, broken],
        output,
        ready_discovery(),
        review_groups=False,
        series_labels=["Study A", "study a", "Study A"],
    )
    result = runner.execute_batch(batch_id)

    assert result is not None
    assert result["workbooks"] == 1
    assert result["source_files"] == 3
    workbook_path = output / "Study A_FormSight.xlsx"
    assert list(output.glob("*.xlsx")) == [workbook_path]
    workbook = load_workbook(workbook_path, data_only=True)
    qa = {row[0]: row[1] for row in workbook["QA_Summary"].iter_rows(min_row=2, values_only=True)}
    assert qa["Series_Label"] == "Study A"
    assert qa["Source_Files"] == 3
    assert qa["Questionnaires"] == 2
    assert qa["Failed_Inputs"] == 1
    headers = [cell.value for cell in workbook["Long_Answers"][1]]
    assert "Series_Label" in headers
    assert "Series_Questionnaire_Index" in headers
    source_column = headers.index("Source_File")
    sources = {
        row[source_column]
        for row in workbook["Long_Answers"].iter_rows(min_row=2, values_only=True)
    }
    assert sources == {first.name, second.name}
    failed = list(workbook["Failed_Jobs"].iter_rows(min_row=2, values_only=True))
    assert len(failed) == 1
    assert failed[0][0] == "Study A"
    assert failed[0][1] == broken.name
