"""
PDF -> Excel (NO duplicate PID + PedEyeQ1-40) [FAST + ZERO manual, PID from filename]

Fixes based on your log
-----------------------
1) BUG FIX:
   - DEFAULT_ANSWER_IF_UNCLEAR was missing -> caused ALL [PED] failed.

2) PID robustness:
   - Supports ranges: A151-A165, A170-233, etc.
   - Supports comma lists: B315,331,332.pdf / B390,394,397.pdf / B430,444,448.pdf
     -> PID list = [B315, B331, B332] (same letter as first one)
   - If PDF contains more questionnaires than IDs in filename:
     -> continue sequentially upward after last ID (still no unknown)
   - If fewer questionnaires than IDs:
     -> only use first N ids

3) No duplicates:
   - Global seen_pids enforced, auto-shift upward if collision

Answers:
- OpenCV passes = 4
- Row-band recheck = ON
- Row "OCR" fallback (density-based) = up to 4 tries
- Remaining unclear -> deterministic RANDOM among 0/1/2

Install
-------
pip install pymupdf pillow numpy pandas openpyxl opencv-python pytesseract
"""

import os
import re
import threading
import traceback
import hashlib
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple, Set

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import numpy as np
import pandas as pd
import cv2
from PIL import Image

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

try:
    import pytesseract
except Exception:
    pytesseract = None


# ============================================================
# SETTINGS
# ============================================================

SHEET_NAME = "All Participants"

PAGES_PER_QUESTIONNAIRE = 8

# Only pages 6-8 contain PedEyeQ answers (0-based indices)
PED_PAGE_IDX = [5, 6, 7]
PED_RANGES = [
    (1, 9, False),    # page 6: Q1-9
    (10, 26, False),  # page 7: Q10-26
    (27, 40, True),   # page 8: Q27-40 (often rotated)
]

ANS_COLUMNS = ["Participant ID"] + [f"PedEyeQ{i}" for i in range(1, 41)]
REF_COLUMNS = ["Source PDF", "Source Pages"]

# Answer policy
ANS_PASSES = 4
ROW_BAND_RECHECK = True
ROW_FALLBACK_TRIES = 4

# IMPORTANT BUG FIX: define this
DEFAULT_ANSWER_IF_UNCLEAR = "2"

# You asked random fallback instead of fixed 2
DEFAULT_RANDOM = True

# Tesseract path override (optional)
TESSERACT_CMD = ""

# Excel behavior
DEFAULT_OVERWRITE_OUTPUT = True


# ============================================================
# LOGGING
# ============================================================

def log_message(msg: str, log_widget: tk.Text = None):
    print(msg)
    if log_widget is not None:
        try:
            log_widget.configure(state="normal")
            log_widget.insert("end", msg + "\n")
            log_widget.see("end")
            log_widget.configure(state="disabled")
        except Exception:
            pass


# ============================================================
# EXCEL
# ============================================================

def build_columns(include_ref: bool) -> List[str]:
    return (REF_COLUMNS + list(ANS_COLUMNS)) if include_ref else list(ANS_COLUMNS)


def ensure_excel_with_header(excel_path: str, columns: List[str], sheet_name: str = SHEET_NAME):
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        if ws.max_row < 1 or ws["A1"].value is None:
            ws.append(columns)
        wb.save(excel_path)
        wb.close()
        return
    df0 = pd.DataFrame([], columns=columns)
    df0.to_excel(excel_path, index=False, sheet_name=sheet_name)


def append_rows_to_excel(excel_path: str, rows: List[Dict[str, str]], columns: List[str], sheet_name: str = SHEET_NAME):
    if not rows:
        return
    ensure_excel_with_header(excel_path, columns=columns, sheet_name=sheet_name)

    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    if ws.max_row < 1 or ws["A1"].value is None:
        ws.append(columns)

    df = pd.DataFrame(rows).reindex(columns=columns).fillna("N/A")
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    wb.save(excel_path)
    wb.close()


# ============================================================
# PDF RENDER
# ============================================================

def render_pdf_page(doc: "fitz.Document", page_index: int, dpi: int) -> Image.Image:
    page = doc.load_page(page_index)
    zoom = dpi / 72.0
    mat = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=mat, alpha=False)
    return Image.frombytes("RGB", (pix.width, pix.height), pix.samples)


def iter_questionnaire_groups(pdf_path: str, dpi: int, pages_per_group: int, log=None):
    if fitz is None:
        raise RuntimeError("PyMuPDF not installed. Install: pip install pymupdf")

    doc = fitz.open(pdf_path)
    n = doc.page_count
    if log:
        log(f" -> {n} page(s) found (streaming)")

    buf: List[Image.Image] = []
    start = 0
    for i in range(n):
        if not buf:
            start = i
        buf.append(render_pdf_page(doc, i, dpi))
        if len(buf) == pages_per_group:
            yield buf, start, i
            buf = []
    if buf:
        yield buf, start, start + len(buf) - 1

    doc.close()


# ============================================================
# PID FROM FILENAME (range OR comma list)
# ============================================================

RANGE_RE = re.compile(r"([AB])\s*0*([0-9]{1,6})\s*[-_]\s*(?:([AB])\s*)?0*([0-9]{1,6})", re.IGNORECASE)
SINGLE_RE = re.compile(r"\b([AB])\s*0*([0-9]{1,6})\b", re.IGNORECASE)

def normalize_pid(letter: str, num: int) -> str:
    return f"{letter.upper()}{num:03d}"

def pid_num(pid: str) -> int:
    m = SINGLE_RE.search(pid)
    return int(m.group(2)) if m else -1

def parse_pid_plan_from_filename(pdf_path: str) -> List[str]:
    """
    Returns a PID plan list extracted from filename:
    - Range A151-A165 -> [A151..A165]
    - A170-233 -> [A170..A233]
    - Single A760 -> [A760]
    - Comma list B315,331,332 -> [B315, B331, B332] (same letter as first)
    If both range and comma exist, range wins.
    """
    base = os.path.basename(pdf_path)
    name, _ = os.path.splitext(base)

    # 1) Range
    m = RANGE_RE.search(name)
    if m:
        letter = m.group(1).upper()
        start = int(m.group(2))
        letter2 = (m.group(3) or letter).upper()
        end = int(m.group(4))
        if letter2 != letter:
            letter2 = letter
        if start > end:
            start, end = end, start
        return [normalize_pid(letter, i) for i in range(start, end + 1)]

    # 2) Comma list / multiple singles: collect all numbers, keep first letter
    singles = list(SINGLE_RE.finditer(name))
    if singles:
        letter = singles[0].group(1).upper()
        nums = []
        for sm in singles:
            nums.append(int(sm.group(2)))
        # remove duplicates and sort by numeric
        nums = sorted(set(nums))
        return [normalize_pid(letter, n) for n in nums]

    raise ValueError(f"Filename has no PID info: {base}")

def pick_unique_pid_from_plan(pid_plan: List[str], group_idx0: int, seen: Set[str]) -> str:
    """
    Use pid_plan[group_idx0] if exists; else continue sequentially upward from last plan id.
    Enforce global uniqueness by shifting upward if collision.
    """
    if not pid_plan:
        raise ValueError("Empty pid_plan")

    # base candidate
    if group_idx0 < len(pid_plan):
        cand = pid_plan[group_idx0]
        letter = cand[0]
        n = pid_num(cand)
    else:
        last = pid_plan[-1]
        letter = last[0]
        n = pid_num(last) + (group_idx0 - (len(pid_plan) - 1))

    # enforce uniqueness
    while True:
        pid = normalize_pid(letter, n)
        if pid not in seen:
            return pid
        n += 1


# ============================================================
# ANSWERS (OpenCV + row-band + fallback)
# ============================================================

PE_ROI = (0.03, 0.35, 0.99, 0.82)

def clamp(v: int, lo: int, hi: int) -> int:
    return max(lo, min(hi, v))

def pil_to_gray(im_pil: Image.Image) -> np.ndarray:
    im = np.array(im_pil.convert("RGB"))
    return cv2.cvtColor(im, cv2.COLOR_RGB2GRAY)

def ensure_upright(gray: np.ndarray) -> np.ndarray:
    h, w = gray.shape[:2]
    if w > h * 1.1:
        return cv2.rotate(gray, cv2.ROTATE_90_CLOCKWISE)
    return gray

def crop_roi(gray: np.ndarray, roi: Tuple[float, float, float, float]) -> np.ndarray:
    h, w = gray.shape[:2]
    x1 = clamp(int(roi[0] * w), 0, w - 1)
    y1 = clamp(int(roi[1] * h), 0, h - 1)
    x2 = clamp(int(roi[2] * w), 1, w)
    y2 = clamp(int(roi[3] * h), 1, h)
    if x2 <= x1 + 10: x2 = min(w, x1 + 10)
    if y2 <= y1 + 10: y2 = min(h, y1 + 10)
    return gray[y1:y2, x1:x2].copy()

def morph_open(bw: np.ndarray, kx: int, ky: int, it: int = 1) -> np.ndarray:
    ker = cv2.getStructuringElement(cv2.MORPH_RECT, (kx, ky))
    return cv2.morphologyEx(bw, cv2.MORPH_OPEN, ker, iterations=it)

def morph_close(bw: np.ndarray, k: int = 3, it: int = 1) -> np.ndarray:
    ker = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (k, k))
    return cv2.morphologyEx(bw, cv2.MORPH_CLOSE, ker, iterations=it)

def adaptive_binarize(gray: np.ndarray, pass_idx: int) -> np.ndarray:
    g = cv2.GaussianBlur(gray, (3, 3), 0)
    if pass_idx == 0:
        block, C = 41, 13
    elif pass_idx == 1:
        block, C = 51, 11
    elif pass_idx == 2:
        block, C = 61, 9
    else:
        block, C = 71, 7
        g = cv2.fastNlMeansDenoising(g, None, 18, 7, 21)

    bw = cv2.adaptiveThreshold(
        g, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, block, C
    )
    bw = cv2.medianBlur(bw, 3)
    return morph_close(bw, 3, 1)

def line_positions_from_projection(proj: np.ndarray, thr: float, min_gap: int) -> List[int]:
    idxs: List[int] = []
    in_peak = False
    start = 0
    for i, v in enumerate(proj):
        if v > thr and not in_peak:
            in_peak = True
            start = i
        if in_peak and v <= thr:
            in_peak = False
            end = i
            idxs.append((start + end) // 2)
    if in_peak:
        idxs.append((start + (len(proj) - 1)) // 2)
    merged: List[int] = []
    for y in sorted(idxs):
        if not merged or abs(y - merged[-1]) > min_gap:
            merged.append(y)
    return merged

def detect_grid_lines(roi_bw: np.ndarray) -> Tuple[List[int], List[int]]:
    h, w = roi_bw.shape[:2]
    hx = max(20, w // 30)
    horiz = morph_open(roi_bw, hx, 1, it=1)
    proj_h = horiz.sum(axis=1).astype(np.float32)
    if proj_h.max() > 0:
        proj_h /= (proj_h.max() + 1e-6)
    h_lines = line_positions_from_projection(proj_h, thr=0.20, min_gap=8)

    vy = max(20, h // 30)
    vert = morph_open(roi_bw, 1, vy, it=1)
    proj_v = vert.sum(axis=0).astype(np.float32)
    if proj_v.max() > 0:
        proj_v /= (proj_v.max() + 1e-6)
    v_lines = line_positions_from_projection(proj_v, thr=0.20, min_gap=10)
    return h_lines, v_lines

def filled_score(bw: np.ndarray, cx: int, cy: int, r: int) -> float:
    h, w = bw.shape[:2]
    x1, y1 = max(0, cx - r), max(0, cy - r)
    x2, y2 = min(w, cx + r), min(h, cy + r)
    roi = bw[y1:y2, x1:x2]
    if roi.size == 0:
        return 0.0
    mask = np.zeros_like(roi, dtype=np.uint8)
    rr = min(r, roi.shape[0] // 2, roi.shape[1] // 2)
    cv2.circle(mask, (roi.shape[1] // 2, roi.shape[0] // 2), rr, 255, -1)
    ink = cv2.countNonZero(cv2.bitwise_and(roi, mask))
    area = cv2.countNonZero(mask)
    return float(ink) / float(area + 1e-6)

def best_score_local_search(bw: np.ndarray, cx: int, cy: int, r: int, search_px: int, step: int = 2) -> float:
    best = -1.0
    for dy in range(-search_px, search_px + 1, step):
        for dx in range(-search_px, search_px + 1, step):
            s = filled_score(bw, cx + dx, cy + dy, r)
            if s > best:
                best = s
    return float(best)

def pick_one_option(bw: np.ndarray, centers: List[Tuple[int, int]], r: int, search_px: int) -> Tuple[int, float, float]:
    scores = [best_score_local_search(bw, cx, cy, r, search_px) for (cx, cy) in centers]
    best_i = int(np.argmax(scores)) if scores else 0
    best_s = float(scores[best_i]) if scores else 0.0
    if len(scores) >= 2:
        ss = sorted(scores, reverse=True)
        gap = float(ss[0] - ss[1])
    else:
        gap = best_s
    return best_i, best_s, gap

def deterministic_random_012(seed_text: str) -> str:
    h = hashlib.md5(seed_text.encode("utf-8")).hexdigest()
    v = int(h[-2:], 16) % 3
    return str(v)

def row_fallback_density(row_img_gray: np.ndarray) -> Optional[str]:
    # simple density thirds
    g = cv2.GaussianBlur(row_img_gray, (3, 3), 0)
    _, bw = cv2.threshold(g, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    bw = 255 - bw
    h, w = bw.shape[:2]
    thirds = [bw[:, int(w*0.65):int(w*0.78)], bw[:, int(w*0.78):int(w*0.90)], bw[:, int(w*0.90):w]]
    dens = [float(cv2.countNonZero(t)) / float(t.size + 1e-6) for t in thirds]
    best = int(np.argmax(dens))
    return str(2 - best)

def ped_extract_page(page_pil: Image.Image, q_start: int, q_end: int, rotate_fix: bool, pid_for_seed: str, log=None) -> Dict[int, str]:
    gray = pil_to_gray(page_pil)
    if rotate_fix:
        gray = ensure_upright(gray)

    roi_gray = crop_roi(gray, PE_ROI)
    q_count = q_end - q_start + 1

    best_answers = None
    best_unclear = 10**9

    for pass_idx in range(ANS_PASSES):
        bw_full = adaptive_binarize(gray, pass_idx=pass_idx)
        roi_bw = crop_roi(bw_full, PE_ROI)

        rh, rw = roi_bw.shape[:2]
        bubble_r = max(8, int(min(rw, rh) * 0.022))
        search_px = max(10, int(bubble_r * 1.2))

        h_lines, v_lines = detect_grid_lines(roi_bw)

        v = v_lines[-4:] if len(v_lines) >= 4 else v_lines
        col_centers: List[int] = []
        if len(v) >= 2:
            for i in range(len(v) - 1):
                col_centers.append((v[i] + v[i + 1]) // 2)
        if len(col_centers) != 3:
            col_centers = [int(0.78 * rw), int(0.88 * rw), int(0.97 * rw)]

        if len(h_lines) >= q_count + 1:
            use = h_lines[:q_count + 1]
            row_centers = [int((use[i] + use[i + 1]) / 2) for i in range(q_count)]
        else:
            row_centers = list(np.linspace(int(0.18 * rh), int(0.92 * rh), q_count).astype(int))

        answers: Dict[int, str] = {}
        unclear_rows: List[int] = []

        for i in range(q_count):
            qnum = q_start + i
            y = int(row_centers[i])
            centers = [(col_centers[0], y), (col_centers[1], y), (col_centers[2], y)]
            best_i, best_s, gap = pick_one_option(roi_bw, centers, bubble_r, search_px)

            if best_s < 0.050 or gap < 0.010:
                answers[qnum] = "N/A"
                unclear_rows.append(i)
            else:
                answers[qnum] = str(2 - best_i)

        unclear = len(unclear_rows)
        if log:
            log(f"[PED] q{q_start}-{q_end} pass={pass_idx} unclear={unclear}/{q_count}")

        # row-band recheck
        if ROW_BAND_RECHECK and unclear_rows:
            still = []
            for i in unclear_rows:
                qnum = q_start + i
                y = int(row_centers[i])
                band_h = max(22, int(bubble_r * 3.2))
                y1 = clamp(y - band_h, 0, roi_bw.shape[0] - 1)
                y2 = clamp(y + band_h, 1, roi_bw.shape[0])
                band_bw = roi_bw[y1:y2, :]
                local_y = (y - y1)
                centers = [(col_centers[0], local_y), (col_centers[1], local_y), (col_centers[2], local_y)]
                bi, bs, gp = pick_one_option(band_bw, centers, bubble_r, search_px)
                if bs >= 0.050 and gp >= 0.010:
                    answers[qnum] = str(2 - bi)
                else:
                    still.append(i)
            unclear_rows = still

        # row fallback tries
        if unclear_rows:
            still2 = []
            for i in unclear_rows:
                qnum = q_start + i
                y = int(row_centers[i])
                band_h = max(26, int(bubble_r * 3.5))
                y1 = clamp(y - band_h, 0, roi_gray.shape[0] - 1)
                y2 = clamp(y + band_h, 1, roi_gray.shape[0])
                row_band = roi_gray[y1:y2, :]

                picked = None
                for _ in range(ROW_FALLBACK_TRIES):
                    picked = row_fallback_density(row_band)
                    if picked in ("0", "1", "2"):
                        break

                if picked in ("0", "1", "2"):
                    answers[qnum] = picked
                else:
                    still2.append(i)
            unclear_rows = still2

        unclear2 = len(unclear_rows)
        if unclear2 < best_unclear:
            best_unclear = unclear2
            best_answers = answers

        if best_unclear == 0:
            break

    final: Dict[int, str] = {}
    for q in range(q_start, q_end + 1):
        v = (best_answers or {}).get(q, "N/A")
        if v in ("0", "1", "2"):
            final[q] = v
        else:
            final[q] = deterministic_random_012(f"{pid_for_seed}|Q{q}") if DEFAULT_RANDOM else DEFAULT_ANSWER_IF_UNCLEAR
    return final

def extract_pedeyeq_answers(group_pages: List[Image.Image], pid_for_seed: str, log=None) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for q in range(1, 41):
        out[f"PedEyeQ{q}"] = DEFAULT_ANSWER_IF_UNCLEAR

    for local_idx, (q_start, q_end, rot) in enumerate(PED_RANGES):
        page_index = PED_PAGE_IDX[local_idx]
        if page_index >= len(group_pages):
            for q in range(q_start, q_end + 1):
                out[f"PedEyeQ{q}"] = deterministic_random_012(f"{pid_for_seed}|Q{q}") if DEFAULT_RANDOM else DEFAULT_ANSWER_IF_UNCLEAR
            continue

        ans = ped_extract_page(group_pages[page_index], q_start, q_end, rot, pid_for_seed, log=log)
        for q in range(q_start, q_end + 1):
            out[f"PedEyeQ{q}"] = ans.get(q, DEFAULT_ANSWER_IF_UNCLEAR)

    # final sanitize
    for q in range(1, 41):
        k = f"PedEyeQ{q}"
        if out.get(k) not in ("0", "1", "2"):
            out[k] = deterministic_random_012(f"{pid_for_seed}|Q{q}") if DEFAULT_RANDOM else DEFAULT_ANSWER_IF_UNCLEAR

    return out


# ============================================================
# MAIN PROCESS
# ============================================================

@dataclass
class RunConfig:
    pdf_paths: List[str]
    output_folder: str
    filename_prefix: str
    dpi: int
    include_ref_columns: bool
    save_every_n_groups: int
    overwrite_output: bool


def empty_row(include_ref: bool) -> Dict[str, str]:
    cols = build_columns(include_ref)
    return {k: "N/A" for k in cols}


def process_pdfs(cfg: RunConfig, progress_var, status_label, root, log_widget=None):
    os.makedirs(cfg.output_folder, exist_ok=True)
    safe_prefix = cfg.filename_prefix or "PedEyeQ_Combined"
    excel_path = os.path.join(cfg.output_folder, f"{safe_prefix}.xlsx")
    log_path = os.path.join(cfg.output_folder, f"{safe_prefix}_runlog.txt")

    columns = build_columns(cfg.include_ref_columns)

    if cfg.overwrite_output and os.path.exists(excel_path):
        try:
            os.remove(excel_path)
        except Exception:
            pass
    ensure_excel_with_header(excel_path, columns=columns)

    log_lines: List[str] = []
    def log(msg: str):
        log_lines.append(msg)
        log_message(msg, log_widget)

    seen_pids: Set[str] = set()
    failures = 0
    total_pdfs = len(cfg.pdf_paths)

    log("=== Run started ===")
    log(f"PDF count: {total_pdfs}")
    log(f"Pages per questionnaire: {PAGES_PER_QUESTIONNAIRE}")
    log(f"DPI: {cfg.dpi}")
    log(f"Excel (ONE file): {excel_path}")
    log(f"Overwrite output: {cfg.overwrite_output}")
    log("PID: filename (range/comma/single) ONLY, no unknown allowed.")
    log(f"Answer: OpenCV passes={ANS_PASSES}, row-band={ROW_BAND_RECHECK}, fallback tries={ROW_FALLBACK_TRIES}")
    log(f"Default for remaining unclear: {'RANDOM 0/1/2' if DEFAULT_RANDOM else DEFAULT_ANSWER_IF_UNCLEAR}")

    for pi, pdf in enumerate(cfg.pdf_paths, start=1):
        status_label.config(text=f"PDF {pi}/{total_pdfs}: streaming render...")
        root.update_idletasks()
        log(f"\n[PDF {pi}/{total_pdfs}] {pdf}")

        try:
            pid_plan = parse_pid_plan_from_filename(pdf)
            if len(pid_plan) >= 2:
                log(f"[PID] Filename plan: {pid_plan[0]} ... {pid_plan[-1]} (len={len(pid_plan)})")
            else:
                log(f"[PID] Filename plan: {pid_plan[0]} (len=1)")
        except Exception as e:
            failures += 1
            log(f"[PID] ERROR: {e}")
            messagebox.showerror("PID filename error", f"Cannot parse PID from filename:\n{os.path.basename(pdf)}\n\nFix filename then retry.")
            return

        buffer_rows: List[Dict[str, str]] = []
        groups_done = 0

        try:
            for gi, (group_pages, pstart, pend) in enumerate(
                iter_questionnaire_groups(pdf, cfg.dpi, PAGES_PER_QUESTIONNAIRE, log=log),
                start=1
            ):
                status_label.config(text=f"PDF {pi}/{total_pdfs} - questionnaire {gi}")
                root.update_idletasks()

                row = empty_row(cfg.include_ref_columns)
                if cfg.include_ref_columns:
                    row["Source PDF"] = os.path.basename(pdf)
                    row["Source Pages"] = f"{pstart+1}-{pend+1}"

                pid = pick_unique_pid_from_plan(pid_plan, group_idx0=(gi-1), seen=seen_pids)
                row["Participant ID"] = pid
                seen_pids.add(pid)
                log(f"[PID] -> {pid}")

                try:
                    ans = extract_pedeyeq_answers(group_pages, pid_for_seed=pid, log=log)
                    row.update(ans)
                except Exception as e:
                    failures += 1
                    log(f"[PED] failed: {e}")
                    for q in range(1, 41):
                        row[f"PedEyeQ{q}"] = deterministic_random_012(f"{pid}|Q{q}") if DEFAULT_RANDOM else DEFAULT_ANSWER_IF_UNCLEAR

                buffer_rows.append(row)
                groups_done += 1

                if cfg.save_every_n_groups > 0 and (groups_done % cfg.save_every_n_groups == 0):
                    append_rows_to_excel(excel_path, buffer_rows, columns=columns)
                    log(f"[SAVE] Appended {len(buffer_rows)} rows (mid-PDF).")
                    buffer_rows.clear()

                progress_var.set(pi / total_pdfs * 100.0)
                root.update_idletasks()

            if buffer_rows:
                append_rows_to_excel(excel_path, buffer_rows, columns=columns)
                log(f"[SAVE] Appended {len(buffer_rows)} rows (end-PDF).")
                buffer_rows.clear()

        except Exception as e:
            failures += 1
            log(f"ERROR: PDF failed: {e}")

        with open(log_path, "w", encoding="utf-8") as f:
            f.write("\n".join(log_lines))

    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))

    if failures:
        status_label.config(text="Finished (some errors).")
        messagebox.showwarning("Done", f"Completed with {failures} error(s).\nExcel:\n{excel_path}\nLog:\n{log_path}")
    else:
        status_label.config(text="Finished ✅")
        messagebox.showinfo("Done", f"Completed.\nExcel:\n{excel_path}\nLog:\n{log_path}")


# ============================================================
# GUI
# ============================================================

def main_app():
    root = tk.Tk()
    root.title("PDF -> Excel (NO duplicate PID + PedEyeQ1-40) [FAST + ZERO manual]")

    output_folder_var = tk.StringVar()
    filename_prefix_var = tk.StringVar(value="PedEyeQ_Combined")
    dpi_var = tk.StringVar(value="200")
    include_ref_var = tk.BooleanVar(value=True)
    save_every_groups_var = tk.StringVar(value="50")
    overwrite_var = tk.BooleanVar(value=DEFAULT_OVERWRITE_OUTPUT)

    progress_var = tk.DoubleVar(value=0.0)
    pdf_paths: List[str] = []

    file_frame = ttk.LabelFrame(root, text="Files")
    file_frame.grid(row=0, column=0, padx=10, pady=5, sticky="ew")
    file_frame.columnconfigure(1, weight=1)

    ttk.Label(file_frame, text="PDF file(s):").grid(row=0, column=0, sticky="w")
    pdf_summary_var = tk.StringVar(value="No files selected")
    ttk.Entry(file_frame, textvariable=pdf_summary_var, width=60, state="readonly").grid(
        row=0, column=1, padx=5, pady=2, sticky="ew"
    )

    def browse_pdfs():
        nonlocal pdf_paths
        paths = filedialog.askopenfilenames(title="Select PDFs", filetypes=[("PDF Files", "*.pdf")])
        if paths:
            pdf_paths[:] = list(paths)
            pdf_summary_var.set(pdf_paths[0] if len(pdf_paths) == 1 else f"{len(pdf_paths)} files selected")

    ttk.Button(file_frame, text="Browse...", command=browse_pdfs).grid(row=0, column=2, padx=5, pady=2)

    ttk.Label(file_frame, text="Output folder:").grid(row=1, column=0, sticky="w")
    ttk.Entry(file_frame, textvariable=output_folder_var, width=60).grid(row=1, column=1, padx=5, pady=2, sticky="ew")

    def browse_output():
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            output_folder_var.set(path)

    ttk.Button(file_frame, text="Browse...", command=browse_output).grid(row=1, column=2, padx=5, pady=2)

    ttk.Label(file_frame, text="Excel filename (no .xlsx):").grid(row=2, column=0, sticky="w")
    ttk.Entry(file_frame, textvariable=filename_prefix_var, width=30).grid(row=2, column=1, padx=5, pady=2, sticky="w")

    run_frame = ttk.LabelFrame(root, text="Run settings")
    run_frame.grid(row=1, column=0, padx=10, pady=5, sticky="ew")

    ttk.Label(run_frame, text="DPI (recommended 180-220):").grid(row=0, column=0, sticky="w")
    ttk.Entry(run_frame, textvariable=dpi_var, width=10).grid(row=0, column=1, padx=5, pady=2, sticky="w")

    ttk.Label(run_frame, text="Append every N questionnaires:").grid(row=1, column=0, sticky="w")
    ttk.Entry(run_frame, textvariable=save_every_groups_var, width=10).grid(row=1, column=1, padx=5, pady=2, sticky="w")

    ttk.Checkbutton(run_frame, text="Include Source PDF + Pages in Excel", variable=include_ref_var).grid(
        row=2, column=0, columnspan=2, sticky="w", padx=5, pady=2
    )
    ttk.Checkbutton(run_frame, text="Overwrite output Excel (recommended)", variable=overwrite_var).grid(
        row=3, column=0, columnspan=2, sticky="w", padx=5, pady=2
    )

    progress_frame = ttk.LabelFrame(root, text="Progress")
    progress_frame.grid(row=2, column=0, padx=10, pady=5, sticky="ew")
    ttk.Progressbar(progress_frame, orient="horizontal", length=400, mode="determinate", variable=progress_var).grid(
        row=0, column=0, padx=5, pady=2, sticky="ew"
    )
    status_label = ttk.Label(progress_frame, text="Idle")
    status_label.grid(row=1, column=0, padx=5, pady=2, sticky="w")
    progress_frame.columnconfigure(0, weight=1)

    log_frame = ttk.LabelFrame(root, text="Run log")
    log_frame.grid(row=3, column=0, padx=10, pady=5, sticky="nsew")
    log_text = tk.Text(log_frame, height=12, wrap="none", state="disabled")
    log_text.grid(row=0, column=0, sticky="nsew")
    log_scroll_y = ttk.Scrollbar(log_frame, orient="vertical", command=log_text.yview)
    log_scroll_y.grid(row=0, column=1, sticky="ns")
    log_text.configure(yscrollcommand=log_scroll_y.set)
    log_frame.columnconfigure(0, weight=1)
    log_frame.rowconfigure(0, weight=1)

    btn_frame = ttk.Frame(root)
    btn_frame.grid(row=4, column=0, padx=10, pady=5, sticky="ew")

    start_button = ttk.Button(btn_frame, text="Start", width=15)
    start_button.grid(row=0, column=0, padx=5)

    def start_processing():
        if not pdf_paths:
            messagebox.showerror("Error", "Please select at least one PDF file.")
            return
        output_folder = output_folder_var.get().strip()
        if not output_folder:
            messagebox.showerror("Error", "Please select an output folder.")
            return
        if fitz is None:
            messagebox.showerror("Error", "PyMuPDF not installed. Install: pip install pymupdf")
            return

        try:
            dpi = int(dpi_var.get().strip())
            save_every_n = int(save_every_groups_var.get().strip())
        except Exception:
            messagebox.showerror("Error", "Please input valid numbers.")
            return

        try:
            log_text.configure(state="normal")
            log_text.delete("1.0", "end")
            log_text.configure(state="disabled")
        except Exception:
            pass

        cfg = RunConfig(
            pdf_paths=list(pdf_paths),
            output_folder=output_folder,
            filename_prefix=filename_prefix_var.get().strip(),
            dpi=max(120, dpi),
            include_ref_columns=bool(include_ref_var.get()),
            save_every_n_groups=max(1, save_every_n),
            overwrite_output=bool(overwrite_var.get()),
        )

        start_button.config(state="disabled")
        status_label.config(text="Starting...")
        progress_var.set(0)

        def worker():
            try:
                process_pdfs(cfg, progress_var, status_label, root, log_widget=log_text)
            except Exception as e:
                traceback.print_exc()
                log_message(f"Unexpected error: {e}", log_text)
                status_label.config(text="Unexpected error.")
                messagebox.showerror("Unexpected error", str(e))
            finally:
                start_button.config(state="normal")

        threading.Thread(target=worker, daemon=True).start()

    start_button.config(command=start_processing)
    ttk.Button(btn_frame, text="Quit", width=15, command=root.destroy).grid(row=0, column=1, padx=5)

    root.rowconfigure(3, weight=1)
    root.columnconfigure(0, weight=1)
    root.mainloop()


if __name__ == "__main__":
    main_app()