import ast
import base64
import io
import json
import os
import queue
import re
import threading
import traceback
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import cv2
import fitz  # PyMuPDF
import numpy as np
import pandas as pd
import requests
from PIL import Image
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# ============================================================
# SETTINGS / ROIS
# ============================================================
DPI_DEFAULT = 220
TARGET_W = 1654
TARGET_H = 2339
SHEET_NAME = "Questionnaires"
REQUEST_TIMEOUT = 180


def nr(x1, y1, x2, y2):
    return (x1, y1, x2, y2)


# Page 1 = consent page
P1 = {
    "pid_crop": nr(0.72, 0.00, 0.99, 0.11),
    "consent_crop": nr(0.62, 0.58, 0.98, 0.75),
    "bottom_crop": nr(0.04, 0.77, 0.76, 0.97),
}

# Page 2 = student / guardian / date page
P2 = {
    "top_crop": nr(0.03, 0.00, 0.80, 0.36),
    "pid_crop": nr(0.78, 0.00, 0.98, 0.12),
}

# Page 3 = questionnaire page with school/name/dob and family history
P3 = {
    "header_crop": nr(0.01, 0.04, 0.99, 0.46),
    "lower_crop": nr(0.01, 0.44, 0.80, 0.99),
    "crossed_box": nr(0.035, 0.475, 0.060, 0.495),
    "lazy_box": nr(0.035, 0.535, 0.060, 0.555),
    "high_box": nr(0.035, 0.595, 0.060, 0.615),
    "retinal_box": nr(0.035, 0.655, 0.060, 0.675),
    "macular_box": nr(0.035, 0.715, 0.060, 0.735),
    "glaucoma_box": nr(0.035, 0.775, 0.060, 0.795),
    "others_box": nr(0.035, 0.835, 0.060, 0.855),
    "normal_box": nr(0.035, 0.892, 0.060, 0.912),
    "medical_box": nr(0.035, 0.936, 0.060, 0.956),
}

# Page 4 = eye exam / symptoms / spectacles / medicine page
P4 = {
    "exam_never": nr(0.035, 0.155, 0.060, 0.175),
    "exam_past_year": nr(0.035, 0.205, 0.060, 0.225),
    "exam_1_2": nr(0.035, 0.255, 0.060, 0.275),
    "exam_2_4": nr(0.035, 0.315, 0.060, 0.335),
    "exam_gt4": nr(0.035, 0.375, 0.060, 0.395),
    "symptoms_no": nr(0.035, 0.500, 0.060, 0.520),
    "symptoms_yes": nr(0.035, 0.548, 0.060, 0.568),
    "blurred": nr(0.035, 0.612, 0.060, 0.632),
    "distance": nr(0.035, 0.720, 0.060, 0.740),
    "intermediate": nr(0.125, 0.720, 0.150, 0.740),
    "near": nr(0.215, 0.720, 0.240, 0.740),
    "squinting": nr(0.035, 0.772, 0.060, 0.792),
    "close_tv": nr(0.035, 0.840, 0.060, 0.860),
    "poor_light": nr(0.035, 0.920, 0.060, 0.940),
    "eye_disease": nr(0.56, 0.215, 0.585, 0.235),
    "eye_infection": nr(0.56, 0.305, 0.585, 0.325),
    "eye_injury": nr(0.56, 0.415, 0.585, 0.435),
    "eye_surgery": nr(0.56, 0.525, 0.585, 0.545),
    "spec_no": nr(0.56, 0.655, 0.585, 0.675),
    "spec_full": nr(0.56, 0.715, 0.585, 0.735),
    "spec_occ": nr(0.56, 0.790, 0.585, 0.810),
    "med_no": nr(0.56, 0.905, 0.585, 0.925),
    "med_yes": nr(0.56, 0.960, 0.585, 0.980),
    "right_text_crop": nr(0.57, 0.14, 0.995, 0.995),
}

# Page 5 = q8-q13 table
P5_ROWS = [
    nr(0.54, 0.262, 0.995, 0.329),
    nr(0.54, 0.329, 0.995, 0.423),
    nr(0.54, 0.423, 0.995, 0.490),
    nr(0.54, 0.490, 0.995, 0.600),
    nr(0.54, 0.600, 0.995, 0.715),
    nr(0.54, 0.715, 0.995, 0.845),
]
P5_COLS = [0.54, 0.635, 0.72, 0.805, 0.89, 0.955, 0.995]
P5_TABLE = nr(0.015, 0.10, 0.995, 0.90)

PID_RE = re.compile(r"\b([ABC])\s*0*(\d{1,4})\b", re.I)
DIGITS_RE = re.compile(r"(\d{1,4})")


# ============================================================
# COLUMNS
# ============================================================
def headers() -> List[str]:
    return [
        "participant_id", "source_pdf", "questionnaire_index_in_pdf", "source_pages",
        "consent_school", "consent_grade_level", "consent_section", "consent_vision_screening", "consent_eye_photos",
        "student_name", "parent_guardian_name", "consent_date",
        "school", "name", "gender", "date_of_birth", "age", "grade", "class_no", "id_code",
        "family_crossed_eyes", "family_crossed_eyes_relation", "family_lazy_eye", "family_lazy_eye_relation",
        "family_high_eye_power", "family_high_eye_power_relation", "family_retinal_disease", "family_retinal_disease_relation",
        "family_macular_disease", "family_macular_disease_relation", "family_glaucoma", "family_glaucoma_relation",
        "family_others", "family_others_relation", "general_normal_no_known_conditions", "general_medical_conditions_present",
        "general_medical_conditions_text", "last_eye_exam", "current_symptoms_no", "current_symptoms_yes",
        "blurred_vision", "blurred_distance", "blurred_intermediate", "blurred_near", "squinting_when_viewing_objects",
        "working_at_close_distance_homework_tv", "using_devices_or_reading_in_poor_lighting",
        "event_eye_disease", "event_eye_disease_remarks", "event_eye_infection_inflammation", "event_eye_infection_inflammation_remarks",
        "event_eye_injury_trauma", "event_eye_injury_trauma_remarks", "event_eye_surgery", "event_eye_surgery_remarks",
        "wear_spectacles", "eye_medicine_use", "eye_medicine_text", "q8", "q9", "q10", "q11", "q12", "q13",
    ]


def blank_row() -> Dict[str, object]:
    return {h: "N/A" for h in headers()}


# ============================================================
# BASIC HELPERS
# ============================================================
def safe_mkdir(path: str):
    if path:
        os.makedirs(path, exist_ok=True)


def normalize_pid(text: str) -> str:
    if not text:
        return "N/A"
    s = str(text).strip()
    if s.upper().startswith("REVIEW_PID"):
        return "REVIEW_PID"
    m = PID_RE.search(s.replace("-", " ").replace("_", " "))
    if not m:
        return "N/A"
    n = int(m.group(2))
    return f"{m.group(1).upper()}{n:03d}" if n <= 999 else f"{m.group(1).upper()}{n}"


def normalize_pid_letter(text: object) -> str:
    s = clean_text_value(text) if text is not None else "N/A"
    if s == "N/A":
        return "N/A"
    s = str(s).strip().upper()
    if s.startswith("A"):
        return "A"
    if s.startswith("B"):
        return "B"
    if s.startswith("C"):
        return "C"
    return "N/A"


def normalize_pid_digits(text: object) -> str:
    if text is None:
        return "N/A"
    s = str(text).strip()
    m = DIGITS_RE.search(s)
    if not m:
        return "N/A"
    n = int(m.group(1))
    return f"{n:03d}" if n <= 999 else str(n)


def clean_text_value(v: object) -> str:
    if v is None:
        return "N/A"
    s = str(v).replace("\x0c", " ").strip()
    s = re.sub(r"\s+", " ", s)
    if not s:
        return "N/A"
    if s.lower() in {"n/a", "na", "none", "null", "unknown", "unreadable", "nil", "blank"}:
        return "N/A"
    if s in {"...", ".", "..", "-", "--", "___"}:
        return "N/A"
    return s


def clean_choice(v: object, allowed: List[str]) -> str:
    s = clean_text_value(v)
    if s == "N/A":
        return s
    s2 = s.lower().replace(" ", "_").replace("-", "_")
    mapping = {
        "yes": "yes", "no": "no",
        "m": "M", "male": "M", "f": "F", "female": "F",
        "full": "full_time", "fulltime": "full_time", "full_time": "full_time",
        "occasionally": "occasional", "occasional": "occasional",
        "pastyear": "past_year", "past_year": "past_year",
        "1_2_years": "1_to_2_years", "1_to_2": "1_to_2_years",
        "2_4_years": "2_to_4_years", "2_to_4": "2_to_4_years",
        "gt4": "more_than_4_years", "more_than_4_years": "more_than_4_years",
        "never": "never",
    }
    s3 = mapping.get(s2, s)
    return s3 if s3 in allowed else "N/A"


def clean_int01(v: object) -> int:
    if isinstance(v, bool):
        return 1 if v else 0
    s = str(v).strip().lower()
    if s in {"1", "yes", "true", "checked", "present", "y"}:
        return 1
    return 0


def rect_from_norm(img: np.ndarray, box: Tuple[float, float, float, float]) -> Tuple[int, int, int, int]:
    h, w = img.shape[:2]
    x1 = max(0, min(w - 1, int(box[0] * w)))
    y1 = max(0, min(h - 1, int(box[1] * h)))
    x2 = max(x1 + 1, min(w, int(box[2] * w)))
    y2 = max(y1 + 1, min(h, int(box[3] * h)))
    return x1, y1, x2, y2


def crop(img: np.ndarray, box: Tuple[float, float, float, float]) -> np.ndarray:
    x1, y1, x2, y2 = rect_from_norm(img, box)
    return img[y1:y2, x1:x2].copy()


def resize_page(gray: np.ndarray) -> np.ndarray:
    if gray.ndim == 3:
        gray = cv2.cvtColor(gray, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)
    gray = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX)
    return cv2.resize(gray, (TARGET_W, TARGET_H), interpolation=cv2.INTER_CUBIC)


def make_vlm_image(gray: np.ndarray) -> np.ndarray:
    if gray.ndim == 3:
        gray = cv2.cvtColor(gray, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.2, tileGridSize=(8, 8))
    gray = clahe.apply(gray)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)
    return gray


def np_to_b64_png(img: np.ndarray) -> str:
    if img.ndim == 2:
        pil = Image.fromarray(img)
    else:
        pil = Image.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
    buf = io.BytesIO()
    pil.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def upscale(img: np.ndarray, scale: float = 2.5) -> np.ndarray:
    h, w = img.shape[:2]
    return cv2.resize(img, (max(1, int(w * scale)), max(1, int(h * scale))), interpolation=cv2.INTER_CUBIC)


def threshold_text(img: np.ndarray) -> np.ndarray:
    if img.ndim == 3:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    img = cv2.GaussianBlur(img, (3, 3), 0)
    _, bw = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    if np.mean(bw) < 127:
        bw = 255 - bw
    return bw


def pid_variant_set(img: np.ndarray) -> Dict[str, np.ndarray]:
    base = make_vlm_image(img)
    zoom = upscale(base, 2.8)
    bw = threshold_text(zoom)
    soft = cv2.GaussianBlur(zoom, (0, 0), 0.8)
    return {
        "raw": zoom,
        "bw": bw,
        "soft": soft,
    }


def vote_weighted(pairs: List[Tuple[str, float]]) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for value, wt in pairs:
        if not value or value == "N/A":
            continue
        out[value] = out.get(value, 0.0) + float(wt)
    return out


def top_two(scores: Dict[str, float]) -> Tuple[str, float, float]:
    if not scores:
        return "N/A", 0.0, 0.0
    items = sorted(scores.items(), key=lambda kv: (-kv[1], kv[0]))
    best_v, best_s = items[0]
    second_s = items[1][1] if len(items) > 1 else 0.0
    return best_v, best_s, second_s


def save_debug_image(path: str, img: np.ndarray):
    safe_mkdir(os.path.dirname(path))
    cv2.imwrite(path, img)


def save_raw_text(path: str, text: str):
    safe_mkdir(os.path.dirname(path))
    with open(path, "w", encoding="utf-8") as f:
        f.write(text or "")




def isolate_pid_line(region: np.ndarray, keep_top_ratio: float = 0.62) -> np.ndarray:
    gray = make_vlm_image(region)
    zoom = upscale(gray, 3.0)
    inv = threshold_inv(zoom)
    h, w = inv.shape[:2]
    top_h = max(1, int(h * keep_top_ratio))
    work = inv[:top_h, :].copy()

    hk = max(25, w // 4)
    horiz = cv2.morphologyEx(work, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_RECT, (hk, 1)))
    work = cv2.subtract(work, horiz)
    work = cv2.morphologyEx(work, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3)))

    contours, _ = cv2.findContours(work, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    boxes = []
    for c in contours:
        x, y, bw, bh = cv2.boundingRect(c)
        area = bw * bh
        if area < max(30, (h * w) // 3000):
            continue
        if bh < 10 or bw < 4:
            continue
        if bw > int(w * 0.90) and bh < int(h * 0.18):
            continue
        boxes.append((x, y, bw, bh))

    if not boxes:
        return zoom

    min_y = min(y for _, y, _, _ in boxes)
    line_cut = min_y + max(40, int(h * 0.18))
    line_boxes = [b for b in boxes if b[1] <= line_cut]
    if not line_boxes:
        line_boxes = boxes

    x1 = min(x for x, _, _, _ in line_boxes)
    y1 = min(y for _, y, _, _ in line_boxes)
    x2 = max(x + bw for x, _, bw, _ in line_boxes)
    y2 = max(y + bh for _, y, _, bh in line_boxes)

    pad_x = max(12, int((x2 - x1) * 0.20))
    pad_y = max(12, int((y2 - y1) * 0.45))
    x1 = max(0, x1 - pad_x)
    y1 = max(0, y1 - pad_y)
    x2 = min(w, x2 + pad_x)
    y2 = min(h, y2 + pad_y)

    out = zoom[y1:y2, x1:x2].copy()
    return out if out.size else zoom


def pid_crop_variants(region: np.ndarray, keep_top_ratio: float) -> Dict[str, np.ndarray]:
    iso = isolate_pid_line(region, keep_top_ratio=keep_top_ratio)
    raw = upscale(make_vlm_image(iso), 1.15)
    soft = cv2.GaussianBlur(raw, (0, 0), 0.8)
    bw = threshold_text(raw)
    return {
        "raw": raw,
        "soft": soft,
        "bw": bw,
    }


def clean_confidence(v: object) -> float:
    try:
        x = float(v)
    except Exception:
        return 0.0
    return max(0.0, min(1.0, x))


def pid_is_valid(pid: str) -> bool:
    pid = normalize_pid(pid)
    if pid == "N/A":
        return False
    m = re.match(r"^[ABC](\d{3,4})$", pid)
    if not m:
        return False
    return m.group(1) != "000"


# ============================================================
# JSON / LM STUDIO
# ============================================================
def clean_json_text(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"^\s*```(?:json)?\s*", "", s, flags=re.I)
    s = re.sub(r"\s*```\s*$", "", s)
    if "{" in s and "}" in s:
        s = s[s.find("{"):s.rfind("}") + 1]
    s = s.replace("\u201c", '"').replace("\u201d", '"').replace("\u2018", "'").replace("\u2019", "'")
    s = re.sub(r",\s*([}\]])", r"\1", s)
    return s.strip()


def parse_json_response(text: str) -> dict:
    s = clean_json_text(text)
    if not s:
        raise ValueError("Empty model response")
    try:
        return json.loads(s)
    except Exception:
        pass
    # Sometimes the model returns python-ish dicts
    s2 = s.replace("null", "None").replace("true", "True").replace("false", "False")
    try:
        obj = ast.literal_eval(s2)
        if isinstance(obj, dict):
            return obj
    except Exception:
        pass
    raise ValueError(f"Invalid JSON from model: {text[:500]}")


class LMStudioClient:
    def __init__(self, base_url: str, model: str, api_key: str, timeout_s: int = REQUEST_TIMEOUT):
        b = (base_url or "").strip().rstrip("/")
        if not b:
            b = "http://127.0.0.1:1234"
        self.base_url = b if b.endswith("/v1") else b + "/v1"
        self.model = (model or "").strip()
        self.api_key = (api_key or "").strip()
        self.timeout_s = max(15, int(timeout_s))

    def _headers(self) -> dict:
        h = {"Content-Type": "application/json"}
        if self.api_key:
            h["Authorization"] = f"Bearer {self.api_key}"
        return h

    def list_models(self) -> List[str]:
        r = requests.get(f"{self.base_url}/models", headers=self._headers(), timeout=20)
        if r.status_code != 200:
            raise RuntimeError(f"HTTP {r.status_code}: {r.text[:300]}")
        data = r.json()
        models = []
        for item in data.get("data", []):
            if isinstance(item, dict) and item.get("id"):
                models.append(str(item["id"]))
        return models

    def vision_json(self, prompt: str, images: List[np.ndarray], max_tokens: int = 500, temperature: float = 0.0, retries: int = 2) -> Tuple[dict, str]:
        if not self.model:
            raise RuntimeError("Model id is empty")
        content = [{"type": "text", "text": prompt}]
        for img in images:
            content.append({"type": "image_url", "image_url": {"url": f"data:image/png;base64,{np_to_b64_png(img)}"}})

        payload = {
            "model": self.model,
            "messages": [{"role": "user", "content": content}],
            "temperature": float(temperature),
            "max_tokens": int(max_tokens),
        }

        last_err = None
        for _ in range(retries + 1):
            try:
                r = requests.post(f"{self.base_url}/chat/completions", headers=self._headers(), json=payload, timeout=self.timeout_s)
                ct = (r.headers.get("Content-Type") or "").lower()
                if "application/json" not in ct:
                    raise RuntimeError(f"Non-JSON response HTTP {r.status_code}: {r.text[:500]}")
                data = r.json()
                if r.status_code >= 400:
                    raise RuntimeError(f"HTTP {r.status_code}: {json.dumps(data)[:500]}")
                choices = data.get("choices")
                if not choices:
                    raise RuntimeError(f"No choices: {json.dumps(data)[:500]}")
                msg = choices[0].get("message", {})
                content_out = msg.get("content", "")
                if isinstance(content_out, list):
                    raw = "\n".join(str(x.get("text", "")) for x in content_out if isinstance(x, dict))
                else:
                    raw = str(content_out)
                return parse_json_response(raw), raw
            except Exception as e:
                last_err = e
        raise RuntimeError(str(last_err))


# ============================================================
# PROMPTS
# ============================================================
def prompt_pid_full(keys: List[str]) -> str:
    keys_txt = ", ".join([f'\"{k}\"' for k in keys] + ['\"final\"'])
    return (
        "All images show the SAME participant ID from one questionnaire. Some images are enhanced variants of the same crop. "
        "Read the handwritten participant ID carefully. There may be other nearby writing, but only the participant ID counts. "
        "Return ONLY JSON with these keys: " + keys_txt + ". "
        "For each image key, output A### or B### or C### or N/A. final must be the best overall ID. No extra text."
    )


def prompt_pid_letter(keys: List[str]) -> str:
    keys_txt = ", ".join([f'\"{k}\"' for k in keys] + ['\"final\"'])
    return (
        "All images show the SAME participant ID from one questionnaire. Read ONLY the FIRST handwritten LETTER of the ID. "
        "The letter must be A, B, or C. Ignore the digits. Return ONLY JSON with keys: " + keys_txt + ". "
        "Each value must be A, B, or N/A. final must be the best overall letter. No extra text."
    )


def prompt_pid_digits(keys: List[str]) -> str:
    keys_txt = ", ".join([f'\"{k}\"' for k in keys] + ['\"final\"'])
    return (
        "All images show the SAME participant ID from one questionnaire. Read ONLY the NUMERIC part of the handwritten ID. "
        "Ignore the first letter. Return ONLY JSON with keys: " + keys_txt + ". "
        "Each value must be digits only like 001 or 4, or N/A. final must be the best overall digits. No extra text."
    )


def prompt_pid_single() -> str:
    return (
        "This image shows a short handwritten participant ID from the top-right area of a questionnaire. "
        "The participant ID format is exactly one uppercase letter A, B, or C followed by exactly three digits, "
        "for example A001 or B389. Ignore any longer personal number, punctuation, underline, printed text, or stray marks. "
        "Prefer the short 4-character handwritten ID closest to the top. "
        "Return ONLY valid JSON with exact keys pid, letter, digits, confidence. "
        "pid must be A### or B### or C### or N/A. letter must be A, B, or C or N/A. digits must be exactly three digits like 001 or N/A. confidence must be a number between 0 and 1."
    )


def prompt_pid_chars() -> str:
    return (
        "This image shows one short handwritten participant ID. Read the ID character by character from left to right. "
        "The ID has exactly four characters: c1 is A, B, or C, and c2 c3 c4 are digits 0-9. "
        "Ignore any longer number, punctuation, underline, printed text, or stray marks. "
        "Return ONLY valid JSON with exact keys c1, c2, c3, c4, confidence. "
        "Example: {\"c1\":\"A\",\"c2\":\"0\",\"c3\":\"0\",\"c4\":\"1\",\"confidence\":0.98}"
    )


def prompt_pid_digits_single(letter_hint: str = "") -> str:
    extra = f" The first letter is definitely {letter_hint}." if letter_hint in {"A", "B", "C"} else ""
    return (
        "This image shows a short handwritten participant ID from a questionnaire."
        + extra +
        " Read ONLY the three digits after the first letter. Ignore any longer personal number, punctuation, underline, printed text, or stray marks. "
        "Return ONLY valid JSON with exact keys digits, confidence. digits must be exactly three digits like 001 or N/A. confidence must be a number between 0 and 1."
    )


def prompt_page1() -> str:
    return (
        "These images are from the consent page of one questionnaire. "
        "Read only these fields and return ONLY valid JSON with these exact keys: "
        "consent_school, consent_grade_level, consent_section, consent_vision_screening, consent_eye_photos. "
        "Use values yes/no/N/A for consent_vision_screening and consent_eye_photos. "
        "Use N/A when blank. Do not use placeholders like ..."
    )


def prompt_page2() -> str:
    return (
        "This image is the top of page 2 of one questionnaire. "
        "Return ONLY valid JSON with exact keys: student_name, parent_guardian_name, consent_date. "
        "Use N/A when blank. Do not use placeholders."
    )


def prompt_page3_header() -> str:
    return (
        "This image is the top section of page 3 of one questionnaire. "
        "Return ONLY valid JSON with exact keys: school, name, gender, date_of_birth, age, grade, class_no, id_code. "
        "gender must be M, F, or N/A. Use N/A when blank. Do not guess class_no if it is empty."
    )


def prompt_page3_lower() -> str:
    return (
        "This image is the lower section of page 3 of one questionnaire. "
        "Read only the handwritten relation / remark lines. "
        "Return ONLY valid JSON with exact keys: "
        "family_crossed_eyes_relation, family_lazy_eye_relation, family_high_eye_power_relation, family_retinal_disease_relation, "
        "family_macular_disease_relation, family_glaucoma_relation, family_others_relation, general_medical_conditions_text. "
        "Use N/A when blank. Do not use placeholders."
    )


def prompt_page4_text() -> str:
    return (
        "This image is the right-hand text / remarks side of page 4 of one questionnaire. "
        "Return ONLY valid JSON with exact keys: event_eye_disease_remarks, event_eye_infection_inflammation_remarks, "
        "event_eye_injury_trauma_remarks, event_eye_surgery_remarks, eye_medicine_text. "
        "Use N/A when blank. Do not use placeholders."
    )


def prompt_page5() -> str:
    return (
        "This image is the q8-q13 answer table. Each row has one marked choice among 0,1,2,3,4,5. "
        "Return ONLY valid JSON with exact integer keys q8,q9,q10,q11,q12,q13. "
        "Each value must be an integer 0 to 5. No extra text."
    )


# ============================================================
# IMAGE / CV EXTRACTION
# ============================================================
def threshold_inv(gray: np.ndarray) -> np.ndarray:
    blur = cv2.GaussianBlur(gray, (3, 3), 0)
    return cv2.adaptiveThreshold(blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 31, 11)


def mark_ratio(gray: np.ndarray) -> float:
    bw = threshold_inv(gray)
    h, w = bw.shape[:2]
    mx, my = max(1, int(w * 0.18)), max(1, int(h * 0.18))
    inner = bw[my:h - my, mx:w - mx]
    if inner.size == 0:
        return 0.0
    return float(np.count_nonzero(inner)) / float(inner.size)


def checked(gray: np.ndarray, threshold: float = 0.015) -> int:
    return 1 if mark_ratio(gray) >= threshold else 0


def extract_choice(page: np.ndarray, choices: List[Tuple[str, str]], threshold: float = 0.008) -> str:
    scores = [(label, mark_ratio(crop(page, P4[roi]))) for label, roi in choices]
    scores.sort(key=lambda x: x[1], reverse=True)
    if not scores or scores[0][1] < threshold:
        return "N/A"
    if len(scores) > 1 and scores[0][1] < scores[1][1] * 1.12:
        return "N/A"
    return scores[0][0]


def extract_page3_checkbox_fields(page3: np.ndarray) -> Dict[str, object]:
    out = {
        "family_crossed_eyes": checked(crop(page3, P3["crossed_box"])),
        "family_lazy_eye": checked(crop(page3, P3["lazy_box"])),
        "family_high_eye_power": checked(crop(page3, P3["high_box"])),
        "family_retinal_disease": checked(crop(page3, P3["retinal_box"])),
        "family_macular_disease": checked(crop(page3, P3["macular_box"])),
        "family_glaucoma": checked(crop(page3, P3["glaucoma_box"])),
        "family_others": checked(crop(page3, P3["others_box"])),
        "general_normal_no_known_conditions": checked(crop(page3, P3["normal_box"])),
        "general_medical_conditions_present": checked(crop(page3, P3["medical_box"])),
    }
    # consistency fix: if medical conditions present, normal usually should be 0
    if out["general_medical_conditions_present"] == 1:
        out["general_normal_no_known_conditions"] = 0
    return out


def extract_page4_checkbox_fields(page4: np.ndarray) -> Dict[str, object]:
    out = {
        "last_eye_exam": extract_choice(page4, [
            ("never", "exam_never"),
            ("past_year", "exam_past_year"),
            ("1_to_2_years", "exam_1_2"),
            ("2_to_4_years", "exam_2_4"),
            ("more_than_4_years", "exam_gt4"),
        ]),
        "current_symptoms_no": checked(crop(page4, P4["symptoms_no"])),
        "current_symptoms_yes": checked(crop(page4, P4["symptoms_yes"])),
        "blurred_vision": checked(crop(page4, P4["blurred"])),
        "blurred_distance": checked(crop(page4, P4["distance"])),
        "blurred_intermediate": checked(crop(page4, P4["intermediate"])),
        "blurred_near": checked(crop(page4, P4["near"])),
        "squinting_when_viewing_objects": checked(crop(page4, P4["squinting"])),
        "working_at_close_distance_homework_tv": checked(crop(page4, P4["close_tv"])),
        "using_devices_or_reading_in_poor_lighting": checked(crop(page4, P4["poor_light"])),
        "event_eye_disease": checked(crop(page4, P4["eye_disease"])),
        "event_eye_infection_inflammation": checked(crop(page4, P4["eye_infection"])),
        "event_eye_injury_trauma": checked(crop(page4, P4["eye_injury"])),
        "event_eye_surgery": checked(crop(page4, P4["eye_surgery"])),
        "wear_spectacles": extract_choice(page4, [
            ("no", "spec_no"),
            ("full_time", "spec_full"),
            ("occasional", "spec_occ"),
        ], threshold=0.010),
        "eye_medicine_use": extract_choice(page4, [
            ("no", "med_no"),
            ("yes", "med_yes"),
        ], threshold=0.010),
    }
    if out["current_symptoms_yes"] == 1 and out["current_symptoms_no"] == 1:
        # keep the stronger one? Here prefer yes because symptom sub-boxes often imply yes.
        if any(out[k] == 1 for k in [
            "blurred_vision", "squinting_when_viewing_objects",
            "working_at_close_distance_homework_tv", "using_devices_or_reading_in_poor_lighting"
        ]):
            out["current_symptoms_no"] = 0
    return out


def q8_q13_from_cv(page5: np.ndarray) -> Tuple[Dict[str, int], float]:
    out: Dict[str, int] = {}
    scores_all = []
    for idx, row_box in enumerate(P5_ROWS, start=8):
        row_scores = []
        for c in range(6):
            cell = crop(page5, (P5_COLS[c], row_box[1], P5_COLS[c + 1], row_box[3]))
            row_scores.append(mark_ratio(cell))
        best = int(np.argmax(row_scores))
        best_s = float(row_scores[best])
        row_scores_sorted = sorted(row_scores, reverse=True)
        gap = row_scores_sorted[0] - (row_scores_sorted[1] if len(row_scores_sorted) > 1 else 0.0)
        scores_all.append(best_s if best_s > 0 else 0.0)
        # stricter threshold avoids wrong forced picks
        if best_s < 0.010 or gap < 0.0025:
            out[f"q{idx}"] = -1
        else:
            out[f"q{idx}"] = best
    conf = float(np.mean(scores_all)) if scores_all else 0.0
    return out, conf


# ============================================================
# PDF / EXCEL
# ============================================================
def render_page(doc: fitz.Document, page_index0: int, dpi: int) -> np.ndarray:
    page = doc.load_page(page_index0)
    mat = fitz.Matrix(dpi / 72.0, dpi / 72.0)
    pix = page.get_pixmap(matrix=mat, alpha=False)
    img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
    if pix.n >= 3:
        img = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    return resize_page(img)


def ensure_excel(path: str):
    if os.path.exists(path):
        wb = load_workbook(path)
        if SHEET_NAME not in wb.sheetnames:
            wb.create_sheet(SHEET_NAME)
        ws = wb[SHEET_NAME]
        if ws.max_row == 1 and ws["A1"].value is None:
            ws.append(headers())
        wb.save(path)
        wb.close()
        return
    df = pd.DataFrame([], columns=headers())
    df.to_excel(path, index=False, sheet_name=SHEET_NAME)


def append_rows_excel(path: str, rows: List[Dict[str, object]]):
    if not rows:
        return
    ensure_excel(path)
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    df = pd.DataFrame(rows).reindex(columns=headers()).fillna("N/A")
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    wb.save(path)
    wb.close()


# ============================================================
# PAGE EXTRACTOR
# ============================================================
@dataclass
class RunConfig:
    pdf_paths: List[str]
    output_excel: str
    base_url: str
    model_id: str
    api_key: str
    dpi: int = DPI_DEFAULT
    first_page_1based: int = 1
    pages_per_questionnaire: int = 8
    save_debug: bool = True


@dataclass
class FilenamePIDHint:
    exact_letter: str = "N/A"
    min_digits: Optional[int] = None
    max_digits: Optional[int] = None
    exact_pid: str = "N/A"


def parse_filename_pid_hint(pdf_name: str) -> FilenamePIDHint:
    name = os.path.splitext(os.path.basename(pdf_name))[0].upper()
    name = re.sub(r"\s+", "", name)

    # Exact single PID like B389
    m_single = re.search(r"\b([ABC])0*(\d{1,4})\b", name)
    if m_single and not re.search(r"[\-_~]", name):
        pid = normalize_pid(f"{m_single.group(1)}{m_single.group(2)}")
        n = int(m_single.group(2))
        return FilenamePIDHint(exact_letter=m_single.group(1), min_digits=n, max_digits=n, exact_pid=pid)

    # Range like A001-A004 or A237-259
    m_range = re.search(r"\b([ABC])0*(\d{1,4})\s*[-_~]\s*([ABC])?0*(\d{1,4})\b", name)
    if m_range:
        l1 = m_range.group(1)
        l2 = m_range.group(3) or l1
        n1 = int(m_range.group(2))
        n2 = int(m_range.group(4))
        if l1 == l2:
            return FilenamePIDHint(exact_letter=l1, min_digits=min(n1, n2), max_digits=max(n1, n2), exact_pid="N/A")
        return FilenamePIDHint(exact_letter="N/A", min_digits=min(n1, n2), max_digits=max(n1, n2), exact_pid="N/A")

    # Fallback: any detected leading letter gives a soft exact_letter hint
    m_letter = re.search(r"\b([ABC])0*\d{1,4}", name)
    if m_letter:
        return FilenamePIDHint(exact_letter=m_letter.group(1))
    return FilenamePIDHint()


def in_hint_digit_range(digits: str, hint: FilenamePIDHint) -> bool:
    if not digits or digits == "N/A":
        return False
    m = DIGITS_RE.search(str(digits))
    if not m:
        return False
    n = int(m.group(1))
    if hint.min_digits is not None and n < hint.min_digits:
        return False
    if hint.max_digits is not None and n > hint.max_digits:
        return False
    return True


def apply_pid_hint(value: str, wt: float, hint: FilenamePIDHint) -> Tuple[str, float]:
    if not value or value == "N/A":
        return "N/A", 0.0
    if len(value) == 1 and value in {"A", "B", "C"}:
        if hint.exact_letter != "N/A":
            if value != hint.exact_letter:
                return "N/A", 0.0
            return value, wt * 2.25
        return value, wt

    if re.fullmatch(r"\d{3,4}", value):
        if hint.min_digits is not None or hint.max_digits is not None:
            if not in_hint_digit_range(value, hint):
                return "N/A", 0.0
            return value, wt * 1.75
        return value, wt

    pid = normalize_pid(value)
    if pid == "N/A":
        return "N/A", 0.0
    if hint.exact_letter != "N/A" and pid[0] != hint.exact_letter:
        return "N/A", 0.0
    if (hint.min_digits is not None or hint.max_digits is not None) and not in_hint_digit_range(pid[1:], hint):
        return "N/A", 0.0
    boost = wt
    if hint.exact_letter != "N/A":
        boost *= 1.8
    if hint.min_digits is not None or hint.max_digits is not None:
        boost *= 1.6
    return pid, boost


def positional_pid_guess(hint: FilenamePIDHint, qidx: int, total_q: int) -> str:
    if hint.exact_pid != "N/A":
        return hint.exact_pid
    if hint.exact_letter == "N/A":
        return "REVIEW_PID"
    if hint.min_digits is None and hint.max_digits is None:
        return "REVIEW_PID"
    lo = hint.min_digits if hint.min_digits is not None else hint.max_digits
    hi = hint.max_digits if hint.max_digits is not None else hint.min_digits
    if lo is None or hi is None:
        return "REVIEW_PID"
    if total_q <= 1 or lo == hi:
        n = lo
    else:
        frac = (max(1, qidx) - 1) / max(1, total_q - 1)
        n = int(round(lo + frac * (hi - lo)))
    if n <= 999:
        return f"{hint.exact_letter}{n:03d}"
    return f"{hint.exact_letter}{n}"


class QuestionnaireExtractor:
    def __init__(self, cfg: RunConfig, log):
        self.cfg = cfg
        self.log = log
        self.client = LMStudioClient(cfg.base_url, cfg.model_id, cfg.api_key)
        self.debug_dir = os.path.join(os.path.dirname(cfg.output_excel), "debug")
        if cfg.save_debug:
            safe_mkdir(self.debug_dir)

    def test_connection(self) -> List[str]:
        return self.client.list_models()

    def extract_pid(self, page1: np.ndarray, page2: np.ndarray, page3: np.ndarray, stem: str, pdf_name: str, qidx: int = 1, total_q: int = 1) -> str:
        hint = parse_filename_pid_hint(pdf_name)
        if hint.exact_pid != "N/A":
            self.log(f"[PID] {stem} filename exact hint -> {hint.exact_pid}")
            return hint.exact_pid

        candidates = [
            ("p3_tight", crop(page3, (0.80, 0.00, 0.985, 0.10)), 0.82, 8.0),
            ("p3_wide", crop(page3, (0.74, 0.00, 0.995, 0.14)), 0.72, 7.0),
            ("p2_tight", crop(page2, (0.80, 0.00, 0.985, 0.10)), 0.72, 5.0),
            ("p2_wide", crop(page2, (0.74, 0.00, 0.995, 0.14)), 0.68, 4.0),
            ("p1_tight", crop(page1, (0.80, 0.00, 0.985, 0.09)), 0.55, 1.4),
            ("p1_wide", crop(page1, (0.72, 0.00, 0.995, 0.12)), 0.50, 1.0),
        ]

        per_image_logs: List[str] = []
        full_votes: List[Tuple[str, float]] = []
        letter_votes: List[Tuple[str, float]] = []
        digit_votes: List[Tuple[str, float]] = []
        char_votes: List[Tuple[str, float]] = []

        for base_name, region, top_ratio, base_w in candidates:
            variants = pid_crop_variants(region, keep_top_ratio=top_ratio)
            # Very aggressive bw crops are often harmful on noisy scans; keep them but downweight later.
            for variant_name, img in variants.items():
                key = f"{base_name}_{variant_name}"
                variant_factor = 1.0 if variant_name in {"raw", "soft"} else 0.45
                if self.cfg.save_debug:
                    save_debug_image(os.path.join(self.debug_dir, f"{stem}_pid_{key}.png"), img)

                try:
                    data, raw = self.client.vision_json(prompt_pid_single(), [img], max_tokens=120)
                except Exception as e:
                    per_image_logs.append(f"{key}: single failed: {e}")
                    continue

                if self.cfg.save_debug:
                    save_raw_text(os.path.join(self.debug_dir, f"{stem}_pid_{key}_single_raw.txt"), raw)

                conf = clean_confidence(data.get("confidence", 0.0))
                wt = base_w * variant_factor * (0.55 + conf)

                pid = normalize_pid(data.get("pid", ""))
                letter = normalize_pid_letter(data.get("letter", ""))
                digits = normalize_pid_digits(data.get("digits", ""))

                pid, pid_w = apply_pid_hint(pid, wt * 1.35, hint)
                if pid_is_valid(pid):
                    full_votes.append((pid, pid_w))
                    letter_votes.append((pid[0], pid_w * 0.55))
                    digit_votes.append((pid[1:], pid_w * 0.80))

                letter, letter_w = apply_pid_hint(letter, wt, hint)
                if letter != "N/A":
                    letter_votes.append((letter, letter_w))

                digits, digit_w = apply_pid_hint(digits, wt * 1.05, hint)
                if digits not in {"N/A", "000"}:
                    digit_votes.append((digits, digit_w))

                per_image_logs.append(f"{key}: single pid={pid} letter={letter} digits={digits} conf={conf:.2f}")

                need_digits = hint.exact_letter != "N/A" and digits in {"N/A", "000"}
                if need_digits and variant_name in {"raw", "soft"}:
                    try:
                        ddata, draw = self.client.vision_json(prompt_pid_digits_single(hint.exact_letter), [img], max_tokens=80)
                        if self.cfg.save_debug:
                            save_raw_text(os.path.join(self.debug_dir, f"{stem}_pid_{key}_digits_raw.txt"), draw)
                        dd = normalize_pid_digits(ddata.get("digits", ""))
                        dconf = clean_confidence(ddata.get("confidence", 0.0))
                        dd, dd_w = apply_pid_hint(dd, base_w * variant_factor * (0.50 + dconf) * 1.45, hint)
                        if dd not in {"N/A", "000"}:
                            digit_votes.append((dd, dd_w))
                            if hint.exact_letter != "N/A":
                                pid2 = normalize_pid(f"{hint.exact_letter}{dd}")
                                if pid_is_valid(pid2):
                                    full_votes.append((pid2, dd_w * 1.15))
                        per_image_logs.append(f"{key}: digits_only digits={dd} conf={dconf:.2f}")
                    except Exception as e:
                        per_image_logs.append(f"{key}: digits_only failed: {e}")

                need_chars = variant_name in {"raw", "soft"} and (not pid_is_valid(pid) or digits in {"N/A", "000"})
                if need_chars:
                    try:
                        cdata, craw = self.client.vision_json(prompt_pid_chars(), [img], max_tokens=120)
                        if self.cfg.save_debug:
                            save_raw_text(os.path.join(self.debug_dir, f"{stem}_pid_{key}_chars_raw.txt"), craw)
                        c1 = normalize_pid_letter(cdata.get("c1", ""))
                        d2 = clean_text_value(cdata.get("c2", ""))
                        d3 = clean_text_value(cdata.get("c3", ""))
                        d4 = clean_text_value(cdata.get("c4", ""))
                        cconf = clean_confidence(cdata.get("confidence", 0.0))
                        cwt = base_w * variant_factor * (0.48 + cconf)
                        c1, c1w = apply_pid_hint(c1, cwt * 1.10, hint)
                        if c1 != "N/A":
                            letter_votes.append((c1, c1w))
                        if d2 in list("0123456789") and d3 in list("0123456789") and d4 in list("0123456789"):
                            d_all = f"{d2}{d3}{d4}"
                            d_all, dw = apply_pid_hint(d_all, cwt * 1.25, hint)
                            if d_all != "N/A" and d_all != "000":
                                digit_votes.append((d_all, dw))
                                combined = normalize_pid(f"{c1}{d_all}") if c1 != "N/A" else (normalize_pid(f"{hint.exact_letter}{d_all}") if hint.exact_letter != "N/A" else "N/A")
                                combined, cw = apply_pid_hint(combined, cwt * 1.20, hint)
                                if pid_is_valid(combined):
                                    full_votes.append((combined, cw))
                                    char_votes.append((combined, cw * 1.15))
                        per_image_logs.append(f"{key}: chars c1={c1} c2={d2} c3={d3} c4={d4} conf={cconf:.2f}")
                    except Exception as e:
                        per_image_logs.append(f"{key}: chars failed: {e}")

        full_scores = vote_weighted(full_votes)
        letter_scores = vote_weighted(letter_votes)
        digit_scores = vote_weighted(digit_votes)
        char_scores = vote_weighted(char_votes)

        best_full, best_full_s, second_full_s = top_two(full_scores)
        best_letter, best_letter_s, second_letter_s = top_two(letter_scores)
        best_digits, best_digits_s, second_digit_s = top_two(digit_scores)
        best_char, best_char_s, second_char_s = top_two(char_scores)

        # Prefer filename letter when available.
        if hint.exact_letter != "N/A":
            best_letter = hint.exact_letter
            best_letter_s = max(best_letter_s, 999.0)

        combined = normalize_pid(f"{best_letter}{best_digits}") if best_letter != "N/A" and best_digits != "N/A" else "N/A"
        combined, _ = apply_pid_hint(combined, 1.0, hint)
        chosen = "REVIEW_PID"

        if pid_is_valid(best_char) and best_char_s >= max(5.0, second_char_s + 1.5):
            chosen = best_char
        elif pid_is_valid(best_full) and best_full_s >= max(6.0, second_full_s + 1.8):
            chosen = best_full
        elif pid_is_valid(combined) and best_digits_s >= max(5.0, second_digit_s + 1.2):
            chosen = combined
        elif pid_is_valid(best_full) and pid_is_valid(combined) and best_full == combined:
            chosen = best_full
        elif pid_is_valid(best_char) and pid_is_valid(combined) and best_char == combined:
            chosen = best_char

        # Hard fallback chain: never return N/A. Use filename constraints aggressively.
        if not pid_is_valid(chosen):
            if hint.exact_letter != "N/A" and best_digits != "N/A" and best_digits != "000":
                guess = normalize_pid(f"{hint.exact_letter}{best_digits}")
                if pid_is_valid(guess):
                    chosen = guess
            if not pid_is_valid(chosen) and pid_is_valid(best_full):
                chosen = best_full
            if not pid_is_valid(chosen) and hint.exact_pid != "N/A":
                chosen = hint.exact_pid
            if not pid_is_valid(chosen):
                chosen = positional_pid_guess(hint, qidx, total_q)
            if chosen == "N/A":
                chosen = "REVIEW_PID"

        self.log(f"[PID] {stem} hint={hint} full={full_scores} letter={letter_scores} digits={digit_scores} char={char_scores} -> {chosen}")
        for line in per_image_logs:
            self.log(f"[PIDDBG] {stem} {line}")
        return chosen

    def extract_page1(self, page1: np.ndarray, stem: str) -> Dict[str, object]:
        imgs = [make_vlm_image(crop(page1, P1["consent_crop"])), make_vlm_image(crop(page1, P1["bottom_crop"]))]
        data, raw = self.client.vision_json(prompt_page1(), imgs, max_tokens=260)
        if self.cfg.save_debug:
            save_raw_text(os.path.join(self.debug_dir, f"{stem}_page1_raw.txt"), raw)
        return data

    def extract_page2(self, page2: np.ndarray, stem: str) -> Dict[str, object]:
        imgs = [make_vlm_image(crop(page2, P2["top_crop"]))]
        data, raw = self.client.vision_json(prompt_page2(), imgs, max_tokens=220)
        if self.cfg.save_debug:
            save_raw_text(os.path.join(self.debug_dir, f"{stem}_page2_raw.txt"), raw)
        return data

    def extract_page3_text(self, page3: np.ndarray, stem: str) -> Dict[str, object]:
        header_img = make_vlm_image(crop(page3, P3["header_crop"]))
        lower_img = make_vlm_image(crop(page3, P3["lower_crop"]))

        d1, raw1 = self.client.vision_json(prompt_page3_header(), [header_img], max_tokens=320)
        d2, raw2 = self.client.vision_json(prompt_page3_lower(), [lower_img], max_tokens=320)
        if self.cfg.save_debug:
            save_raw_text(os.path.join(self.debug_dir, f"{stem}_page3_header_raw.txt"), raw1)
            save_raw_text(os.path.join(self.debug_dir, f"{stem}_page3_lower_raw.txt"), raw2)
            save_debug_image(os.path.join(self.debug_dir, f"{stem}_page3.png"), page3)

        out = {}
        out.update(d1)
        out.update(d2)

        # repair pass for critical fields if still weak
        critical = ["school", "name", "date_of_birth", "age", "grade"]
        if sum(clean_text_value(out.get(k)) == "N/A" for k in critical) >= 3:
            d3, raw3 = self.client.vision_json(prompt_page3_header(), [header_img], max_tokens=280)
            out.update({k: v for k, v in d3.items() if clean_text_value(v) != "N/A"})
            if self.cfg.save_debug:
                save_raw_text(os.path.join(self.debug_dir, f"{stem}_page3_header_retry_raw.txt"), raw3)
        return out

    def extract_page4_text(self, page4: np.ndarray, stem: str) -> Dict[str, object]:
        img = make_vlm_image(crop(page4, P4["right_text_crop"]))
        data, raw = self.client.vision_json(prompt_page4_text(), [img], max_tokens=240)
        if self.cfg.save_debug:
            save_raw_text(os.path.join(self.debug_dir, f"{stem}_page4_raw.txt"), raw)
            save_debug_image(os.path.join(self.debug_dir, f"{stem}_page4.png"), page4)
        return data

    def extract_page5(self, page5: np.ndarray, stem: str) -> Dict[str, int]:
        cv_out, conf = q8_q13_from_cv(page5)
        need_fallback = any(v < 0 for v in cv_out.values()) or conf < 0.013
        if not need_fallback:
            return cv_out

        img = make_vlm_image(crop(page5, P5_TABLE))
        data, raw = self.client.vision_json(prompt_page5(), [img], max_tokens=120)
        if self.cfg.save_debug:
            save_raw_text(os.path.join(self.debug_dir, f"{stem}_page5_raw.txt"), raw)
            save_debug_image(os.path.join(self.debug_dir, f"{stem}_page5_table.png"), img)

        out = {}
        for q in range(8, 14):
            key = f"q{q}"
            v = cv_out.get(key, -1)
            if isinstance(v, int) and 0 <= v <= 5:
                out[key] = v
                continue
            try:
                n = int(data.get(key, 0))
            except Exception:
                n = 0
            out[key] = max(0, min(5, n))
        return out

    def _normalize_row(self, row: Dict[str, object]) -> Dict[str, object]:
        text_fields = [
            "consent_school", "consent_grade_level", "consent_section", "student_name", "parent_guardian_name", "consent_date",
            "school", "name", "date_of_birth", "age", "grade", "class_no", "id_code",
            "family_crossed_eyes_relation", "family_lazy_eye_relation", "family_high_eye_power_relation", "family_retinal_disease_relation",
            "family_macular_disease_relation", "family_glaucoma_relation", "family_others_relation", "general_medical_conditions_text",
            "event_eye_disease_remarks", "event_eye_infection_inflammation_remarks", "event_eye_injury_trauma_remarks", "event_eye_surgery_remarks",
            "eye_medicine_text",
        ]
        for k in text_fields:
            row[k] = clean_text_value(row.get(k))

        row["participant_id"] = normalize_pid(row.get("participant_id", ""))
        row["consent_vision_screening"] = clean_choice(row.get("consent_vision_screening"), ["yes", "no"])
        row["consent_eye_photos"] = clean_choice(row.get("consent_eye_photos"), ["yes", "no"])
        row["gender"] = clean_choice(row.get("gender"), ["M", "F"])
        row["last_eye_exam"] = clean_choice(row.get("last_eye_exam"), ["never", "past_year", "1_to_2_years", "2_to_4_years", "more_than_4_years"])
        row["wear_spectacles"] = clean_choice(row.get("wear_spectacles"), ["no", "full_time", "occasional"])
        row["eye_medicine_use"] = clean_choice(row.get("eye_medicine_use"), ["yes", "no"])

        for k in [
            "family_crossed_eyes", "family_lazy_eye", "family_high_eye_power", "family_retinal_disease", "family_macular_disease",
            "family_glaucoma", "family_others", "general_normal_no_known_conditions", "general_medical_conditions_present",
            "current_symptoms_no", "current_symptoms_yes", "blurred_vision", "blurred_distance", "blurred_intermediate", "blurred_near",
            "squinting_when_viewing_objects", "working_at_close_distance_homework_tv", "using_devices_or_reading_in_poor_lighting",
            "event_eye_disease", "event_eye_infection_inflammation", "event_eye_injury_trauma", "event_eye_surgery",
        ]:
            row[k] = clean_int01(row.get(k))

        # If class_no clearly contains section words rather than a number, blank it.
        class_no = clean_text_value(row.get("class_no"))
        if class_no != "N/A" and not re.search(r"\d", class_no):
            row["class_no"] = "N/A"

        if row["general_medical_conditions_present"] == 1:
            row["general_normal_no_known_conditions"] = 0
        if row["current_symptoms_yes"] == 1:
            row["current_symptoms_no"] = 0
        if row["eye_medicine_use"] == "no":
            row["eye_medicine_text"] = "N/A"

        for q in range(8, 14):
            try:
                n = int(row.get(f"q{q}", 0))
            except Exception:
                n = 0
            row[f"q{q}"] = max(0, min(5, n))

        return row

    def extract_questionnaire(self, pdf_name: str, qidx: int, total_q: int, block_start0: int, block: List[np.ndarray]) -> Dict[str, object]:
        row = blank_row()
        row["source_pdf"] = pdf_name
        row["questionnaire_index_in_pdf"] = qidx
        row["source_pages"] = f"{block_start0 + 1}-{block_start0 + 5}"
        stem = f"{os.path.splitext(pdf_name)[0]}_q{qidx:04d}_p{block_start0 + 1}-{block_start0 + 5}"
        page1, page2, page3, page4, page5 = block[0], block[1], block[2], block[3], block[4]

        try:
            row["participant_id"] = self.extract_pid(page1, page2, page3, stem, pdf_name, qidx=qidx, total_q=total_q)
        except Exception as e:
            self.log(f"[WARN] {pdf_name} q{qidx} pid extract failed: {e}")

        try:
            row.update(self.extract_page1(page1, stem))
        except Exception as e:
            self.log(f"[WARN] {pdf_name} q{qidx} page1 extract failed: {e}")

        try:
            row.update(self.extract_page2(page2, stem))
        except Exception as e:
            self.log(f"[WARN] {pdf_name} q{qidx} page2 extract failed: {e}")

        row.update(extract_page3_checkbox_fields(page3))
        try:
            row.update(self.extract_page3_text(page3, stem))
        except Exception as e:
            self.log(f"[WARN] {pdf_name} q{qidx} page3 text extract failed: {e}")

        row.update(extract_page4_checkbox_fields(page4))
        try:
            row.update(self.extract_page4_text(page4, stem))
        except Exception as e:
            self.log(f"[WARN] {pdf_name} q{qidx} page4 text extract failed: {e}")

        try:
            row.update(self.extract_page5(page5, stem))
        except Exception as e:
            self.log(f"[WARN] {pdf_name} q{qidx} page5 extract failed: {e}")

        return self._normalize_row(row)


# ============================================================
# GUI APP
# ============================================================
class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Questionnaire Extractor - LM Studio PID V4")
        self.root.geometry("1200x820")
        self.msg_q: queue.Queue = queue.Queue()
        self.pdf_paths: List[str] = []

        self.base_url_var = tk.StringVar(value="http://127.0.0.1:1234/v1")
        self.model_var = tk.StringVar(value="qwen2.5-vl-7b-instruct")
        self.api_key_var = tk.StringVar(value="")
        self.output_var = tk.StringVar(value=str(Path.cwd() / "questionnaire_output.xlsx"))
        self.dpi_var = tk.StringVar(value=str(DPI_DEFAULT))
        self.first_page_var = tk.StringVar(value="1")
        self.pages_per_q_var = tk.StringVar(value="8")
        self.save_debug_var = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="Ready")
        self.progress_var = tk.DoubleVar(value=0.0)

        self._build_ui()
        self.root.after(120, self.poll_messages)

    def _build_ui(self):
        top = tk.Frame(self.root)
        top.pack(fill="x", padx=10, pady=8)

        r1 = tk.Frame(top)
        r1.pack(fill="x", pady=3)
        tk.Label(r1, text="LM Studio Base URL", width=18, anchor="w").pack(side="left")
        tk.Entry(r1, textvariable=self.base_url_var, width=45).pack(side="left", padx=4)
        tk.Label(r1, text="Model ID", width=10, anchor="w").pack(side="left", padx=(10, 0))
        tk.Entry(r1, textvariable=self.model_var, width=30).pack(side="left", padx=4)
        tk.Label(r1, text="Token", width=8, anchor="w").pack(side="left", padx=(10, 0))
        tk.Entry(r1, textvariable=self.api_key_var, width=28, show="*").pack(side="left", padx=4)
        tk.Button(r1, text="Test LM Studio", command=self.test_lmstudio).pack(side="left", padx=8)

        r2 = tk.Frame(top)
        r2.pack(fill="x", pady=3)
        tk.Label(r2, text="Output Excel", width=18, anchor="w").pack(side="left")
        tk.Entry(r2, textvariable=self.output_var, width=90).pack(side="left", padx=4, fill="x", expand=True)
        tk.Button(r2, text="Browse", command=self.pick_output).pack(side="left", padx=6)

        r3 = tk.Frame(top)
        r3.pack(fill="x", pady=3)
        tk.Label(r3, text="DPI", width=18, anchor="w").pack(side="left")
        tk.Entry(r3, textvariable=self.dpi_var, width=8).pack(side="left", padx=4)
        tk.Label(r3, text="First questionnaire page", width=20, anchor="w").pack(side="left")
        tk.Entry(r3, textvariable=self.first_page_var, width=8).pack(side="left", padx=4)
        tk.Label(r3, text="Pages per questionnaire", width=20, anchor="w").pack(side="left")
        tk.Entry(r3, textvariable=self.pages_per_q_var, width=8).pack(side="left", padx=4)
        ttk.Checkbutton(r3, text="Save debug raw/image files", variable=self.save_debug_var).pack(side="left", padx=10)

        files = tk.LabelFrame(self.root, text="PDF files")
        files.pack(fill="both", expand=False, padx=10, pady=8)
        btns = tk.Frame(files)
        btns.pack(fill="x", pady=4)
        tk.Button(btns, text="Add PDFs", command=self.add_pdfs).pack(side="left", padx=4)
        tk.Button(btns, text="Remove Selected", command=self.remove_selected).pack(side="left", padx=4)
        tk.Button(btns, text="Clear", command=self.clear_pdfs).pack(side="left", padx=4)

        self.files_list = tk.Listbox(files, selectmode=tk.EXTENDED, height=8)
        self.files_list.pack(fill="both", expand=True, padx=6, pady=6)

        run = tk.Frame(self.root)
        run.pack(fill="x", padx=10, pady=8)
        self.start_btn = tk.Button(run, text="Start Extraction", command=self.start_worker, width=18)
        self.start_btn.pack(side="left", padx=4)

        self.pb = ttk.Progressbar(self.root, orient="horizontal", mode="determinate", variable=self.progress_var)
        self.pb.pack(fill="x", padx=10, pady=6)
        tk.Label(self.root, textvariable=self.status_var, anchor="w").pack(fill="x", padx=10)

        logf = tk.LabelFrame(self.root, text="Log")
        logf.pack(fill="both", expand=True, padx=10, pady=8)
        self.log_list = tk.Listbox(logf, width=180, height=22)
        ys = tk.Scrollbar(logf, orient="vertical", command=self.log_list.yview)
        self.log_list.configure(yscrollcommand=ys.set)
        self.log_list.pack(side="left", fill="both", expand=True)
        ys.pack(side="right", fill="y")

    def log(self, text: str):
        self.msg_q.put(("log", text))

    def poll_messages(self):
        while True:
            try:
                kind, payload = self.msg_q.get_nowait()
            except queue.Empty:
                break
            if kind == "log":
                self.log_list.insert(tk.END, payload)
                self.log_list.yview_moveto(1.0)
            elif kind == "status":
                self.status_var.set(payload)
            elif kind == "progress":
                self.progress_var.set(payload)
            elif kind == "done_ok":
                self.start_btn.config(state="normal")
                messagebox.showinfo("Done", payload)
            elif kind == "done_err":
                self.start_btn.config(state="normal")
                messagebox.showerror("Error", payload)
        self.root.after(120, self.poll_messages)

    def add_pdfs(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
        for p in paths:
            if p not in self.pdf_paths:
                self.pdf_paths.append(p)
                self.files_list.insert(tk.END, p)

    def remove_selected(self):
        idxs = list(self.files_list.curselection())[::-1]
        for i in idxs:
            self.files_list.delete(i)
            del self.pdf_paths[i]

    def clear_pdfs(self):
        self.files_list.delete(0, tk.END)
        self.pdf_paths = []

    def pick_output(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if p:
            self.output_var.set(p)

    def test_lmstudio(self):
        try:
            client = LMStudioClient(self.base_url_var.get().strip(), self.model_var.get().strip(), self.api_key_var.get().strip())
            models = client.list_models()
            msg = "Models found:\n" + ("\n".join(models) if models else "(none)")
            messagebox.showinfo("LM Studio", msg)
        except Exception as e:
            messagebox.showerror("LM Studio test failed", str(e))

    def start_worker(self):
        if not self.pdf_paths:
            messagebox.showerror("Error", "Please add at least one PDF")
            return
        try:
            cfg = RunConfig(
                pdf_paths=list(self.pdf_paths),
                output_excel=self.output_var.get().strip(),
                base_url=self.base_url_var.get().strip(),
                model_id=self.model_var.get().strip(),
                api_key=self.api_key_var.get().strip(),
                dpi=max(120, int(self.dpi_var.get().strip())),
                first_page_1based=max(1, int(self.first_page_var.get().strip())),
                pages_per_questionnaire=max(5, int(self.pages_per_q_var.get().strip())),
                save_debug=bool(self.save_debug_var.get()),
            )
        except Exception:
            messagebox.showerror("Error", "Please check DPI / first page / pages per questionnaire")
            return
        self.start_btn.config(state="disabled")
        self.progress_var.set(0.0)
        self.status_var.set("Starting...")
        threading.Thread(target=self.run_extraction, args=(cfg,), daemon=True).start()

    def run_extraction(self, cfg: RunConfig):
        try:
            safe_mkdir(os.path.dirname(cfg.output_excel) or ".")
            ensure_excel(cfg.output_excel)
            log_path = os.path.splitext(cfg.output_excel)[0] + "_log.txt"
            all_log_lines: List[str] = []

            def log_local(msg: str):
                all_log_lines.append(msg)
                self.log(msg)
                with open(log_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(all_log_lines))

            extractor = QuestionnaireExtractor(cfg, log_local)
            models = extractor.test_connection()
            log_local(f"LM Studio reachable. Models: {', '.join(models) if models else '(none)'}")

            total_questionnaires_est = 0
            for pdf in cfg.pdf_paths:
                doc = fitz.open(pdf)
                total_questionnaires_est += max(0, (doc.page_count - (cfg.first_page_1based - 1)) // cfg.pages_per_questionnaire)
                doc.close()
            total_questionnaires_est = max(total_questionnaires_est, 1)

            done = 0
            for pdf_path in cfg.pdf_paths:
                pdf_name = os.path.basename(pdf_path)
                log_local(f"Starting {pdf_name}")
                doc = fitz.open(pdf_path)
                page_count = doc.page_count
                start0 = cfg.first_page_1based - 1
                rows_to_append = []
                qidx = 0
                for block_start0 in range(start0, page_count, cfg.pages_per_questionnaire):
                    if block_start0 + 4 >= page_count:
                        break
                    qidx += 1
                    self.msg_q.put(("status", f"{pdf_name} - questionnaire {qidx}"))
                    block = [render_page(doc, block_start0 + i, cfg.dpi) for i in range(5)]
                    total_q_pdf = max(1, (page_count - start0) // cfg.pages_per_questionnaire)
                    row = extractor.extract_questionnaire(pdf_name, qidx, total_q_pdf, block_start0, block)
                    rows_to_append.append(row)
                    done += 1
                    self.msg_q.put(("progress", min(100.0, done * 100.0 / total_questionnaires_est)))
                    log_local(f"Extracted {pdf_name} q{qidx}: participant_id={row.get('participant_id', 'N/A')}")
                doc.close()
                append_rows_excel(cfg.output_excel, rows_to_append)
                log_local(f"Finished {pdf_name} -> {len(rows_to_append)} questionnaire(s)")

            self.msg_q.put(("status", "Finished"))
            self.msg_q.put(("progress", 100.0))
            self.msg_q.put(("done_ok", f"Done. Output:\n{cfg.output_excel}\n\nLog:\n{os.path.splitext(cfg.output_excel)[0] + '_log.txt'}"))
        except Exception as e:
            tb = traceback.format_exc()
            self.msg_q.put(("done_err", f"{e}\n\n{tb}"))


# ============================================================
# MAIN
# ============================================================
def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
