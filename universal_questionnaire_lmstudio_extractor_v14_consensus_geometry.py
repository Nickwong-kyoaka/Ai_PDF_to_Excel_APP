# -*- coding: utf-8 -*-
"""
Universal Questionnaire Full-Page Local VLM Extractor v14
-------------------------------------------------------
For LM Studio local Vision LLM / VLM.

Goal:
    A universal extractor for scanned questionnaires / forms.
    It does NOT need page-specific questionnaire instructions, column templates, or fixed ROIs.
    It sends each full page to a local VLM, lets the model discover the form structure,
    and outputs a dynamic Excel workbook.

Key behavior:
    - Works with different questionnaire layouts.
    - Extracts identity fields, handwritten text, selected checkboxes, circled answers,
      radio options, tables, dates, numbers, and free-text answers.
    - No final "UNCLEAR" values. If truly blank/unanswered, N/A is allowed.
    - If ambiguous but not blank, the VLM must choose the most likely visible answer;
      conflicts are recorded in the Conflicts sheet, but final cells remain clean.
    - Optional SEC / second-pass verification with item-level reconciliation.
    - Dedicated selection-mark verification for circled/ticked gender, yes/no, and checkboxes.
    - Dedicated scale-table verification with table-geometry localization, first/SEC consensus locks, and independent visual confirmation.
    - Optional Data_Analysis dashboard with descriptive statistics, QA metrics, charts, and reference-based error metrics.
    - Robust JSON recovery: compact retry, finish-reason detection, and region fallback.
    - A failed page no longer aborts the whole questionnaire.
    - Resume, checkpoint NDJSON, failed jobs, QA summary.

Install:
    pip install pymupdf pillow requests openpyxl

Package:
    auto-py-to-exe -> One File -> Window Based

Recommended LM Studio:
    Base URL: http://127.0.0.1:1234/v1
    Temperature: 0
    Context Length: 32768
    Max Concurrent: 1
    Flash Attention: ON
    Structured Output: OFF
"""

import base64
import hashlib
import io
import json
import math
import os
import queue
import re
import statistics
import threading
import time
import traceback
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import fitz  # PyMuPDF
import requests
from PIL import Image, ImageOps
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
import tkinter as tk
from tkinter import ttk, filedialog, messagebox


APP_TITLE = "Universal Questionnaire LM Studio Extractor v14 - Consensus Geometry"
DEFAULT_BASE_URL = "http://127.0.0.1:1234/v1"
DEFAULT_MODEL_ID = "qwen2.5-vl-7b-instruct"
DEFAULT_DPI = 240
DEFAULT_IMAGE_MAX_SIDE = 3000
DEFAULT_MAX_TOKENS = 8192
DEFAULT_TIMEOUT = 600
DEFAULT_PAGES_PER_QUESTIONNAIRE = 6
DEFAULT_SAVE_EVERY_N = 5

SHEET_FORMS = "Questionnaires"
SHEET_LONG = "Long_Answers"
SHEET_PAGES = "Page_Extracts"
SHEET_CONFLICTS = "Conflicts"
SHEET_FAILED = "Failed_Jobs"
SHEET_QA = "QA_Summary"
SHEET_ANALYSIS = "Data_Analysis"
SHEET_ANALYSIS_DATA = "_Analysis_Data"
SHEET_LOG = "Run_Log"

CONTROL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
UNCLEAR_RE = re.compile(r"\b(unclear|unknown|unreadable|not sure|can't tell|cannot tell|ambiguous)\b", re.I)
PID_RE = re.compile(r"\b(CSA|[ABCabc])\s*[-_ ]?\s*0*(\d{1,4})\b", re.I)

SELECTION_MARK_TYPES = {"checkbox", "checkbox_multi_select", "radio", "yes_no", "consent", "circle_scale", "table_scale", "likert_scale"}

SCALE_MARK_TYPES = {"circle_scale", "table_scale", "likert_scale"}

OUTPUT_LANGUAGE_OPTIONS = [
    "Preserve source language",
    "English",
    "Traditional Chinese",
    "Simplified Chinese",
    "Custom",
]

OUTPUT_LANGUAGE_MAP = {
    "Preserve source language": "",
    "English": "English",
    "Traditional Chinese": "Traditional Chinese",
    "Simplified Chinese": "Simplified Chinese",
    "Custom": "",
}


# -----------------------------------------------------------------------------
# Utility
# -----------------------------------------------------------------------------
def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def resolve_output_language(mode: Any, custom: Any = "") -> str:
    mode_s = clean_text(mode, allow_blank=True)
    custom_s = clean_text(custom, allow_blank=True)
    if mode_s == "Custom":
        return custom_s
    return OUTPUT_LANGUAGE_MAP.get(mode_s, mode_s if mode_s not in {"", "N/A"} else "")


def output_language_prompt(target_language: str) -> str:
    target = clean_text(target_language, allow_blank=True)
    if not target:
        return """
OUTPUT LANGUAGE:
- Preserve the source-language wording of labels and section titles.
- For bilingual pages, prefer the clearest short label and do not duplicate the same item in two languages.
""".strip()
    return f"""
OUTPUT LANGUAGE FOR THE FINAL EXCEL CONTENT: {target}
- Write document titles, section names, item labels, categorical option labels, and short reasons in {target}.
- Translate printed questionnaire wording into {target} when needed so the Excel content is consistent.
- Preserve personal names, participant IDs, school/proper names, addresses, dates, times, medication names, and free-text handwriting exactly as written.
- Preserve numeric scale values exactly as printed; never translate or renumber scores.
- Do not duplicate an item once in the source language and again in {target}.
""".strip()


def language_normalization_prompt(compact: Dict[str, Any], target_language: str) -> str:
    target = clean_text(target_language, allow_blank=True)
    return f"""
Normalize the LANGUAGE of this already-extracted questionnaire page JSON to {target}.
This is a text-only normalization step. DO NOT change any answer meaning or numeric score.

STRICT RULES:
1. Return one valid JSON object only.
2. Keep meta.page_language as the detected SOURCE language. Add meta.output_language = "{target}".
3. Translate only: meta.document_title, item.section, item.label, item.reason, categorical item.options, and categorical selected item.value.
4. Preserve exactly: participant_id, all identity values, personal names, school/proper names, dates, times, medication names, free-text handwriting, numeric answers, item ids, confidence, blank flags, and item types.
5. Numeric scale answers and printed score values must remain exactly unchanged.
6. Do not add or remove items. Do not merge different questions.
7. For duplicated bilingual wording inside one label, keep one concise {target} label.
8. N/A remains N/A.

JSON:
""".strip() + "\n" + json.dumps(compact, ensure_ascii=False, separators=(",", ":"))



def apply_language_normalization_safely(
    original: Dict[str, Any],
    translated: Dict[str, Any],
    target_language: str,
) -> Dict[str, Any]:
    """Copy only language-bearing fields from a translated JSON; preserve answers."""
    orig = cleanup_compact_page(original)
    trans = cleanup_compact_page(translated)
    out = cleanup_compact_page(orig)

    # Preserve detected source language; only title becomes target-language text.
    out["meta"]["document_title"] = trans.get("meta", {}).get("document_title", out["meta"].get("document_title", "N/A"))
    out["meta"]["output_language"] = clean_text(target_language)

    trans_items = trans.get("items", [])
    used = set()
    for idx, item in enumerate(out.get("items", [])):
        match_idx = None
        iid = clean_text(item.get("id"))
        # Prefer stable id and position. Repeated ids are matched in sequence.
        for j, t in enumerate(trans_items):
            if j in used:
                continue
            if clean_text(t.get("id")) == iid:
                match_idx = j
                break
        if match_idx is None and idx < len(trans_items):
            match_idx = idx
        if match_idx is None or match_idx >= len(trans_items):
            continue
        t = trans_items[match_idx]
        used.add(match_idx)

        item["section"] = clean_text(t.get("section")) if clean_text(t.get("section")) != "N/A" else item.get("section", "N/A")
        item["label"] = clean_text(t.get("label")) if clean_text(t.get("label")) != "N/A" else item.get("label", "N/A")
        item["reason"] = clean_text(t.get("reason"), allow_blank=True)

        item_type = clean_text(item.get("type")).lower()
        original_options = item.get("options", []) if isinstance(item.get("options"), list) else []
        translated_options = t.get("options", []) if isinstance(t.get("options"), list) else []
        numeric_scale = (
            item_type in SCALE_MARK_TYPES
            and original_options
            and all(re.fullmatch(r"[-+]?\d+(?:\.\d+)?", clean_text(x)) for x in original_options)
        )
        if translated_options and len(translated_options) == len(original_options) and not numeric_scale:
            item["options"] = [clean_text(x) for x in translated_options]

        # Translate categorical answer labels only; never touch numeric answers or free text.
        old_value = clean_text(item.get("value"))
        new_value = clean_text(t.get("value"))
        if (
            old_value != "N/A"
            and new_value != "N/A"
            and not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", old_value)
            and item_type in {"checkbox", "checkbox_multi_select", "radio", "yes_no", "consent"}
        ):
            item["value"] = new_value

    out["quality_flags"] = sorted(set(out.get("quality_flags", []) + [f"output_language:{clean_text(target_language)}"]))
    out = dedupe_compact_items(out)
    out = remove_malformed_duplicate_items(out)
    out = repair_scale_sections_compact(out)
    return cleanup_compact_page(out)


def _strip_translation_parentheticals(value: Any) -> str:
    s = clean_text(value)
    if s == "N/A":
        return s
    # Remove long bilingual parenthetical repeats but preserve short clinical qualifiers.
    s = re.sub(r"\s*[\(\[][^)\]]{18,}[\)\]]\s*", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def dedupe_compact_items(compact: Dict[str, Any]) -> Dict[str, Any]:
    """Collapse exact/translated duplicates without merging genuinely different repeated IDs."""
    c = cleanup_compact_page(compact)
    kept: Dict[Tuple[str, str], Dict[str, Any]] = {}
    order: List[Tuple[str, str]] = []
    for item in c.get("items", []):
        iid = clean_text(item.get("id"))
        label_key = norm_compare(_strip_translation_parentheticals(item.get("label")))
        key = (iid, label_key)
        if key not in kept:
            kept[key] = dict(item)
            order.append(key)
            continue
        old = kept[key]
        same_answer = _item_signature(old) == _item_signature(item)
        if same_answer:
            # Keep the higher-confidence / cleaner language version.
            old_score = float(old.get("confidence", 0.0))
            new_score = float(item.get("confidence", 0.0))
            if new_score > old_score or len(clean_text(item.get("section"))) < len(clean_text(old.get("section"))):
                kept[key] = dict(item)
            continue
        # Same ID/label but different answers: keep both by adding section to the key.
        key2 = (iid + "|" + norm_compare(item.get("section")), label_key)
        if key2 not in kept:
            kept[key2] = dict(item)
            order.append(key2)
    c["items"] = [kept[k] for k in order if k in kept]
    return cleanup_compact_page(c)


def _looks_like_scale_option_header(value: Any, options: Optional[List[Any]] = None) -> bool:
    s = norm_compare(value)
    if s in {"", "n/a"}:
        return False
    common = [
        "never", "sometimes", "all the time", "none of the time", "a little of the time",
        "some of the time", "a lot of the time", "most of the time",
        "從不", "偶爾", "經常", "很少", "間中", "頻繁", "總是",
        "每週少於一次", "每週一或二次", "每週三次或以上", "從未發生過",
        "hindi kailanman", "paminsan-minsan", "palagi",
    ]
    if any(norm_compare(x) == s or norm_compare(x) in s for x in common):
        return True
    for x in options or []:
        nx = norm_compare(x)
        if nx and (s == nx or s in nx or nx in s):
            return True
    return False


def remove_malformed_duplicate_items(compact: Dict[str, Any]) -> Dict[str, Any]:
    """Remove a duplicate whose label belongs to one question but answer signature duplicates another.

    Example fixed: a spurious Q3b 'Last Eye Examination = No symptoms' duplicated the real
    Current Symptoms item while Q3a already represented Last Eye Examination.
    """
    c = cleanup_compact_page(compact)
    items = [dict(x) for x in c.get("items", [])]
    sig_to_indices: Dict[str, List[int]] = defaultdict(list)
    label_to_indices: Dict[str, List[int]] = defaultdict(list)
    for i, item in enumerate(items):
        sig_to_indices[_item_signature(item)].append(i)
        label_to_indices[norm_compare(item.get("label"))].append(i)

    drop = set()
    for label_key, idxs in label_to_indices.items():
        if label_key in {"", "n/a"} or len(idxs) < 2:
            continue
        for i in idxs:
            item = items[i]
            sig = _item_signature(item)
            outside_same_sig = [j for j in sig_to_indices.get(sig, []) if j not in idxs]
            iid = clean_text(item.get("id"))
            has_suffix = bool(re.fullmatch(r"[QqRr]?\d+[A-Za-z]", iid) or "_detail" in iid.lower())
            if outside_same_sig and has_suffix:
                drop.add(i)

    if drop:
        c["items"] = [x for i, x in enumerate(items) if i not in drop]
        c["quality_flags"] = sorted(set(c.get("quality_flags", []) + ["malformed_duplicate_removed"]))
    return cleanup_compact_page(c)


def repair_scale_sections_compact(compact: Dict[str, Any], inherited_title: str = "") -> Dict[str, Any]:
    c = cleanup_compact_page(compact)
    if not _looks_like_scale_table_page(c):
        return c
    title = clean_text(c.get("meta", {}).get("document_title"))
    if title == "N/A":
        title = clean_text(inherited_title)
    if title == "N/A":
        title = "Scale table"
    for item in c.get("items", []):
        section = clean_text(item.get("section"))
        opts = item.get("options", []) if isinstance(item.get("options"), list) else []
        if section == "N/A" or _looks_like_scale_option_header(section, opts):
            item["section"] = title
    return cleanup_compact_page(c)


def apply_previous_page_context_compact(compact: Dict[str, Any], previous_page: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    c = cleanup_compact_page(compact)
    if not previous_page:
        return c
    prev_type = clean_text(previous_page.get("page_type")).lower()
    prev_title = clean_text(previous_page.get("document_title"))
    if "scale_table" in prev_type and _looks_like_scale_table_page(c):
        c.setdefault("meta", {})["page_type"] = "scale_table"
        if clean_text(c["meta"].get("document_title")) == "N/A" and prev_title != "N/A":
            c["meta"]["document_title"] = prev_title
    return repair_scale_sections_compact(c, inherited_title=prev_title)


def postprocess_questionnaire_pages(pages: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Questionnaire-level context repair after all pages are extracted."""
    out: List[Dict[str, Any]] = []
    last_scale_title = "N/A"
    for page in pages:
        p = dict(page)
        answers = p.get("answers", []) if isinstance(p.get("answers"), list) else []
        scaleish = sum(1 for a in answers if clean_text(a.get("answer_type")).lower() in SCALE_MARK_TYPES | {"yes_no", "radio"})
        numeric = sum(1 for a in answers if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", clean_text(a.get("selected_value"))))
        looks_scale = "scale_table" in clean_text(p.get("page_type")).lower() or (len(answers) >= 6 and scaleish >= math.ceil(len(answers) * 0.60) and numeric >= math.ceil(len(answers) * 0.35))
        if looks_scale:
            p["page_type"] = "scale_table"
            if clean_text(p.get("document_title")) == "N/A" and last_scale_title != "N/A":
                p["document_title"] = last_scale_title
            if clean_text(p.get("document_title")) != "N/A":
                last_scale_title = clean_text(p.get("document_title"))
            for a in answers:
                section = clean_text(a.get("section"))
                opts = a.get("selected_options", []) if isinstance(a.get("selected_options"), list) else []
                if section == "N/A" or _looks_like_scale_option_header(section, opts):
                    a["section"] = last_scale_title if last_scale_title != "N/A" else "Scale table"
                if clean_text(a.get("answer_type")).lower() in {"yes_no", "radio"} and re.fullmatch(r"[-+]?\d+(?:\.\d+)?", clean_text(a.get("selected_value"))):
                    a["answer_type"] = "circle_scale"
        out.append(p)
    return out


def safe_mkdir(path: str) -> None:
    if path:
        os.makedirs(path, exist_ok=True)


def clean_text(value: Any, allow_blank: bool = False) -> str:
    if value is None:
        return "" if allow_blank else "N/A"
    if isinstance(value, (dict, list)):
        s = json.dumps(value, ensure_ascii=False)
    else:
        s = str(value)
    s = CONTROL_CHARS_RE.sub(" ", s)
    s = s.replace("\u0000", " ").strip()
    s = re.sub(r"\s+", " ", s)
    if len(s) > 32000:
        s = s[:32000] + " ...[TRUNCATED]"
    if allow_blank:
        return s
    if not s:
        return "N/A"
    if s.lower() in {"n/a", "na", "none", "null", "blank", "empty", "not applicable"}:
        return "N/A"
    if UNCLEAR_RE.search(s):
        return "N/A"
    return s


def normalize_pid(value: Any) -> str:
    """Normalize common participant IDs while preserving the useful zero padding.

    Examples:
        CSA83  -> CSA083
        CSA083 -> CSA083
        A1     -> A001
    """
    s = clean_text(value)
    if s == "N/A":
        return "N/A"
    s2 = s.upper().replace("Ｃ", "C").replace("Ｓ", "S").replace("Ａ", "A")
    s2 = s2.replace("－", "-").replace("_", " ").replace("-", " ")
    m = PID_RE.search(s2)
    if not m:
        return s
    prefix = m.group(1).upper()
    n = int(m.group(2))
    if n <= 0:
        return "N/A"
    if prefix == "CSA":
        return f"CSA{n:03d}" if n <= 999 else f"CSA{n}"
    return f"{prefix}{n:03d}" if n <= 999 else f"{prefix}{n}"


def sanitize_key(value: Any, max_len: int = 90) -> str:
    s = clean_text(value)
    if s == "N/A":
        return "unknown"
    s = re.sub(r"[\s\n\r\t]+", "_", s)
    s = re.sub(r"[^0-9A-Za-z_\u4e00-\u9fff\u3400-\u4dbf]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return (s or "unknown")[:max_len]


def norm_compare(value: Any) -> str:
    if value is None:
        return "N/A"
    if isinstance(value, list):
        vals = [norm_compare(v) for v in value]
        vals = [v for v in vals if v != "N/A"]
        return "|".join(sorted(vals)) if vals else "N/A"
    if isinstance(value, dict):
        d = {str(k): norm_compare(v) for k, v in value.items() if norm_compare(v) != "N/A"}
        return json.dumps(d, ensure_ascii=False, sort_keys=True) if d else "N/A"
    s = str(value).strip().lower()
    s = s.replace(" ", "").replace("：", ":").replace("／", "/")
    trad_map = str.maketrans({
        "鍾": "钟", "鐘": "钟", "锺": "钟",
        "視": "视", "學": "学", "藥": "药", "醫": "医",
        "號": "号", "歲": "岁", "歲": "岁", "時": "时",
        "間": "间", "頭": "头", "讀": "读", "網": "网",
        "無": "无", "與": "与", "體": "体", "會": "会",
        "嗎": "吗", "這": "这", "為": "为", "個": "个",
    })
    s = s.translate(trad_map)
    s = re.sub(r"[。．,，\s]+", "", s)
    if not s or s in {"n/a", "na", "none", "null", "blank", "empty"}:
        return "N/A"
    if UNCLEAR_RE.search(s):
        return "N/A"
    return s


def flatten_json(prefix: str, obj: Any, out: Dict[str, str], max_items: int = 200) -> None:
    """Flatten arbitrary JSON into Excel columns safely."""
    if len(out) > 2000:
        return
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = sanitize_key(k)
            flatten_json(f"{prefix}_{key}" if prefix else key, v, out, max_items=max_items)
    elif isinstance(obj, list):
        # Keep small scalar lists compact; expand dict lists.
        if all(not isinstance(x, (dict, list)) for x in obj):
            out[prefix] = clean_text("; ".join(clean_text(x) for x in obj if clean_text(x) != "N/A"))
        else:
            for i, v in enumerate(obj[:max_items], start=1):
                flatten_json(f"{prefix}_{i:02d}", v, out, max_items=max_items)
    else:
        out[prefix] = clean_text(obj)


def extract_text_values(obj: Any) -> List[str]:
    vals: List[str] = []
    if isinstance(obj, dict):
        for v in obj.values():
            vals.extend(extract_text_values(v))
    elif isinstance(obj, list):
        for v in obj:
            vals.extend(extract_text_values(v))
    else:
        s = clean_text(obj)
        if s != "N/A":
            vals.append(s)
    return vals


def stable_choice(candidates: List[str], key: str) -> str:
    vals = [clean_text(c) for c in candidates if clean_text(c) != "N/A"]
    if not vals:
        return "N/A"
    counts = Counter(vals)
    top_count = max(counts.values())
    tops = sorted([v for v, c in counts.items() if c == top_count])
    if len(tops) == 1:
        return tops[0]
    h = int(hashlib.sha256((key + "|" + "|".join(tops)).encode("utf-8")).hexdigest(), 16)
    return tops[h % len(tops)]


# -----------------------------------------------------------------------------
# Image / PDF
# -----------------------------------------------------------------------------
def image_to_base64_png(img: Image.Image) -> str:
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def resize_max_side(img: Image.Image, max_side: int) -> Image.Image:
    if max_side <= 0:
        return img
    w, h = img.size
    m = max(w, h)
    if m <= max_side:
        return img
    scale = max_side / float(m)
    return img.resize((max(1, int(w * scale)), max(1, int(h * scale))), Image.Resampling.LANCZOS)




def make_zoom_tiles(img: Image.Image, max_tiles: int = 4) -> List[Image.Image]:
    """Return zoomed crops of the same page to help VLM read dense tables/handwriting.
    These are not used for computer-vision detection; the AI still decides all answers.
    """
    w, h = img.size
    tiles: List[Image.Image] = []
    if w <= 0 or h <= 0:
        return tiles
    # Portrait forms: top/middle/bottom are usually most useful.
    if h >= w:
        crops = [
            (0, 0, w, int(h * 0.38)),
            (0, int(h * 0.28), w, int(h * 0.72)),
            (0, int(h * 0.60), w, h),
            (int(w * 0.45), 0, w, h),  # right column / answer columns
        ]
    else:
        # Landscape/rotated wide pages: left/middle/right plus lower band.
        crops = [
            (0, 0, int(w * 0.42), h),
            (int(w * 0.28), 0, int(w * 0.72), h),
            (int(w * 0.58), 0, w, h),
            (0, int(h * 0.45), w, h),
        ]
    for box in crops[:max_tiles]:
        x1, y1, x2, y2 = box
        x1, y1 = max(0, x1), max(0, y1)
        x2, y2 = min(w, x2), min(h, y2)
        if x2 - x1 < 50 or y2 - y1 < 50:
            continue
        crop = img.crop((x1, y1, x2, y2))
        # Upscale crop slightly so small text/marks become easier for the VLM.
        cw, ch = crop.size
        scale = 1.35
        crop = crop.resize((int(cw * scale), int(ch * scale)), Image.Resampling.LANCZOS)
        tiles.append(crop)
    return tiles



def make_mark_focus_tiles(img: Image.Image, max_tiles: int = 5) -> List[Image.Image]:
    """High-detail crops for physical selection marks.

    The VLM still decides the answer. These crops only enlarge likely answer areas:
    top identity bands (gender/sex), right-side answer columns, and dense table bands.
    """
    w, h = img.size
    if w <= 0 or h <= 0:
        return []
    boxes: List[Tuple[int, int, int, int]] = []
    if h >= w:
        boxes = [
            (0, 0, w, int(h * 0.48)),                              # identity + upper questions
            (int(w * 0.42), int(h * 0.10), w, int(h * 0.52)),      # upper-right: gender / yes-no
            (int(w * 0.42), 0, w, h),                              # right answer columns
            (0, int(h * 0.28), w, int(h * 0.70)),                  # middle dense rows
            (0, int(h * 0.58), w, h),                              # lower rows
        ]
    else:
        boxes = [
            (0, 0, int(w * 0.52), h),
            (int(w * 0.38), 0, w, h),
            (0, 0, w, int(h * 0.58)),
            (0, int(h * 0.38), w, h),
            (int(w * 0.20), int(h * 0.15), int(w * 0.85), int(h * 0.85)),
        ]
    out: List[Image.Image] = []
    for box in boxes[:max_tiles]:
        x1, y1, x2, y2 = box
        x1, y1 = max(0, x1), max(0, y1)
        x2, y2 = min(w, x2), min(h, y2)
        if x2 - x1 < 80 or y2 - y1 < 80:
            continue
        crop = img.crop((x1, y1, x2, y2))
        cw, ch = crop.size
        scale = 1.6
        out.append(crop.resize((max(1, int(cw * scale)), max(1, int(ch * scale))), Image.Resampling.LANCZOS))
    return out



def make_scale_table_tiles(img: Image.Image, max_tiles: int = 6) -> List[Image.Image]:
    """High-detail crops for dense matrix/Likert tables.

    The AI still decides every answer. These crops preserve the row label together
    with the answer columns, which is critical for continuation pages where column
    values may run in a non-obvious order such as 2, 1, 0 from left to right.
    """
    w, h = img.size
    if w <= 0 or h <= 0:
        return []

    boxes: List[Tuple[int, int, int, int]] = []
    if h >= w:
        # Portrait table pages: horizontal bands retain row id/text and the answer cell.
        boxes = [
            (0, 0, w, int(h * 0.40)),
            (0, int(h * 0.22), w, int(h * 0.60)),
            (0, int(h * 0.43), w, int(h * 0.81)),
            (0, int(h * 0.64), w, h),
            (int(w * 0.55), 0, w, h),  # answer columns / headers
        ]
    else:
        # Upright landscape table pages: split by rows, not by columns.
        boxes = [
            (0, 0, w, int(h * 0.43)),
            (0, int(h * 0.24), w, int(h * 0.67)),
            (0, int(h * 0.48), w, int(h * 0.88)),
            (0, int(h * 0.70), w, h),
            (int(w * 0.62), 0, w, h),  # score columns and their printed digits
        ]

    out: List[Image.Image] = []
    for box in boxes[:max_tiles]:
        x1, y1, x2, y2 = box
        x1, y1 = max(0, x1), max(0, y1)
        x2, y2 = min(w, x2), min(h, y2)
        if x2 - x1 < 100 or y2 - y1 < 100:
            continue
        crop = img.crop((x1, y1, x2, y2))
        cw, ch = crop.size
        scale = 1.85
        out.append(crop.resize((max(1, int(cw * scale)), max(1, int(ch * scale))), Image.Resampling.LANCZOS))
    return out

def make_region_tiles(img: Image.Image) -> List[Image.Image]:
    """Three broad non-destructive regions for last-resort extraction recovery."""
    w, h = img.size
    if w <= 0 or h <= 0:
        return []
    boxes: List[Tuple[int, int, int, int]] = []
    overlap = 0.06
    if h >= w:
        # Horizontal bands for portrait forms.
        starts = [0.0, 0.31, 0.62]
        ends = [0.38, 0.69, 1.0]
        for a, b in zip(starts, ends):
            boxes.append((0, int(max(0, a - overlap) * h), w, int(min(1, b + overlap) * h)))
    else:
        starts = [0.0, 0.31, 0.62]
        ends = [0.38, 0.69, 1.0]
        for a, b in zip(starts, ends):
            boxes.append((int(max(0, a - overlap) * w), 0, int(min(1, b + overlap) * w), h))
    out: List[Image.Image] = []
    for box in boxes:
        crop = img.crop(box)
        cw, ch = crop.size
        scale = 1.45
        out.append(crop.resize((max(1, int(cw * scale)), max(1, int(ch * scale))), Image.Resampling.LANCZOS))
    return out


def enhance_image(img: Image.Image) -> Image.Image:
    img = ImageOps.grayscale(img)
    img = ImageOps.autocontrast(img, cutoff=1)
    return img.convert("RGB")


def render_pdf_page(doc: fitz.Document, page_index0: int, dpi: int, max_side: int, enhance: bool) -> Image.Image:
    page = doc.load_page(page_index0)
    pix = page.get_pixmap(matrix=fitz.Matrix(dpi / 72.0, dpi / 72.0), alpha=False)
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    if enhance:
        img = enhance_image(img)
    img = resize_max_side(img, max_side)
    return img


# -----------------------------------------------------------------------------
# LM Studio client
# -----------------------------------------------------------------------------
class ModelJSONError(RuntimeError):
    """Carries the raw model output so the caller can save it for debugging."""

    def __init__(self, message: str, raw: str = "", finish_reason: str = "", status_code: Optional[int] = None):
        super().__init__(message)
        self.raw = raw
        self.finish_reason = finish_reason
        self.status_code = status_code


def _extract_first_balanced_object(text: str) -> Optional[str]:
    """Return the first complete JSON object, respecting quoted strings."""
    start = text.find("{")
    if start < 0:
        return None
    depth = 0
    in_string = False
    escape = False
    for i in range(start, len(text)):
        ch = text[i]
        if in_string:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_string = False
            continue
        if ch == '"':
            in_string = True
        elif ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return text[start:i + 1]
    return None


def parse_json_response(text: str) -> Dict[str, Any]:
    """Parse a model response without trusting markdown fences or trailing prose.

    A truncated object is deliberately NOT guessed/repaired into fake data. It is
    surfaced to the resilient extraction layer, which retries with a smaller schema
    and can fall back to page regions.
    """
    original = str(text or "")
    s = original.strip().lstrip("\ufeff")
    s = re.sub(r"^\s*```(?:json|javascript|js)?\s*", "", s, flags=re.I)
    s = re.sub(r"\s*```\s*$", "", s)
    s = s.replace("\u201c", '"').replace("\u201d", '"').replace("\u2018", "'").replace("\u2019", "'")
    s = CONTROL_CHARS_RE.sub(" ", s)

    candidate = _extract_first_balanced_object(s)
    if candidate is None:
        if "{" in s:
            raise ValueError("Truncated JSON object: no matching closing brace")
        raise ValueError("No JSON object found in model response")

    candidate = re.sub(r",\s*([}\]])", r"\1", candidate)
    candidate = re.sub(r"\bNaN\b|\bInfinity\b|-\bInfinity\b", "null", candidate)
    try:
        obj = json.loads(candidate)
    except Exception as e:
        # A second conservative cleanup: remove comments that some models emit.
        cleaned = re.sub(r"/\*.*?\*/", "", candidate, flags=re.S)
        cleaned = re.sub(r"(^|\s)//[^\n\r]*", r"\1", cleaned)
        cleaned = re.sub(r",\s*([}\]])", r"\1", cleaned)
        try:
            obj = json.loads(cleaned)
        except Exception as e2:
            raise ValueError(f"Invalid JSON syntax: {str(e2)[:240]}") from e
    if not isinstance(obj, dict):
        raise ValueError("Model JSON is not an object")
    return obj


class LMStudioClient:
    def __init__(self, base_url: str, model: str, api_key: str = "", timeout: int = DEFAULT_TIMEOUT, temperature: float = 0.0):
        base = (base_url or DEFAULT_BASE_URL).strip().rstrip("/")
        self.base_url = base if base.endswith("/v1") else base + "/v1"
        self.model = (model or DEFAULT_MODEL_ID).strip()
        self.api_key = (api_key or "").strip()
        self.timeout = max(30, int(timeout))
        self.temperature = float(temperature)
        # None = not tested, True = works, False = server/model rejected it.
        self._json_mode_supported: Optional[bool] = None
        self.last_finish_reason = ""

    def headers(self) -> Dict[str, str]:
        h = {"Content-Type": "application/json"}
        if self.api_key:
            h["Authorization"] = f"Bearer {self.api_key}"
        return h

    def list_models(self) -> List[str]:
        r = requests.get(f"{self.base_url}/models", headers=self.headers(), timeout=20)
        if r.status_code != 200:
            raise RuntimeError(f"HTTP {r.status_code}: {r.text[:500]}")
        data = r.json()
        return [str(x.get("id")) for x in data.get("data", []) if isinstance(x, dict) and x.get("id")]

    @staticmethod
    def _content_to_text(raw: Any) -> str:
        if isinstance(raw, str):
            return raw
        if isinstance(raw, list):
            parts: List[str] = []
            for x in raw:
                if isinstance(x, dict):
                    if x.get("text") is not None:
                        parts.append(str(x.get("text")))
                    elif x.get("content") is not None:
                        parts.append(str(x.get("content")))
                elif x is not None:
                    parts.append(str(x))
            return "\n".join(parts)
        return str(raw or "")

    def vision_json(self, prompt: str, images: List[Image.Image], max_tokens: int = DEFAULT_MAX_TOKENS, retries: int = 1) -> Tuple[Dict[str, Any], str]:
        """Call an OpenAI-compatible LM Studio endpoint and parse one JSON object.

        Newer models/LM Studio builds often support response_format=json_object.
        v10 tries it automatically and falls back without it if unsupported, so the
        same EXE can run Qwen2.5-VL, newer Qwen-VL models, and other compatible VLMs.
        """
        content: List[Dict[str, Any]] = [{"type": "text", "text": prompt}]
        for img in images:
            content.append({
                "type": "image_url",
                "image_url": {"url": "data:image/png;base64," + image_to_base64_png(img)},
            })

        last_err: Optional[Exception] = None
        last_raw = ""
        last_finish = ""
        attempts = max(1, int(retries) + 1)
        for _ in range(attempts):
            payload: Dict[str, Any] = {
                "model": self.model,
                "messages": [{"role": "user", "content": content}],
                "temperature": self.temperature,
                "max_tokens": max(256, int(max_tokens)),
            }
            use_json_mode = self._json_mode_supported is not False
            if use_json_mode:
                payload["response_format"] = {"type": "json_object"}

            try:
                r = requests.post(
                    f"{self.base_url}/chat/completions",
                    headers=self.headers(),
                    json=payload,
                    timeout=self.timeout,
                )
                if r.status_code >= 400:
                    body = r.text[:1500]
                    # Some LM Studio/model combinations reject response_format.
                    if use_json_mode and r.status_code in {400, 404, 422}:
                        lower = body.lower()
                        if any(x in lower for x in ["response_format", "json_object", "structured", "unsupported"]):
                            self._json_mode_supported = False
                            payload.pop("response_format", None)
                            r = requests.post(
                                f"{self.base_url}/chat/completions",
                                headers=self.headers(),
                                json=payload,
                                timeout=self.timeout,
                            )
                    if r.status_code >= 400:
                        raise ModelJSONError(f"HTTP {r.status_code}: {r.text[:1200]}", status_code=r.status_code)
                elif use_json_mode and self._json_mode_supported is None:
                    self._json_mode_supported = True

                data = r.json()
                choices = data.get("choices") or []
                if not choices:
                    raise ModelJSONError(f"No choices in response: {str(data)[:800]}")
                choice = choices[0] if isinstance(choices[0], dict) else {}
                msg = choice.get("message", {}) if isinstance(choice, dict) else {}
                raw_s = self._content_to_text(msg.get("content", ""))
                # Some reasoning-capable servers expose only a final field.
                if not raw_s.strip() and isinstance(msg, dict):
                    raw_s = self._content_to_text(msg.get("final", ""))
                finish_reason = clean_text(choice.get("finish_reason"), allow_blank=True).lower()
                self.last_finish_reason = finish_reason
                last_raw, last_finish = raw_s, finish_reason
                try:
                    obj = parse_json_response(raw_s)
                    return obj, raw_s
                except Exception as pe:
                    last_err = pe
                    # Repeating the same request after a hard length truncation usually
                    # produces the same failure; let the caller switch strategy.
                    if finish_reason in {"length", "max_tokens"}:
                        break
            except ModelJSONError as e:
                last_err = e
                last_raw = e.raw or last_raw
                last_finish = e.finish_reason or last_finish
            except Exception as e:
                last_err = e
            time.sleep(0.5)

        reason = str(last_err) if last_err else "Unknown model/JSON error"
        if last_finish:
            reason += f" (finish_reason={last_finish})"
        raise ModelJSONError(reason, raw=last_raw, finish_reason=last_finish)

    def text_json(self, prompt: str, max_tokens: int = 4096, retries: int = 1) -> Tuple[Dict[str, Any], str]:
        """Text-only JSON call for language normalization and other non-visual post-processing."""
        last_err: Optional[Exception] = None
        last_raw = ""
        attempts = max(1, int(retries) + 1)
        for _ in range(attempts):
            payload: Dict[str, Any] = {
                "model": self.model,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": self.temperature,
                "max_tokens": max(256, int(max_tokens)),
            }
            use_json_mode = self._json_mode_supported is not False
            if use_json_mode:
                payload["response_format"] = {"type": "json_object"}
            try:
                r = requests.post(
                    f"{self.base_url}/chat/completions",
                    headers=self.headers(),
                    json=payload,
                    timeout=self.timeout,
                )
                if r.status_code >= 400 and use_json_mode and r.status_code in {400, 404, 422}:
                    body = r.text[:1200].lower()
                    if any(x in body for x in ["response_format", "json_object", "structured", "unsupported"]):
                        self._json_mode_supported = False
                        payload.pop("response_format", None)
                        r = requests.post(
                            f"{self.base_url}/chat/completions",
                            headers=self.headers(),
                            json=payload,
                            timeout=self.timeout,
                        )
                if r.status_code >= 400:
                    raise ModelJSONError(f"HTTP {r.status_code}: {r.text[:1200]}", status_code=r.status_code)
                data = r.json()
                choices = data.get("choices") or []
                if not choices:
                    raise ModelJSONError(f"No choices in response: {str(data)[:800]}")
                choice = choices[0] if isinstance(choices[0], dict) else {}
                msg = choice.get("message", {}) if isinstance(choice, dict) else {}
                raw_s = self._content_to_text(msg.get("content", ""))
                if not raw_s.strip() and isinstance(msg, dict):
                    raw_s = self._content_to_text(msg.get("final", ""))
                last_raw = raw_s
                return parse_json_response(raw_s), raw_s
            except Exception as e:
                last_err = e
                time.sleep(0.4)
        raise ModelJSONError(str(last_err) if last_err else "Unknown text JSON error", raw=last_raw)


# -----------------------------------------------------------------------------
# Generic prompts
# -----------------------------------------------------------------------------
def universal_extraction_prompt(
    page_no: int,
    total_pages_in_group: int,
    sec: bool = False,
    existing_json: Optional[Dict[str, Any]] = None,
    use_tiles: bool = False,
    compact_retry: bool = False,
    target_language: str = "",
) -> str:
    """Universal compact-schema prompt.

    v9 asked for fields + answers + tables simultaneously, which duplicated the same
    data and made dense pages hit the response token limit. v10 uses one canonical
    item list. The rest of the program converts it to the familiar Excel structure.
    """
    existing = ""
    if sec and existing_json is not None:
        existing_text = json.dumps(existing_json, ensure_ascii=False, separators=(",", ":"))
        existing = (
            "\nFIRST-PASS JSON TO VERIFY. Keep the same item ids whenever possible. "
            "Correct visible mistakes, add genuinely missing filled items, and do not blindly copy it:\n"
            + existing_text[:22000]
        )
    mode = "SECOND PASS VERIFICATION" if sec else ("COMPACT RECOVERY RETRY" if compact_retry else "FIRST PASS EXTRACTION")
    image_note = ""
    if use_tiles:
        image_note = """
IMAGE INPUT:
- Image 1 is the full page.
- Additional images are zoomed views of the SAME page.
- Use zooms only as evidence for small handwriting, marks, and dense rows.
- Never duplicate an item because it appears in more than one view.
"""
    compact_note = """
COMPACT RETRY MODE:
- The previous response was too long or invalid.
- Be concise. Use short labels (max about 120 characters).
- Do not repeat printed instructions or bilingual translations unless needed to identify the item.
- Output minified valid JSON if possible.
""" if compact_retry else ""

    lang_note = output_language_prompt(target_language)
    return f"""
You are doing {mode} for a scanned questionnaire/form page.
This is page {page_no} of {total_pages_in_group} in the current document group.
{image_note}{compact_note}
{lang_note}
Automatically adapt to English, Traditional/Simplified Chinese, Filipino/Tagalog, bilingual/mixed pages, and unknown layouts.
Question types may include identity data, consent yes/no, checkbox multi-select, radio/single choice, circled numeric scale, matrix/Likert rows, handwritten text, date, time, number, signature, and remarks.

GOAL:
Extract all user-filled information and every visible numbered matrix/scale row, without needing a fixed template.

STRICT RULES:
1. Return ONE valid JSON object only. No markdown fences, comments, or explanation.
2. Never output UNCLEAR/UNKNOWN/UNREADABLE/AMBIGUOUS or similar.
3. A genuinely blank/unanswered/not-applicable item must be value "N/A" and blank true.
4. A visible but difficult mark must get the most likely concrete value, blank false, lower confidence, and a short reason.
5. Never treat all printed options as selected. Only visible ticks/checks/circles/crosses/fills/underlines/handwriting count.
6. For a scale/radio/table row, return exactly one value. If duplicate marks exist, choose the most intentional/final visible mark and mention it in reason.
7. For checkbox multi-select, put selected option labels in options; value may be a concise semicolon-separated summary.
8. For matrix/Likert tables, create ONE item per printed row. Do not also duplicate those rows elsewhere.
9. Use exactly ONE canonical items list. Do not create separate fields/answers/tables copies of the same data.
10. Stable item ids: use printed section/question/row numbers when visible (examples Q1, Q5a, S2_Q3, R18). Keep ids stable in SEC.
11. Participant IDs may look like A001, B023, CSA083, CSA122 or another handwritten code. Extract the most likely visible ID.
12. Keep labels concise but identifiable. Follow the OUTPUT LANGUAGE block above. Do not create duplicate source-language and translated copies.
13. Confidence must be numeric 0.0-1.0.
14. For gender/sex and every single-choice item, decide ONLY from the physical mark. Never infer from a name, handwriting style, age, or context. Examples: if "男 / 女" is printed and the circle encloses 女, the answer is 女; if M/F is printed, select only the physically circled/ticked letter.
15. For checkbox items, value must contain only the actually selected option label(s). If no box is marked, use value "N/A" and blank true. Never use ["N/A"].
16. Do not merge independent subquestions into one answer. Example: eye examination yes/no and pupil-dilation yes/no are separate items; medicine type, medicine name, and medicine source are separate items.
17. If a selected answer has an adjacent write-in line, capture BOTH the selection and the written detail. Put the detail in item.text or a separate stable subitem such as Q2_detail. Never drop a medication name, date, source, relationship, or "Other:" text that is visibly filled.
18. Compound questions with several independently filled values must be split into separate stable subitems (for example TV hours, phone/tablet hours, and reading hours), not one semicolon string.
19. Printed option lists are evidence only. A nearby underline, pen stroke, or crossed-out mark is not a selection unless it clearly marks the corresponding checkbox/radio/circle.
20. A question inside a matrix/scale table is NOT yes/no merely because its English text starts with Do/Does/Are/Is. Use type circle_scale or table_scale and read the marked score column.
21. For scale tables, preserve the printed score values and their visual mapping. Never invent a score outside the printed columns, and never assume the left-to-right order is ascending.
22. meta.page_language must be ONE concrete detected value such as English, Chinese, Filipino/Tagalog, Bilingual, Mixed, or N/A. Do not copy an option list.
23. meta.page_type must be ONE concrete type (or a short genuine compound such as consent/basic_info). Do not copy the whole option list from the schema.

JSON SCHEMA EXACTLY:
{{
  "meta": {{
    "document_title": "... or N/A",
    "page_language": "one detected value, e.g. English or Chinese",
    "page_type": "one detected type, e.g. health_history or scale_table",
    "participant_id": "... or N/A"
  }},
  "identity": {{
    "name": "N/A",
    "student_name": "N/A",
    "parent_guardian_name": "N/A",
    "school": "N/A",
    "gender": "N/A",
    "date_of_birth": "N/A",
    "age": "N/A",
    "grade": "N/A",
    "class_no": "N/A",
    "date": "N/A",
    "section": "N/A"
  }},
  "items": [
    {{
      "id": "stable id",
      "section": "section/table title or N/A",
      "label": "short field/question/row label",
      "type": "text/date/time/number/id/signature/checkbox/radio/yes_no/consent/circle_scale/table_scale/free_text/other",
      "value": "concrete answer or N/A",
      "options": [],
      "text": "handwritten/free text or N/A",
      "blank": false,
      "confidence": 0.0,
      "reason": ""
    }}
  ],
  "handwriting": [],
  "quality_flags": []
}}
{existing}
""".strip()


def region_extraction_prompt(region_no: int, region_count: int, page_no: int, total_pages: int, target_language: str = "") -> str:
    lang_note = output_language_prompt(target_language)
    return f"""
This image is region {region_no} of {region_count} from page {page_no}/{total_pages} of a scanned questionnaire.
{lang_note}
Extract ONLY filled data and visible numbered scale/table rows in this crop.
Return one compact JSON object using exactly this schema:
{{"meta":{{"document_title":"N/A","page_language":"N/A","page_type":"other","participant_id":"N/A"}},"identity":{{}},"items":[{{"id":"stable id","section":"N/A","label":"short label","type":"text/date/time/number/id/signature/checkbox/radio/yes_no/consent/circle_scale/table_scale/free_text/other","value":"answer or N/A","options":[],"text":"N/A","blank":false,"confidence":0.0,"reason":""}}],"handwriting":[],"quality_flags":[]}}
Rules: valid JSON only; no markdown; no UNCLEAR words; blank truly unanswered items may be N/A; do not invent marks; keep labels short; do not duplicate the same row within this crop.
""".strip()


def conflict_resolution_prompt(conflicts: List[Dict[str, Any]]) -> str:
    compact = []
    for i, c in enumerate(conflicts, start=1):
        compact.append({
            "resolution_id": f"C{i}",
            "item_id": c.get("item_id", "N/A"),
            "label": c.get("label", "N/A"),
            "type": c.get("type", "other"),
            "first": c.get("first", "N/A"),
            "sec": c.get("sec", "N/A"),
        })
    return """
Re-check ONLY the listed conflicting questionnaire items against the image(s).
Return one concrete final answer for each resolution_id. N/A is allowed only if the item is truly blank/unanswered.
Never output UNCLEAR or an explanation outside JSON.
If two marks are visible, choose the most intentional/final mark and mention duplicate marks briefly in reason.
Return exactly:
{"resolutions":[{"resolution_id":"C1","value":"answer or N/A","options":[],"text":"N/A","blank":false,"confidence":0.0,"reason":""}]}
CONFLICTS:
""".strip() + "\n" + json.dumps(compact, ensure_ascii=False, separators=(",", ":"))


def _is_gender_item(item: Dict[str, Any]) -> bool:
    label = norm_compare(item.get("label"))
    section = norm_compare(item.get("section"))
    values = [norm_compare(item.get("value"))]
    values.extend(norm_compare(x) for x in item.get("options", []) if clean_text(x) != "N/A")
    joined = "|".join(values)
    return (
        any(tok in label for tok in ["gender", "sex", "性别", "性別"])
        or any(tok in section for tok in ["gender", "sex", "性别", "性別"])
        or ("男" in joined and "女" in joined)
        or ("male" in joined and "female" in joined)
    )


def _selection_candidate_items(compact: Dict[str, Any]) -> List[Dict[str, Any]]:
    c = cleanup_compact_page(compact)
    out: List[Dict[str, Any]] = []
    page_is_scale_table = _looks_like_scale_table_page(c)

    for item in c.get("items", []):
        item_type = clean_text(item.get("type")).lower()
        if item_type not in SELECTION_MARK_TYPES:
            continue

        # Any dense repeated scale page must be handled by the dedicated geometry
        # verifier, not by the generic mark verifier. v13 only skipped literal
        # page_type=scale_table, which corrupted PSQI/sleep pages and continuation pages.
        if page_is_scale_table:
            # Dense matrices are authoritative geometry tasks. Even if the first model
            # mislabels a row as checkbox_multi_select/yes_no, the generic verifier must
            # not touch it; that was the main cause of CN PSQI regressions in v13.
            continue

        options = [clean_text(x) for x in item.get("options", []) if clean_text(x) != "N/A"]
        allowed_values: List[str] = []
        if item_type in SCALE_MARK_TYPES:
            numeric_options = [x for x in options if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", x)]
            if numeric_options and len(numeric_options) == len(options):
                allowed_values = numeric_options
        elif item_type in {"radio", "yes_no", "consent"} and options:
            allowed_values = options[:]
        elif item_type in {"checkbox", "checkbox_multi_select"}:
            allowed_values = options[:]

        out.append({
            "item_key": compact_item_key(item),
            "id": clean_text(item.get("id")),
            "section": clean_text(item.get("section")),
            "label": clean_text(item.get("label")),
            "type": item_type,
            "allowed_values": allowed_values,
            "choice_labels": options,
            "is_gender": _is_gender_item(item),
            "is_identity_gender": False,
        })

    ident = c.get("identity", {}) if isinstance(c.get("identity"), dict) else {}
    gender_value = clean_text(ident.get("gender"))
    page_type = clean_text(c.get("meta", {}).get("page_type")).lower() if isinstance(c.get("meta"), dict) else ""
    has_identity_context = any(clean_text(ident.get(k)) != "N/A" for k in ["name", "student_name", "date_of_birth", "age", "grade"])
    if gender_value != "N/A" or ("basic_info" in page_type and has_identity_context):
        gv = norm_compare(gender_value)
        if any(x in gv for x in ["男", "女"]):
            allowed = ["男", "女"]
        elif any(x in gv for x in ["male", "female"]):
            allowed = ["Male", "Female"]
        else:
            allowed = ["M", "F"]
        out.insert(0, {
            "item_key": "__IDENTITY_GENDER__",
            "id": "IDENTITY_GENDER",
            "section": "Identity",
            "label": "Gender / Sex",
            "type": "radio",
            "allowed_values": allowed,
            "choice_labels": allowed,
            "is_gender": True,
            "is_identity_gender": True,
        })
    return out

def selection_mark_verification_prompt(candidates: List[Dict[str, Any]], target_language: str = "") -> str:
    payload = []
    for i, c in enumerate(candidates, start=1):
        payload.append({
            "key": f"M{i}",
            "id": c.get("id", "N/A"),
            "section": c.get("section", "N/A"),
            "label": c.get("label", "N/A"),
            "type": c.get("type", "other"),
            "allowed_values": c.get("allowed_values", []),
            "choice_labels": c.get("choice_labels", []),
            "is_gender": bool(c.get("is_gender")),
        })
    lang_note = output_language_prompt(target_language)
    return lang_note + "\n" + """
You are a specialist visual verifier for PHYSICAL SELECTION MARKS on a scanned questionnaire.
Do not re-extract prose. Inspect only circles, ticks, checks, filled boxes, crosses, underlines used as selections, and final corrected marks.

CRITICAL RULES:
1. Return valid JSON only. No markdown.
2. Do NOT use prior answers or infer from names/context. Read the physical mark.
3. Gender/sex: if 男/女 or M/F is printed, choose only the option actually circled/ticked. A circle around 女 means 女. Never infer gender from the name.
4. Scale/table/radio/yes-no: output exactly one selected value. Follow the row and column carefully.
5. Checkbox: output every actually selected option in values. Do not include unmarked printed choices.
6. If no mark is present, use value "N/A", values [], blank true.
7. A stray pen stroke, underline, printed glyph, or neighboring row's circle is not a selection.
8. If an answer was corrected, choose the most intentional final visible mark.
9. Use the zoomed images as evidence for small marks; all images show the same page.
10. confidence must be 0.0-1.0.

Return exactly:
{"marks":[{"key":"M1","value":"single value or N/A","values":[],"blank":false,"confidence":0.0,"reason":"short physical-mark reason"}]}

ITEMS:
""".strip() + "\n" + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def _map_choice_to_allowed(value: Any, allowed: List[str], labels: List[str]) -> str:
    v = clean_text(value)
    if v == "N/A":
        return "N/A"
    if not allowed and not labels:
        return v
    nv = norm_compare(v)
    # Direct match against allowed values.
    for a in allowed:
        if norm_compare(a) == nv:
            return clean_text(a)
    # A scale verifier may return the option label instead of its numeric index.
    for idx, label in enumerate(labels):
        if norm_compare(label) == nv:
            if allowed and idx < len(allowed):
                return clean_text(allowed[idx])
            return clean_text(label)
    # Common gender synonyms.
    gender_map = {
        "male": "男", "m": "男", "boy": "男",
        "female": "女", "f": "女", "girl": "女",
        "男": "男", "女": "女",
    }
    if nv in gender_map:
        gv = gender_map[nv]
        for a in allowed + labels:
            if norm_compare(a) in {norm_compare(gv), norm_compare("male" if gv == "男" else "female"), norm_compare("m" if gv == "男" else "f")}:
                return clean_text(a)
        return gv
    return "N/A"


def apply_selection_mark_verification(
    compact: Dict[str, Any],
    client: LMStudioClient,
    images: List[Image.Image],
    source_pdf: str,
    qidx: int,
    page_no: int,
    max_tokens: int,
    target_language: str = "",
) -> Tuple[Dict[str, Any], List[Dict[str, Any]], Dict[str, Any]]:
    """Run an independent, geometry-focused verifier over all selection-mark items.

    This is deliberately different from SEC: it sees no first/SEC answers, only item
    definitions and the page images. High-confidence mark results can override generic
    extraction, which fixes cases where both first and SEC make the same visual mistake.
    """
    merged = cleanup_compact_page(compact)
    candidates = _selection_candidate_items(merged)
    if not candidates:
        return merged, [], {"marks": []}

    item_map = {compact_item_key(x): x for x in merged.get("items", [])}
    all_debug_marks: List[Dict[str, Any]] = []
    conflicts: List[Dict[str, Any]] = []

    # Keep each response compact enough for local 7B VLMs while still efficient.
    batch_size = 24
    for start in range(0, len(candidates), batch_size):
        batch = candidates[start:start + batch_size]
        prompt = selection_mark_verification_prompt(batch, target_language=target_language)
        obj, _raw = client.vision_json(
            prompt,
            images,
            max_tokens=max(1024, min(int(max_tokens), 4096)),
            retries=1,
        )
        marks = obj.get("marks", []) if isinstance(obj.get("marks"), list) else []
        all_debug_marks.extend(marks)
        by_key = {clean_text(m.get("key")): m for m in marks if isinstance(m, dict)}

        for local_idx, cand in enumerate(batch, start=1):
            mark = by_key.get(f"M{local_idx}")
            if not isinstance(mark, dict):
                continue
            conf = _as_confidence(mark.get("confidence"))
            blank = _as_bool(mark.get("blank"))
            item_key = cand["item_key"]
            is_identity_gender = bool(cand.get("is_identity_gender"))
            item = item_map.get(item_key)
            if is_identity_gender:
                old_value = clean_text(merged.get("identity", {}).get("gender"))
                old_blank = old_value == "N/A"
            else:
                if not isinstance(item, dict):
                    continue
                old_value = clean_text(item.get("value"))
                old_blank = bool(item.get("blank"))
            item_type = cand.get("type")
            new_value = "N/A"

            if item_type == "checkbox":
                raw_values = mark.get("values", [])
                if not isinstance(raw_values, list):
                    raw_values = [raw_values]
                mapped: List[str] = []
                for rv in raw_values:
                    mv = _map_choice_to_allowed(rv, cand.get("allowed_values", []), cand.get("choice_labels", []))
                    if mv != "N/A" and mv not in mapped:
                        mapped.append(mv)
                # Some models put a single checkbox selection in value instead of values.
                if not mapped:
                    mv = _map_choice_to_allowed(mark.get("value"), cand.get("allowed_values", []), cand.get("choice_labels", []))
                    if mv != "N/A":
                        mapped.append(mv)
                new_value = "; ".join(mapped) if mapped else "N/A"
                blank = not bool(mapped)
            else:
                new_value = _map_choice_to_allowed(mark.get("value"), cand.get("allowed_values", []), cand.get("choice_labels", []))
                if new_value == "N/A":
                    blank = True

            # Specialized mark reading is authoritative when confident. A claimed blank
            # requires stronger evidence before erasing a concrete answer.
            should_override = False
            if new_value != "N/A" and conf >= 0.80:
                should_override = True
            elif new_value == "N/A" and blank and conf >= 0.97 and (old_value == "N/A" or old_blank):
                should_override = True

            if not should_override:
                continue

            if norm_compare(old_value) != norm_compare(new_value) or old_blank != blank:
                conflicts.append({
                    "source_pdf": source_pdf,
                    "questionnaire_index_in_pdf": qidx,
                    "page_no_in_questionnaire": page_no,
                    "field": f"selection_mark:{cand.get('id')}:{cand.get('label')}",
                    "first_value": old_value,
                    "sec_value": clean_text(mark),
                    "chosen_value": new_value,
                    "reason": "specialized physical selection-mark verifier override",
                })

            if is_identity_gender:
                if new_value != "N/A":
                    merged.setdefault("identity", {})["gender"] = new_value
            else:
                item["value"] = new_value
                item["blank"] = bool(blank)
                item["confidence"] = conf
                item["reason"] = clean_text(mark.get("reason"), allow_blank=True)

                # Keep identity gender synchronized with a visually verified item.
                if cand.get("is_gender") and new_value != "N/A":
                    merged.setdefault("identity", {})["gender"] = new_value

    merged["items"] = list(item_map.values())
    merged["quality_flags"] = sorted(set(merged.get("quality_flags", []) + ["selection_mark_verification_used"]))
    return cleanup_compact_page(merged), conflicts, {"marks": all_debug_marks}



def _looks_like_scale_table_page(compact: Dict[str, Any]) -> bool:
    c = cleanup_compact_page(compact)
    meta = c.get("meta", {}) if isinstance(c.get("meta"), dict) else {}
    page_type = clean_text(meta.get("page_type")).lower()
    if "scale_table" in page_type:
        return True
    items = [x for x in c.get("items", []) if isinstance(x, dict)]
    if len(items) < 6:
        return False

    selection_like = 0
    numeric_answers = 0
    scale_types = 0
    ids = []
    for item in items:
        t = clean_text(item.get("type")).lower()
        v = clean_text(item.get("value"))
        iid = clean_text(item.get("id"))
        ids.append(iid)
        if t in SCALE_MARK_TYPES:
            scale_types += 1
            selection_like += 1
        elif t in {"yes_no", "radio", "consent"}:
            selection_like += 1
        if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", v):
            numeric_answers += 1

    # Dense repeated selection rows with numeric answers are very often a scale
    # continuation page even when the model labels them "other" or "yes_no".
    dense_numeric = selection_like >= max(6, math.ceil(len(items) * 0.60)) and numeric_answers >= max(4, math.ceil(len(items) * 0.40))
    dense_scale = scale_types >= max(5, math.ceil(len(items) * 0.45))
    sequential_ids = sum(1 for x in ids if re.fullmatch(r"[QqRr]?\d+[A-Za-z]?", x)) >= max(6, math.ceil(len(items) * 0.60))
    return bool(dense_scale or (dense_numeric and sequential_ids))

def _scale_table_candidate_items(compact: Dict[str, Any]) -> List[Dict[str, Any]]:
    c = cleanup_compact_page(compact)
    if not _looks_like_scale_table_page(c):
        return []

    items = [x for x in c.get("items", []) if isinstance(x, dict)]
    # Detect lettered row families such as Q5a..Q5j. A row can be misclassified as
    # free_text when it also has a scale mark (for example PSQI "Other reason").
    scale_prefix_counts: Counter = Counter()
    for item in items:
        t = clean_text(item.get("type")).lower()
        iid = clean_text(item.get("id")).lower()
        m = re.fullmatch(r"([qr]?\d+)[a-z]", iid)
        if m and t in SCALE_MARK_TYPES:
            scale_prefix_counts[m.group(1)] += 1

    out: List[Dict[str, Any]] = []
    for item in items:
        item_id = clean_text(item.get("id"))
        label = clean_text(item.get("label"))
        item_type = clean_text(item.get("type")).lower()
        if item_id == "N/A" or label == "N/A":
            continue

        include = item_type in SCALE_MARK_TYPES | {"yes_no", "radio", "consent"}
        if not include and item_type in {"free_text", "text", "other"}:
            m = re.fullmatch(r"([qr]?\d+)[a-z]", item_id.lower())
            include = bool(m and scale_prefix_counts.get(m.group(1), 0) >= 3)
        if not include:
            continue

        out.append({
            "item_key": compact_item_key(item),
            "id": item_id,
            "section": clean_text(item.get("section")),
            "label": label,
            "old_type": item_type,
        })
    return out

def build_item_consensus_map(first: Dict[str, Any], sec: Optional[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """Record first/SEC agreement before any visual verifier overwrites values."""
    if not sec:
        return {}
    a = cleanup_compact_page(first)
    b = cleanup_compact_page(sec)
    amap = {compact_item_key(x): x for x in a.get("items", []) if isinstance(x, dict)}
    bmap = {compact_item_key(x): x for x in b.get("items", []) if isinstance(x, dict)}
    out: Dict[str, Dict[str, Any]] = {}
    for key in set(amap) | set(bmap):
        ia = amap.get(key)
        ib = bmap.get(key)
        if not ia or not ib:
            continue
        va = clean_text(ia.get("value"))
        vb = clean_text(ib.get("value"))
        ba = bool(ia.get("blank"))
        bb = bool(ib.get("blank"))
        agreed = norm_compare(va) == norm_compare(vb) and ba == bb
        out[key] = {
            "agreed": agreed,
            "value": va if agreed else "N/A",
            "blank": ba if agreed else False,
            "first_value": va,
            "sec_value": vb,
            "confidence": min(float(ia.get("confidence", 0.0)), float(ib.get("confidence", 0.0))),
        }
    return out


def _columns_valid(columns: List[str]) -> bool:
    cols = [clean_text(x) for x in columns if clean_text(x) != "N/A"]
    if not (2 <= len(cols) <= 7) or len(set(norm_compare(x) for x in cols)) != len(cols):
        return False
    # Prefer a compact score map. Very long prose entries are option labels, not
    # reliable physical score-column keys for geometry mapping.
    return all(len(x) <= 24 for x in cols)


def _same_columns(a: List[str], b: List[str]) -> bool:
    return len(a) == len(b) and all(norm_compare(x) == norm_compare(y) for x, y in zip(a, b))


def scale_table_verification_prompt(candidates: List[Dict[str, Any]], target_language: str = "") -> str:
    payload = []
    for i, c in enumerate(candidates, start=1):
        payload.append({
            "key": f"T{i}",
            "id": c.get("id", "N/A"),
            "section": c.get("section", "N/A"),
            "label": c.get("label", "N/A"),
        })
    lang_note = output_language_prompt(target_language)
    return lang_note + "\n" + """
You are a specialist visual verifier for a MATRIX / LIKERT / CIRCLED-SCORE TABLE.
The page may be a continuation page, may be bilingual, and may have been rotated upright before you received it.
Do NOT use any previous extracted answer. Read the printed score columns and the physical marks directly from the images.

CRITICAL METHOD:
A. First identify the printed score values at the answer columns in their ACTUAL visual left-to-right order.
   Examples can be 0,1,2; 2,1,0; 0,1,2,3,4; 5,4,3,2,1,0; or another order.
   Never assume scores increase from left to right.
B. Then, for each requested row, follow that row horizontally to the answer columns.
C. The answer is the PRINTED SCORE VALUE inside the column containing the user's circle/tick/check/final mark.
D. A printed digit by itself is not a mark. Look for a hand-drawn circle, tick, fill, cross, or obvious final correction around/in that cell.
E. Even if the English question starts with Do/Does/Are/Is, it is still a SCALE ROW when it is inside this table. Never convert it to yes/no.
F. Continuation pages may omit the column header. In that case, infer the mapping from the repeated printed digits visible in the row cells.
G. Use row ids and labels only to locate the row. Do not infer the answer semantically from the question text.
H. Return one row result per requested key. If truly no mark is present, use N/A and blank true. Never use UNCLEAR.
I. If two marks exist, choose the most intentional/final visible mark and mention it briefly.
J. Confidence must be 0.0-1.0.

Return valid JSON only:
{"column_values_left_to_right":["2","1","0"],"rows":[{"key":"T1","value":"2","blank":false,"confidence":0.99,"reason":"circle in the left score column printed 2"}]}

REQUESTED ROWS:
""".strip() + "\n" + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))



def scale_orientation_ensemble_prompt() -> str:
    return """
Four images show the SAME questionnaire page at four rotations:
Image 1 = current orientation
Image 2 = 90 degrees clockwise
Image 3 = 180 degrees
Image 4 = 270 degrees clockwise

Choose the image where:
- printed text is upright and easiest to read;
- table rows run horizontally;
- question numbers progress naturally from top to bottom;
- score columns are vertical columns.

Return JSON only:
{"best_image_index":1,"confidence":0.0,"reason":"short reason"}
""".strip()


def choose_scale_table_orientation(
    client: LMStudioClient,
    img: Image.Image,
    max_tokens: int = 256,
) -> Tuple[Image.Image, Dict[str, Any]]:
    variants = [
        img,
        img.rotate(-90, expand=True),
        img.rotate(180, expand=True),
        img.rotate(-270, expand=True),
    ]
    try:
        obj, _ = client.vision_json(
            scale_orientation_ensemble_prompt(),
            variants,
            max_tokens=max(128, min(int(max_tokens), 512)),
            retries=1,
        )
        idx = int(obj.get("best_image_index", 1))
        if idx not in {1, 2, 3, 4}:
            idx = 1
        return variants[idx - 1], obj
    except Exception as e:
        return img, {"best_image_index": 1, "confidence": 0.0, "reason": f"orientation ensemble failed: {e}"}


def scale_table_geometry_prompt(candidates: List[Dict[str, Any]]) -> str:
    rows = [{"id": clean_text(x.get("id")), "label": clean_text(x.get("label"))[:100]} for x in candidates[:80]]
    return """
Locate the MATRIX / LIKERT / CIRCLED-SCORE TABLE geometry on this upright questionnaire page.
Do not answer the questions. Only locate the table and score columns.

Return JSON only:
{
  "is_scale_table": true,
  "table_bbox_0_1000": [left, top, right, bottom],
  "score_bbox_0_1000": [left, top, right, bottom],
  "column_values_left_to_right": ["2","1","0"],
  "confidence": 0.0,
  "reason": "short reason"
}

RULES:
- Coordinates are integers from 0 to 1000 relative to the page.
- table_bbox must cover the repeated questionnaire rows and answer columns, not page margins or logos.
- score_bbox must cover the physical answer columns.
- If printed numeric score digits exist, return those digits in actual visual left-to-right order.
- Never assume ascending order.
- A continuation page can still be a scale table even without a title/header.
- If several scale blocks share the page, table_bbox may cover all of them and score_bbox may cover their common answer-column area.

EXPECTED ROW IDS (for locating only):
""".strip() + "\n" + json.dumps(rows, ensure_ascii=False, separators=(",", ":"))


def locate_scale_table_geometry(
    client: LMStudioClient,
    img: Image.Image,
    candidates: List[Dict[str, Any]],
    max_tokens: int = 1024,
) -> Dict[str, Any]:
    try:
        obj, _ = client.vision_json(
            scale_table_geometry_prompt(candidates),
            [img],
            max_tokens=max(512, min(int(max_tokens), 1536)),
            retries=1,
        )
    except Exception as e:
        return {"valid": False, "reason": f"geometry locator failed: {e}"}

    def bbox(name: str) -> Optional[List[float]]:
        raw = obj.get(name, [])
        if not isinstance(raw, list) or len(raw) != 4:
            return None
        try:
            vals = [max(0.0, min(1000.0, float(x))) for x in raw]
        except Exception:
            return None
        x1, y1, x2, y2 = vals
        if x2 - x1 < 180 or y2 - y1 < 120:
            return None
        return vals

    tb = bbox("table_bbox_0_1000")
    sb = bbox("score_bbox_0_1000")
    cols_raw = obj.get("column_values_left_to_right", [])
    if not isinstance(cols_raw, list):
        cols_raw = [cols_raw]
    cols = []
    for x in cols_raw:
        v = clean_text(x)
        if v != "N/A" and norm_compare(v) not in {norm_compare(y) for y in cols}:
            cols.append(v)
    if not _columns_valid(cols):
        cols = []
    conf = _as_confidence(obj.get("confidence"))
    valid = bool(tb and sb and conf >= 0.55)
    return {
        "valid": valid,
        "table_bbox_0_1000": tb,
        "score_bbox_0_1000": sb,
        "column_values_left_to_right": cols,
        "confidence": conf,
        "reason": clean_text(obj.get("reason"), allow_blank=True),
        "raw": obj,
    }


def _bbox_norm_to_pixels(img: Image.Image, bbox_0_1000: Optional[List[float]], fallback: Tuple[float, float, float, float]) -> Tuple[int, int, int, int]:
    w, h = img.size
    if bbox_0_1000 and len(bbox_0_1000) == 4:
        x1, y1, x2, y2 = bbox_0_1000
        vals = (x1 / 1000.0, y1 / 1000.0, x2 / 1000.0, y2 / 1000.0)
    else:
        vals = fallback
    fx1, fy1, fx2, fy2 = vals
    return (
        max(0, min(w - 1, int(w * fx1))),
        max(0, min(h - 1, int(h * fy1))),
        max(1, min(w, int(w * fx2))),
        max(1, min(h, int(h * fy2))),
    )


def scale_row_anchor_prompt(batch: List[Dict[str, Any]]) -> str:
    rows = [
        {"key": f"T{i}", "id": clean_text(c.get("id")), "label": clean_text(c.get("label"))[:120]}
        for i, c in enumerate(batch, start=1)
    ]
    return """
You are locating row positions inside ONE already-cropped questionnaire scale table.
Do NOT read or answer any circled score. Only locate the horizontal center of each requested printed row.

Return JSON only:
{"rows":[{"key":"T1","y_center_0_1000":123,"confidence":0.99}]}

RULES:
- y_center_0_1000 is the row center relative to this cropped table image: top=0, bottom=1000.
- Match by printed question ID first, then label text.
- Keep rows in the requested key mapping; do not shift a neighboring row.
- Omit a row if you cannot locate it confidently.

REQUESTED ROWS:
""".strip() + "\n" + json.dumps(rows, ensure_ascii=False, separators=(",", ":"))


def locate_scale_batch_rows(
    client: LMStudioClient,
    img: Image.Image,
    batch: List[Dict[str, Any]],
    geometry: Optional[Dict[str, Any]],
    max_tokens: int = 1024,
) -> Dict[str, float]:
    geom = geometry or {}
    table_box = _bbox_norm_to_pixels(img, geom.get("table_bbox_0_1000"), (0.03, 0.05, 0.97, 0.97))
    x1, y1, x2, y2 = table_box
    if x2 <= x1 or y2 <= y1:
        return {}
    table = resize_max_side(img.crop((x1, y1, x2, y2)), 2200)
    try:
        obj, _ = client.vision_json(
            scale_row_anchor_prompt(batch),
            [table],
            max_tokens=max(512, min(int(max_tokens), 1536)),
            retries=1,
        )
    except Exception:
        return {}
    rows = obj.get("rows", []) if isinstance(obj.get("rows"), list) else []
    out: Dict[str, float] = {}
    for r in rows:
        if not isinstance(r, dict):
            continue
        key = clean_text(r.get("key"), allow_blank=True)
        try:
            y = float(r.get("y_center_0_1000"))
        except Exception:
            continue
        conf = _as_confidence(r.get("confidence"))
        if key and 0 <= y <= 1000 and conf >= 0.72:
            out[key] = y
    return out


def make_scale_batch_images(
    img: Image.Image,
    start_index: int,
    end_index: int,
    total_rows: int,
    tight: bool = False,
    geometry: Optional[Dict[str, Any]] = None,
    row_anchor_y_0_1000: Optional[List[float]] = None,
) -> List[Image.Image]:
    """Create evidence views using the detected table bbox instead of the whole page.

    v13 divided 4%-97% of the entire page into equal row bands. That caused severe
    row shifts whenever a table occupied only the top or middle of a page. v14 first
    locates the actual table and score-column geometry, then slices inside that bbox.
    """
    w, h = img.size
    if w <= 0 or h <= 0 or total_rows <= 0:
        return [img]

    geom = geometry or {}
    table_box = _bbox_norm_to_pixels(img, geom.get("table_bbox_0_1000"), (0.03, 0.05, 0.97, 0.97))
    score_box = _bbox_norm_to_pixels(img, geom.get("score_bbox_0_1000"), (0.62, 0.05, 0.98, 0.97))
    tx1, ty1, tx2, ty2 = table_box
    sx1, sy1, sx2, sy2 = score_box
    if tx2 <= tx1 or ty2 <= ty1:
        return [img]

    table = img.crop((tx1, ty1, tx2, ty2))
    table_big = resize_max_side(table, 2400 if tight else 2100)

    table_h = ty2 - ty1
    anchors = []
    for y in row_anchor_y_0_1000 or []:
        try:
            yf = float(y)
        except Exception:
            continue
        if 0 <= yf <= 1000:
            anchors.append(yf)
    if anchors:
        # AI-only row localization handles variable-height wrapped rows. Equal slicing is
        # kept only as a fallback when the row locator cannot find the printed IDs.
        centers = sorted(ty1 + table_h * (y / 1000.0) for y in anchors)
        if len(centers) >= 2:
            gaps = [b - a for a, b in zip(centers, centers[1:]) if b > a]
            local_gap = min(gaps) if gaps else table_h / max(1, total_rows)
        else:
            local_gap = table_h / max(1, total_rows)
        pad = max(18.0, local_gap * (0.62 if not tight else 0.42))
        y1 = max(ty1, int(min(centers) - pad))
        y2 = min(ty2, int(max(centers) + pad))
    else:
        margin_rows = 0.85 if not tight else 0.35
        rel1 = max(0.0, (start_index - margin_rows) / max(1, total_rows))
        rel2 = min(1.0, (end_index + margin_rows) / max(1, total_rows))
        y1 = max(ty1, int(ty1 + table_h * rel1))
        y2 = min(ty2, int(ty1 + table_h * rel2))
    min_h = max(90, int(table_h * 0.05))
    if y2 - y1 < min_h:
        pad = max(20, min_h // 2)
        mid = (y1 + y2) // 2
        y1, y2 = max(ty1, mid - pad), min(ty2, mid + pad)

    band = img.crop((tx1, y1, tx2, y2))
    band_big = resize_max_side(band, 2400 if tight else 2200)

    # Preserve the same vertical row band but use the detected physical score area.
    score_y1 = max(y1, sy1)
    score_y2 = min(y2, sy2)
    if score_y2 <= score_y1:
        score_y1, score_y2 = y1, y2
    score = img.crop((max(tx1, sx1), score_y1, min(w, max(sx2, sx1 + 1)), score_y2))
    score_big = resize_max_side(score, 2000 if tight else 1800)

    return [table_big, band_big, score_big]

def scale_table_band_prompt(
    candidates: List[Dict[str, Any]],
    target_language: str = "",
    pass_name: str = "primary",
    canonical_columns: Optional[List[str]] = None,
) -> str:
    payload = []
    for i, c in enumerate(candidates, start=1):
        payload.append({
            "key": f"T{i}",
            "id": c.get("id", "N/A"),
            "label": c.get("label", "N/A"),
        })
    lang_note = output_language_prompt(target_language)
    known = [clean_text(x) for x in (canonical_columns or []) if clean_text(x) != "N/A"]
    known_note = (
        "PAGE-LEVEL VERIFIED COLUMN MAP (left-to-right): " + json.dumps(known, ensure_ascii=False) + "\nUse this map exactly; do not invent or omit a column."
        if _columns_valid(known) else
        "No trusted page-level column map is available; read the printed score columns carefully."
    )
    return lang_note + "\n" + f"""
You are doing the {pass_name} PHYSICAL-MARK verification for a small consecutive band of a matrix/Likert table.
The images show the SAME page/table:
- Image 1 is the whole detected table for global row alignment.
- Image 2 enlarges the approximate consecutive row band.
- Image 3 enlarges the physical score columns for that band.

{known_note}

DO NOT use question meaning. DO NOT use previous extracted answers.
Your only job is to map each hand-drawn circle/tick/final mark to the PRINTED SCORE COLUMN.

METHOD:
1. Locate each requested row by its printed question id and top-to-bottom order using Image 1.
2. Confirm the same row in Images 2-3. Never shift a mark from a neighboring row.
3. Return BOTH column_index_1based and the printed value in that column.
4. A printed digit alone is not a user mark. Require a visible circle/tick/fill/cross/final correction.
5. A question beginning Do/Does/Are/Is is still a scale row, NOT yes/no.
6. If two marks exist, look again and choose the most intentional/final visible mark; mention duplicate/correction briefly.
7. If truly blank, return N/A with blank true. Never output UNCLEAR.
8. confidence must be 0.0-1.0.

Return JSON only:
{{"column_values_left_to_right":["2","1","0"],"rows":[{{"key":"T1","column_index_1based":1,"value":"2","blank":false,"confidence":0.99,"reason":"visible circle around printed 2"}}]}}

REQUESTED ROWS:
""".strip() + "\n" + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))

def _parse_scale_band_result(
    obj: Dict[str, Any],
    batch: List[Dict[str, Any]],
    canonical_columns: Optional[List[str]] = None,
) -> Tuple[List[str], Dict[str, Dict[str, Any]]]:
    canonical = [clean_text(x) for x in (canonical_columns or []) if clean_text(x) != "N/A"]
    raw_cols = obj.get("column_values_left_to_right", [])
    if not isinstance(raw_cols, list):
        raw_cols = [raw_cols]
    detected: List[str] = []
    for x in raw_cols:
        v = clean_text(x)
        if v != "N/A" and norm_compare(v) not in {norm_compare(y) for y in detected}:
            detected.append(v)
    columns = canonical if _columns_valid(canonical) else detected
    if not _columns_valid(columns):
        columns = []

    rows = obj.get("rows", []) if isinstance(obj.get("rows"), list) else []
    by_key: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        if not isinstance(r, dict):
            continue
        key = clean_text(r.get("key"))
        if not key:
            continue
        value = clean_text(r.get("value"))
        try:
            idx = int(r.get("column_index_1based"))
        except Exception:
            idx = 0

        blank = _as_bool(r.get("blank"))
        if 1 <= idx <= len(columns):
            # A valid physical column index is stronger than the model's textual value.
            value = columns[idx - 1]
            blank = False
        elif blank:
            value = "N/A"
            idx = 0
        elif value != "N/A" and columns:
            mapped = next((c for c in columns if norm_compare(c) == norm_compare(value)), None)
            if mapped is None:
                continue
            value = mapped
        elif value != "N/A" and not columns:
            # Without a stable map, do not trust a free-form score guess.
            continue

        by_key[key] = {
            "value": value,
            "blank": bool(blank or value == "N/A"),
            "confidence": _as_confidence(r.get("confidence")),
            "reason": clean_text(r.get("reason"), allow_blank=True),
            "column_index_1based": idx,
        }
    return columns, by_key

def _choose_scale_result(
    a: Optional[Dict[str, Any]],
    b: Optional[Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    if a and b:
        if norm_compare(a.get("value")) == norm_compare(b.get("value")):
            return a if a.get("confidence", 0.0) >= b.get("confidence", 0.0) else b
        # No agreement: do not silently average. Caller should target-recheck.
        return None
    return a or b


def apply_scale_table_verification(
    compact: Dict[str, Any],
    client: LMStudioClient,
    images: List[Image.Image],
    source_pdf: str,
    qidx: int,
    page_no: int,
    max_tokens: int,
    target_language: str = "",
    verification_level: str = "Careful",
    use_orientation_ensemble: bool = True,
    consensus_map: Optional[Dict[str, Dict[str, Any]]] = None,
) -> Tuple[Dict[str, Any], List[Dict[str, Any]], Dict[str, Any]]:
    """Verify scale marks with table geometry + first/SEC consensus protection.

    v13 could overwrite two agreeing correct passes with one bad row-band read because it
    divided the entire page into equal-height rows. v14 first locates the actual table,
    uses a canonical page-level column map, and requires independent confirmation before
    overriding a strong first/SEC agreement.
    """
    merged = cleanup_compact_page(compact)
    candidates = _scale_table_candidate_items(merged)
    if not candidates:
        return merged, [], {"orientation": {}, "geometry": {}, "batches": []}

    base_img = images[0] if images else None
    if base_img is None:
        return merged, [], {"orientation": {}, "geometry": {}, "batches": []}

    level = clean_text(verification_level, allow_blank=True).lower()
    consensus_map = consensus_map or {}

    orientation_debug: Dict[str, Any] = {}
    oriented = base_img
    if use_orientation_ensemble and level in {"careful", "maximum"}:
        oriented, orientation_debug = choose_scale_table_orientation(client, base_img, max_tokens=256)

    geometry = locate_scale_table_geometry(client, oriented, candidates, max_tokens=max_tokens)
    canonical_columns = geometry.get("column_values_left_to_right", []) if isinstance(geometry, dict) else []
    if not _columns_valid(canonical_columns):
        canonical_columns = []

    item_map = {compact_item_key(x): x for x in merged.get("items", []) if isinstance(x, dict)}
    conflicts: List[Dict[str, Any]] = []
    debug_batches: List[Dict[str, Any]] = []
    total_rows = len(candidates)
    batch_size = 8 if level == "fast" else (5 if level == "careful" else 4)

    def result_agree(a: Optional[Dict[str, Any]], b: Optional[Dict[str, Any]]) -> bool:
        return bool(a and b and norm_compare(a.get("value")) == norm_compare(b.get("value")) and bool(a.get("blank")) == bool(b.get("blank")))

    def strong_consensus_for(cand: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        rec = consensus_map.get(cand.get("item_key"), {})
        if rec and rec.get("agreed") and clean_text(rec.get("value")) != "N/A":
            return rec
        return None

    for start in range(0, total_rows, batch_size):
        batch = candidates[start:start + batch_size]
        end = start + len(batch)
        row_anchors: Dict[str, float] = {}
        if level in {"careful", "maximum"}:
            row_anchors = locate_scale_batch_rows(
                client, oriented, batch, geometry, max_tokens=max_tokens
            )
        anchor_values = [row_anchors.get(f"T{i}") for i in range(1, len(batch) + 1)]
        anchor_values = [x for x in anchor_values if x is not None]
        batch_images = make_scale_batch_images(
            oriented, start, end, total_rows, tight=False, geometry=geometry,
            row_anchor_y_0_1000=anchor_values,
        )
        prompt_a = scale_table_band_prompt(batch, target_language=target_language, pass_name="primary", canonical_columns=canonical_columns)
        obj_a, _ = client.vision_json(
            prompt_a,
            batch_images,
            max_tokens=max(1536, min(int(max_tokens), 4096)),
            retries=1,
        )
        cols_a, rows_a = _parse_scale_band_result(obj_a, batch, canonical_columns=canonical_columns)
        columns = canonical_columns if _columns_valid(canonical_columns) else cols_a

        # Decide whether an independent confirmation is required. Maximum always uses
        # it. Careful uses it whenever the primary would overturn a strong SEC consensus,
        # would turn a concrete answer into N/A, or when the page-level map is weak.
        need_confirmation = level == "maximum"
        if level == "careful":
            if not _columns_valid(columns):
                need_confirmation = True
            for local_idx, cand in enumerate(batch, start=1):
                r = rows_a.get(f"T{local_idx}")
                if not r:
                    continue
                old_item = item_map.get(cand.get("item_key"), {})
                old_value = clean_text(old_item.get("value")) if isinstance(old_item, dict) else "N/A"
                cons = strong_consensus_for(cand)
                if cons and norm_compare(cons.get("value")) != norm_compare(r.get("value")):
                    need_confirmation = True
                if old_value != "N/A" and clean_text(r.get("value")) == "N/A":
                    need_confirmation = True

        obj_b: Dict[str, Any] = {}
        rows_b: Dict[str, Dict[str, Any]] = {}
        cols_b: List[str] = []
        if need_confirmation:
            tight_images = make_scale_batch_images(
                oriented, start, end, total_rows, tight=True, geometry=geometry,
                row_anchor_y_0_1000=anchor_values,
            )
            prompt_b = scale_table_band_prompt(batch, target_language=target_language, pass_name="independent confirmation", canonical_columns=columns)
            obj_b, _ = client.vision_json(
                prompt_b,
                list(reversed(tight_images)),
                max_tokens=max(1536, min(int(max_tokens), 4096)),
                retries=1,
            )
            cols_b, rows_b = _parse_scale_band_result(obj_b, batch, canonical_columns=columns)
            if not _columns_valid(columns) and _columns_valid(cols_b):
                columns = cols_b

        if not _columns_valid(columns):
            debug_batches.append({"start": start, "end": end, "primary": obj_a, "confirmation": obj_b, "status": "no-stable-column-map"})
            continue

        # Repair scale type/options only after a stable physical column map exists.
        for cand in batch:
            item = item_map.get(cand.get("item_key"))
            if isinstance(item, dict):
                item["type"] = "circle_scale"
                item["options"] = columns[:]

        unresolved: List[Tuple[int, Dict[str, Any], Optional[Dict[str, Any]], Optional[Dict[str, Any]]]] = []
        for local_idx, cand in enumerate(batch, start=1):
            key = f"T{local_idx}"
            a = rows_a.get(key)
            b = rows_b.get(key)
            item = item_map.get(cand.get("item_key"))
            if not isinstance(item, dict):
                continue
            old_value = clean_text(item.get("value"))
            old_blank = bool(item.get("blank"))
            cons = strong_consensus_for(cand)

            result: Optional[Dict[str, Any]] = None
            two_pass = result_agree(a, b)
            if two_pass:
                result = a if float(a.get("confidence", 0.0)) >= float(b.get("confidence", 0.0)) else b
            elif need_confirmation and a and b:
                unresolved.append((local_idx, cand, a, b))
                continue
            else:
                result = a or b

            if not result or float(result.get("confidence", 0.0)) < 0.84:
                unresolved.append((local_idx, cand, a, b))
                continue

            new_value = clean_text(result.get("value"))
            new_blank = bool(result.get("blank"))
            if new_value != "N/A" and not any(norm_compare(new_value) == norm_compare(c) for c in columns):
                unresolved.append((local_idx, cand, a, b))
                continue

            # Never let one visual pass delete a concrete answer. N/A is allowed only
            # for a truly blank row and requires two very-high-confidence confirmations.
            if new_value == "N/A" and old_value != "N/A":
                if not (two_pass and float(a.get("confidence", 0.0)) >= 0.98 and float(b.get("confidence", 0.0)) >= 0.98):
                    continue

            # Protect first+SEC agreement. Two visual reads made from the same detected
            # geometry can repeat the same row-shift error, so they are NOT enough to
            # overturn two agreeing extraction passes. Careful/Maximum must send the row
            # to an independent single-row tiebreak first.
            if cons and norm_compare(cons.get("value")) != norm_compare(new_value):
                if level in {"careful", "maximum"}:
                    unresolved.append((local_idx, cand, a, b))
                    continue
                if not (two_pass and float(a.get("confidence", 0.0)) >= 0.94 and float(b.get("confidence", 0.0)) >= 0.94):
                    continue

            # On a misclassified/illegal original value, a single strong geometry read is
            # enough. Otherwise require higher confidence for a change.
            old_legal = any(norm_compare(old_value) == norm_compare(c) for c in columns)
            threshold = 0.80 if not old_legal else 0.88
            if float(result.get("confidence", 0.0)) < threshold:
                continue

            if norm_compare(old_value) != norm_compare(new_value) or old_blank != new_blank:
                conflicts.append({
                    "source_pdf": source_pdf,
                    "questionnaire_index_in_pdf": qidx,
                    "page_no_in_questionnaire": page_no,
                    "field": f"scale_table:{cand.get('id')}:{cand.get('label')}",
                    "first_value": old_value,
                    "sec_value": json.dumps(result, ensure_ascii=False),
                    "chosen_value": new_value,
                    "reason": "geometry verifier override with consensus protection" if cons else "geometry verifier override",
                })
            item["value"] = new_value
            item["blank"] = bool(new_blank)
            item["confidence"] = float(result.get("confidence", 0.0))
            item["reason"] = clean_text(result.get("reason"), allow_blank=True)

        # Careful/Maximum: independently recheck only rows that remained unresolved.
        # This third view is mandatory before changing a strong first+SEC consensus.
        if level in {"careful", "maximum"} and unresolved:
            for local_idx, cand, a, b in unresolved:
                global_idx = start + local_idx - 1
                one_anchor = row_anchors.get(f"T{local_idx}")
                one_images = make_scale_batch_images(
                    oriented, global_idx, global_idx + 1, total_rows, tight=True, geometry=geometry,
                    row_anchor_y_0_1000=[one_anchor] if one_anchor is not None else None,
                )
                prompt_t = scale_table_band_prompt([cand], target_language=target_language, pass_name="single-row tiebreak", canonical_columns=columns)
                try:
                    obj_t, _ = client.vision_json(
                        prompt_t,
                        one_images,
                        max_tokens=max(1024, min(int(max_tokens), 2048)),
                        retries=1,
                    )
                    _, rows_t = _parse_scale_band_result(obj_t, [cand], canonical_columns=columns)
                    t = rows_t.get("T1")
                    votes = [x for x in [a, b, t] if isinstance(x, dict) and clean_text(x.get("value")) != "N/A"]
                    if not votes:
                        continue
                    counts = Counter(norm_compare(x.get("value")) for x in votes)
                    winner_norm, winner_count = counts.most_common(1)[0]
                    if winner_count < 2:
                        continue
                    result = max((x for x in votes if norm_compare(x.get("value")) == winner_norm), key=lambda x: float(x.get("confidence", 0.0)))
                    if float(result.get("confidence", 0.0)) < 0.86:
                        continue
                    item = item_map.get(cand.get("item_key"))
                    if not isinstance(item, dict):
                        continue
                    old_value = clean_text(item.get("value"))
                    old_blank = bool(item.get("blank"))
                    cons = strong_consensus_for(cand)
                    new_value = clean_text(result.get("value"))
                    new_blank = bool(result.get("blank"))

                    # A strong first+SEC agreement counts as two independent votes. To
                    # overturn it, the single-row tiebreak itself must explicitly agree
                    # with the visual winner at high confidence. This prevents two bad
                    # crops from defeating two correct full-page reads.
                    if cons and norm_compare(cons.get("value")) != norm_compare(new_value):
                        if not t or norm_compare(t.get("value")) != winner_norm or float(t.get("confidence", 0.0)) < 0.90:
                            continue

                    if new_value == "N/A" and old_value != "N/A":
                        continue
                    if norm_compare(old_value) != norm_compare(new_value) or old_blank != new_blank:
                        conflicts.append({
                            "source_pdf": source_pdf,
                            "questionnaire_index_in_pdf": qidx,
                            "page_no_in_questionnaire": page_no,
                            "field": f"scale_table_tiebreak:{cand.get('id')}:{cand.get('label')}",
                            "first_value": old_value,
                            "sec_value": json.dumps(result, ensure_ascii=False),
                            "chosen_value": new_value,
                            "reason": "2-of-3 single-row visual majority",
                        })
                    item["type"] = "circle_scale"
                    item["options"] = columns[:]
                    item["value"] = new_value
                    item["blank"] = new_blank
                    item["confidence"] = float(result.get("confidence", 0.0))
                    item["reason"] = clean_text(result.get("reason"), allow_blank=True)
                except Exception:
                    continue

        debug_batches.append({
            "start": start,
            "end": end,
            "primary": obj_a,
            "confirmation": obj_b,
            "columns": columns,
            "confirmation_used": need_confirmation,
            "unresolved_count": len(unresolved),
            "row_anchors_0_1000": row_anchors,
        })

    merged["items"] = list(item_map.values())
    merged["quality_flags"] = sorted(set(merged.get("quality_flags", []) + ["scale_table_geometry_consensus_verification_used"]))
    return cleanup_compact_page(merged), conflicts, {
        "orientation": orientation_debug,
        "geometry": geometry,
        "verification_level": verification_level,
        "batches": debug_batches,
    }

def critical_identity_verification_prompt(compact: Dict[str, Any]) -> str:
    c = cleanup_compact_page(compact)
    current = {
        "participant_id": c.get("meta", {}).get("participant_id", "N/A"),
        "identity": c.get("identity", {}),
    }
    return """
You are an independent high-value IDENTITY / HANDWRITING verifier for a scanned questionnaire page.
Read the actual handwriting and physical marks. Do not trust the existing values below; they are only field-location hints.

Verify only fields that are visibly present on this page:
participant_id, name, student_name, parent_guardian_name, school, gender, date_of_birth, age, grade, class_no, date, section.

STRICT RULES:
- Return valid JSON only.
- Transcribe names and handwriting exactly; do not translate personal names.
- Gender/sex must come only from the physical M/F, Male/Female, or 男/女 mark. Never infer from a name.
- A selected option such as 女 or M must be one concrete value, never "男 / 女" or "M / F".
- If a field is not visible on this page, omit it from the result instead of guessing.
- confidence must be 0.0-1.0.

Return:
{"fields":[{"field":"parent_guardian_name","value":"exact visible value","confidence":0.0,"reason":"short visual reason"}]}

CURRENT FIELD HINTS:
""".strip() + "\n" + json.dumps(current, ensure_ascii=False, separators=(",", ":"))


def apply_critical_identity_verification(
    compact: Dict[str, Any],
    client: LMStudioClient,
    images: List[Image.Image],
    source_pdf: str,
    qidx: int,
    page_no: int,
    max_tokens: int,
) -> Tuple[Dict[str, Any], List[Dict[str, Any]], Dict[str, Any]]:
    c = cleanup_compact_page(compact)
    ident = c.get("identity", {}) if isinstance(c.get("identity"), dict) else {}
    has_context = c.get("meta", {}).get("participant_id") != "N/A" or any(
        clean_text(ident.get(k)) != "N/A"
        for k in ["name", "student_name", "parent_guardian_name", "school", "gender", "date_of_birth", "age", "grade", "class_no", "date", "section"]
    )
    if not has_context:
        return c, [], {"primary": {"fields": []}, "confirmation": {"fields": []}}

    prompt = critical_identity_verification_prompt(c)
    obj, _ = client.vision_json(
        prompt,
        images,
        max_tokens=max(1024, min(int(max_tokens), 3072)),
        retries=1,
    )
    rows = obj.get("fields", []) if isinstance(obj.get("fields"), list) else []
    allowed = {"participant_id", "name", "student_name", "parent_guardian_name", "school", "gender", "date_of_birth", "age", "grade", "class_no", "date", "section"}

    primary: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        field = clean_text(row.get("field"), allow_blank=True).strip()
        value = clean_text(row.get("value"))
        if field in allowed and value != "N/A":
            primary[field] = row

    # Important identity changes are confirmed independently. This is cheap because it
    # runs only on identity pages, and it prevents a single targeted pass from replacing
    # a correct name while still allowing two targeted reads to fix persistent same-error
    # cases such as Joy/Jay or an unresolved M/F / 男/女 mark.
    need_confirm = False
    for field, row in primary.items():
        value = clean_text(row.get("value"))
        if field == "participant_id":
            old = normalize_pid(c.get("meta", {}).get("participant_id"))
            newv = normalize_pid(value)
        else:
            old = clean_text(c.setdefault("identity", {}).get(field))
            newv = value
        if newv != "N/A" and norm_compare(old) != norm_compare(newv) and field in {
            "participant_id", "name", "student_name", "parent_guardian_name", "gender", "date_of_birth"
        }:
            need_confirm = True
            break

    confirm_obj: Dict[str, Any] = {"fields": []}
    confirm: Dict[str, Dict[str, Any]] = {}
    if need_confirm:
        try:
            confirm_obj, _ = client.vision_json(
                prompt + "\nINDEPENDENT CONFIRMATION: re-read the physical handwriting/mark from scratch; do not copy a previous answer.",
                list(reversed(images)),
                max_tokens=max(1024, min(int(max_tokens), 3072)),
                retries=1,
            )
            for row in confirm_obj.get("fields", []) if isinstance(confirm_obj.get("fields"), list) else []:
                if not isinstance(row, dict):
                    continue
                field = clean_text(row.get("field"), allow_blank=True).strip()
                value = clean_text(row.get("value"))
                if field in allowed and value != "N/A":
                    confirm[field] = row
        except Exception:
            confirm_obj = {"fields": []}

    conflicts: List[Dict[str, Any]] = []
    for field, row in primary.items():
        value = clean_text(row.get("value"))
        conf = _as_confidence(row.get("confidence"))
        threshold = 0.84 if field == "gender" else (0.90 if field in {"participant_id", "age", "grade", "class_no", "date", "date_of_birth"} else 0.94)
        if conf < threshold:
            continue

        if field == "participant_id":
            value = normalize_pid(value)
            if value == "N/A":
                continue
            old = normalize_pid(c.get("meta", {}).get("participant_id"))
            changing = norm_compare(old) != norm_compare(value)
            if changing and field in confirm:
                cv = normalize_pid(confirm[field].get("value"))
                if norm_compare(cv) != norm_compare(value) or _as_confidence(confirm[field].get("confidence")) < threshold:
                    continue
            elif changing and need_confirm:
                continue
            if changing:
                conflicts.append({"source_pdf": source_pdf, "questionnaire_index_in_pdf": qidx, "page_no_in_questionnaire": page_no, "field": "critical_identity:participant_id", "first_value": old, "sec_value": json.dumps(row, ensure_ascii=False), "chosen_value": value, "reason": "two-view targeted identity handwriting verifier"})
            c.setdefault("meta", {})["participant_id"] = value
            continue

        old = clean_text(c.setdefault("identity", {}).get(field))
        if field == "gender" and any(sep in value for sep in ["/", "|"]) and len(value) <= 12:
            continue
        changing = norm_compare(old) != norm_compare(value)
        if changing and field in {"name", "student_name", "parent_guardian_name", "gender", "date_of_birth"}:
            crow = confirm.get(field)
            if not crow:
                continue
            cval = clean_text(crow.get("value"))
            if field == "gender" and any(sep in cval for sep in ["/", "|"]) and len(cval) <= 12:
                continue
            if norm_compare(cval) != norm_compare(value) or _as_confidence(crow.get("confidence")) < threshold:
                continue
        if changing:
            conflicts.append({"source_pdf": source_pdf, "questionnaire_index_in_pdf": qidx, "page_no_in_questionnaire": page_no, "field": f"critical_identity:{field}", "first_value": old, "sec_value": json.dumps(row, ensure_ascii=False), "chosen_value": value, "reason": "two-view targeted identity handwriting/mark verifier" if field in confirm else "targeted identity verifier"})
        c["identity"][field] = value

    c["quality_flags"] = sorted(set(c.get("quality_flags", []) + ["critical_identity_verification_used"]))
    return cleanup_compact_page(c), conflicts, {"primary": obj, "confirmation": confirm_obj}


def missing_subfield_recovery_prompt(compact: Dict[str, Any], target_language: str = "") -> str:
    c = cleanup_compact_page(compact)
    existing = [{"id": x.get("id"), "section": x.get("section"), "label": x.get("label"), "value": x.get("value"), "text": x.get("text")} for x in c.get("items", [])]
    lang_note = output_language_prompt(target_language)
    return lang_note + "\n" + """
You are a supplemental verifier for OMITTED FILLED SUBFIELDS on a scanned questionnaire page.
The main extraction already found the items listed below. Find ONLY visibly filled information that it missed.

HIGH-PRIORITY omissions:
- a selected region/address option;
- handwriting beside a selected Yes/No option, such as medication name or relationship;
- a nested subquestion, such as pupil-dilation drops Yes/No;
- medicine type, medicine name, and medicine source as separate filled subitems;
- several independently filled activity/time fields that were wrongly merged into one string;
- an "Other:" write-in that contains handwriting.

STRICT RULES:
- Do not repeat an existing item.
- Do not add blank fields.
- Add only concrete visible information with confidence >= 0.80.
- Preserve personal names, medication names, dates, times, numbers, and free-text handwriting exactly.
- Use stable ids derived from the printed question, e.g. Q2_detail, Q8_dilation, Q10_source, Q11_tv_hours.
- Return valid JSON only; never output UNCLEAR.

Return:
{"missing_items":[{"id":"Q2_detail","section":"...","label":"Medication name","type":"free_text","value":"Folic Acid","options":[],"text":"Folic Acid","blank":false,"confidence":0.95,"reason":"visible handwriting beside selected Yes"}]}

EXISTING ITEMS:
""".strip() + "\n" + json.dumps(existing, ensure_ascii=False, separators=(",", ":"))


def apply_missing_subfield_recovery(
    compact: Dict[str, Any],
    client: LMStudioClient,
    images: List[Image.Image],
    source_pdf: str,
    qidx: int,
    page_no: int,
    max_tokens: int,
    target_language: str = "",
) -> Tuple[Dict[str, Any], List[Dict[str, Any]], Dict[str, Any]]:
    c = cleanup_compact_page(compact)
    if _looks_like_scale_table_page(c):
        return c, [], {"missing_items": []}
    items = c.get("items", [])
    if not items or len(items) > 40:
        return c, [], {"missing_items": []}
    obj, _ = client.vision_json(
        missing_subfield_recovery_prompt(c, target_language=target_language),
        images,
        max_tokens=max(1536, min(int(max_tokens), 4096)),
        retries=1,
    )
    rows = obj.get("missing_items", []) if isinstance(obj.get("missing_items"), list) else []
    existing_keys = {compact_item_key(x) for x in items if isinstance(x, dict)}
    existing_label_values = {(norm_compare(x.get("label")), norm_compare(x.get("value")), norm_compare(x.get("text"))) for x in items if isinstance(x, dict)}
    used_ids = {clean_text(x.get("id")) for x in items if isinstance(x, dict)}
    added: List[Dict[str, Any]] = []
    conflicts: List[Dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        value = clean_text(row.get("value"))
        text = clean_text(row.get("text"))
        conf = _as_confidence(row.get("confidence"))
        if conf < 0.80 or (value == "N/A" and text == "N/A"):
            continue
        item = cleanup_compact_page({"meta": {}, "identity": {}, "items": [row]}).get("items", [])
        if not item:
            continue
        item = item[0]
        sig = (norm_compare(item.get("label")), norm_compare(item.get("value")), norm_compare(item.get("text")))
        if sig in existing_label_values or compact_item_key(item) in existing_keys:
            continue
        base_id = clean_text(item.get("id"))
        if base_id in used_ids:
            n = 2
            while f"{base_id}_detail{n}" in used_ids:
                n += 1
            item["id"] = f"{base_id}_detail{n}"
        used_ids.add(clean_text(item.get("id")))
        existing_keys.add(compact_item_key(item))
        existing_label_values.add(sig)
        added.append(item)
        conflicts.append({"source_pdf": source_pdf, "questionnaire_index_in_pdf": qidx, "page_no_in_questionnaire": page_no, "field": f"missing_subfield:{item.get('id')}:{item.get('label')}", "first_value": "N/A", "sec_value": json.dumps(row, ensure_ascii=False), "chosen_value": clean_text(item.get("value")), "reason": "supplemental omitted-filled-subfield recovery"})
    c["items"] = items + added
    if added:
        c["quality_flags"] = sorted(set(c.get("quality_flags", []) + ["missing_subfield_recovery_used"]))
    return cleanup_compact_page(c), conflicts, obj


def orientation_prompt() -> str:
    return """
The image may be rotated. Decide which rotation makes the printed text upright and easiest to read.
Return ONLY JSON: {"rotation_degrees": 0/90/180/270, "confidence": 0.0}
""".strip()


# -----------------------------------------------------------------------------
# Compact schema normalization / reconciliation
# -----------------------------------------------------------------------------
def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    s = str(value or "").strip().lower()
    return s in {"1", "true", "yes", "y", "blank", "n/a"}


def _as_confidence(value: Any) -> float:
    try:
        x = float(value)
    except Exception:
        return 0.5
    return max(0.0, min(1.0, x))


def _legacy_to_compact(data: Dict[str, Any]) -> Dict[str, Any]:
    """Accept old-schema model output as a compatibility fallback."""
    meta = {
        "document_title": clean_text(data.get("document_title")),
        "page_language": clean_text(data.get("page_language")),
        "page_type": clean_text(data.get("page_type")),
        "participant_id": normalize_pid(data.get("participant_id")),
    }
    identity = data.get("identity_fields", {}) if isinstance(data.get("identity_fields"), dict) else {}
    items: List[Dict[str, Any]] = []
    for i, fld in enumerate(data.get("fields", []) if isinstance(data.get("fields"), list) else [], start=1):
        if not isinstance(fld, dict):
            continue
        items.append({
            "id": clean_text(fld.get("field_id") or f"F{i}"),
            "section": "N/A",
            "label": clean_text(fld.get("label")),
            "type": clean_text(fld.get("field_type")),
            "value": clean_text(fld.get("value")),
            "options": [],
            "text": "N/A",
            "blank": clean_text(fld.get("value")) == "N/A",
            "confidence": _as_confidence(fld.get("confidence")),
            "reason": clean_text(fld.get("low_confidence_reason"), allow_blank=True),
        })
    for i, ans in enumerate(data.get("answers", []) if isinstance(data.get("answers"), list) else [], start=1):
        if not isinstance(ans, dict):
            continue
        items.append({
            "id": clean_text(ans.get("question_id") or f"Q{i}"),
            "section": "N/A",
            "label": clean_text(ans.get("question_text")),
            "type": clean_text(ans.get("answer_type")),
            "value": clean_text(ans.get("selected_value")),
            "options": ans.get("selected_options", []) if isinstance(ans.get("selected_options"), list) else [],
            "text": clean_text(ans.get("written_text")),
            "blank": _as_bool(ans.get("is_blank")) or clean_text(ans.get("selected_value")) == "N/A",
            "confidence": _as_confidence(ans.get("confidence")),
            "reason": clean_text(ans.get("low_confidence_reason"), allow_blank=True),
        })
    for ti, tbl in enumerate(data.get("tables", []) if isinstance(data.get("tables"), list) else [], start=1):
        if not isinstance(tbl, dict):
            continue
        section = clean_text(tbl.get("table_title") or tbl.get("table_id") or f"T{ti}")
        for ri, row in enumerate(tbl.get("rows", []) if isinstance(tbl.get("rows"), list) else [], start=1):
            if not isinstance(row, dict):
                continue
            items.append({
                "id": clean_text(row.get("row_id") or f"R{ri}"),
                "section": section,
                "label": clean_text(row.get("row_label")),
                "type": "table_scale",
                "value": clean_text(row.get("selected_value")),
                "options": [],
                "text": clean_text(row.get("cells")),
                "blank": clean_text(row.get("selected_value")) == "N/A",
                "confidence": _as_confidence(row.get("confidence")),
                "reason": clean_text(row.get("low_confidence_reason"), allow_blank=True),
            })
    return {
        "meta": meta,
        "identity": identity,
        "items": items,
        "handwriting": data.get("visible_handwriting", []) if isinstance(data.get("visible_handwriting"), list) else [],
        "quality_flags": data.get("quality_flags", []) if isinstance(data.get("quality_flags"), list) else [],
    }


def _sanitize_page_language(value: Any) -> str:
    s = clean_text(value)
    low = s.lower()
    # Models sometimes copy the whole enum placeholder literally from the schema.
    if low.count("/") >= 4 and all(tok in low for tok in ["english", "chinese", "filipino", "bilingual"]):
        return "N/A"
    aliases = {
        "traditional chinese": "Traditional Chinese",
        "simplified chinese": "Simplified Chinese",
        "chinese": "Chinese",
        "english": "English",
        "filipino": "Filipino/Tagalog",
        "tagalog": "Filipino/Tagalog",
        "filipino/tagalog": "Filipino/Tagalog",
        "bilingual": "Bilingual",
        "mixed": "Mixed",
    }
    return aliases.get(low, s)


def _sanitize_page_type(value: Any) -> str:
    s = clean_text(value)
    low = s.lower()
    parts = [x.strip() for x in low.split("/") if x.strip()]
    allowed = {"consent", "basic_info", "health_history", "scale_table", "sleep_questionnaire", "signature_page", "other", "extraction_error"}
    # Reject the full schema enum copied verbatim, but preserve short compound types.
    if len(parts) >= 5 and set(parts).issubset(allowed):
        return "other"
    if parts and set(parts).issubset(allowed):
        return "/".join(parts)
    return low if low != "n/a" else "other"


def cleanup_compact_page(data: Dict[str, Any]) -> Dict[str, Any]:
    d = dict(data) if isinstance(data, dict) else {}
    if "meta" not in d or "items" not in d:
        d = _legacy_to_compact(d)

    meta = d.get("meta", {}) if isinstance(d.get("meta"), dict) else {}
    meta_out = {
        "document_title": clean_text(meta.get("document_title")),
        "page_language": _sanitize_page_language(meta.get("page_language")),
        "page_type": _sanitize_page_type(meta.get("page_type")),
        "participant_id": normalize_pid(meta.get("participant_id")),
        "output_language": clean_text(meta.get("output_language")),
    }
    identity_in = d.get("identity", {}) if isinstance(d.get("identity"), dict) else {}
    identity = {str(k): clean_text(v) for k, v in identity_in.items()}

    quality_flags = [clean_text(x) for x in d.get("quality_flags", []) if clean_text(x) != "N/A"] if isinstance(d.get("quality_flags"), list) else []
    handwriting = [clean_text(x) for x in d.get("handwriting", []) if clean_text(x) != "N/A"] if isinstance(d.get("handwriting"), list) else []
    items_out: List[Dict[str, Any]] = []
    for i, item in enumerate(d.get("items", []) if isinstance(d.get("items"), list) else [], start=1):
        if not isinstance(item, dict):
            continue
        raw_value = item.get("value")
        if isinstance(raw_value, list):
            value_parts = [clean_text(x) for x in raw_value]
            value_parts = [x for x in value_parts if x != "N/A"]
            value = "; ".join(dict.fromkeys(value_parts)) if value_parts else "N/A"
        else:
            value = clean_text(raw_value)
        options = [clean_text(x) for x in item.get("options", []) if clean_text(x) != "N/A"] if isinstance(item.get("options"), list) else []
        text = clean_text(item.get("text"))
        blank = _as_bool(item.get("blank"))
        if value == "N/A" and text == "N/A":
            blank = True
        elif value != "N/A":
            # Never keep the internally inconsistent state value=<answer>, blank=true.
            blank = False
        reason = clean_text(item.get("reason"), allow_blank=True)
        if raw_value is not None and UNCLEAR_RE.search(str(raw_value)):
            quality_flags.append(f"item_{i}_unclear_word_replaced_with_NA")
            if not reason:
                reason = "model used a prohibited unclear-like value; recheck recommended"
        item_type = clean_text(item.get("type")).lower()
        if item_type == "n/a":
            item_type = "other"
        items_out.append({
            "id": clean_text(item.get("id") or f"I{i}"),
            "section": clean_text(item.get("section")),
            "label": clean_text(item.get("label")),
            "type": item_type,
            "value": value,
            "options": options,
            "text": text,
            "blank": bool(blank),
            "confidence": _as_confidence(item.get("confidence")),
            "reason": reason,
        })
    return {
        "meta": meta_out,
        "identity": identity,
        "items": items_out,
        "handwriting": handwriting,
        "quality_flags": sorted(set(quality_flags)),
    }

def compact_item_key(item: Dict[str, Any]) -> str:
    section = norm_compare(item.get("section"))
    iid = clean_text(item.get("id"))
    label = norm_compare(item.get("label"))
    if iid != "N/A":
        iid_norm = re.sub(r"[^0-9A-Za-z]+", "", iid).lower() or sanitize_key(iid).lower()
        if section != "N/A":
            return f"{section}|{iid_norm}"
        return f"{iid_norm}|{label[:80]}"
    return f"{section}|{label[:120]}"


def _item_signature(item: Dict[str, Any]) -> str:
    payload = {
        "value": norm_compare(item.get("value")),
        "options": norm_compare(item.get("options", [])),
        "text": norm_compare(item.get("text")),
        "blank": bool(item.get("blank")),
    }
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


def merge_compact_regions(regions: List[Dict[str, Any]]) -> Dict[str, Any]:
    cleaned = [cleanup_compact_page(x) for x in regions if isinstance(x, dict)]
    if not cleaned:
        raise RuntimeError("Region fallback produced no valid JSON regions")
    meta: Dict[str, Any] = {}
    for k in ["document_title", "page_language", "page_type", "participant_id"]:
        vals = [r.get("meta", {}).get(k, "N/A") for r in cleaned]
        meta[k] = stable_choice(vals, f"region_meta|{k}")
    identity_keys = sorted({k for r in cleaned for k in r.get("identity", {}).keys()})
    identity: Dict[str, str] = {}
    for k in identity_keys:
        vals = [r.get("identity", {}).get(k, "N/A") for r in cleaned]
        identity[k] = stable_choice(vals, f"region_identity|{k}")

    merged_items: Dict[str, Dict[str, Any]] = {}
    for region in cleaned:
        for item in region.get("items", []):
            key = compact_item_key(item)
            if key not in merged_items:
                merged_items[key] = dict(item)
                continue
            old = merged_items[key]
            if _item_signature(old) == _item_signature(item):
                if item.get("confidence", 0.0) > old.get("confidence", 0.0):
                    merged_items[key] = dict(item)
                continue
            # Overlap conflict: prefer concrete over N/A, then higher confidence.
            old_concrete = clean_text(old.get("value")) != "N/A" or bool(old.get("options")) or clean_text(old.get("text")) != "N/A"
            new_concrete = clean_text(item.get("value")) != "N/A" or bool(item.get("options")) or clean_text(item.get("text")) != "N/A"
            if new_concrete and not old_concrete:
                chosen = dict(item)
            elif old_concrete and not new_concrete:
                chosen = dict(old)
            elif item.get("confidence", 0.0) > old.get("confidence", 0.0):
                chosen = dict(item)
            else:
                chosen = dict(old)
            chosen["reason"] = "; ".join(x for x in [clean_text(chosen.get("reason"), allow_blank=True), "region overlap conflict"] if x)
            chosen["confidence"] = min(float(chosen.get("confidence", 0.5)), 0.75)
            merged_items[key] = chosen

    flags = {"region_fallback_used"}
    handwriting: List[str] = []
    for r in cleaned:
        flags.update(r.get("quality_flags", []))
        handwriting.extend(r.get("handwriting", []))
    return cleanup_compact_page({
        "meta": meta,
        "identity": identity,
        "items": list(merged_items.values()),
        "handwriting": sorted(set(handwriting)),
        "quality_flags": sorted(flags),
    })


def compact_to_page_json(compact: Dict[str, Any]) -> Dict[str, Any]:
    c = cleanup_compact_page(compact)
    answers: List[Dict[str, Any]] = []
    for i, item in enumerate(c.get("items", []), start=1):
        answers.append({
            "question_id": clean_text(item.get("id") or f"I{i}"),
            "question_text": clean_text(item.get("label")),
            "answer_type": clean_text(item.get("type")),
            "selected_value": clean_text(item.get("value")),
            "selected_options": item.get("options", []) if isinstance(item.get("options"), list) else [],
            "written_text": clean_text(item.get("text")),
            "is_blank": bool(item.get("blank")),
            "confidence": _as_confidence(item.get("confidence")),
            "low_confidence_reason": clean_text(item.get("reason"), allow_blank=True),
            "section": clean_text(item.get("section")),
        })
    return {
        "document_title": c.get("meta", {}).get("document_title", "N/A"),
        "page_language": c.get("meta", {}).get("page_language", "N/A"),
        "output_language": c.get("meta", {}).get("output_language", "N/A"),
        "page_type": c.get("meta", {}).get("page_type", "N/A"),
        "participant_id": c.get("meta", {}).get("participant_id", "N/A"),
        "identity_fields": c.get("identity", {}),
        "fields": [],
        "answers": answers,
        "tables": [],
        "visible_handwriting": c.get("handwriting", []),
        "quality_flags": c.get("quality_flags", []),
        "page_notes": "",
        "items": c.get("items", []),
    }


def _fallback_choose_item(first: Dict[str, Any], sec: Dict[str, Any]) -> Dict[str, Any]:
    a = dict(first)
    b = dict(sec)
    a_concrete = clean_text(a.get("value")) != "N/A" or bool(a.get("options")) or clean_text(a.get("text")) != "N/A"
    b_concrete = clean_text(b.get("value")) != "N/A" or bool(b.get("options")) or clean_text(b.get("text")) != "N/A"
    if a_concrete and not b_concrete:
        return a
    if b_concrete and not a_concrete:
        return b
    if not a_concrete and not b_concrete:
        return a if a.get("confidence", 0.0) >= b.get("confidence", 0.0) else b
    # Both concrete and disagree: avoid the old v9 behavior of blindly replacing
    # the first pass with SEC. Prefer a clearly higher-confidence pass; otherwise
    # keep first until a targeted tiebreak says otherwise.
    if b.get("confidence", 0.0) >= a.get("confidence", 0.0) + 0.20:
        return b
    return a


def reconcile_compact_pages(
    first: Dict[str, Any],
    sec: Optional[Dict[str, Any]],
    source_pdf: str,
    qidx: int,
    page_no: int,
    client: Optional[LMStudioClient] = None,
    images: Optional[List[Image.Image]] = None,
    max_tokens: int = 2048,
) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    first_c = cleanup_compact_page(first)
    if not sec:
        return first_c, []
    sec_c = cleanup_compact_page(sec)
    conflicts: List[Dict[str, Any]] = []

    merged = cleanup_compact_page(first_c)
    # Metadata and PID.
    for k in ["document_title", "page_language", "page_type"]:
        a = first_c["meta"].get(k, "N/A")
        b = sec_c["meta"].get(k, "N/A")
        if a == "N/A" and b != "N/A":
            merged["meta"][k] = b
    a_pid = normalize_pid(first_c["meta"].get("participant_id"))
    b_pid = normalize_pid(sec_c["meta"].get("participant_id"))
    if a_pid == "N/A" and b_pid != "N/A":
        merged["meta"]["participant_id"] = b_pid
    elif a_pid != "N/A" and b_pid != "N/A" and a_pid != b_pid:
        conflicts.append({
            "source_pdf": source_pdf, "questionnaire_index_in_pdf": qidx,
            "page_no_in_questionnaire": page_no, "field": "participant_id",
            "first_value": a_pid, "sec_value": b_pid, "chosen_value": a_pid,
            "reason": "PID differs; normalized first pass retained pending page-level consensus",
        })

    # Identity: keep nonblank first, fill blanks from SEC, record true conflicts.
    identity_keys = sorted(set(first_c.get("identity", {})) | set(sec_c.get("identity", {})))
    for k in identity_keys:
        a = clean_text(first_c.get("identity", {}).get(k))
        b = clean_text(sec_c.get("identity", {}).get(k))
        if a == "N/A" and b != "N/A":
            merged["identity"][k] = b
        elif a != "N/A" and b != "N/A" and norm_compare(a) != norm_compare(b):
            conflicts.append({
                "source_pdf": source_pdf, "questionnaire_index_in_pdf": qidx,
                "page_no_in_questionnaire": page_no, "field": f"identity.{k}",
                "first_value": a, "sec_value": b, "chosen_value": a,
                "reason": "identity differs; first retained unless later page consensus differs",
            })

    first_map = {compact_item_key(x): x for x in first_c.get("items", [])}
    sec_map = {compact_item_key(x): x for x in sec_c.get("items", [])}
    all_keys = list(dict.fromkeys(list(first_map.keys()) + list(sec_map.keys())))
    chosen_map: Dict[str, Dict[str, Any]] = {}
    resolver_inputs: List[Dict[str, Any]] = []
    conflict_key_order: List[str] = []

    for key in all_keys:
        a = first_map.get(key)
        b = sec_map.get(key)
        if a is None:
            chosen_map[key] = dict(b)
            continue
        if b is None:
            chosen_map[key] = dict(a)
            continue
        if _item_signature(a) == _item_signature(b):
            chosen_map[key] = dict(a if a.get("confidence", 0.0) >= b.get("confidence", 0.0) else b)
            continue
        fallback = _fallback_choose_item(a, b)
        chosen_map[key] = dict(fallback)
        resolver_inputs.append({
            "item_id": clean_text(a.get("id") if clean_text(a.get("id")) != "N/A" else b.get("id")),
            "label": clean_text(a.get("label") if clean_text(a.get("label")) != "N/A" else b.get("label")),
            "type": clean_text(a.get("type") if clean_text(a.get("type")) != "N/A" else b.get("type")),
            "first": {"value": a.get("value"), "options": a.get("options"), "text": a.get("text"), "blank": a.get("blank")},
            "sec": {"value": b.get("value"), "options": b.get("options"), "text": b.get("text"), "blank": b.get("blank")},
        })
        conflict_key_order.append(key)

    resolutions: Dict[str, Dict[str, Any]] = {}
    if resolver_inputs and client is not None and images:
        try:
            prompt = conflict_resolution_prompt(resolver_inputs)
            resolved_obj, _ = client.vision_json(prompt, images, max_tokens=max(768, min(int(max_tokens), 3072)), retries=1)
            rows = resolved_obj.get("resolutions", []) if isinstance(resolved_obj.get("resolutions"), list) else []
            for row in rows:
                if isinstance(row, dict):
                    rid = clean_text(row.get("resolution_id"))
                    if rid != "N/A":
                        resolutions[rid] = row
        except Exception:
            resolutions = {}

    for i, (key, info) in enumerate(zip(conflict_key_order, resolver_inputs), start=1):
        a = first_map[key]
        b = sec_map[key]
        chosen = dict(chosen_map[key])
        res = resolutions.get(f"C{i}")
        reason = "first/SEC differ; fallback rule used"
        if isinstance(res, dict):
            rv = clean_text(res.get("value"))
            ropts = [clean_text(x) for x in res.get("options", []) if clean_text(x) != "N/A"] if isinstance(res.get("options"), list) else []
            rtext = clean_text(res.get("text"))
            rblank = _as_bool(res.get("blank"))
            # Do not turn two concrete candidates into N/A unless resolver explicitly
            # says blank and both original passes were also non-concrete.
            originals_concrete = any(
                clean_text(x.get("value")) != "N/A" or bool(x.get("options")) or clean_text(x.get("text")) != "N/A"
                for x in [a, b]
            )
            if not (rv == "N/A" and not ropts and rtext == "N/A" and originals_concrete):
                chosen.update({
                    "value": rv,
                    "options": ropts,
                    "text": rtext,
                    "blank": rblank,
                    "confidence": _as_confidence(res.get("confidence")),
                    "reason": clean_text(res.get("reason"), allow_blank=True),
                })
                reason = "targeted visual tiebreak"
        chosen_map[key] = chosen
        conflicts.append({
            "source_pdf": source_pdf,
            "questionnaire_index_in_pdf": qidx,
            "page_no_in_questionnaire": page_no,
            "field": f"item:{info.get('item_id')}:{info.get('label')}",
            "first_value": clean_text(info.get("first")),
            "sec_value": clean_text(info.get("sec")),
            "chosen_value": clean_text({"value": chosen.get("value"), "options": chosen.get("options"), "text": chosen.get("text")}),
            "reason": reason,
        })

    merged["items"] = list(chosen_map.values())
    merged["handwriting"] = sorted(set(first_c.get("handwriting", []) + sec_c.get("handwriting", [])))
    merged["quality_flags"] = sorted(set(first_c.get("quality_flags", []) + sec_c.get("quality_flags", [])))
    return cleanup_compact_page(merged), conflicts


def cleanup_page_json(data: Dict[str, Any]) -> Dict[str, Any]:
    """Compatibility helper used by older call sites and checkpoint migrations."""
    if isinstance(data, dict) and "meta" in data and "items" in data:
        return compact_to_page_json(data)
    d = dict(data) if isinstance(data, dict) else {}
    d["document_title"] = clean_text(d.get("document_title"))
    d["page_language"] = clean_text(d.get("page_language"))
    d["page_type"] = clean_text(d.get("page_type"))
    d["participant_id"] = normalize_pid(d.get("participant_id"))
    if not isinstance(d.get("identity_fields"), dict):
        d["identity_fields"] = {}
    d["identity_fields"] = {str(k): clean_text(v) for k, v in d["identity_fields"].items()}
    for k in ["fields", "answers", "tables", "visible_handwriting", "quality_flags"]:
        if not isinstance(d.get(k), list):
            d[k] = []
    d["page_notes"] = clean_text(d.get("page_notes"), allow_blank=True)
    return d


# -----------------------------------------------------------------------------
# Workbook writing
# -----------------------------------------------------------------------------
def make_form_row(record: Dict[str, Any]) -> Dict[str, str]:
    row: Dict[str, str] = {
        "source_pdf": clean_text(record.get("source_pdf")),
        "questionnaire_index_in_pdf": clean_text(record.get("questionnaire_index_in_pdf")),
        "source_pages": clean_text(record.get("source_pages")),
        "participant_id": clean_text(record.get("participant_id")),
        "page_count": clean_text(len(record.get("pages", []))),
        "needs_review": clean_text(record.get("needs_review", False)),
        "conflict_count": clean_text(record.get("conflict_count", 0)),
        "page_error_count": clean_text(record.get("page_error_count", 0)),
        "error": clean_text(record.get("error", "N/A")),
    }
    # Merge identity fields across pages, choosing most common non-N/A value.
    ident_candidates: Dict[str, List[str]] = defaultdict(list)
    for p in record.get("pages", []):
        if isinstance(p, dict):
            ident = p.get("identity_fields", {}) if isinstance(p.get("identity_fields"), dict) else {}
            for k, v in ident.items():
                sv = clean_text(v)
                if sv != "N/A":
                    ident_candidates[sanitize_key(k)].append(sv)
    for k, vals in ident_candidates.items():
        row[f"identity_{k}"] = stable_choice(vals, f"{record.get('source_pdf')}|{record.get('questionnaire_index_in_pdf')}|{k}")

    # Flatten per page with dynamic generic columns.
    for pidx, p in enumerate(record.get("pages", []), start=1):
        if not isinstance(p, dict):
            continue
        row[f"p{pidx}_page_type"] = clean_text(p.get("page_type"))
        row[f"p{pidx}_language"] = clean_text(p.get("page_language"))
        row[f"p{pidx}_output_language"] = clean_text(p.get("output_language"))
        row[f"p{pidx}_title"] = clean_text(p.get("document_title"))
        # identity repeated per page only when visible.
        ident = p.get("identity_fields", {}) if isinstance(p.get("identity_fields"), dict) else {}
        for k, v in ident.items():
            sv = clean_text(v)
            if sv != "N/A":
                row[f"p{pidx}_identity_{sanitize_key(k)}"] = sv
        # Generic fields.
        for i, fld in enumerate(p.get("fields", []) if isinstance(p.get("fields"), list) else [], start=1):
            if not isinstance(fld, dict):
                continue
            label = sanitize_key(fld.get("label") or fld.get("field_id") or f"field{i}")
            val = clean_text(fld.get("value"))
            if val != "N/A":
                row[f"p{pidx}_field_{label}"] = val
        # Answers.
        for i, ans in enumerate(p.get("answers", []) if isinstance(p.get("answers"), list) else [], start=1):
            if not isinstance(ans, dict):
                continue
            qid = sanitize_key(ans.get("question_id") or ans.get("question_text") or f"Q{i}")
            section = sanitize_key(ans.get("section"))
            answer_key = qid if section == "unknown" else f"{section}_{qid}"
            val = clean_text(ans.get("selected_value"))
            opts = ans.get("selected_options")
            wt = clean_text(ans.get("written_text"))
            parts = []
            if val != "N/A":
                parts.append(val)
            if isinstance(opts, list) and opts:
                opts_clean = [clean_text(x) for x in opts if clean_text(x) != "N/A"]
                if opts_clean:
                    parts.append("options=" + "; ".join(opts_clean))
            if wt != "N/A":
                parts.append("text=" + wt)
            if parts:
                col = f"p{pidx}_answer_{answer_key}"
                if col in row:
                    n = 2
                    while f"{col}_{n}" in row:
                        n += 1
                    col = f"{col}_{n}"
                row[col] = " | ".join(parts)
        # Tables rows.
        for ti, tbl in enumerate(p.get("tables", []) if isinstance(p.get("tables"), list) else [], start=1):
            if not isinstance(tbl, dict):
                continue
            tid = sanitize_key(tbl.get("table_id") or tbl.get("table_title") or f"T{ti}")
            rows = tbl.get("rows", [])
            if not isinstance(rows, list):
                continue
            for ri, tr in enumerate(rows, start=1):
                if not isinstance(tr, dict):
                    continue
                rid = sanitize_key(tr.get("row_id") or tr.get("row_label") or f"R{ri}")
                val = clean_text(tr.get("selected_value"))
                if val != "N/A":
                    row[f"p{pidx}_table_{tid}_{rid}"] = val
    return row


def build_long_rows(record: Dict[str, Any]) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    base = {
        "source_pdf": clean_text(record.get("source_pdf")),
        "questionnaire_index_in_pdf": clean_text(record.get("questionnaire_index_in_pdf")),
        "participant_id": clean_text(record.get("participant_id")),
    }
    for pidx, p in enumerate(record.get("pages", []), start=1):
        if not isinstance(p, dict):
            continue
        for ans in p.get("answers", []) if isinstance(p.get("answers"), list) else []:
            if not isinstance(ans, dict):
                continue
            r = dict(base)
            r.update({
                "page_no_in_questionnaire": str(pidx),
                "item_type": "answer",
                "item_id": clean_text(ans.get("question_id")),
                "item_label": clean_text(ans.get("question_text")),
                "section": clean_text(ans.get("section")),
                "answer_type": clean_text(ans.get("answer_type")),
                "value": clean_text(ans.get("selected_value")),
                "selected_options": clean_text(ans.get("selected_options")),
                "written_text": clean_text(ans.get("written_text")),
                "confidence": clean_text(ans.get("confidence")),
                "low_confidence_reason": clean_text(ans.get("low_confidence_reason")),
            })
            out.append(r)
        for fld in p.get("fields", []) if isinstance(p.get("fields"), list) else []:
            if not isinstance(fld, dict):
                continue
            r = dict(base)
            r.update({
                "page_no_in_questionnaire": str(pidx),
                "item_type": "field",
                "item_id": clean_text(fld.get("field_id")),
                "item_label": clean_text(fld.get("label")),
                "answer_type": clean_text(fld.get("field_type")),
                "value": clean_text(fld.get("value")),
                "selected_options": "N/A",
                "written_text": "N/A",
                "confidence": clean_text(fld.get("confidence")),
                "low_confidence_reason": clean_text(fld.get("low_confidence_reason")),
            })
            out.append(r)
        for tbl in p.get("tables", []) if isinstance(p.get("tables"), list) else []:
            if not isinstance(tbl, dict):
                continue
            rows = tbl.get("rows", [])
            if not isinstance(rows, list):
                continue
            for tr in rows:
                if not isinstance(tr, dict):
                    continue
                r = dict(base)
                r.update({
                    "page_no_in_questionnaire": str(pidx),
                    "item_type": "table_row",
                    "item_id": clean_text(tr.get("row_id")),
                    "item_label": clean_text(tr.get("row_label")),
                    "answer_type": "table",
                    "value": clean_text(tr.get("selected_value")),
                    "selected_options": "N/A",
                    "written_text": clean_text(tr.get("cells")),
                    "confidence": "N/A",
                    "low_confidence_reason": "N/A",
                })
                out.append(r)
    return out


def write_sheet(wb: Workbook, title: str, rows: List[Dict[str, Any]], freeze: bool = True) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["No rows"])
        return
    headers = []
    seen = set()
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                headers.append(k)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for r in rows:
        ws.append([clean_text(r.get(h)) for h in headers])
    if freeze:
        ws.freeze_panes = "A2"
    # width
    for idx, h in enumerate(headers, start=1):
        width = min(45, max(10, len(str(h)) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width
    # basic highlighting
    if "needs_review" in headers:
        col = headers.index("needs_review") + 1
        fill = PatternFill("solid", fgColor="FFF2CC")
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, col).value).lower() == "true":
                for c in range(1, ws.max_column + 1):
                    ws.cell(row, c).fill = fill



def _numeric_value(value: Any) -> Optional[float]:
    s = clean_text(value)
    if s == "N/A":
        return None
    s = s.replace(",", "").strip()
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", s):
        try:
            return float(s)
        except Exception:
            return None
    return None


def _safe_mean(values: List[float]) -> Optional[float]:
    return statistics.mean(values) if values else None


def _safe_median(values: List[float]) -> Optional[float]:
    return statistics.median(values) if values else None


def _safe_stdev(values: List[float]) -> Optional[float]:
    return statistics.stdev(values) if len(values) >= 2 else None


def _pearson(xs: List[float], ys: List[float]) -> Optional[float]:
    if len(xs) < 2 or len(xs) != len(ys):
        return None
    mx, my = statistics.mean(xs), statistics.mean(ys)
    sx = sum((x - mx) ** 2 for x in xs)
    sy = sum((y - my) ** 2 for y in ys)
    if sx <= 0 or sy <= 0:
        return None
    return sum((x - mx) * (y - my) for x, y in zip(xs, ys)) / math.sqrt(sx * sy)


def _rank_values(values: List[float]) -> List[float]:
    order = sorted(range(len(values)), key=lambda i: values[i])
    ranks = [0.0] * len(values)
    i = 0
    while i < len(order):
        j = i + 1
        while j < len(order) and values[order[j]] == values[order[i]]:
            j += 1
        rank = (i + 1 + j) / 2.0
        for k in range(i, j):
            ranks[order[k]] = rank
        i = j
    return ranks


def _spearman(xs: List[float], ys: List[float]) -> Optional[float]:
    if len(xs) < 2 or len(xs) != len(ys):
        return None
    return _pearson(_rank_values(xs), _rank_values(ys))


def _cohen_kappa(pred: List[str], truth: List[str]) -> Optional[float]:
    if not pred or len(pred) != len(truth):
        return None
    n = len(pred)
    po = sum(1 for a, b in zip(pred, truth) if norm_compare(a) == norm_compare(b)) / n
    cp = Counter(norm_compare(x) for x in pred)
    ct = Counter(norm_compare(x) for x in truth)
    labels = set(cp) | set(ct)
    pe = sum((cp.get(k, 0) / n) * (ct.get(k, 0) / n) for k in labels)
    if pe >= 1.0:
        return 1.0 if po >= 1.0 else None
    return (po - pe) / (1.0 - pe)


def _analysis_key(row: Dict[str, Any]) -> Tuple[str, str, str, str]:
    return (
        norm_compare(row.get("participant_id")),
        norm_compare(row.get("page_no_in_questionnaire")),
        norm_compare(row.get("section")),
        norm_compare(row.get("item_id")),
    )


def _load_reference_long_answers(path: str) -> Tuple[Dict[Tuple[str, str, str, str], Dict[str, Any]], str]:
    if not path:
        return {}, "Reference Excel not provided. MAE/RMSE/accuracy metrics require a ground-truth workbook."
    if not os.path.exists(path):
        return {}, f"Reference Excel not found: {path}"
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        if SHEET_LONG not in wb.sheetnames:
            return {}, f"Reference workbook has no '{SHEET_LONG}' sheet."
        ws = wb[SHEET_LONG]
        rows = ws.iter_rows(values_only=True)
        headers = [clean_text(x) for x in next(rows)]
        idx = {h: i for i, h in enumerate(headers)}
        required = ["participant_id", "page_no_in_questionnaire", "item_id", "section", "value"]
        missing = [x for x in required if x not in idx]
        if missing:
            return {}, "Reference Long_Answers is missing columns: " + ", ".join(missing)
        out: Dict[Tuple[str, str, str, str], Dict[str, Any]] = {}
        for vals in rows:
            row = {h: vals[i] if i < len(vals) else None for h, i in idx.items()}
            key = _analysis_key(row)
            if key[0] and key[3]:
                out[key] = row
        return out, f"Loaded {len(out)} reference answers from {os.path.basename(path)}."
    except Exception as e:
        return {}, f"Reference workbook could not be read: {e}"


def _cronbach_alpha_rows(answer_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    # Section-level complete-case alpha for numeric questionnaire items.
    section_data: Dict[str, Dict[Tuple[str, str, str, str], Dict[Tuple[str, str], float]]] = defaultdict(lambda: defaultdict(dict))
    section_items: Dict[str, set] = defaultdict(set)
    for r in answer_rows:
        v = _numeric_value(r.get("value"))
        if v is None:
            continue
        section = clean_text(r.get("section"))
        if section == "N/A":
            section = f"Page {clean_text(r.get('page_no_in_questionnaire'))}"
        participant = (
            clean_text(r.get("source_pdf")),
            clean_text(r.get("questionnaire_index_in_pdf")),
            clean_text(r.get("participant_id")),
            section,
        )
        item_key = (clean_text(r.get("page_no_in_questionnaire")), clean_text(r.get("item_id")))
        section_data[section][participant][item_key] = v
        section_items[section].add(item_key)

    out: List[Dict[str, Any]] = []
    for section, pmap in section_data.items():
        items = sorted(section_items[section])
        if len(items) < 2:
            continue
        complete = []
        for vals in pmap.values():
            if all(k in vals for k in items):
                complete.append([vals[k] for k in items])
        if len(complete) < 3:
            continue
        item_vars = []
        for j in range(len(items)):
            col = [row[j] for row in complete]
            item_vars.append(statistics.variance(col) if len(set(col)) > 1 else 0.0)
        totals = [sum(row) for row in complete]
        total_var = statistics.variance(totals) if len(set(totals)) > 1 else 0.0
        if total_var <= 0:
            alpha = None
        else:
            k = len(items)
            alpha = (k / (k - 1.0)) * (1.0 - sum(item_vars) / total_var)
        out.append({
            "section": section,
            "items": len(items),
            "complete_participants": len(complete),
            "cronbach_alpha": alpha,
        })
    return sorted(out, key=lambda x: (x["section"]))


def build_analysis_payload(
    records: List[Dict[str, Any]],
    conflicts: List[Dict[str, Any]],
    failed: List[Dict[str, Any]],
    reference_excel: str = "",
    top_n: int = 15,
) -> Dict[str, Any]:
    long_rows: List[Dict[str, Any]] = []
    for rec in records:
        long_rows.extend(build_long_rows(rec))
    answer_rows = [r for r in long_rows if clean_text(r.get("item_type")).lower() == "answer"]

    total_answers = len(answer_rows)
    answered_rows = [r for r in answer_rows if clean_text(r.get("value")) != "N/A"]
    missing_rows = total_answers - len(answered_rows)
    numeric_rows = [(r, _numeric_value(r.get("value"))) for r in answered_rows]
    numeric_rows = [(r, v) for r, v in numeric_rows if v is not None]
    confidences = [_numeric_value(r.get("confidence")) for r in answer_rows]
    confidences = [x for x in confidences if x is not None]

    metrics = [
        ("Questionnaires", len(records)),
        ("Pages processed", sum(len(r.get("pages", [])) for r in records)),
        ("Answer rows", total_answers),
        ("Answered rows", len(answered_rows)),
        ("N/A rows", missing_rows),
        ("Completion rate", (len(answered_rows) / total_answers) if total_answers else None),
        ("N/A rate", (missing_rows / total_answers) if total_answers else None),
        ("Numeric answer rows", len(numeric_rows)),
        ("Mean confidence", _safe_mean(confidences)),
        ("Low-confidence rate (<0.80)", (sum(1 for x in confidences if x < 0.80) / len(confidences)) if confidences else None),
        ("Needs-review questionnaires", sum(1 for r in records if bool(r.get("needs_review")))),
        ("Conflicts", len(conflicts)),
        ("Failed jobs", len(failed)),
    ]

    # Per-item descriptive statistics.
    grouped: Dict[Tuple[str, str, str, str], List[float]] = defaultdict(list)
    grouped_missing: Counter = Counter()
    grouped_total: Counter = Counter()
    for r in answer_rows:
        key = (
            clean_text(r.get("page_no_in_questionnaire")),
            clean_text(r.get("section")),
            clean_text(r.get("item_id")),
            clean_text(r.get("item_label")),
        )
        grouped_total[key] += 1
        v = _numeric_value(r.get("value"))
        if v is None:
            grouped_missing[key] += 1
        else:
            grouped[key].append(v)

    item_stats: List[Dict[str, Any]] = []
    for key in sorted(grouped_total):
        vals = grouped.get(key, [])
        if not vals:
            continue
        page_no, section, item_id, label = key
        item_stats.append({
            "page": page_no,
            "section": section,
            "item_id": item_id,
            "item_label": label,
            "n": len(vals),
            "mean": _safe_mean(vals),
            "median": _safe_median(vals),
            "stdev": _safe_stdev(vals),
            "min": min(vals),
            "max": max(vals),
            "n_a_rate": grouped_missing[key] / grouped_total[key] if grouped_total[key] else None,
        })
    item_stats.sort(key=lambda x: (-x["n"], x["page"], x["item_id"]))

    # Participant/questionnaire mean score for a trend line.
    participant_values: Dict[Tuple[str, str, str], List[float]] = defaultdict(list)
    for r, v in numeric_rows:
        pk = (
            clean_text(r.get("source_pdf")),
            clean_text(r.get("questionnaire_index_in_pdf")),
            clean_text(r.get("participant_id")),
        )
        participant_values[pk].append(v)
    participant_scores: List[Dict[str, Any]] = []
    for pk, vals in participant_values.items():
        source_pdf, qidx, pid = pk
        participant_scores.append({
            "questionnaire": f"{source_pdf} q{qidx} {pid}",
            "source_pdf": source_pdf,
            "questionnaire_index": qidx,
            "participant_id": pid,
            "mean_score": _safe_mean(vals),
            "numeric_items": len(vals),
        })
    participant_scores.sort(key=lambda x: (x["source_pdf"], int(x["questionnaire_index"]) if str(x["questionnaire_index"]).isdigit() else 0))

    numeric_distribution = Counter()
    for _r, v in numeric_rows:
        label = str(int(v)) if float(v).is_integer() else str(round(v, 4))
        numeric_distribution[label] += 1

    language_counter = Counter()
    page_type_counter = Counter()
    for rec in records:
        for page in rec.get("pages", []):
            if isinstance(page, dict):
                language_counter[clean_text(page.get("page_language"))] += 1
                page_type_counter[clean_text(page.get("page_type"))] += 1

    confidence_buckets = Counter()
    for c in confidences:
        if c >= 0.95:
            confidence_buckets[">=0.95"] += 1
        elif c >= 0.80:
            confidence_buckets["0.80-0.949"] += 1
        else:
            confidence_buckets["<0.80"] += 1

    # QA override counters from conflict reasons.
    scale_overrides = sum(1 for c in conflicts if "scale-table" in clean_text(c.get("reason")).lower())
    mark_overrides = sum(1 for c in conflicts if "selection-mark" in clean_text(c.get("reason")).lower())
    sec_conflicts = max(0, len(conflicts) - scale_overrides - mark_overrides)
    qa_metrics = [
        ("SEC/merge conflicts", sec_conflicts),
        ("Selection-mark overrides", mark_overrides),
        ("Scale-table overrides", scale_overrides),
        ("Conflict rate per answer", len(conflicts) / total_answers if total_answers else None),
        ("Failed-job rate", len(failed) / len(records) if records else None),
    ]

    # Optional reference-based validation metrics.
    ref_map, ref_note = _load_reference_long_answers(reference_excel)
    pred_pairs: List[str] = []
    truth_pairs: List[str] = []
    pred_num: List[float] = []
    truth_num: List[float] = []
    scatter_rows: List[Dict[str, float]] = []
    for r in answer_rows:
        ref = ref_map.get(_analysis_key(r))
        if not ref:
            continue
        pv = clean_text(r.get("value"))
        tv = clean_text(ref.get("value"))
        if pv == "N/A" or tv == "N/A":
            continue
        pred_pairs.append(pv)
        truth_pairs.append(tv)
        pn, tn = _numeric_value(pv), _numeric_value(tv)
        if pn is not None and tn is not None:
            pred_num.append(pn)
            truth_num.append(tn)
            scatter_rows.append({"reference": tn, "extracted": pn})

    ref_metrics: List[Tuple[str, Any]] = [("Reference status", ref_note)]
    if pred_pairs:
        exact = sum(1 for a, b in zip(pred_pairs, truth_pairs) if norm_compare(a) == norm_compare(b)) / len(pred_pairs)
        ref_metrics.extend([
            ("Comparable answers", len(pred_pairs)),
            ("Exact accuracy", exact),
            ("Cohen's kappa", _cohen_kappa(pred_pairs, truth_pairs)),
        ])
    if pred_num:
        errors = [p - t for p, t in zip(pred_num, truth_num)]
        abs_errors = [abs(x) for x in errors]
        sq_errors = [x * x for x in errors]
        mean_truth = statistics.mean(truth_num)
        ss_res = sum((p - t) ** 2 for p, t in zip(pred_num, truth_num))
        ss_tot = sum((t - mean_truth) ** 2 for t in truth_num)
        ref_metrics.extend([
            ("Comparable numeric answers", len(pred_num)),
            ("Numeric exact accuracy", sum(1 for x in abs_errors if x <= 1e-12) / len(abs_errors)),
            ("MAE", statistics.mean(abs_errors)),
            ("MSE", statistics.mean(sq_errors)),
            ("RMSE", math.sqrt(statistics.mean(sq_errors))),
            ("Mean error / bias", statistics.mean(errors)),
            ("Median absolute error", statistics.median(abs_errors)),
            ("Max absolute error", max(abs_errors)),
            ("Within +/-1 accuracy", sum(1 for x in abs_errors if x <= 1.0) / len(abs_errors)),
            ("Pearson r", _pearson(pred_num, truth_num)),
            ("Spearman rho", _spearman(pred_num, truth_num)),
            ("R-squared", (1.0 - ss_res / ss_tot) if ss_tot > 0 else None),
        ])

    reliability_rows = _cronbach_alpha_rows(answer_rows)
    return {
        "metrics": metrics,
        "qa_metrics": qa_metrics,
        "reference_metrics": ref_metrics,
        "item_stats": item_stats[:max(1, int(top_n))],
        "participant_scores": participant_scores,
        "numeric_distribution": numeric_distribution,
        "language_distribution": language_counter,
        "page_type_distribution": page_type_counter,
        "confidence_distribution": confidence_buckets,
        "reliability_rows": reliability_rows,
        "scatter_rows": scatter_rows,
    }



ANALYSIS_TEXT = {
    "English": {
        "title": "Questionnaire Data Analysis",
        "subtitle": "Automatically generated from extracted questionnaire data. Reference-based error metrics require an optional ground-truth Excel workbook.",
        "descriptive": "Descriptive Statistics",
        "qa": "Extraction QA / Reliability",
        "reference": "Reference Error Metrics",
        "metric": "Metric",
        "value": "Value",
        "numeric_summary": "Numeric Item Summary",
        "cronbach": "Cronbach's Alpha by Section",
        "bar_title": "Mean Score by Numeric Item",
        "bar_y": "Mean score",
        "bar_x": "Item",
        "pie_title": "Answer Completeness",
        "line_title": "Mean Numeric Score by Questionnaire",
        "line_y": "Mean score",
        "line_x": "Questionnaire",
        "scatter_title": "Extracted Value vs Ground Truth",
        "scatter_x": "Ground truth",
        "scatter_y": "Extracted",
    },
    "Traditional Chinese": {
        "title": "問卷數據分析",
        "subtitle": "由問卷提取結果自動生成。參考答案誤差指標需要可選的人工標準 Excel。",
        "descriptive": "描述性統計",
        "qa": "提取品質 / 信度",
        "reference": "參考答案誤差指標",
        "metric": "指標",
        "value": "數值",
        "numeric_summary": "數值題目摘要",
        "cronbach": "各部分 Cronbach's Alpha",
        "bar_title": "各數值題目平均分",
        "bar_y": "平均分",
        "bar_x": "題目",
        "pie_title": "答案完整度",
        "line_title": "各問卷平均數值分數",
        "line_y": "平均分",
        "line_x": "問卷",
        "scatter_title": "提取值與人工標準值",
        "scatter_x": "人工標準",
        "scatter_y": "提取值",
    },
    "Simplified Chinese": {
        "title": "问卷数据分析",
        "subtitle": "由问卷提取结果自动生成。参考答案误差指标需要可选的人工标准 Excel。",
        "descriptive": "描述性统计",
        "qa": "提取质量 / 信度",
        "reference": "参考答案误差指标",
        "metric": "指标",
        "value": "数值",
        "numeric_summary": "数值题目摘要",
        "cronbach": "各部分 Cronbach's Alpha",
        "bar_title": "各数值题目平均分",
        "bar_y": "平均分",
        "bar_x": "题目",
        "pie_title": "答案完整度",
        "line_title": "各问卷平均数值分数",
        "line_y": "平均分",
        "line_x": "问卷",
        "scatter_title": "提取值与人工标准值",
        "scatter_x": "人工标准",
        "scatter_y": "提取值",
    },
}

METRIC_LABEL_TRANSLATIONS = {
    "Traditional Chinese": {
        "Questionnaires": "問卷數量",
        "Pages processed": "已處理頁數",
        "Answer rows": "答案行數",
        "Answered rows": "已回答行數",
        "N/A rows": "N/A 行數",
        "Completion rate": "完成率",
        "N/A rate": "N/A 比例",
        "Numeric answer rows": "數值答案行數",
        "Mean confidence": "平均信心分數",
        "Low-confidence rate (<0.80)": "低信心比例（<0.80）",
        "Needs-review questionnaires": "需要覆核的問卷",
        "Conflicts": "衝突數",
        "Failed jobs": "失敗工作數",
        "SEC/merge conflicts": "SEC / 合併衝突",
        "Selection-mark overrides": "選擇標記修正次數",
        "Scale-table overrides": "量表圈選修正次數",
        "Conflict rate per answer": "每答案衝突率",
        "Failed-job rate": "失敗工作率",
    },
    "Simplified Chinese": {
        "Questionnaires": "问卷数量",
        "Pages processed": "已处理页数",
        "Answer rows": "答案行数",
        "Answered rows": "已回答行数",
        "N/A rows": "N/A 行数",
        "Completion rate": "完成率",
        "N/A rate": "N/A 比例",
        "Numeric answer rows": "数值答案行数",
        "Mean confidence": "平均置信度",
        "Low-confidence rate (<0.80)": "低置信度比例（<0.80）",
        "Needs-review questionnaires": "需要复核的问卷",
        "Conflicts": "冲突数",
        "Failed jobs": "失败工作数",
        "SEC/merge conflicts": "SEC / 合并冲突",
        "Selection-mark overrides": "选择标记修正次数",
        "Scale-table overrides": "量表圈选修正次数",
        "Conflict rate per answer": "每答案冲突率",
        "Failed-job rate": "失败工作率",
    },
}


def analysis_text_for_cfg(cfg: Any) -> Tuple[str, Dict[str, str]]:
    target = resolve_output_language(
        getattr(cfg, "output_language_mode", ""),
        getattr(cfg, "custom_output_language", ""),
    )
    key = target if target in ANALYSIS_TEXT else "English"
    return key, ANALYSIS_TEXT[key]


def translate_metric_label(metric: str, language_key: str) -> str:
    return METRIC_LABEL_TRANSLATIONS.get(language_key, {}).get(metric, metric)


def _write_metric_table(ws, start_row: int, start_col: int, title: str, rows: List[Tuple[str, Any]], metric_header: str = "Metric", value_header: str = "Value", language_key: str = "English") -> int:
    ws.cell(start_row, start_col, title)
    ws.cell(start_row, start_col).font = Font(bold=True, color="FFFFFF", size=12)
    ws.cell(start_row, start_col).fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
    ws.cell(start_row + 1, start_col, metric_header)
    ws.cell(start_row + 1, start_col + 1, value_header)
    for c in range(start_col, start_col + 2):
        ws.cell(start_row + 1, c).font = Font(bold=True)
        ws.cell(start_row + 1, c).fill = PatternFill("solid", fgColor="D9EAF7")
    r = start_row + 2
    for metric, value in rows:
        ws.cell(r, start_col, translate_metric_label(metric, language_key))
        ws.cell(r, start_col + 1, value if value is not None else "N/A")
        if isinstance(value, float):
            if "rate" in metric.lower() or "accuracy" in metric.lower():
                ws.cell(r, start_col + 1).number_format = "0.00%"
            else:
                ws.cell(r, start_col + 1).number_format = "0.0000"
        r += 1
    return r


def write_data_analysis_sheet(
    wb: Workbook,
    records: List[Dict[str, Any]],
    conflicts: List[Dict[str, Any]],
    failed: List[Dict[str, Any]],
    cfg: Any,
) -> None:
    payload = build_analysis_payload(
        records,
        conflicts,
        failed,
        reference_excel=clean_text(getattr(cfg, "reference_excel", ""), allow_blank=True),
        top_n=max(5, int(getattr(cfg, "analysis_top_n", 15))),
    )

    ws = wb.create_sheet(SHEET_ANALYSIS)
    data_ws = wb.create_sheet(SHEET_ANALYSIS_DATA)
    data_ws.sheet_state = "hidden"
    language_key, atext = analysis_text_for_cfg(cfg)

    # Dashboard title.
    ws.merge_cells("A1:K2")
    ws["A1"] = atext["title"]
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=18)
    ws["A1"].fill = PatternFill("solid", fgColor="17365D")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws["A3"] = atext["subtitle"]
    ws.merge_cells("A3:K3")
    ws["A3"].alignment = Alignment(wrap_text=True)

    if bool(getattr(cfg, "analysis_descriptive", True)):
        _write_metric_table(ws, 5, 1, atext["descriptive"], payload["metrics"], atext["metric"], atext["value"], language_key)
    if bool(getattr(cfg, "analysis_qa", True)):
        _write_metric_table(ws, 5, 4, atext["qa"], payload["qa_metrics"], atext["metric"], atext["value"], language_key)
        _write_metric_table(ws, 5, 7, atext["reference"], payload["reference_metrics"], atext["metric"], atext["value"], language_key)

    # Numeric item summary.
    start = 24
    ws.cell(start, 1, atext["numeric_summary"])
    ws.cell(start, 1).font = Font(bold=True, color="FFFFFF", size=12)
    ws.cell(start, 1).fill = PatternFill("solid", fgColor="1F4E78")
    if language_key == "Traditional Chinese":
        headers = ["頁面", "部分", "題目 ID", "題目名稱", "樣本數", "平均值", "中位數", "標準差", "最小值", "最大值", "N/A 比例"]
    elif language_key == "Simplified Chinese":
        headers = ["页面", "部分", "题目 ID", "题目名称", "样本数", "平均值", "中位数", "标准差", "最小值", "最大值", "N/A 比例"]
    else:
        headers = ["Page", "Section", "Item ID", "Item label", "N", "Mean", "Median", "Std dev", "Min", "Max", "N/A rate"]
    for j, h in enumerate(headers, start=1):
        ws.cell(start + 1, j, h)
        ws.cell(start + 1, j).font = Font(bold=True)
        ws.cell(start + 1, j).fill = PatternFill("solid", fgColor="D9EAF7")
    for i, row in enumerate(payload["item_stats"], start=start + 2):
        vals = [row["page"], row["section"], row["item_id"], row["item_label"], row["n"], row["mean"], row["median"], row["stdev"], row["min"], row["max"], row["n_a_rate"]]
        for j, v in enumerate(vals, start=1):
            ws.cell(i, j, v if v is not None else "N/A")
        for j in [6, 7, 8, 9, 10]:
            ws.cell(i, j).number_format = "0.000"
        ws.cell(i, 11).number_format = "0.00%"

    # Reliability table.
    if bool(getattr(cfg, "analysis_reliability", True)):
        r0 = 24
        c0 = 13
        ws.cell(r0, c0, atext["cronbach"])
        ws.cell(r0, c0).font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(r0, c0).fill = PatternFill("solid", fgColor="1F4E78")
        if language_key == "Traditional Chinese":
            rh = ["部分", "題目數", "完整參與者", "Cronbach alpha"]
        elif language_key == "Simplified Chinese":
            rh = ["部分", "题目数", "完整参与者", "Cronbach alpha"]
        else:
            rh = ["Section", "Items", "Complete participants", "Cronbach alpha"]
        for j, h in enumerate(rh, start=c0):
            ws.cell(r0 + 1, j, h)
            ws.cell(r0 + 1, j).font = Font(bold=True)
            ws.cell(r0 + 1, j).fill = PatternFill("solid", fgColor="D9EAF7")
        for i, row in enumerate(payload["reliability_rows"], start=r0 + 2):
            ws.cell(i, c0, row["section"])
            ws.cell(i, c0 + 1, row["items"])
            ws.cell(i, c0 + 2, row["complete_participants"])
            ws.cell(i, c0 + 3, row["cronbach_alpha"] if row["cronbach_alpha"] is not None else "N/A")
            ws.cell(i, c0 + 3).number_format = "0.000"

    # Helper tables for charts.
    helper_col = 1
    data_ws.cell(1, helper_col, "Item")
    data_ws.cell(1, helper_col + 1, "Mean")
    for i, row in enumerate(payload["item_stats"], start=2):
        data_ws.cell(i, helper_col, f"P{row['page']} {row['item_id']}")
        data_ws.cell(i, helper_col + 1, row["mean"])
    bar_end = max(2, 1 + len(payload["item_stats"]))

    data_ws["D1"] = "Status"
    data_ws["E1"] = "Count"
    total_answers = next((v for k, v in payload["metrics"] if k == "Answer rows"), 0) or 0
    answered = next((v for k, v in payload["metrics"] if k == "Answered rows"), 0) or 0
    data_ws["D2"], data_ws["E2"] = "Answered", answered
    data_ws["D3"], data_ws["E3"] = "N/A", max(0, total_answers - answered)

    data_ws["G1"] = "Questionnaire"
    data_ws["H1"] = "Mean score"
    for i, row in enumerate(payload["participant_scores"], start=2):
        data_ws.cell(i, 7, row["questionnaire"])
        data_ws.cell(i, 8, row["mean_score"])
    line_end = max(2, 1 + len(payload["participant_scores"]))

    data_ws["J1"] = "Language"
    data_ws["K1"] = "Pages"
    for i, (k, v) in enumerate(payload["language_distribution"].most_common(), start=2):
        data_ws.cell(i, 10, k)
        data_ws.cell(i, 11, v)

    data_ws["M1"] = "Page type"
    data_ws["N1"] = "Pages"
    for i, (k, v) in enumerate(payload["page_type_distribution"].most_common(), start=2):
        data_ws.cell(i, 13, k)
        data_ws.cell(i, 14, v)

    data_ws["P1"] = "Confidence"
    data_ws["Q1"] = "Count"
    for i, (k, v) in enumerate(payload["confidence_distribution"].items(), start=2):
        data_ws.cell(i, 16, k)
        data_ws.cell(i, 17, v)

    data_ws["S1"] = "Reference"
    data_ws["T1"] = "Extracted"
    for i, row in enumerate(payload["scatter_rows"], start=2):
        data_ws.cell(i, 19, row["reference"])
        data_ws.cell(i, 20, row["extracted"])
    scatter_end = max(2, 1 + len(payload["scatter_rows"]))

    if bool(getattr(cfg, "analysis_charts", True)):
        chart_anchor_row = 5
        if bool(getattr(cfg, "analysis_bar", True)) and payload["item_stats"]:
            chart = BarChart()
            chart.title = atext["bar_title"]
            chart.y_axis.title = atext["bar_y"]
            chart.x_axis.title = atext["bar_x"]
            chart.style = 10
            chart.height = 8
            chart.width = 15
            data = Reference(data_ws, min_col=2, min_row=1, max_row=bar_end)
            cats = Reference(data_ws, min_col=1, min_row=2, max_row=bar_end)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "T5")

        if bool(getattr(cfg, "analysis_pie", True)) and total_answers:
            chart = PieChart()
            chart.title = atext["pie_title"]
            data = Reference(data_ws, min_col=5, min_row=1, max_row=3)
            labels = Reference(data_ws, min_col=4, min_row=2, max_row=3)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 7.5
            chart.width = 11
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            ws.add_chart(chart, "T21")

        if bool(getattr(cfg, "analysis_line", True)) and len(payload["participant_scores"]) >= 2:
            chart = LineChart()
            chart.title = atext["line_title"]
            chart.y_axis.title = "Mean score"
            chart.x_axis.title = atext["line_x"]
            chart.style = 13
            chart.height = 8
            chart.width = 16
            data = Reference(data_ws, min_col=8, min_row=1, max_row=line_end)
            cats = Reference(data_ws, min_col=7, min_row=2, max_row=line_end)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "A44")

        if bool(getattr(cfg, "analysis_scatter", True)) and len(payload["scatter_rows"]) >= 2:
            chart = ScatterChart()
            chart.title = "Extracted vs Reference Scores"
            chart.x_axis.title = "Reference"
            chart.y_axis.title = atext["scatter_y"]
            chart.style = 13
            chart.height = 8
            chart.width = 14
            xvalues = Reference(data_ws, min_col=19, min_row=2, max_row=scatter_end)
            yvalues = Reference(data_ws, min_col=20, min_row=2, max_row=scatter_end)
            series = Series(yvalues, xvalues, title="Extracted vs reference")
            chart.series.append(series)
            ws.add_chart(chart, "T44")

    # Professional worksheet formatting.
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max(18, ws.max_column)):
        for cell in row:
            if cell.value is not None and cell.row > 3:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in {"A":16,"B":16,"C":14,"D":16,"E":14,"F":14,"G":18,"H":18,"I":14,"J":14,"K":14,"L":18,"M":18,"N":16,"O":16,"P":16}.items():
        ws.column_dimensions[col].width = width
    ws.column_dimensions["D"].width = 36
    ws.freeze_panes = "A4"


def write_workbook(path: str, records: List[Dict[str, Any]], conflicts: List[Dict[str, Any]], failed: List[Dict[str, Any]], log_lines: List[str], cfg: Optional[Any] = None) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    form_rows = [make_form_row(r) for r in records]
    long_rows: List[Dict[str, str]] = []
    page_rows: List[Dict[str, str]] = []
    for r in records:
        long_rows.extend(build_long_rows(r))
        for pidx, p in enumerate(r.get("pages", []), start=1):
            if not isinstance(p, dict):
                continue
            pr = {
                "source_pdf": clean_text(r.get("source_pdf")),
                "questionnaire_index_in_pdf": clean_text(r.get("questionnaire_index_in_pdf")),
                "participant_id": clean_text(r.get("participant_id")),
                "page_no_in_questionnaire": str(pidx),
                "page_type": clean_text(p.get("page_type")),
                "page_language": clean_text(p.get("page_language")),
                "output_language": clean_text(p.get("output_language")),
                "document_title": clean_text(p.get("document_title")),
                "raw_json": json.dumps(p, ensure_ascii=False),
            }
            page_rows.append(pr)

    write_sheet(wb, SHEET_FORMS, form_rows)
    write_sheet(wb, SHEET_LONG, long_rows)
    write_sheet(wb, SHEET_PAGES, page_rows)
    write_sheet(wb, SHEET_CONFLICTS, conflicts)
    write_sheet(wb, SHEET_FAILED, failed)
    write_sheet(wb, SHEET_QA, build_qa_rows(records, conflicts, failed))
    if cfg is not None and bool(getattr(cfg, "analysis_sheet", True)):
        write_data_analysis_sheet(wb, records, conflicts, failed, cfg)
    write_sheet(wb, SHEET_LOG, [{"log": x} for x in log_lines])
    wb.save(path)


def build_qa_rows(records: List[Dict[str, Any]], conflicts: List[Dict[str, Any]], failed: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    total = len(records)
    review = sum(1 for r in records if bool(r.get("needs_review")))
    pid_missing = sum(1 for r in records if clean_text(r.get("participant_id")) == "N/A")
    total_pages = sum(len(r.get("pages", [])) for r in records)
    page_error_count = sum(int(r.get("page_error_count", 0) or 0) for r in records)
    partial_questionnaires = sum(1 for r in records if int(r.get("page_error_count", 0) or 0) > 0)
    total_long_items = 0
    language_counter = Counter()
    page_type_counter = Counter()
    recovery_counter = Counter()
    for r in records:
        for p in r.get("pages", []):
            if isinstance(p, dict):
                language_counter[clean_text(p.get("page_language"))] += 1
                page_type_counter[clean_text(p.get("page_type"))] += 1
                total_long_items += len(p.get("fields", []) if isinstance(p.get("fields"), list) else [])
                total_long_items += len(p.get("answers", []) if isinstance(p.get("answers"), list) else [])
                for tbl in p.get("tables", []) if isinstance(p.get("tables"), list) else []:
                    if isinstance(tbl, dict) and isinstance(tbl.get("rows"), list):
                        total_long_items += len(tbl.get("rows", []))
                for flag in p.get("quality_flags", []) if isinstance(p.get("quality_flags"), list) else []:
                    recovery_counter[clean_text(flag)] += 1
    selection_mark_overrides = sum(
        1 for c in conflicts
        if "selection-mark verifier" in clean_text(c.get("reason"), allow_blank=True).lower()
    )
    scale_table_overrides = sum(
        1 for c in conflicts
        if "scale-table" in clean_text(c.get("reason"), allow_blank=True).lower()
    )
    rows = [
        {"metric": "total_questionnaires", "value": total},
        {"metric": "total_pages", "value": total_pages},
        {"metric": "extracted_items_fields_answers_table_rows", "value": total_long_items},
        {"metric": "needs_review_count", "value": review},
        {"metric": "needs_review_rate", "value": f"{(review / total * 100):.2f}%" if total else "0%"},
        {"metric": "missing_participant_id", "value": pid_missing},
        {"metric": "conflict_rows", "value": len(conflicts)},
        {"metric": "selection_mark_overrides", "value": selection_mark_overrides},
        {"metric": "scale_table_overrides", "value": scale_table_overrides},
        {"metric": "page_extraction_failures", "value": page_error_count},
        {"metric": "partial_questionnaires", "value": partial_questionnaires},
        {"metric": "failed_job_rows", "value": len(failed)},
    ]
    for k, v in recovery_counter.most_common(20):
        rows.append({"metric": "recovery_or_quality_flag", "value": k, "count": v})
    for k, v in language_counter.most_common(10):
        rows.append({"metric": "page_language", "value": k, "count": v})
    for k, v in page_type_counter.most_common(20):
        rows.append({"metric": "page_type", "value": k, "count": v})
    cnt = Counter(clean_text(c.get("field")) for c in conflicts)
    for k, v in cnt.most_common(20):
        rows.append({"metric": "top_conflict_field", "value": k, "count": v})
    return rows


# -----------------------------------------------------------------------------
# Config / extractor
# -----------------------------------------------------------------------------
@dataclass
class RunConfig:
    pdf_paths: List[str]
    output_excel: str
    base_url: str
    model_id: str
    api_key: str
    dpi: int = DEFAULT_DPI
    image_max_side: int = DEFAULT_IMAGE_MAX_SIDE
    max_tokens: int = DEFAULT_MAX_TOKENS
    timeout: int = DEFAULT_TIMEOUT
    temperature: float = 0.0
    first_page_1based: int = 1
    pages_per_questionnaire: int = DEFAULT_PAGES_PER_QUESTIONNAIRE  # 0 = whole PDF or auto group by ID
    auto_group_by_id: bool = False
    enhance: bool = True
    auto_orientation: bool = True
    use_zoom_tiles: bool = True
    sec_pass: bool = True
    selection_mark_verify: bool = True
    critical_identity_verify: bool = True
    missing_subfield_recovery: bool = True
    scale_table_verify: bool = True
    scale_verification_level: str = "Careful"  # Fast / Careful / Maximum
    scale_orientation_ensemble: bool = True
    output_language_mode: str = "Preserve source language"
    custom_output_language: str = ""
    final_language_normalization: bool = True
    resume: bool = True
    save_debug_json: bool = True
    save_debug_images: bool = False
    save_every_n: int = DEFAULT_SAVE_EVERY_N
    analysis_sheet: bool = True
    analysis_descriptive: bool = True
    analysis_qa: bool = True
    analysis_reliability: bool = True
    analysis_charts: bool = True
    analysis_bar: bool = True
    analysis_line: bool = True
    analysis_pie: bool = True
    analysis_scatter: bool = True
    analysis_top_n: int = 15
    reference_excel: str = ""


class UniversalExtractor:
    def __init__(self, cfg: RunConfig, log_func):
        self.cfg = cfg
        self.log = log_func
        self.client = LMStudioClient(cfg.base_url, cfg.model_id, cfg.api_key, cfg.timeout, cfg.temperature)
        self.target_output_language = resolve_output_language(cfg.output_language_mode, cfg.custom_output_language)
        out_dir = os.path.dirname(cfg.output_excel) or "."
        stem = os.path.splitext(os.path.basename(cfg.output_excel))[0]
        self.checkpoint_path = os.path.join(out_dir, stem + "_checkpoint.ndjson")
        self.failed_path = os.path.join(out_dir, stem + "_failed_jobs.jsonl")
        self.debug_dir = os.path.join(out_dir, stem + "_debug")
        if cfg.save_debug_json or cfg.save_debug_images:
            safe_mkdir(self.debug_dir)

    def test_connection(self) -> List[str]:
        return self.client.list_models()

    def save_debug(self, name: str, obj_or_text: Any) -> None:
        if not self.cfg.save_debug_json:
            return
        safe_mkdir(self.debug_dir)
        path = os.path.join(self.debug_dir, name)
        with open(path, "w", encoding="utf-8") as f:
            if isinstance(obj_or_text, (dict, list)):
                json.dump(obj_or_text, f, ensure_ascii=False, indent=2)
            else:
                f.write(str(obj_or_text))

    def maybe_orient(self, img: Image.Image, stem: str) -> Image.Image:
        if not self.cfg.auto_orientation:
            return img
        try:
            data, raw = self.client.vision_json(orientation_prompt(), [img], max_tokens=80, retries=1)
            self.save_debug(f"{stem}_orientation.json", data)
            rot = int(data.get("rotation_degrees", 0))
            if rot in {90, 180, 270}:
                return img.rotate(-rot, expand=True)
        except Exception as e:
            self.log(f"[WARN] orientation failed for {stem}: {e}")
        return img

    def _extract_compact_resilient(
        self,
        img: Image.Image,
        page_no: int,
        total_pages: int,
        stem: str,
        tag: str,
        sec: bool = False,
        existing_json: Optional[Dict[str, Any]] = None,
    ) -> Tuple[Dict[str, Any], str]:
        """Three-layer recovery: normal -> compact retry -> region fallback."""
        # Normal pass: full page only. This avoids visual duplication and is faster.
        prompt = universal_extraction_prompt(
            page_no, total_pages, sec=sec, existing_json=existing_json,
            use_tiles=False, compact_retry=False,
            target_language=self.target_output_language,
        )
        try:
            obj, raw = self.client.vision_json(prompt, [img], max_tokens=self.cfg.max_tokens, retries=1)
            return cleanup_compact_page(obj), raw
        except ModelJSONError as e:
            self.save_debug(f"{stem}_p{page_no:02d}_{tag}_invalid_raw.txt", e.raw or str(e))
            self.log(f"[WARN] {tag} JSON failed on page {page_no}; compact retry: {e}")

        # Compact retry: add a small number of zoom views, ask for the same compact
        # schema, and avoid duplicated printed text.
        retry_images = [img]
        if self.cfg.use_zoom_tiles:
            retry_images.extend(make_zoom_tiles(img, max_tiles=2))
        prompt2 = universal_extraction_prompt(
            page_no, total_pages, sec=sec, existing_json=existing_json,
            use_tiles=len(retry_images) > 1, compact_retry=True,
            target_language=self.target_output_language,
        )
        try:
            obj, raw = self.client.vision_json(
                prompt2, retry_images,
                max_tokens=max(self.cfg.max_tokens, 8192),
                retries=1,
            )
            c = cleanup_compact_page(obj)
            c["quality_flags"] = sorted(set(c.get("quality_flags", []) + ["compact_retry_used"]))
            return c, raw
        except ModelJSONError as e:
            self.save_debug(f"{stem}_p{page_no:02d}_{tag}_compact_invalid_raw.txt", e.raw or str(e))
            self.log(f"[WARN] {tag} compact retry failed on page {page_no}; using region fallback: {e}")

        # Last resort: split the page into broad regions. Each response is much
        # smaller, so dense pages cannot kill the whole questionnaire by truncation.
        region_results: List[Dict[str, Any]] = []
        region_raws: List[str] = []
        regions = make_region_tiles(img)
        for ri, region in enumerate(regions, start=1):
            try:
                rp = region_extraction_prompt(ri, len(regions), page_no, total_pages, target_language=self.target_output_language)
                robj, rraw = self.client.vision_json(
                    rp, [region], max_tokens=min(max(3072, self.cfg.max_tokens // 2), 5120), retries=1
                )
                region_results.append(cleanup_compact_page(robj))
                region_raws.append(rraw)
                self.save_debug(f"{stem}_p{page_no:02d}_{tag}_region{ri}.json", robj)
            except ModelJSONError as e:
                self.save_debug(f"{stem}_p{page_no:02d}_{tag}_region{ri}_invalid_raw.txt", e.raw or str(e))
                self.log(f"[WARN] region {ri}/{len(regions)} failed on page {page_no}: {e}")
        if not region_results:
            raise RuntimeError(f"All JSON recovery strategies failed for page {page_no} ({tag})")
        merged = merge_compact_regions(region_results)
        merged["quality_flags"] = sorted(set(merged.get("quality_flags", []) + [f"{tag}_region_fallback"]))
        return merged, "\n\n--- REGION ---\n\n".join(region_raws)

    def extract_page(
        self,
        img: Image.Image,
        source_pdf: str,
        qidx: int,
        page_no: int,
        total_pages: int,
        stem: str,
        previous_page: Optional[Dict[str, Any]] = None,
    ) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        first_c, raw = self._extract_compact_resilient(
            img, page_no, total_pages, stem, tag="first", sec=False, existing_json=None
        )
        self.save_debug(f"{stem}_p{page_no:02d}_first_compact.json", first_c)
        self.save_debug(f"{stem}_p{page_no:02d}_first_raw.txt", raw)

        conflicts: List[Dict[str, Any]] = []
        merged_c = cleanup_compact_page(first_c)
        scale_consensus: Dict[str, Dict[str, Any]] = {}

        if self.cfg.sec_pass:
            sec_c, raw2 = self._extract_compact_resilient(
                img, page_no, total_pages, stem, tag="sec", sec=True, existing_json=first_c
            )
            self.save_debug(f"{stem}_p{page_no:02d}_sec_compact.json", sec_c)
            self.save_debug(f"{stem}_p{page_no:02d}_sec_raw.txt", raw2)
            scale_consensus = build_item_consensus_map(first_c, sec_c)

            evidence_images = [img]
            if self.cfg.use_zoom_tiles:
                evidence_images.extend(make_zoom_tiles(img, max_tiles=3))
            merged_c, sec_conflicts = reconcile_compact_pages(
                first_c, sec_c, source_pdf, qidx, page_no,
                client=self.client,
                images=evidence_images,
                max_tokens=self.cfg.max_tokens,
            )
            conflicts.extend(sec_conflicts)

        # Use the previous page as structural context before any mark verifier. This is
        # crucial for rotated continuation pages whose first pass calls them "other".
        merged_c = apply_previous_page_context_compact(merged_c, previous_page)

        # High-value identity/handwriting verification fixes persistent same-error cases
        # such as 男/女 left unresolved or Joy read as Jay by both first and SEC.
        if self.cfg.critical_identity_verify:
            identity_images = [img]
            identity_images.extend(make_zoom_tiles(img, max_tiles=4))
            try:
                merged_c, id_conflicts, id_debug = apply_critical_identity_verification(
                    merged_c,
                    self.client,
                    identity_images,
                    source_pdf,
                    qidx,
                    page_no,
                    self.cfg.max_tokens,
                )
                conflicts.extend(id_conflicts)
                self.save_debug(f"{stem}_p{page_no:02d}_critical_identity_verify.json", id_debug)
            except Exception as e:
                self.log(f"[WARN] critical identity verification failed on page {page_no}: {e}")

        # Recover filled subfields that a generic one-item-per-question extraction often
        # drops: medication names, nested dilation yes/no, medicine source, region, etc.
        if self.cfg.missing_subfield_recovery:
            supplemental_images = [img]
            if self.cfg.use_zoom_tiles:
                supplemental_images.extend(make_zoom_tiles(img, max_tiles=4))
            try:
                merged_c, sub_conflicts, sub_debug = apply_missing_subfield_recovery(
                    merged_c,
                    self.client,
                    supplemental_images,
                    source_pdf,
                    qidx,
                    page_no,
                    self.cfg.max_tokens,
                    target_language=self.target_output_language,
                )
                conflicts.extend(sub_conflicts)
                self.save_debug(f"{stem}_p{page_no:02d}_missing_subfields.json", sub_debug)
            except Exception as e:
                self.log(f"[WARN] missing-subfield recovery failed on page {page_no}: {e}")

        # Independent geometry-focused verification for non-table marks. Dense scale
        # tables/sleep matrices are skipped here and handled by the dedicated verifier.
        if self.cfg.selection_mark_verify:
            mark_images = [img]
            mark_images.extend(make_mark_focus_tiles(img, max_tiles=4))
            try:
                merged_c, mark_conflicts, mark_debug = apply_selection_mark_verification(
                    merged_c,
                    self.client,
                    mark_images,
                    source_pdf,
                    qidx,
                    page_no,
                    self.cfg.max_tokens,
                    target_language=self.target_output_language,
                )
                conflicts.extend(mark_conflicts)
                self.save_debug(f"{stem}_p{page_no:02d}_mark_verify.json", mark_debug)
            except Exception as e:
                self.log(f"[WARN] selection-mark verification failed on page {page_no}: {e}")

        # Final authoritative scale pass. v14 localizes the actual table bbox, keeps a
        # canonical column map, and protects first+SEC agreement from one bad verifier.
        if self.cfg.scale_table_verify and _looks_like_scale_table_page(merged_c):
            try:
                merged_c, scale_conflicts, scale_debug = apply_scale_table_verification(
                    merged_c,
                    self.client,
                    [img],
                    source_pdf,
                    qidx,
                    page_no,
                    self.cfg.max_tokens,
                    target_language=self.target_output_language,
                    verification_level=self.cfg.scale_verification_level,
                    use_orientation_ensemble=self.cfg.scale_orientation_ensemble,
                    consensus_map=scale_consensus,
                )
                conflicts.extend(scale_conflicts)
                self.save_debug(f"{stem}_p{page_no:02d}_scale_table_verify.json", scale_debug)
            except Exception as e:
                self.log(f"[WARN] scale-table verification failed on page {page_no}: {e}")

        merged_c = remove_malformed_duplicate_items(merged_c)
        merged_c = repair_scale_sections_compact(merged_c, inherited_title=clean_text(previous_page.get("document_title")) if previous_page else "")

        # Optional final language normalization. Only language-bearing fields can change.
        if self.target_output_language and self.cfg.final_language_normalization:
            try:
                lang_obj, lang_raw = self.client.text_json(
                    language_normalization_prompt(merged_c, self.target_output_language),
                    max_tokens=max(2048, min(self.cfg.max_tokens, 6144)),
                    retries=1,
                )
                merged_c = apply_language_normalization_safely(
                    merged_c, lang_obj, self.target_output_language
                )
                self.save_debug(f"{stem}_p{page_no:02d}_language_normalized_raw.txt", lang_raw)
                self.save_debug(f"{stem}_p{page_no:02d}_language_normalized.json", merged_c)
            except Exception as e:
                self.log(f"[WARN] final language normalization failed on page {page_no}: {e}")
        else:
            merged_c = dedupe_compact_items(merged_c)
            merged_c = remove_malformed_duplicate_items(merged_c)
            merged_c = repair_scale_sections_compact(merged_c, inherited_title=clean_text(previous_page.get("document_title")) if previous_page else "")

        page_json = compact_to_page_json(merged_c)
        self.save_debug(f"{stem}_p{page_no:02d}_merged_compact.json", merged_c)
        self.save_debug(f"{stem}_p{page_no:02d}_merged.json", page_json)
        return page_json, conflicts

    def make_jobs(self) -> List[Dict[str, Any]]:
        jobs: List[Dict[str, Any]] = []
        for pdf_path in self.cfg.pdf_paths:
            pdf_name = os.path.basename(pdf_path)
            try:
                doc = fitz.open(pdf_path)
                page_count = doc.page_count
                doc.close()
            except Exception as e:
                jobs.append({"pdf_path": pdf_path, "pdf_name": pdf_name, "qidx": 1, "start0": 0, "end0": 0, "error": str(e)})
                continue
            start0 = max(0, self.cfg.first_page_1based - 1)
            if self.cfg.pages_per_questionnaire and self.cfg.pages_per_questionnaire > 0:
                ppq = self.cfg.pages_per_questionnaire
                qidx = 0
                for block_start in range(start0, page_count, ppq):
                    block_end = min(page_count, block_start + ppq)
                    if block_end <= block_start:
                        continue
                    qidx += 1
                    jobs.append({"pdf_path": pdf_path, "pdf_name": pdf_name, "qidx": qidx, "start0": block_start, "end0": block_end})
            else:
                # Universal fallback: one whole PDF as one questionnaire.
                jobs.append({"pdf_path": pdf_path, "pdf_name": pdf_name, "qidx": 1, "start0": start0, "end0": page_count})
        return jobs

    def read_checkpoint(self) -> Tuple[List[Dict[str, Any]], set, List[Dict[str, Any]], List[Dict[str, Any]]]:
        records: List[Dict[str, Any]] = []
        done = set()
        conflicts: List[Dict[str, Any]] = []
        failed_all: List[Dict[str, Any]] = []
        if not self.cfg.resume or not os.path.exists(self.checkpoint_path):
            return records, done, conflicts, failed_all
        with open(self.checkpoint_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except Exception:
                    continue
                if obj.get("record_type") == "questionnaire":
                    rec = obj.get("record", {})
                    key = (rec.get("source_pdf"), int(rec.get("questionnaire_index_in_pdf", 0)))
                    # Partial records are intentionally reprocessed on resume.
                    if int(rec.get("page_error_count", 0) or 0) == 0:
                        records.append(rec)
                        done.add(key)
                        conflicts.extend(obj.get("conflicts", []) if isinstance(obj.get("conflicts"), list) else [])
                elif obj.get("record_type") == "failed":
                    failed_all.append(obj.get("failed", obj))
        # Do not keep stale failures for questionnaires that later completed.
        failed = [
            x for x in failed_all
            if (x.get("source_pdf"), int(x.get("questionnaire_index_in_pdf", 0) or 0)) not in done
        ]
        return records, done, conflicts, failed

    def append_checkpoint(self, record_type: str, payload: Dict[str, Any]) -> None:
        safe_mkdir(os.path.dirname(self.checkpoint_path) or ".")
        obj = {"record_type": record_type, "timestamp": now_str()}
        obj.update(payload)
        with open(self.checkpoint_path, "a", encoding="utf-8") as f:
            f.write(json.dumps(obj, ensure_ascii=False) + "\n")

    def process_job(self, job: Dict[str, Any]) -> Tuple[Optional[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
        pdf_path = job["pdf_path"]
        pdf_name = job["pdf_name"]
        qidx = int(job["qidx"])
        start0 = int(job["start0"])
        end0 = int(job["end0"])
        if job.get("error"):
            return None, [], [{"source_pdf": pdf_name, "questionnaire_index_in_pdf": qidx, "error": job.get("error")}]

        pages: List[Dict[str, Any]] = []
        conflicts_all: List[Dict[str, Any]] = []
        page_failures: List[Dict[str, Any]] = []
        doc: Optional[fitz.Document] = None
        try:
            doc = fitz.open(pdf_path)
            total_pages = end0 - start0
            stem = f"{os.path.splitext(pdf_name)[0]}_q{qidx:04d}_pages_{start0+1}-{end0}"
            for local_idx, page0 in enumerate(range(start0, end0), start=1):
                self.log(f"[INFO] Extracting {pdf_name} q{qidx} page {local_idx}/{total_pages}")
                try:
                    img = render_pdf_page(doc, page0, self.cfg.dpi, self.cfg.image_max_side, self.cfg.enhance)
                    img = self.maybe_orient(img, f"{stem}_p{local_idx:02d}")
                    if self.cfg.save_debug_images:
                        safe_mkdir(self.debug_dir)
                        img.save(os.path.join(self.debug_dir, f"{stem}_p{local_idx:02d}.png"))
                    page_json, conflicts = self.extract_page(
                        img, pdf_name, qidx, local_idx, total_pages, stem,
                        previous_page=pages[-1] if pages else None,
                    )
                    pages.append(page_json)
                    conflicts_all.extend(conflicts)
                except Exception as e:
                    err = {
                        "source_pdf": pdf_name,
                        "questionnaire_index_in_pdf": qidx,
                        "source_pages": f"{start0+1}-{end0}",
                        "page_no_in_questionnaire": local_idx,
                        "absolute_pdf_page": page0 + 1,
                        "error": str(e),
                        "traceback": traceback.format_exc(),
                    }
                    page_failures.append(err)
                    self.log(f"[ERROR] Page {local_idx}/{total_pages} failed but questionnaire will continue: {e}")
                    pages.append({
                        "document_title": "N/A",
                        "page_language": "N/A",
                        "page_type": "extraction_error",
                        "participant_id": "N/A",
                        "identity_fields": {},
                        "fields": [],
                        "answers": [],
                        "tables": [],
                        "visible_handwriting": [],
                        "quality_flags": ["page_extraction_failed"],
                        "page_notes": str(e),
                    })

            # Questionnaire-level context repair: propagate scale-table titles/types to
            # continuation pages and remove option-header text from section names.
            pages = postprocess_questionnaire_pages(pages)

            # Choose participant ID from all successful pages. normalize_pid makes
            # CSA83 and CSA083 the same canonical CSA083 before voting.
            pids = [normalize_pid(p.get("participant_id")) for p in pages if normalize_pid(p.get("participant_id")) != "N/A"]
            participant_id = stable_choice(pids, f"{pdf_name}|{qidx}|pid") if pids else "N/A"
            needs_review = bool(conflicts_all) or participant_id == "N/A" or bool(page_failures)
            record = {
                "source_pdf": pdf_name,
                "questionnaire_index_in_pdf": qidx,
                "source_pages": f"{start0+1}-{end0}",
                "participant_id": participant_id,
                "pages": pages,
                "conflict_count": len(conflicts_all),
                "page_error_count": len(page_failures),
                "needs_review": needs_review,
                "error": "N/A" if not page_failures else f"{len(page_failures)} page(s) failed; partial results preserved",
            }
            return record, conflicts_all, page_failures
        except Exception as e:
            err = {
                "source_pdf": pdf_name,
                "questionnaire_index_in_pdf": qidx,
                "source_pages": f"{start0+1}-{end0}",
                "error": str(e),
                "traceback": traceback.format_exc(),
            }
            return None, [], [err]
        finally:
            if doc is not None:
                try:
                    doc.close()
                except Exception:
                    pass

    def run(self, progress_callback=None) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
        records, done, conflicts, failed = self.read_checkpoint()
        jobs = self.make_jobs()
        todo = [j for j in jobs if (j.get("pdf_name"), int(j.get("qidx", 0))) not in done]
        self.log(f"[INFO] Output language: {self.target_output_language or 'Preserve source language'}; "
                 f"scale verification={self.cfg.scale_verification_level}; "
                 f"scale orientation ensemble={self.cfg.scale_orientation_ensemble}")
        self.log(f"[INFO] Total jobs: {len(jobs)}, to process: {len(todo)}, skipped by resume: {len(jobs) - len(todo)}")
        for idx, job in enumerate(todo, start=1):
            if progress_callback:
                progress_callback(idx - 1, max(1, len(todo)))
            pdf_name, qidx = job["pdf_name"], int(job["qidx"])
            self.log(f"[INFO] Rendering {pdf_name} questionnaire {qidx}, pages {int(job['start0'])+1}-{int(job['end0'])}")
            rec, conf, job_failures = self.process_job(job)
            if rec is not None:
                records.append(rec)
                conflicts.extend(conf)
                self.append_checkpoint("questionnaire", {"record": rec, "conflicts": conf})
                self.log(
                    f"[INFO] Saved q{qidx} from {pdf_name} as participant_id={rec.get('participant_id')} "
                    f"conflicts={len(conf)} page_errors={rec.get('page_error_count', 0)} review={rec.get('needs_review')}"
                )
            for err in job_failures:
                failed.append(err)
                self.append_checkpoint("failed", {"failed": err})
                with open(self.failed_path, "a", encoding="utf-8") as f:
                    f.write(json.dumps(err, ensure_ascii=False) + "\n")
                self.log(f"[ERROR] Recorded failure {pdf_name} q{qidx}: {err.get('error')}")
            if self.cfg.save_every_n > 0 and idx % self.cfg.save_every_n == 0:
                self.log("[INFO] Writing intermediate Excel...")
                write_workbook(self.cfg.output_excel, records, conflicts, failed, [], None)
        if progress_callback:
            progress_callback(len(todo), max(1, len(todo)))
        return records, conflicts, failed


# -----------------------------------------------------------------------------
# GUI
# -----------------------------------------------------------------------------
class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1450x980")
        self.pdf_paths: List[str] = []
        self.msg_q: queue.Queue = queue.Queue()
        self.log_lines: List[str] = []

        self.base_url_var = tk.StringVar(value=DEFAULT_BASE_URL)
        self.model_var = tk.StringVar(value=DEFAULT_MODEL_ID)
        self.api_key_var = tk.StringVar(value="")
        self.output_var = tk.StringVar(value=str(Path.cwd() / "universal_questionnaire_output.xlsx"))
        self.dpi_var = tk.StringVar(value=str(DEFAULT_DPI))
        self.max_side_var = tk.StringVar(value=str(DEFAULT_IMAGE_MAX_SIDE))
        self.max_tokens_var = tk.StringVar(value=str(DEFAULT_MAX_TOKENS))
        self.timeout_var = tk.StringVar(value=str(DEFAULT_TIMEOUT))
        self.temperature_var = tk.StringVar(value="0")
        self.first_page_var = tk.StringVar(value="1")
        self.pages_per_q_var = tk.StringVar(value=str(DEFAULT_PAGES_PER_QUESTIONNAIRE))
        self.save_every_var = tk.StringVar(value=str(DEFAULT_SAVE_EVERY_N))
        self.output_language_var = tk.StringVar(value="English")
        self.custom_output_language_var = tk.StringVar(value="")
        self.final_language_normalization_var = tk.BooleanVar(value=True)
        self.scale_verification_level_var = tk.StringVar(value="Careful")
        self.scale_orientation_ensemble_var = tk.BooleanVar(value=True)

        self.enhance_var = tk.BooleanVar(value=True)
        self.orientation_var = tk.BooleanVar(value=True)
        self.zoom_tiles_var = tk.BooleanVar(value=True)
        self.sec_var = tk.BooleanVar(value=True)
        self.mark_verify_var = tk.BooleanVar(value=True)
        self.critical_identity_var = tk.BooleanVar(value=True)
        self.missing_subfield_var = tk.BooleanVar(value=True)
        self.scale_table_var = tk.BooleanVar(value=True)
        self.resume_var = tk.BooleanVar(value=True)
        self.debug_json_var = tk.BooleanVar(value=True)
        self.debug_img_var = tk.BooleanVar(value=False)

        # Excel analysis/dashboard options.
        self.analysis_sheet_var = tk.BooleanVar(value=True)
        self.analysis_descriptive_var = tk.BooleanVar(value=True)
        self.analysis_qa_var = tk.BooleanVar(value=True)
        self.analysis_reliability_var = tk.BooleanVar(value=True)
        self.analysis_charts_var = tk.BooleanVar(value=True)
        self.analysis_bar_var = tk.BooleanVar(value=True)
        self.analysis_line_var = tk.BooleanVar(value=True)
        self.analysis_pie_var = tk.BooleanVar(value=True)
        self.analysis_scatter_var = tk.BooleanVar(value=True)
        self.analysis_top_n_var = tk.StringVar(value="15")
        self.reference_excel_var = tk.StringVar(value="")

        self.progress_var = tk.DoubleVar(value=0.0)
        self.status_var = tk.StringVar(value="Ready")
        self._build_ui()
        self.root.after(100, self.poll_messages)

    def _build_ui(self):
        top = ttk.Frame(self.root)
        top.pack(fill="x", padx=10, pady=8)

        r = 0
        ttk.Label(top, text="LM Studio Base URL").grid(row=r, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.base_url_var, width=38).grid(row=r, column=1, sticky="we", padx=4)
        ttk.Label(top, text="Model ID").grid(row=r, column=2, sticky="w")
        ttk.Entry(top, textvariable=self.model_var, width=32).grid(row=r, column=3, sticky="we", padx=4)
        ttk.Label(top, text="API key/token").grid(row=r, column=4, sticky="w")
        ttk.Entry(top, textvariable=self.api_key_var, show="*", width=20).grid(row=r, column=5, sticky="we", padx=4)
        ttk.Button(top, text="Test Connection", command=self.test_connection).grid(row=r, column=6, padx=4)

        r += 1
        ttk.Label(top, text="Output Excel").grid(row=r, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.output_var, width=90).grid(row=r, column=1, columnspan=5, sticky="we", padx=4, pady=3)
        ttk.Button(top, text="Browse", command=self.pick_output).grid(row=r, column=6, padx=4)

        r += 1
        lang_frame = ttk.Frame(top)
        lang_frame.grid(row=r, column=0, columnspan=7, sticky="we", pady=3)
        ttk.Label(lang_frame, text="Excel content language").pack(side="left", padx=(4, 2))
        lang_combo = ttk.Combobox(
            lang_frame,
            textvariable=self.output_language_var,
            values=OUTPUT_LANGUAGE_OPTIONS,
            state="readonly",
            width=26,
        )
        lang_combo.pack(side="left", padx=(0, 8))
        ttk.Label(lang_frame, text="Custom language").pack(side="left", padx=(4, 2))
        ttk.Entry(lang_frame, textvariable=self.custom_output_language_var, width=24).pack(side="left", padx=(0, 8))
        ttk.Checkbutton(
            lang_frame,
            text="Final AI language normalization",
            variable=self.final_language_normalization_var,
        ).pack(side="left", padx=6)
        ttk.Label(lang_frame, text="Scale verification").pack(side="left", padx=(12, 2))
        ttk.Combobox(
            lang_frame,
            textvariable=self.scale_verification_level_var,
            values=["Fast", "Careful", "Maximum"],
            state="readonly",
            width=11,
        ).pack(side="left", padx=(0, 6))
        ttk.Checkbutton(
            lang_frame,
            text="4-rotation scale orientation ensemble",
            variable=self.scale_orientation_ensemble_var,
        ).pack(side="left", padx=6)

        r += 1
        settings = ttk.Frame(top)
        settings.grid(row=r, column=0, columnspan=7, sticky="we", pady=3)
        for label, var, width in [
            ("DPI", self.dpi_var, 6),
            ("Image max side", self.max_side_var, 8),
            ("Max tokens", self.max_tokens_var, 8),
            ("Timeout", self.timeout_var, 7),
            ("Temp", self.temperature_var, 5),
            ("First page", self.first_page_var, 6),
            ("Pages/questionnaire", self.pages_per_q_var, 6),
            ("Save every N", self.save_every_var, 6),
        ]:
            ttk.Label(settings, text=label).pack(side="left", padx=(4, 1))
            ttk.Entry(settings, textvariable=var, width=width).pack(side="left", padx=(0, 8))

        r += 1
        checks = ttk.Frame(top)
        checks.grid(row=r, column=0, columnspan=7, sticky="we", pady=3)
        ttk.Checkbutton(checks, text="Enhance image", variable=self.enhance_var).pack(side="left", padx=4)
        ttk.Checkbutton(checks, text="Auto orientation", variable=self.orientation_var).pack(side="left", padx=4)
        ttk.Checkbutton(checks, text="AI zoom tiles", variable=self.zoom_tiles_var).pack(side="left", padx=4)
        ttk.Checkbutton(checks, text="SEC / second-pass verification", variable=self.sec_var).pack(side="left", padx=4)
        ttk.Checkbutton(checks, text="Selection-mark verification", variable=self.mark_verify_var).pack(side="left", padx=4)
        ttk.Checkbutton(checks, text="Critical identity/handwriting verification", variable=self.critical_identity_var).pack(side="left", padx=4)
        ttk.Checkbutton(checks, text="Recover omitted filled subfields", variable=self.missing_subfield_var).pack(side="left", padx=4)
        ttk.Checkbutton(checks, text="Scale-table verification", variable=self.scale_table_var).pack(side="left", padx=4)
        ttk.Checkbutton(checks, text="Resume from checkpoint", variable=self.resume_var).pack(side="left", padx=4)
        ttk.Checkbutton(checks, text="Save debug JSON/raw", variable=self.debug_json_var).pack(side="left", padx=4)
        ttk.Checkbutton(checks, text="Save debug images", variable=self.debug_img_var).pack(side="left", padx=4)

        r += 1
        analysis = ttk.LabelFrame(top, text="Excel Data Analysis")
        analysis.grid(row=r, column=0, columnspan=7, sticky="we", pady=5)
        row1 = ttk.Frame(analysis)
        row1.pack(fill="x", padx=4, pady=3)
        ttk.Checkbutton(row1, text="Generate Data_Analysis sheet", variable=self.analysis_sheet_var).pack(side="left", padx=4)
        ttk.Checkbutton(row1, text="Descriptive stats", variable=self.analysis_descriptive_var).pack(side="left", padx=4)
        ttk.Checkbutton(row1, text="QA metrics", variable=self.analysis_qa_var).pack(side="left", padx=4)
        ttk.Checkbutton(row1, text="Cronbach alpha", variable=self.analysis_reliability_var).pack(side="left", padx=4)
        ttk.Checkbutton(row1, text="Charts", variable=self.analysis_charts_var).pack(side="left", padx=4)
        ttk.Checkbutton(row1, text="Bar", variable=self.analysis_bar_var).pack(side="left", padx=4)
        ttk.Checkbutton(row1, text="Line", variable=self.analysis_line_var).pack(side="left", padx=4)
        ttk.Checkbutton(row1, text="Pie", variable=self.analysis_pie_var).pack(side="left", padx=4)
        ttk.Checkbutton(row1, text="Scatter", variable=self.analysis_scatter_var).pack(side="left", padx=4)
        ttk.Label(row1, text="Top N items").pack(side="left", padx=(12, 2))
        ttk.Entry(row1, textvariable=self.analysis_top_n_var, width=5).pack(side="left", padx=2)

        row2 = ttk.Frame(analysis)
        row2.pack(fill="x", padx=4, pady=3)
        ttk.Label(row2, text="Optional ground-truth/reference Excel (for MAE, RMSE, accuracy, correlation):").pack(side="left", padx=4)
        ttk.Entry(row2, textvariable=self.reference_excel_var, width=70).pack(side="left", fill="x", expand=True, padx=4)
        ttk.Button(row2, text="Browse", command=self.pick_reference_excel).pack(side="left", padx=4)

        files = ttk.LabelFrame(self.root, text="PDF files")
        files.pack(fill="both", expand=False, padx=10, pady=8)
        btns = ttk.Frame(files)
        btns.pack(fill="x", pady=4)
        ttk.Button(btns, text="Add PDFs", command=self.add_pdfs).pack(side="left", padx=4)
        ttk.Button(btns, text="Remove selected", command=self.remove_selected).pack(side="left", padx=4)
        ttk.Button(btns, text="Clear", command=self.clear_pdfs).pack(side="left", padx=4)
        ttk.Label(btns, text="Universal multilingual mode. Choose one Excel content language. Careful/Maximum uses table geometry, row anchors, first+SEC consensus protection, and targeted tiebreaks for circled marks.").pack(side="left", padx=12)
        self.files_list = tk.Listbox(files, height=8, selectmode=tk.EXTENDED)
        self.files_list.pack(fill="both", expand=True, padx=6, pady=6)

        run_frame = ttk.Frame(self.root)
        run_frame.pack(fill="x", padx=10, pady=6)
        self.start_btn = ttk.Button(run_frame, text="Start Extraction", command=self.start_worker)
        self.start_btn.pack(side="left", padx=4)
        self.pb = ttk.Progressbar(run_frame, orient="horizontal", mode="determinate", variable=self.progress_var)
        self.pb.pack(side="left", fill="x", expand=True, padx=8)
        ttk.Label(run_frame, textvariable=self.status_var).pack(side="left", padx=4)

        logf = ttk.LabelFrame(self.root, text="Log")
        logf.pack(fill="both", expand=True, padx=10, pady=8)
        self.log_list = tk.Listbox(logf, height=24)
        ys = ttk.Scrollbar(logf, orient="vertical", command=self.log_list.yview)
        self.log_list.configure(yscrollcommand=ys.set)
        self.log_list.pack(side="left", fill="both", expand=True)
        ys.pack(side="right", fill="y")

        top.columnconfigure(1, weight=1)
        top.columnconfigure(3, weight=1)

    def log(self, msg: str):
        line = f"{now_str()} {msg}"
        self.log_lines.append(line)
        self.msg_q.put(("log", line))

    def poll_messages(self):
        while True:
            try:
                kind, payload = self.msg_q.get_nowait()
            except queue.Empty:
                break
            if kind == "log":
                self.log_list.insert(tk.END, payload)
                self.log_list.yview_moveto(1.0)
            elif kind == "status":
                self.status_var.set(payload)
            elif kind == "progress":
                self.progress_var.set(float(payload))
            elif kind == "done":
                self.start_btn.config(state="normal")
                messagebox.showinfo("Done", payload)
            elif kind == "error":
                self.start_btn.config(state="normal")
                messagebox.showerror("Error", payload)
        self.root.after(100, self.poll_messages)

    def add_pdfs(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
        for p in paths:
            if p not in self.pdf_paths:
                self.pdf_paths.append(p)
                self.files_list.insert(tk.END, p)

    def remove_selected(self):
        for i in reversed(self.files_list.curselection()):
            self.files_list.delete(i)
            del self.pdf_paths[i]

    def clear_pdfs(self):
        self.files_list.delete(0, tk.END)
        self.pdf_paths = []

    def pick_output(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if p:
            self.output_var.set(p)

    def pick_reference_excel(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p:
            self.reference_excel_var.set(p)

    def test_connection(self):
        try:
            client = LMStudioClient(self.base_url_var.get(), self.model_var.get(), self.api_key_var.get(), int(self.timeout_var.get()), float(self.temperature_var.get()))
            models = client.list_models()
            messagebox.showinfo("LM Studio", "Models found:\n" + ("\n".join(models) if models else "(none)"))
        except Exception as e:
            messagebox.showerror("LM Studio test failed", str(e))

    def build_config(self) -> RunConfig:
        return RunConfig(
            pdf_paths=list(self.pdf_paths),
            output_excel=self.output_var.get().strip(),
            base_url=self.base_url_var.get().strip(),
            model_id=self.model_var.get().strip(),
            api_key=self.api_key_var.get().strip(),
            dpi=max(100, int(self.dpi_var.get().strip())),
            image_max_side=max(1000, int(self.max_side_var.get().strip())),
            max_tokens=max(512, int(self.max_tokens_var.get().strip())),
            timeout=max(60, int(self.timeout_var.get().strip())),
            temperature=float(self.temperature_var.get().strip()),
            first_page_1based=max(1, int(self.first_page_var.get().strip())),
            pages_per_questionnaire=max(0, int(self.pages_per_q_var.get().strip())),
            enhance=bool(self.enhance_var.get()),
            auto_orientation=bool(self.orientation_var.get()),
            use_zoom_tiles=bool(self.zoom_tiles_var.get()),
            sec_pass=bool(self.sec_var.get()),
            selection_mark_verify=bool(self.mark_verify_var.get()),
            critical_identity_verify=bool(self.critical_identity_var.get()),
            missing_subfield_recovery=bool(self.missing_subfield_var.get()),
            scale_table_verify=bool(self.scale_table_var.get()),
            scale_verification_level=self.scale_verification_level_var.get().strip() or "Careful",
            scale_orientation_ensemble=bool(self.scale_orientation_ensemble_var.get()),
            output_language_mode=self.output_language_var.get().strip() or "Preserve source language",
            custom_output_language=self.custom_output_language_var.get().strip(),
            final_language_normalization=bool(self.final_language_normalization_var.get()),
            resume=bool(self.resume_var.get()),
            save_debug_json=bool(self.debug_json_var.get()),
            save_debug_images=bool(self.debug_img_var.get()),
            save_every_n=max(1, int(self.save_every_var.get().strip())),
            analysis_sheet=bool(self.analysis_sheet_var.get()),
            analysis_descriptive=bool(self.analysis_descriptive_var.get()),
            analysis_qa=bool(self.analysis_qa_var.get()),
            analysis_reliability=bool(self.analysis_reliability_var.get()),
            analysis_charts=bool(self.analysis_charts_var.get()),
            analysis_bar=bool(self.analysis_bar_var.get()),
            analysis_line=bool(self.analysis_line_var.get()),
            analysis_pie=bool(self.analysis_pie_var.get()),
            analysis_scatter=bool(self.analysis_scatter_var.get()),
            analysis_top_n=max(5, int(self.analysis_top_n_var.get().strip())),
            reference_excel=self.reference_excel_var.get().strip(),
        )

    def start_worker(self):
        if not self.pdf_paths:
            messagebox.showerror("Error", "Please add at least one PDF.")
            return
        try:
            cfg = self.build_config()
        except Exception as e:
            messagebox.showerror("Invalid settings", str(e))
            return
        self.start_btn.config(state="disabled")
        self.progress_var.set(0)
        threading.Thread(target=self.worker, args=(cfg,), daemon=True).start()

    def worker(self, cfg: RunConfig):
        try:
            safe_mkdir(os.path.dirname(cfg.output_excel) or ".")
            extractor = UniversalExtractor(cfg, self.log)
            models = extractor.test_connection()
            self.log("[INFO] LM Studio reachable. Models: " + (", ".join(models) if models else "(none)"))

            def progress(done: int, total: int):
                self.msg_q.put(("progress", 100.0 * done / max(1, total)))
                self.msg_q.put(("status", f"{done}/{total}"))

            records, conflicts, failed = extractor.run(progress_callback=progress)
            self.log("[INFO] Writing final Excel...")
            write_workbook(cfg.output_excel, records, conflicts, failed, self.log_lines, cfg)
            self.msg_q.put(("progress", 100.0))
            self.msg_q.put(("done", f"Finished. Excel saved:\n{cfg.output_excel}"))
        except Exception as e:
            tb = traceback.format_exc()
            self.log("[ERROR] " + str(e))
            self.log(tb)
            self.msg_q.put(("error", str(e)))


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
